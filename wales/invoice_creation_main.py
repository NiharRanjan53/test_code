import pandas as pd
import os
from openpyxl.styles import Border, Side
from openpyxl import Workbook, load_workbook
from pathlib import Path
import shutil
import time
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill, numbers
import numpy as np

from invoice_creation_log_ui import log_ui_progress_message, read_ui_progress_messages

from invoice_creation_utils import (
    user_input_file_checks,
    get_dataframe_by_file_and_sheet, 
    extract_q_year_from_timeliness,
    delete_files_in_folder,
    find_first_data_row_in_df,
    get_latest_clp_file,
    create_pn_clp_lookup_column,
    merge_billing_with_clp,
    compute_extended_price,
    apply_cost_category_flags, #
    apply_module_identification,
    apply_discount_from_contract,
    compute_discount_amount,
    process_to_bill_customer_material,
    apply_handling_fees,
    compute_total_price,
    compute_vendor_total_price,
    enrich_with_part_keywords,
    load_cleaned_eipc_dataframe,
    enrich_with_ata,
    enrich_with_ata_long,
    assign_llp_flag,
    apply_llp_exclusion,
    apply_blade_inclusion,
    apply_missing_receipts_exclusion,
    apply_pma_der_exclusion,
    apply_sb_logic,
    apply_workscope_inclusion_exclusion,
    apply_qec_lru_logic,
    append_to_excel,
    write_single_df_to_excel,
 
    # Repair Utils
    read_crd_any_format,
    get_latest_crd_file,
    compute_repair_total_price,
    apply_repair_sb_logic,
 
    # Vendor Utils
    extract_vendor_cost_category,
    filter_billing_for_vendor,

    # Generate invoice
    compare_excel_sheets_contains,
    create_in_scope,
    create_llp_replacement,
    create_qec_lru,
    create_pma_der,
    create_missing_on_receipts,
    # create_taxes,
    create_material_repair,
    create_service_bulletin,

    #Invoice UDFs
    generate_cover_sheet,
    create_sheets_from_range,
    write_multiple_dfs_to_sheet_with_gap,
    write_dfs_with_gaps,
    write_sb_df_with_gaps,
    fill_cover_sheet,
    fill_inscope_sheet,
    delete_empty_rows_and_return_metadata,
    fill_llp_replacement_sheet,
    fill_qec_lru_replacement_sheet,
    fill_pma_der_sheet,
    fill_mor_sheet,
    fill_mat_sheet,
    fill_rep_sheet,
    fill_vendor_sheet,
    fill_sb_oa_summary,
    generate_invoice_summary,
    find_invoice_template_in_current_folder,
    to_number,
    write_cover_sheet_cost_items,
    write_formulas_back,
    extract_inscope_outscope_cost,
    delete_empty_sheets,
    check_rbu_catlog,
)

from extract_used_llp_file import bake_formulas_and_remove_hidden_sheets_final

from invoice_creation_extract_draft_invoice import (
    extract_cover_data,
    transform_excel_with_headers
)

from invoice_creation_download_invoice_logic import (
    compare_files, 
    update_cover_sheet_from_comparison
)
from s3connect import read_excel_from_s3, list_files_in_s3,save_excel_to_s3,s3_file_exists

from workscope_classify import (
    extract_customer_workscope_levels,
    resolve_matching_workscope,
    resolve_matching_workscope_avianca,
    # apply_scrap_exclusion,
)



def replace_if_exists(source_file, target_file):
    if os.path.exists(source_file):
        print(f"🔁 Overwriting {target_file} with {source_file}")
        df = pd.read_excel(source_file)
        df.to_excel(target_file, index=False)
    else:
        print(f"⚠️ {source_file} not found. Skipping overwrite.")
 
# === === === === === === === Material Pipeline === === === === === === === === === ===
def run_material_pipeline(master_folder: str, result: dict, esn):
    print("\n=== === === === === === === Starting Material Pipeline...=== === === === === === ===")
 
    # === Step 1: Load CLP File and Merge with Billing ===
    q_year, induction_date, cust_name, customer_first_word,q_year_g3, g3_date = extract_q_year_from_timeliness(result) 
    # customer_df = extract_customer_workscope_levels(result)
    # workscope = resolve_matching_workscope(customer_df,result)
    workscope = resolve_matching_workscope_avianca(result)
    print(workscope)
    print(q_year_g3)

    latest_clp_path = get_latest_clp_file(master_folder, q_year_g3, induction_date) #Identifying the latest CLP based on the given induction_date (file path)
    
    latest_clp_df = read_excel_from_s3(latest_clp_path)
    # clp_continuous_df = find_first_data_row_in_df(latest_clp_df, "price")
    clp_continuous_df = find_first_data_row_in_df(latest_clp_df, "clp")

    clp_continuous_df = create_pn_clp_lookup_column(clp_continuous_df, "clp") # PN, PK CLP
    print(clp_continuous_df)
    
    # === Step 2: Merge with Billing and Compute Extended Price ===
    pipeline_type = "material"
    billing_enriched = merge_billing_with_clp(result, clp_continuous_df, pipeline_type)  #PN, Unit price (With CLP and without CLP (Cost/qty) conditions are handled)
    billing_enriched.to_excel("data1.xlsx", index=False)

    billing_ext_price = compute_extended_price(billing_enriched, pipeline_type)
    billing_ext_price.to_excel("data2.xlsx", index=False)

    # === Step 3: Apply Cost Category and Discounts ===
    contract_cost_category = apply_cost_category_flags(result, billing_ext_price)
    module_identified_df = apply_module_identification(result, contract_cost_category)
    
    # === Step 5: Enrich Part Keywords and ATA_LONG ===
    print(f"Shape before enrich_with_pk : {module_identified_df.shape}")
    gen_part_keywords = enrich_with_part_keywords(result, module_identified_df, "material")
    gen_part_keywords.to_excel('GEN_PART_KEYWORDS.xlsx', index = False)
    
    
    # === Determining Customer Material and Flaging for Handling Fee ====
    print(f"Shape before process_to_bill_customer_material : {gen_part_keywords.shape}")
    df_handling_fee_flag = process_to_bill_customer_material(gen_part_keywords)
    print(f"Shape after process_to_bill_customer_material : {df_handling_fee_flag.shape}")

    # === Apply Discount ====
    contract_discounts = apply_discount_from_contract(result, df_handling_fee_flag)
    contract_discount_amount = compute_discount_amount(contract_discounts)
    
    # === Scrap Exclusion =======
    print(f"Shape before apply_blade_inclusion : {contract_discount_amount.shape}")
    blade_inclusion = apply_blade_inclusion(result, contract_discount_amount, "material",workscope)
    print(f"Shape after apply_blade_inclusion : {blade_inclusion.shape}") 
 
   


    eipc_path = result["file_paths"]["eipc"]
    print(eipc_path)
    eipc_df = load_cleaned_eipc_dataframe(eipc_path)
    eipc_ata_df = enrich_with_ata(result, eipc_df, gen_part_keywords, "material")
    # eipc_ata_long = enrich_with_ata_long(result, eipc_df, gen_part_keywords, "material") #ata_long column

    # contract_discounts = apply_discount_from_contract(result, eipc_ata_long)
    # contract_discount_amount = compute_discount_amount(contract_discounts)
 

    # # === Step 4: Apply Handling Fees and Compute Total Price ===
    # contract_handling_fee_per_part_set = apply_handling_fees(result, contract_discount_amount)

    # # === Step 6: Apply Exclusion/Inclusion Logic ===
    # input("ENTER DUMMY")
    contract_handling_fee_per_part_set = blade_inclusion.copy()
    contract_handling_fee_per_part_set["remarks"] = "MATL UNKNOWN"

    llp_tagging = assign_llp_flag(result, contract_handling_fee_per_part_set)
    llp_tagging.to_excel("LLP_TAGGING_TEST.xlsx", index = False)
   
    # llp_exclusion = apply_llp_exclusion(llp_tagging, result, esn, customer_first_word)
    # pint()
    # llp_exclusion.to_excel("llp_exclusion.xlsx", index=False)
    # pipeline_type = "material"
    # contract_total_price = compute_total_price(llp_exclusion, cust_name, esn, q_year, result, customer_first_word, pipeline_type)
    # contract_total_price.drop_duplicates(inplace=True)
    

    # blade_inclusion = apply_blade_inclusion(result, contract_total_price, "material",workscope)   

    # biiii
    
    # mor_excl = apply_missing_receipts_exclusion(blade_inclusion)

    # pma_der_codes_excl = apply_pma_der_exclusion(mor_excl)
    
    # pma_der_codes_excl.to_excel("pma_der_till.xlsx", index=False)
    
    
    print(f"Shape before apply_sb_logic : {blade_inclusion.shape}")
    sb_matl, sb_filtered = apply_sb_logic(result, blade_inclusion)
    print(f"Shape after apply_sb_logic : {sb_matl.shape}") 
    
    # pma_der_codes_excl.to_excel("pma_der_codes_excl.xlsx", index=False)
    print(f"Shape before apply_workscope_inclusion_exclusion : {sb_matl.shape}")
    cust_ws_level = apply_workscope_inclusion_exclusion(result, llp_tagging, workscope, "material", "MATL UNKNOWN")
    print(f"Shape after apply_workscope_inclusion_exclusion : {cust_ws_level.shape}")
    
    check_rbu_catlog(result,cust_ws_level)
    
    # xyz

    #cust_ws_level.to_excel("after_workn.xlsx", index=False)

    #df_new=pd.read_excel("after_workn.xlsx")

    #df_new['code_no_dash'] = df_new['ata chapter'].str.replace('-', '', regex=False)

    #mask = (df_new['remarks'] == 'Exclusion as WKS = 2') & (df_new['code_no_dash'].str.endswith('00', na=False))

    #df_new.loc[mask, 'remarks'] = 'Inclusion with Logic'
    
    non_remarked_df = cust_ws_level[cust_ws_level['remarks'] == 'MATL UNKNOWN'].copy()
    # non_remarked_df.to_excel("non_remarked_df.xlsx", index=False)

    remarked_df = cust_ws_level[cust_ws_level['remarks'] != 'MATL UNKNOWN'].copy()
    # remarked_df.to_excel("remarked_df.xlsx", index=False)
    #non_remarked_df = df_new[df_new['remarks'] == 'MATL UNKNOWN'].copy()
    # non_remarked_df.to_excel("non_remarked_df.xlsx", index=False)

   # remarked_df = df_new[df_new['remarks'] != 'MATL UNKNOWN'].copy()
    # remarked_df.to_excel("remarked_df.xlsx", index=False)

    acc_lrus = apply_qec_lru_logic(result, non_remarked_df, "material")
    acc_lrus_df = acc_lrus[acc_lrus['remarks'] != 'MATL UNKNOWN'].copy()
    
    # Exclude remaining parts
    direct_exclude_non_remarked_df = acc_lrus[acc_lrus['remarks'] == 'MATL UNKNOWN'].copy()
    direct_exclude_non_remarked_df['remarks'] = "Exclusion" #'Direct Exclusion'

    # 24/8/2025: ADDED THIS NEW LOGIC
    remarked_df = remarked_df.loc[:, ~remarked_df.columns.duplicated()]
    acc_lrus_df = acc_lrus_df.loc[:, ~acc_lrus_df.columns.duplicated()]
    direct_exclude_non_remarked_df = direct_exclude_non_remarked_df.loc[:, ~direct_exclude_non_remarked_df.columns.duplicated()]
    # 24/8/2025: END

    final_matl_df = pd.concat([remarked_df, acc_lrus_df, direct_exclude_non_remarked_df], ignore_index=True)

    final_matl_df_0 = final_matl_df[final_matl_df['total_price'] != 0].copy()
    final_matl_df_0.drop_duplicates(inplace=True)

    print("=== === === === === === === Material Pipeline Completed Successfully!=== === === === === === ===")
    return final_matl_df, cust_name, sb_filtered

 
# === === === === === === === Repair Pipeline === === === === === === === === === ===
def run_repair_pipeline(master_folder: str, result: dict, sb_filtered: pd.DataFrame):
    print("\n=== === === === === === === Starting Repair Pipeline...=== === === === === === ===")

    # === Step 1: Fetch Latest CRD & Convert to .xlsx ===
    q_year, induction_date, *_ = extract_q_year_from_timeliness(result)
    latest_crd_path = get_latest_crd_file(master_folder, induction_date)

    new_crd_df = read_crd_any_format(latest_crd_path, header_row=0)
 
    # === Step 2: Compute Repair Total Price ===
    repair_df = get_dataframe_by_file_and_sheet(result["dataframes"], "internal", "internal repair")
    
    repair_with_keywords = enrich_with_part_keywords(result, repair_df, "repair")
    repair_with_keywords.drop_duplicates(inplace=True)
    
    repair_with_keywords.to_excel("repair_with_keywords.xlsx", index=False)
    
    
    # pint()

    # === Step 3: Enrich with ATA_LONG and Genpact Part Keywords ===
    eipc_path = result["file_paths"]["eipc"]
    repair_eipc_df = load_cleaned_eipc_dataframe(eipc_path)
    repair_with_ata = enrich_with_ata_long(result, repair_eipc_df, repair_with_keywords, "repair")

    repair_total_price, labor_amount, new_parts_non_llp, new_parts_llp, cfe = compute_repair_total_price(result, repair_with_ata, new_crd_df, induction_date)
    
    # === Step 4.1: Blade Inclusion ===
    repair_total_price["remarks"] = "REPAIR UNKNOWN"
    customer_df = extract_customer_workscope_levels(result)
    workscope = resolve_matching_workscope(customer_df,result)
    # print(f"Matched workscope --> {workscope}")
    
    repair_blade_df = apply_blade_inclusion(result, repair_total_price, "repair",workscope)
    repair_blade_df.to_excel("repair_apply_blade_inclusion.xlsx", index=False)

    sb_filtered = apply_repair_sb_logic(result, labor_amount, induction_date, new_crd_df, sb_filtered)
    
    cust_ws_level = apply_workscope_inclusion_exclusion(result, repair_blade_df, "repair", "REPAIR UNKNOWN")
    
    non_remarked_df = cust_ws_level[cust_ws_level['remarks'] == 'REPAIR UNKNOWN'].copy()
    # non_remarked_df.to_excel("repair_non_remarked_df.xlsx", index=False)

    remarked_df = cust_ws_level[cust_ws_level['remarks'] != 'REPAIR UNKNOWN'].copy()
    remarked_df.to_excel("repair_remarked_df.xlsx", index=False)

    acc_lrus = apply_qec_lru_logic(result, non_remarked_df, "repair")
    acc_lrus_df = acc_lrus[acc_lrus['remarks'] != 'REPAIR UNKNOWN'].copy()
    
    # Exclude remaining parts
    direct_exclude_non_remarked_df = acc_lrus[acc_lrus['remarks'] == 'REPAIR UNKNOWN'].copy()
    direct_exclude_non_remarked_df['remarks'] = "Exclusion" #'Direct Exclusion'

    # 24/8/2025: ADDED THIS NEW LOGIC
    remarked_df = remarked_df.loc[:, ~remarked_df.columns.duplicated()]
    acc_lrus_df = acc_lrus_df.loc[:, ~acc_lrus_df.columns.duplicated()]
    direct_exclude_non_remarked_df = direct_exclude_non_remarked_df.loc[:, ~direct_exclude_non_remarked_df.columns.duplicated()]
    # 24/8/2025: END

    final_rep_df = pd.concat([remarked_df, acc_lrus_df, direct_exclude_non_remarked_df], ignore_index=True)

    final_rep_df_0 = final_rep_df[final_rep_df['total_price'] != 0].copy()
    final_rep_df_0.drop_duplicates(inplace=True)

    print("=== === === === === === === Repair Pipeline Completed Successfully!=== === === === === === ===")
    return final_rep_df_0, labor_amount, new_parts_non_llp, new_parts_llp, cfe


# === === === === === === === Vendor Pipeline === === === === === === === === === ===
def run_vendor_pipeline(master_folder: str, result: dict, esn):
    print("\n=== === === === === === === Starting Vendor Pipeline...=== === === === === === ===")
 
    # === Step 1: Extract Vendor Cost Categories from Contract ===
    vendor_cost_cat = extract_vendor_cost_category(result)
 
    # === Step 2: Filter Billing for Vendor Cost Categories ===
    vendor_subcon_cc_df = filter_billing_for_vendor(result, vendor_cost_cat)

    # vendor_cost_df = compute_vendor_total_price(vendor_subcon_cc_df)
 
    # # === Step 3: Enrich with Part Keywords and ATA_LONG ===
    # vendor_gen_part_keywords = enrich_with_part_keywords(result, vendor_subcon_cc_df, "vendor")
    # vendor_gen_part_keywords.drop_duplicates(inplace=True)
    # vendor_gen_part_keywords.to_excel("vendor_gen_part_keywords.xlsx", index=False)
 
    # eipc_path = result["file_paths"]["eipc"]
    # eipc_df = load_cleaned_eipc_dataframe(eipc_path)
    # vendor_eipc_ata_long = enrich_with_ata_long(result, eipc_df, vendor_gen_part_keywords, "vendor")
    # vendor_eipc_ata_long.to_excel("vendor_eipc_ata_long.xlsx", index=False)

    q_year, induction_date, cust_name, customer_first_word,q_year_g3, g3_date = extract_q_year_from_timeliness(result) 

    pipeline_type = "vendor"
    ven_billing_enriched = merge_billing_with_clp(result, vendor_subcon_cc_df, pipeline_type)
    ven_billing_enriched.to_excel("merge_billing_with_clp.xlsx", index=False)

    ven_billing_ext_price = compute_extended_price(ven_billing_enriched, pipeline_type)
    ven_billing_ext_price.rename(columns={'description_x':'description'}, inplace=True)
    ven_billing_ext_price.to_excel("compute_extended_price.xlsx", index=False)
    
    ven_contract_total_price = compute_total_price(ven_billing_ext_price, cust_name, esn, q_year, result, customer_first_word, pipeline_type)
    ven_contract_total_price.drop_duplicates(inplace=True)
    ven_contract_total_price.to_excel("check_ven_total_price_df.xlsx", index=False)

    eipc_path = result["file_paths"]["eipc"]
    eipc_df = load_cleaned_eipc_dataframe(eipc_path)
    vendor_eipc_ata_long = enrich_with_ata_long(result, eipc_df, ven_contract_total_price, "vendor")
    vendor_eipc_ata_long.to_excel("vendor_eipc_ata_long.xlsx", index=False)

    # === Step 3: Enrich with Part Keywords and ATA_LONG ===
    vendor_gen_part_keywords = enrich_with_part_keywords(result, vendor_eipc_ata_long, "vendor")
    vendor_gen_part_keywords.drop_duplicates(inplace=True)
    vendor_gen_part_keywords.to_excel("vendor_gen_part_keywords.xlsx", index=False)

    # === Step 4: Apply Logic for Inclusion/Exclusion ===
    vendor_gen_part_keywords["remarks"] = "VEND UNKNOWN"
    customer_df = extract_customer_workscope_levels(result)
    workscope = resolve_matching_workscope(customer_df,result)
    vendor_blade_inclusion = apply_blade_inclusion(result, vendor_gen_part_keywords, "vendor",workscope)
    vendor_blade_inclusion.to_excel("vendor_blade_inclusion.xlsx", index=False)

    vendor_ws_incl = apply_workscope_inclusion_exclusion(result, vendor_blade_inclusion, "vendor", "VEND UNKNOWN")
    #vendor_ws_incl = apply_workscope_inclusion_exclusion_vendor(result, vendor_blade_inclusion, "VEND UNKNOWN")
    #vendor_ws_incl.to_excel("vendor_ws_incl.xlsx", index=False)
    
    ven_ws_df = vendor_ws_incl[vendor_ws_incl['remarks'] != "VEND UNKNOWN"]
    ven_ata_logic_df = vendor_ws_incl[vendor_ws_incl['remarks'] == "VEND UNKNOWN"]

    vendor_acc_lrus = apply_qec_lru_logic(result, ven_ata_logic_df, "vendor")
    # vendor_acc_lrus.to_excel("vendor_acc_lrus.xlsx", index=False)

    ven_all_remarked_df = vendor_acc_lrus[vendor_acc_lrus['remarks'] != 'VEND UNKNOWN'].copy()
    ven_non_remarked_df = vendor_acc_lrus[vendor_acc_lrus['remarks'] == 'VEND UNKNOWN'].copy()
    # ven_non_remarked_df['remarks'] = "Exclusion" #'Direct Exclusion'
    
    # = = = = = = = = = = = = = = = = = =
    # Exclude remaining parts
    ven_direct_exclude_non_remarked_df = ven_non_remarked_df[ven_non_remarked_df['remarks'].str.upper().eq("VEND UNKNOWN")].copy() #== 'REPAIR UNKNOWN'].copy()
    ven_direct_exclude_non_remarked_df['remarks'] = "Exclusion" #'Direct Exclusion'

    # 24/8/2025: ADDED THIS NEW LOGIC
    ven_ws_df = ven_ws_df.loc[:, ~ven_ws_df.columns.duplicated()]
    ven_all_remarked_df = ven_all_remarked_df.loc[:, ~ven_all_remarked_df.columns.duplicated()]
    ven_direct_exclude_non_remarked_df = ven_direct_exclude_non_remarked_df.loc[:, ~ven_direct_exclude_non_remarked_df.columns.duplicated()]
    # 24/8/2025: END

    final_vend_df = pd.concat([ven_ws_df, ven_all_remarked_df, ven_direct_exclude_non_remarked_df], ignore_index=True)
    final_vend_df.drop_duplicates(inplace=True)


    vend_final_df = final_vend_df[final_vend_df['total_price'] != 0].copy()
    # vend_final_df.to_excel("vend_final_df.xlsx", index=False)
    print("=== === === === === === === Vendor Pipeline Completed Successfully!=== === === === === === ===\n")
    # pint(s)
    return vend_final_df


def build_material_values_file(raw_material_path, values_output_path):
    """
    Read overall_material_data.xlsx, compute all price columns in Python
    using the same logic as fill_mat_sheet, overwrite any existing Extended Price
    with the computed one, write a values-only Excel, and return the material subtotal.
    """
    df = pd.read_excel(raw_material_path)
 
    # 1) Normalize key column names
    df = df.rename(columns={
        "Unit Price": "unit_price",
        "Qty": "qty",
        "Discount %": "discount_%",
        "Discount $": "discount_$",
        "CLP": "clp",
        "Handling Fee %": "handling_fee_%",
        "Handling Fee Price": "handling_fee_price",
        "Excluded [O&A] Total Price": "excluded_[o&a]_total_price",
        "has_CUST": "has_cust",
    })
 
    # 2) Ensure numeric types (replace with 0 if NaN / None)
    for col in ["unit_price", "qty", "discount_%", "clp", "handling_fee_%", "has_cust"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        else:
            # If column is missing, create it as zeros
            df[col] = 0
 
    # 3) ALWAYS compute extended_price in Python (ignore any existing Extended Price values)
    df["extended_price"] = df["unit_price"] * df["qty"]
 
    # 4) ALWAYS compute discount_$ from computed extended_price
    df["discount_$"] = df["extended_price"] * df["discount_%"]
 
    # 5) ALWAYS compute handling_fee_price
    df["handling_fee_price"] = df["clp"] * df["qty"] * df["handling_fee_%"]
 
    # 6) excluded_[o&a]_total_price = IF(has_cust=1, handling_fee_price,
    #                                   extended_price - discount_$ + handling_fee_price)
    df["excluded_[o&a]_total_price"] = np.where(
        df["has_cust"] == 1,
        df["handling_fee_price"],
        df["extended_price"] - df["discount_$"] + df["handling_fee_price"],
    )
 
    # 7) Material subtotal = sum of excluded_[o&a]_total_price
    mat_sum_val = df["excluded_[o&a]_total_price"].sum()
 
    # 8) Map back to user-facing column names for saving
    #    - Overwrite "Extended Price" with computed extended_price
    #    - Rename tech columns back to original labels
    df["Extended Price"] = df["extended_price"]          # overwrite old data with computed
    df["Discount $"] = df["discount_$"]
    df["Handling Fee Price"] = df["handling_fee_price"]
    df["Excluded [O&A] Total Price"] = df["excluded_[o&a]_total_price"]
    df["has_CUST"] = df["has_cust"]
 
    # 9) Drop internal tech columns so they don't appear twice
    df = df.drop(
        columns=[
            "extended_price",      # internal
            "discount_$",
            "handling_fee_price",
            "excluded_[o&a]_total_price",
            "unit_price",          # original "Unit Price" column already exists
            "qty",
            "discount_%",
            "clp",
            "handling_fee_%",
            "has_cust",
        ],
        errors="ignore",
    )
 
    # 10) Save values-only version (no formulas, all numbers)
    df.to_excel(values_output_path, index=False)
 
    return mat_sum_val

def build_repair_values_file(raw_repair_path, values_output_path):
    """
    Read internal_table.xlsx (GE Repair), compute price columns in Python
    using the same logic as fill_rep_sheet, write a values-only Excel file,
    and return the repair subtotal.
 
    Logic (same as fill_rep_sheet):
      extended_price            = unit_price * qty
      discount_$                = extended_price * discount_%
      excluded_[o&a]_total_price= extended_price - discount_$
      subtotal                  = SUM(excluded_[o&a]_total_price)
    """
    df = pd.read_excel(raw_repair_path)
 
    # 1) Rename nicely formatted Excel headers to normalized names we can work with
    df = df.rename(columns={
        "Unit Price": "unit_price",
        "Qty": "qty",
        "Discount %": "discount_%",
        "Discount $": "discount_$",
        "Extended Price": "extended_price",
        "Excluded [O&A] Total Price": "excluded_[o&a]_total_price",
        # other columns we just leave as-is (Item (SD), Part Number, etc.)
    })
 
    # 2) Ensure numeric types (replace with 0 if NaN / None)
    for col in ["unit_price", "qty", "discount_%"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        else:
            df[col] = 0  # if missing, create with zeros so formulas don't break
 
    # 3) ALWAYS compute extended_price in Python (ignore any old values)
    df["extended_price"] = df["unit_price"] * df["qty"]
 
    # 4) ALWAYS compute discount_$ from computed extended_price
    df["discount_$"] = df["extended_price"] * df["discount_%"]
 
    # 5) ALWAYS compute excluded_[o&a]_total_price = extended_price - discount_$
    df["excluded_[o&a]_total_price"] = df["extended_price"] - df["discount_$"]
 
    # 6) Repair subtotal = sum of excluded_[o&a]_total_price
    rep_sum_val = df["excluded_[o&a]_total_price"].sum()
 
    # 7) Map back to user-facing column names for saving
    #    Overwrite existing display columns with computed values
    df["Extended Price"] = df["extended_price"]
    df["Discount $"] = df["discount_$"]
    df["Excluded [O&A] Total Price"] = df["excluded_[o&a]_total_price"]
 
    # 8) Drop internal technical columns so they don’t appear twice
    df = df.drop(
        columns=[
            "extended_price",
            "discount_$",
            "excluded_[o&a]_total_price",
            "unit_price",
            "qty",
            "discount_%",
        ],
        errors="ignore",
    )
 
    # 9) Save values-only version (no formulas, all numbers)
    df.to_excel(values_output_path, index=False)
 
    return rep_sum_val


 
def build_vendor_values_file(raw_vendor_path, values_output_path):
    df = pd.read_excel(raw_vendor_path)
 
    if df.empty:
        df.to_excel(values_output_path, index=False)
        return 0
 
    # 1️⃣ Normalize headers
    df = df.rename(columns={
        "Qty": "qty",
        "Vendor Invoice Amount": "vendor_invoice_amount",
        "Handling Fee %": "handling_fee_%",
        "Handling Fee Price": "handling_fee_price",
        "Excluded [O&A] Total Price": "excluded_[o&a]_total_price",
        "Discount $": "discount_$",
        "Discount %": "discount_%",
        "Repair Source": "repair_source",
    })
 
    # 2️⃣ Ensure numeric safety
    for col in [
        "qty",
        "vendor_invoice_amount",
        "handling_fee_%",
        "discount_%",
        "discount_$",
    ]:
        df[col] = pd.to_numeric(df.get(col, 0), errors="coerce").fillna(0)
 
    # 3️⃣ Normalize Repair Source
    df["repair_source"] = (
        df.get("repair_source", "")
        .astype(str)
        .str.strip()
        .str.upper()
    )
 
    is_subcon_non_ge = df["repair_source"] == "SUBCON-NONGE-ACTUAL"
 
    # 4️⃣ Apply BUSINESS RULES
    # -------------------------------------------------
    # SUBCON-NONGE-ACTUAL → Discount = 0
    df.loc[is_subcon_non_ge, ["discount_%", "discount_$"]] = 0
 
    # NOT SUBCON-NONGE-ACTUAL → Handling Fee = 0
    df.loc[~is_subcon_non_ge, "handling_fee_%"] = 0
    # -------------------------------------------------
 
    # 5️⃣ Recompute prices (authoritative)
    df["handling_fee_price"] = df["qty"] * df["handling_fee_%"]
 
    df["excluded_[o&a]_total_price"] = (
        df["vendor_invoice_amount"]
        + df["handling_fee_price"]
        - df["discount_$"]
    )
 
    # 6️⃣ Vendor subtotal
    ven_sum_val = df["excluded_[o&a]_total_price"].sum()
 
    # 7️⃣ Map back to display columns
    df["Handling Fee Price"] = df["handling_fee_price"]
    df["Excluded [O&A] Total Price"] = df["excluded_[o&a]_total_price"]
 
    # 8️⃣ Cleanup internal columns
    df = df.drop(
        columns=[
            "qty",
            "vendor_invoice_amount",
            "handling_fee_%",
            "handling_fee_price",
            "excluded_[o&a]_total_price",
        ],
        errors="ignore",
    )
 
    # 9️⃣ Save values-only file
    df.to_excel(values_output_path, index=False)
 
    return ven_sum_val

def build_qec_mat_values_file(raw_qec_mat_path: str, values_output_path: str):
    """
    QEC Material (qec_mat.xlsx) – recompute price columns in Python using the
    same logic as the Material/QEC block in fill_qec_lru_replacement_sheet:
 
        extended_price             = unit_price * qty
        discount_$                 = extended_price * discount_%
        handling_fee_price         = clp * qty * handling_fee_%
        excluded_[o&a]_total_price = extended_price - discount_$ + handling_fee_price
        qec_mat_sum                = SUM(excluded_[o&a]_total_price)
 
    Writes a values-only Excel file and returns qec_mat_sum.
    """
    df = pd.read_excel(raw_qec_mat_path)
 
    # Rename to normalized working names
    df = df.rename(columns={
        "Unit Price": "unit_price",
        "Qty": "qty",
        "CLP": "clp",
        "Discount %": "discount_%",
        "Discount $": "discount_$",
        "Handling Fee %": "handling_fee_%",
        "Handling Fee Price": "handling_fee_price",
        "Excluded [O&A] Total Price": "excluded_[o&a]_total_price",
    })
 
    # Ensure numeric columns exist and are numeric
    for col in ["unit_price", "qty", "clp", "discount_%", "handling_fee_%"]:
        if col not in df.columns:
            df[col] = 0
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
 
    # 1) extended_price = unit_price * qty
    df["extended_price"] = df["unit_price"] * df["qty"]
 
    # 2) discount_$ = extended_price * discount_%
    df["discount_$"] = df["extended_price"] * df["discount_%"]
 
    # 3) handling_fee_price = clp * qty * handling_fee_%
    df["handling_fee_price"] = df["clp"] * df["qty"] * df["handling_fee_%"]
 
    # 4) excluded_[o&a]_total_price = extended_price - discount_$ + handling_fee_price
    df["excluded_[o&a]_total_price"] = (
        df["extended_price"] - df["discount_$"] + df["handling_fee_price"]
    )
 
    # 5) subtotal
    qec_mat_sum = df["excluded_[o&a]_total_price"].sum()
 
    # Map computed values back to display column names
    df["Extended Price"] = df["extended_price"]
    df["Discount $"] = df["discount_$"]
    df["Handling Fee Price"] = df["handling_fee_price"]
    df["Excluded [O&A] Total Price"] = df["excluded_[o&a]_total_price"]
 
    # Drop internal working cols (keep the original display headers)
    df = df.drop(
        columns=[
            "unit_price",
            "qty",
            "clp",
            "discount_%",
            "handling_fee_%",
            "extended_price",
            "discount_$",
            "handling_fee_price",
            "excluded_[o&a]_total_price",
        ],
        errors="ignore",
    )
 
    # Save values-only version
    df.to_excel(values_output_path, index=False)
 
    return qec_mat_sum

def build_qec_rep_values_file(raw_qec_rep_path: str, values_output_path: str):
    """
    QEC Repair (qec_rep.xlsx) – recompute price columns in Python using the
    same logic style as the Repair sheet:
 
        extended_price             = unit_price * qty
        discount_$                 = extended_price * discount_%
        excluded_[o&a]_total_price = extended_price - discount_$
        qec_rep_sum                = SUM(excluded_[o&a]_total_price)
 
    Writes a values-only Excel file and returns qec_rep_sum.
    """
    df = pd.read_excel(raw_qec_rep_path)
 
    # Normalize names
    df = df.rename(columns={
        "Unit Price": "unit_price",
        "Qty": "qty",
        "Discount %": "discount_%",
        "Discount $": "discount_$",
        "Extended Price": "extended_price",
        "Excluded [O&A] Total Price": "excluded_[o&a]_total_price",
    })
 
    # Ensure numeric
    for col in ["unit_price", "qty", "discount_%"]:
        if col not in df.columns:
            df[col] = 0
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
 
    # 1) extended_price = unit_price * qty
    df["extended_price"] = df["unit_price"] * df["qty"]
 
    # 2) discount_$ = extended_price * discount_%
    df["discount_$"] = df["extended_price"] * df["discount_%"]
 
    # 3) excluded_[o&a]_total_price = extended_price - discount_$
    df["excluded_[o&a]_total_price"] = df["extended_price"] - df["discount_$"]
 
    # 4) subtotal
    qec_rep_sum = df["excluded_[o&a]_total_price"].sum()
 
    # Map back to display columns
    df["Extended Price"] = df["extended_price"]
    df["Discount $"] = df["discount_$"]
    df["Excluded [O&A] Total Price"] = df["excluded_[o&a]_total_price"]
 
    # Drop internal working cols
    df = df.drop(
        columns=[
            "unit_price",
            "qty",
            "discount_%",
            "extended_price",
            "discount_$",
            "excluded_[o&a]_total_price",
        ],
        errors="ignore",
    )
 
    df.to_excel(values_output_path, index=False)
 
    return qec_rep_sum


def run_create_invoice(output_excel, result: dict, cover_page, cust_name): #mat_result: pd.DataFrame, rep_result: pd.DataFrame, vend_result: pd.DataFrame, cover_page):
    print("\n=== === === === === === === Creating Invoice === === === === === === ===")
    log_ui_progress_message("Generating Billing receipt..", overwrite=True)

    inscope_table1, inscope_table2, inscope_sheet_name, year_str = create_in_scope(result, output_excel)
    #inscope_df_info = write_multiple_dfs_to_sheet_with_gap(output_excel, inscope_sheet_name, dfs=[inscope_table1, inscope_table2], start_row=5, gap_rows=3)
    #col_position = inscope_table2.columns.get_loc("A") + 1  # Position after "A"
    #inscope_table2.insert(col_position, "A+", "")  # Add blank column
    #.rename(columns={"Module Identifier": "Minor Module"}, inplace=True)
    #combined_df = pd.concat([inscope_table1, inscope_table2], ignore_index=True)
    #inscope_df_info = write_single_df_to_excel(output_excel, inscope_sheet_name, combined_df, start_row=1, gap_rows=1)
    #print(inscope_df_info)    
    #fill_inscope_sheet(output_excel, inscope_sheet_name, inscope_df_info, year_str)

    col_position = inscope_table2.columns.get_loc("A") + 1  # Position after "A"
    inscope_table2.insert(col_position, "A+", "")  # Add blank column
    inscope_table2.rename(columns={"Module Identifier": "Minor Module"}, inplace=True)
    combined_df = pd.concat([inscope_table1, inscope_table2], ignore_index=True)

    #inscope_df_info = write_multiple_dfs_to_sheet_with_gap(output_excel, inscope_sheet_name, dfs=[inscope_table1, inscope_table2], start_row=5, gap_rows=3)
    inscope_df_info = write_single_df_to_excel(output_excel, inscope_sheet_name, combined_df, start_row=4, gap_rows=1)

    fill_inscope_sheet(output_excel, inscope_sheet_name, inscope_df_info, year_str)

    # = = LLP Replacement = = 
    used_llp_template_path = "SCC_LLP_Pro_Rata_Template_Filled.xlsx"
    new_llp_df, used_llp_df, overall_llp_df   = create_llp_replacement(result, output_excel, cust_name)

    llp_dfs = ["llp_new_table.xlsx", "llp_used_table.xlsx", "overall_llp_df.xlsx"]
    # metadata_llp = write_dfs_with_gaps(file_paths=["overall_llp_df.xlsx"], output_file=output_excel, sheet_name=mat_sheet_name)

    # llp_df_info = write_dfs_with_gaps(llp_dfs, output_excel, llp_replacement_sheet_name, start_rows=[6], row_gaps=5)
    # fill_llp_replacement_sheet(output_excel, llp_replacement_sheet_name, llp_df_info, year_str)

    # # = = QEC LRU = = 
    # mat_qec_lru_df, rep_qec_lru_df_rep, qec_lru_sheet_name = create_qec_lru(result, output_excel)
    # # old_qec_name_mat, new_qec_name_mat, old_qec_name_rep, new_qec_name_rep = "QEC - LRU Material", "QEC-LRU-ACC", "QEC - LRU Repair", "QEC-LRU-ACC"
    # metadata_qec = write_dfs_with_gaps(file_paths=["qec_mat.xlsx", "qec_rep.xlsx"], output_file=output_excel, sheet_name=qec_lru_sheet_name)
    # fill_qec_lru_replacement_sheet(output_excel, qec_lru_sheet_name, metadata_qec)

    # qer_lru_df_info = write_multiple_dfs_to_sheet_with_gap(output_excel, qec_lru_sheet_name, dfs=[qec_lru_df], start_row=4, gap_rows=0)
    
    # start_row_1 = qer_lru_df_info[0]["row_start"]
    # end_row_1 = qer_lru_df_info[0]["row_end"]
    
    # wb = load_workbook(output_excel)
    # ws = wb[qec_lru_sheet_name]
    
    # row_ranges = [(start_row_1+1, end_row_1+1)] 
    # qec_lru_info_refined = delete_empty_rows_and_return_metadata(ws, row_ranges)
    # wb.save(output_excel)

    # fill_qec_lru_replacement_sheet(output_excel, qec_lru_sheet_name, qec_lru_info_refined, year_str)

    # = = PMA DER = = 
    pma_der_df = "pma_der_df.xlsx"
    pma_der_table = create_pma_der(result, output_excel)
   

    # = = Missing on Receipts = = 
    mor_df = "mor_df.xlsx"
    mor_table = create_missing_on_receipts(result, output_excel)
    
    
    # = = Service Bulleting = = 
    overall_sb_grouped, sb_sheet_name = create_service_bulletin(result, output_excel)
    metadata_sb = write_sb_df_with_gaps(file_paths=["overall_sb_grouped.xlsx"], output_file=output_excel, sheet_name=sb_sheet_name)
    fill_sb_oa_summary(output_excel, sb_sheet_name, metadata_sb)
 

    # = = Material, Repair, Vendor = = 
    mat_table, internal_table, external_table, mat_sheet_name, repair_sheet_name, vendor_sheet_name = create_material_repair(result, output_excel)
    # rrrrrrrrr
    matl_rep_ven_dfs = ["mat_table.xlsx", "internal_table.xlsx", "external_table.xlsx"]
    overall_llp_pma_mor_dfs = pd.concat([overall_llp_df, pma_der_table, mor_table], ignore_index=True)
    overall_material_df = pd.concat([overall_llp_pma_mor_dfs, mat_table], ignore_index=True)
    cust_types = [
    "MATL-CUSTOMER",
    "MATL-CUST-SCRAPREP",
    "LLP-CUST-SCRAPREP"
    ]
    
    overall_material_df.loc[
        overall_material_df["Material Type"].isin(cust_types),
        "has_CUST"
    ] = 1
    overall_material_df.to_excel("overall_material_data2.xlsx", index=False)
    df_new=pd.read_excel("overall_material_data2.xlsx")

    df_new['code_no_dash'] = df_new['ATA Chapter Code'].str.replace('-', '', regex=False)

    #mask = (df_new['Comments'] == 'Exclusion as WKS = 2') & (df_new['code_no_dash'].str.endswith('00', na=False))
    mask = (df_new['Comments'].str.startswith(('Exclusion as WKS = 1', 'Exclusion as WKS = 2'),na=False) & df_new['code_no_dash'].str.endswith('00', na=False))


    df_new.loc[mask, 'Comments'] = 'Inclusion with Logic'
    df_clean = df_new[~df_new['Comments'].str.contains('Inclusion with Logic', case=False, na=False)]
        
    mask_used = df_clean["Material Type"].astype(str).str.startswith("MATL-USED", na=False)

    # Precompute adjusted CLP and the comparison term
    clp_adj = df_clean.loc[mask_used, "CLP"] * 0.7
    compare_val = clp_adj * df_clean.loc[mask_used, "Qty"]

    #    Apply the rule only to the masked rows
    # If CLP*0.7*Qty < Unit Price  -> Unit Price = CLP*0.7
    # Else                          -> Unit Price = Unit Price / Qty
    df_clean.loc[mask_used, "Unit Price"] = np.where(
    compare_val < df_clean.loc[mask_used, "Unit Price"],
    clp_adj,
    df_clean.loc[mask_used, "Unit Price"] / df_clean.loc[mask_used, "Qty"]
    )
        #30-12-2025 
    df_sb_m=pd.read_excel("Sb_test_23.xlsx")
    df_sb_m.drop_duplicates(inplace=True)
    unique_df1_m = df_clean[~df_clean['Part Number'].isin(df_sb_m['Part Number'])]
    overall_sb_sb = pd.concat([unique_df1_m, df_sb_m])
    overall_sb_sb = overall_sb_sb.drop_duplicates(keep='last')
    
    mask = (
    (overall_sb_sb["Unit Price"] == 0) &
    (overall_sb_sb["Qty"].notna()) &
    (overall_sb_sb["Qty"] > 0)
    )
    
    overall_sb_sb.loc[mask, "Unit Price"] = (
        overall_sb_sb.loc[mask, "Extended Price"] /
        overall_sb_sb.loc[mask, "Qty"]
    )
 
    overall_sb_sb.to_excel("overall_material_data.xlsx", index=False)
 

    # df_clean.to_excel("overall_material_data.xlsx", index=False)

    df_rep=pd.read_excel("internal_table.xlsx")

    #$df_new['code_no_dash'] = df_new['ATA Chapter Code'].str.replace('-', '', regex=False)
    #new logic
    #mask = ((df_rep['Comments'] == 'Exclusion as WKS = 2') & (df_rep['Part Description'] == 'CASE-CONTAINMENT'))
    mask = (df_rep['Comments'].str.startswith(('Exclusion as WKS = 1', 'Exclusion as WKS = 2'),na=False) & (df_rep['Part Description'] == 'CASE-CONTAINMENT'))


    df_rep.loc[mask, 'Comments'] = 'Inclusion with Logic'
    df_clean2 = df_rep[~df_rep['Comments'].str.contains('Inclusion with Logic', case=False, na=False)]
    
    df_sb_r=pd.read_excel("Sb_test_23_r.xlsx")
    df_sb_r.drop_duplicates(inplace=True)
    unique_df1 = df_clean2[~df_clean2['Part Number'].isin(df_sb_r['Part Number'])]
    df_not_in_sb = df_clean2[
    ~df_clean2['Part Number'].isin(df_sb_r['Part Number'])
    ].copy()

    unique_df1 = df_not_in_sb[
    ~(
        (df_not_in_sb['Part Description'] == 'CASE-CONTAINMENT') &
        (df_not_in_sb['Comments'] != 'Exclusion as WKS = 0')
    )
    ]

    overall_sb_r = pd.concat([unique_df1, df_sb_r])
    
    
   # -------------------------------------------------
    # Normalize SB72-0119 rows for Excel pricing logic
    # -------------------------------------------------
    
    sb_mask = (
        overall_sb_r["Repair Code / Service Level"]
        .astype(str)
        .str.strip()
        .eq("SB72-0119")
    )
    
    # 1️⃣ Qty → 1
    if "Qty" in overall_sb_r.columns:
        overall_sb_r.loc[sb_mask, "Qty"] = (
            overall_sb_r.loc[sb_mask, "Qty"]
            .fillna(1)
            .replace(0, 1)
        )
    
    # 2️⃣ Discount % and Discount $ → 0
    for col in ["Discount %", "Discount $"]:
        if col in overall_sb_r.columns:
            overall_sb_r.loc[sb_mask, col] = 0
    
    # 3️⃣ Excluded Total = Extended Price
    if {"Excluded [O&A] Total Price", "Extended Price"}.issubset(overall_sb_r.columns):
        overall_sb_r.loc[sb_mask, "Excluded [O&A] Total Price"] = (
            overall_sb_r.loc[sb_mask, "Extended Price"]
        )
    
    # 4️⃣ Any OTHER numeric column → default to 1 if empty
    numeric_cols = overall_sb_r.select_dtypes(include=["number"]).columns
    
    exclude_cols = {
        "Discount %",
        "Discount $",
        "Excluded [O&A] Total Price",
        "Extended Price"
    }
    
    for col in numeric_cols:
        if col not in exclude_cols:
            overall_sb_r.loc[sb_mask, col] = (
                overall_sb_r.loc[sb_mask, col]
                .fillna(0)
                # .replace(0, 1)
            )
 
    #overall_sb_sb.to_excel("overall_material_data.xlsx", index=False)
    overall_sb_r.to_excel("internal_table2.xlsx", index=False)
    
    # df_clean2.to_excel("internal_table2.xlsx", index=False)
    
    

    df_ven=pd.read_excel("external_table.xlsx")
    df_filtered_v =  df_ven[ df_ven['Repair Source'].str.startswith('SUBCON', na=False)]
        
    df_filtered_v["Vendor Invoice Amount"] = pd.to_numeric(df_filtered_v["Vendor Invoice Amount"], errors="coerce")
    df_filtered_v["Discount %"] = 31
    df_filtered_v["Discount $"] = (df_filtered_v["Vendor Invoice Amount"] * 31 / 100).round(2)
    df_filtered_v.to_excel("external_table2.xlsx", index=False)
    
    # mat_sum_val = overall_material_df['Excluded [O&A] Total Price'].sum()
    # rep_sum_val = internal_table['Excluded [O&A] Total Price'].sum()
    # ven_sum_val = external_table['Excluded [O&A] Total Price'].sum()


     #Material Value
    material_values_path = "overall_material_data_values_only.xlsx"
    overall_material_df = "overall_material_data.xlsx"
    
    #GE Repair Value
    rep_raw_path = "internal_table2.xlsx"
    rep_values_path = "internal_table_values_only.xlsx"
    
    #Sub Contracted Repair
    ven_raw_path = "external_table2.xlsx"
    ven_values_path = "external_table_values_only.xlsx"
    
    mat_sum_val = build_material_values_file(overall_material_df, material_values_path)
    print("Material subtotal from values-only file:", mat_sum_val)
    
    # Build values-only repair file + get subtotal
    rep_sum_val = build_repair_values_file(rep_raw_path, rep_values_path)
    print("Repair subtotal (Python, same logic as sheet):", rep_sum_val)
    
    # Build values-only vendor file + get subtotal
    ven_sum_val = build_vendor_values_file(ven_raw_path, ven_values_path)
    print("Vendor subtotal (Python, same logic as vendor sheet):", ven_sum_val)



    metadata_mat = write_dfs_with_gaps(file_paths=["overall_material_data.xlsx"], output_file=output_excel, sheet_name=mat_sheet_name)
    metadata_rep = write_dfs_with_gaps(file_paths=["internal_table2.xlsx"], output_file=output_excel, sheet_name=repair_sheet_name)
    metadata_ven = write_dfs_with_gaps(file_paths=["external_table2.xlsx"], output_file=output_excel, sheet_name=vendor_sheet_name)

    print("Overall Material Data written to Excel.", output_excel)
    
    # Fill sheets (styling, formulas, etc.)
    fill_mat_sheet(output_excel, mat_sheet_name, metadata_mat)
    fill_rep_sheet(output_excel, repair_sheet_name, metadata_rep)
    fill_vendor_sheet(output_excel, vendor_sheet_name, metadata_ven)

    # = = QEC LRU = = 
    # Delete extra QEC LRU Repair and Material Sheet
    sheets_to_del = ['QEC - LRU Repair', 'QEC - LRU Material', 'Misc (test cell, storage, etc)', 'Scrap Table']
    wb = load_workbook(output_excel)
    for sheet_name in sheets_to_del:
        if sheet_name in wb.sheetnames:
            std = wb[sheet_name]
            wb.remove(std)
    wb.save(output_excel)
    
    mat_qec_lru_df, rep_qec_lru_df_rep, qec_lru_sheet_name, qec_mat_sum, qec_rep_sum = create_qec_lru(result, output_excel)
    
    df_qec_m=pd.read_excel("qec_mat.xlsx")    
    filtered_df2 = df_qec_m[~df_qec_m["Part Number"].isin(df_sb_m["Part Number"])]
    filtered_df2.to_excel("qec_m_2.xlsx",index=False)
    
    # Build values-only QEC files + get sums
    qec_mat_values_path = "qec_mat_values_only.xlsx"
    qec_rep_values_path = "qec_rep_values_only.xlsx"
    
    qec_mat_sum = build_qec_mat_values_file("qec_m_2.xlsx", qec_mat_values_path)
    qec_rep_sum = build_qec_rep_values_file("qec_rep.xlsx", qec_rep_values_path)
    
    print("QEC material subtotal:", qec_mat_sum)
    print("QEC repair subtotal  :", qec_rep_sum)

    metadata_qec = write_dfs_with_gaps(file_paths=["qec_m_2.xlsx", "qec_rep.xlsx"], output_file=output_excel, sheet_name=qec_lru_sheet_name)
    fill_qec_lru_replacement_sheet(output_excel, qec_lru_sheet_name, metadata_qec)

    # # = = Service Bulleting = = 
    # overall_sb_grouped, sb_sheet_name = create_service_bulletin(result, output_excel)
    # metadata_sb = write_sb_df_with_gaps(file_paths=["overall_sb_grouped.xlsx"], output_file=output_excel, sheet_name=sb_sheet_name)
    # fill_sb_oa_summary(output_excel, sb_sheet_name, metadata_sb)

    # Compare and Delete not matching or empty sheets
    target_inv_sheets = ["Material", "GE Repair", "Sub Contracted Repair", "QEC-LRU-ACC", "Misc", "Scrap Table", "SB O&A Summary"]
    result = compare_excel_sheets_contains(output_excel, target_inv_sheets)
    
    for target, matches in result.items():
        if matches:
            print(f"[✅] Found matches for '{target}': {matches}")
        else:
            print(f"[❌] No matches for '{target}'")
            delete_empty_sheets(output_excel)
            # print(f"[❌] deleted not matched sheets for '{target}'")
    
    return mat_sum_val, rep_sum_val, ven_sum_val, qec_mat_sum, qec_rep_sum


def invfinalDownload():
    actual_draft_invoice_path = find_invoice_template_in_current_folder("draft ge celma ffp invoice template")

    file1 = "UI_final_invoice_details_temp.xlsx"
    # time.sleep(40)
 
    file2 = "UI_final_invoice_details.xlsx"
 
    # 1. Get updated values from file2
    updated_rows = compare_files(file1, file2)
    print(updated_rows)
 
    attributes_below = [
    "customer:",
    "contract",
    "engineserial",
    "notification",
    "customer_order",
    "invoice",
    "invoice_date",
    "work_performed_at",
    "foreign_affiliate_of",
    "send_payment_to",
    "shipping_address",
    "billing_address",
    "shop_visit_induction_date",
    "shipping_reference",
    "shipped_date",
    "payment_terms",
    "payment_due_date",
    ]
 
    update_cover_sheet_from_comparison(
        cover_path=actual_draft_invoice_path,
        updated_rows=compare_files(file1, file2),
        attributes_below=attributes_below
    )
     
    return True

# Main Function
def invCreationMain(esnNo, recomm_flag):
    import time
    import glob

    start_time = time.time()
    print("Invoice Creation process started...")
    print("## Recomm_Flag received:", recomm_flag)
    
    # Delete previous Draft invoices outside SMBA folder
    # Match files that contain the string anywhere in the name
    target_substring = "Draft_GE Celma_FFP Invoice Template"
    matching_files = glob.glob(f"*{target_substring}*")
    
    if matching_files:
        for file in matching_files:
            try:
                os.remove(file)
                print(f"Deleted: {file}")
            except Exception as e:
                print(f"Failed to delete {file}: {e}")
    else:
        print("No draft files found.")

    # Delete old files in the create_final_Invoice folder:
    deleted, failed = delete_files_in_folder(r"create_final_Invoice", pattern="*", recursive=False)
    print("Deleted: Old Draft Final Invoice from create folder", *deleted, sep="\n")
    if failed:
        print("\nFailed:")
        for f, err in failed:
            print(f"{f} -> {err}")
            
    files_to_del = [
    "qec_m_2.xlsx",
    "qec_rep.xlsx",
    "qec_mat.xlsx"
    ]
    
    for fp in files_to_del:
        if os.path.exists(fp):
            print(f"Removing file: {fp}")
            os.remove(fp)
        else:
            # silently ignore OR log if needed
            print(f"Skipping (not found): {fp}")

    
    esn = esnNo
    print("ESN Frontend",esn)
    
    master_folder = "SMBA_Wales" # MASTER FOLDER
    engfam_esn_files = user_input_file_checks(master_folder,esn) #Files adn dfs ready!
    # print("eng_fam_sheets-alllllll",engfam_esn_files)
    # append_to_excel("Applying Billing Rules and Exclusions...")

    # === Step 1: Initialize the Material Pipeline ===
    mat_result, cust_name, sb_filtered = run_material_pipeline(master_folder, engfam_esn_files, esn)
    # # # Write final outputs
    mat_result.to_excel("final_summary_material.xlsx", index=False)
    # mattttttttttttt
    #cust_ws_level.to_excel("after_workn.xlsx", index=False)

    # sb_filtered = pd.read_excel("sb_filtered_with_ual.xlsx")


    # === Step 2: Initialize the Repair Pipeline ===
    repair_result, labor_amount, new_parts_non_llp, new_parts_llp, cfe = run_repair_pipeline(master_folder, engfam_esn_files, sb_filtered)
    # Write final outputs
    repair_result.to_excel("final_summary_repair.xlsx", index=False)
    
    # rrrrr

    # === Step 3: Initialize the Vendor Pipeline ===
    vendor_result = run_vendor_pipeline(master_folder, engfam_esn_files, esn)
    # Write final outputs
    vendor_result.to_excel("final_summary_vendor.xlsx", index=False)

    if recomm_flag:
        print("✅ Recommendation flag is TRUE. Applying recommendation files.")
 
        replace_if_exists(
            "final_summary_material_recc.xlsx",
            "final_summary_material.xlsx"
        )
 
        replace_if_exists(
            "final_summary_repair_recc.xlsx",
            "final_summary_repair.xlsx"
        )
 
        replace_if_exists(
            "final_summary_vendor_recc.xlsx",
            "final_summary_vendor.xlsx"
        )
 
    else:
        print("ℹ️ Recommendation flag is FALSE. Using original summaries.")

    # # BIlling receipt code
    # append_to_excel("Generating Billing Receipt Details...")
    output_path = "Final_Invoice_Summary.xlsx"  
    mat_result = mat_result[mat_result['remarks'].astype(str).str.contains("exclusion", case=False, na=False)]
    repair_result = repair_result[repair_result['remarks'].astype(str).str.contains("exclusion", case=False, na=False)]
    vendor_result = vendor_result[vendor_result['remarks'].astype(str).str.contains("exclusion", case=False, na=False)]

    gen_invoice = generate_invoice_summary(mat_result, repair_result, vendor_result, output_path)

    # append_to_excel("Creating Draft Invoice...")
    cover_page = generate_cover_sheet("cover_invoice_sheet.xlsx")
    fill_cover_sheet(cover_page, engfam_esn_files, labor_amount, new_parts_non_llp, new_parts_llp, cfe)
        
    input_excel = "cover_invoice_attr.xlsx"
    esn_list = [int(engfam_esn_files['esn'])]
    esn = str(esn_list[0])
    output_excel = f"Draft_GE Celma_FFP Invoice Template_{esn}.xlsx"

    create_sheets_file = "cover_invoice_with_sheets_added.xlsx"
    create_sheets_from_range(input_excel, create_sheets_file)

    # append_to_excel("Appending Time & Materials (T&M) and Fixed Price (FP) Costs...")
    shutil.copy(create_sheets_file, output_excel)
    
    mat_sum_val, rep_sum_val, ven_sum_val, qec_mat_sum, qec_rep_sum = run_create_invoice(output_excel, engfam_esn_files, cover_page, cust_name)    
    mat_sum_val  = to_number(mat_sum_val)
    rep_sum_val  = to_number(rep_sum_val)
    ven_sum_val  = to_number(ven_sum_val)
    qec_mat_sum  = to_number(qec_mat_sum)
    qec_rep_sum  = to_number(qec_rep_sum)
    o_and_above_total_sum_val = (mat_sum_val + rep_sum_val + ven_sum_val + qec_mat_sum + qec_rep_sum)
    other_total_sum  = 0
    grand_total_sum = (o_and_above_total_sum_val + other_total_sum)

    print("\n--- Invoice Summary ---")
    print(f"Material Total            : {mat_sum_val}")
    print(f"Repair Total              : {rep_sum_val}")
    print(f"Vendor Total              : {ven_sum_val}")
    print(f"QEC Material Total        : {qec_mat_sum}")
    print(f"QEC Repair Total          : {qec_rep_sum}")
    print(f"Over & Above Total       : {o_and_above_total_sum_val}")
    print(f"Other Total               : {other_total_sum}")
    print(f"GRAND TOTAL              : {grand_total_sum}")
    print("------------------------\n")

    # # Overwriting the COST ITEMS
    actual_draft_invoice_path = find_invoice_template_in_current_folder("draft ge celma ffp invoice template")
    cover_sheet_with_cost_path = write_cover_sheet_cost_items(actual_draft_invoice_path, mat_sum_val, rep_sum_val, ven_sum_val, qec_mat_sum, qec_rep_sum, o_and_above_total_sum_val, other_total_sum, grand_total_sum)
 
    def safe_copyfile(src, dst, retries=5, delay=1):
        for attempt in range(1, retries + 1):
            try:
                shutil.copyfile(src, dst)
                print(f"[✔️] Successfully copied '{src}' to '{dst}'")
                return True
            except PermissionError as e:
                print(f"[⚠️] Attempt {attempt} failed: Permission denied for '{dst}'. Retrying in {delay} seconds...")
                time.sleep(delay)
        print(f"[❌] Failed to copy '{src}' to '{dst}' after {retries} attempts due to permission errors.")
        return False
    
    def close_excel_and_copy(cover_sheet_with_cost_path, data_only_sheet):
        # Kill Excel processes to release any file locks
        print("[ℹ️] Killing all Excel processes...")
        os.system("taskkill /f /im excel.exe")
        time.sleep(2)  # Give Windows a moment to release locks

        success = safe_copyfile(cover_sheet_with_cost_path, data_only_sheet)
        if not success:
            raise PermissionError(f"Could not copy file '{cover_sheet_with_cost_path}' to '{data_only_sheet}' due to permission issues.")


    data_only_sheet = "save_only_data_remove_formula.xlsx" 
    close_excel_and_copy(cover_sheet_with_cost_path, "save_only_data_remove_formula.xlsx" )   

    # append_to_excel("Preparing Invoice for Download...")

    in_path  = find_invoice_template_in_current_folder("save_only_data_remove_formula")
    print(f"in_path: {in_path}")

    out_path = "baked_values_temp.xlsx"
    if os.path.exists(out_path):
        try:
            print(f"Deleting existing file: {out_path}")
            os.remove(out_path)
        
        except Exception as e:
            print(f"Error deleting file {out_path}: {e}")

    bake_formulas_and_remove_hidden_sheets_final("save_only_data_remove_formula.xlsx", "baked_values_temp.xlsx", "Summary")
    print("out_path :", out_path)

    # Now that file is saved and closed, extract values
    MFP,material, ge_repair, sub_contracted_repair, qec_lru_repair, qec_lru_material, total_sum = extract_inscope_outscope_cost("baked_values_temp.xlsx")
    # print("SUM: i", material, ge_repair, sub_contracted_repair, qec_lru_repair, qec_lru_material, total_sum)

    extract_cover_data(cover_sheet_with_cost_path,MFP)
    UI_final_invoice_details = "UI_final_invoice_details.xlsx"
    # final_result = transform_excel_with_headers("UI_final_invoice_details.xlsx", "UI_final_invoice_details.xlsx")
    # print(final_result)

    # ui_df = pd.read_excel("UI_final_invoice_details.xlsx")
    # ui_df.loc[ui_df["Attribute"].str.strip() == "material_val", "Value"] = material
    # ui_df.loc[ui_df["Attribute"].str.strip() == "GE Repair", "Value"] = ge_repair
    # ui_df.loc[ui_df["Attribute"].str.strip() == "Sub Contracted Repair", "Value"] = sub_contracted_repair
    # ui_df.loc[ui_df["Attribute"].str.strip() == "QEC-LRU-ACC", "Value"] = [qec_lru_repair, qec_lru_material]
    # ui_df.loc[ui_df["Attribute"].str.strip() == "Over and Above (exclusions) Total", "Value"] = total_sum
    # ui_df.loc[ui_df["Attribute"].str.strip() == "Final_Invoice_Total", "Value"] = total_sum

    # ui_df.to_excel("UI_final_invoice_details.xlsx")
    # print("UI_final_invoice_details Content: ", ui_df, "\n")

    # shutil.copy("UI_final_invoice_details.xlsx", "UI_final_invoice_details_temp.xlsx")
    UI_final_invoice_details, UI_final_invoice_details_temp = "UI_final_invoice_details.xlsx", "UI_final_invoice_details_temp.xlsx"
    close_excel_and_copy(UI_final_invoice_details, UI_final_invoice_details_temp)  
    
    # append_to_excel("Your Draft Invoice is Ready to Download...")
    invfinalDownload()

    # cover_sheet_with_cost_path = cover_sheet_with_cost_path = write_formulas_back(actual_draft_invoice_path)

    # Clean files
    file_to_del_1 = "cover_invoice_attr.xlsx"
    file_to_del_2 = "cover_invoice_sheet.xlsx"
    # file_to_del_3 = "final_summary_material.xlsx"
    # file_to_del_4 = "final_summary_repair.xlsx"
    # file_to_del_5 = "final_summary_vendor.xlsx"
    # file_to_del_6 = "repair_service_bulletin_remarks.xlsx"
    file_to_del_7 = "service_bulletin_remarks.xlsx"
    # file_to_del_8 = "baked_values_temp.xlsx"
    # file_to_del_9 = "cover_invoice_with_sheets_added.xlsx"
    # file_to_del_10 = "save_only_data_remove_formula.xlsx" 
    # file_to_del_11 = "external_table.xlsx"
    file_to_del_12 = "inscope_egx_df.xlsx"
    # file_to_del_13 = "inscope_sap_merged_df.xlsx"
    # file_to_del_14 = "internal_table.xlsx"
    # file_to_del_15 = "mat_sb_data.xlsx"
    # file_to_del_16 = "mat_table.xlsx"
    # file_to_del_17 = "rep_sb_data.xlsx"
    # file_to_del_18 = "UI_final_invoice_details_temp.xlsx"
    # file_to_del_19 = f"Draft_GE Celma_FFP Invoice Template_{esnNo}"
 
    files_to_del = [file_to_del_1, file_to_del_2, file_to_del_7, file_to_del_12]

    for fp in files_to_del:
        try:
            os.remove(fp)
        except FileNotFoundError:
            print(f"File not found: {fp}")
        except Exception as e:
            print(f"Error deleteing {fp}: {e}")

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"\nSMBA Invoice Creation took {elapsed_time:.2f}seconds to run.")
    return True

esnNo = "569287" #"802406" #"894117, 874822, 892547, 888836
reccFlag = False
invCreationMain(esnNo,reccFlag)       