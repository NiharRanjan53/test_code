import pandas as pd
import os
import warnings
import re
from datetime import datetime, timedelta
from itertools import permutations
import calendar
from lxml import etree
from pathlib import Path
import sys
import shutil
import openpyxl
import numpy as np
from numbers import Number
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.styles.colors import Color
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill, numbers
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.drawing.image import Image
from collections import defaultdict
#from pycel.excelcompiler import ExcelCompiler
from extract_used_llp_file import extract_tlc_lcr_used_llp, bake_formulas_and_remove_hidden_sheets, find_file_and_sheet#bake_values_in_visible_sheet, delete_hidden_sheets, find_file_in_smba
from s3connect import read_excel_from_s3, list_files_in_s3,save_excel_to_s3,s3_file_exists
from io import BytesIO
import posixpath
from rapidfuzz import fuzz
from typing import List, Tuple
import logging 
from rapidfuzz import fuzz

# from workscope_classify import scrap_exclusion_calculation

# logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)


# File presence checker (by keyword match)
EXPECTED_FILES = {
    "contract": ("Contract", "Contract Summary"),
    "billing": ("Billing", "Billing Request"),
    "timeliness": ("Timeliness Tracker", "Timeliness Tracker"),
    "internal": ("Billing", "Internal Repair"),
    "comp": ("Catalog", "Comp"),
    "clp": ("Catalog", "Parts Catalog"),
    "eipc": ("Contract", "EIPC"),
    "sap": ("Billing", "Workscope SAP"),
    "lru_cat5": ("Billing", "LRU Z14 cat 5"),
    "lru_soft_time": ("Billing", "LRU Z14 soft time result"),
    "lru_ze57": ("Billing", "LRU ZE57"),
    "service bullet": ("Billing", "SBsZE14"),
    "used_llp": ("Billing", "AD Status"),
    "escalation": ("Catalog", "Escalation"),
    "SB":("Billing","SBsZE57")
 
} 

def normalize_string(s: str) -> str:
    return re.sub(r'[\W_]+', '', s.lower())

def get_customer_code_from_filtered_df(df, customer_col='end customer name'):
    
    if df.empty:
        print("[WARNING] No rows matched the condition.")
        return None
 
    # Get the first customer value from the filtered result
    customer_value = df.iloc[0][customer_col]
    customer_first_word = customer_value.strip().split()[0]

    # Generate the 3-letter code
    if isinstance(customer_value, str):
        words = customer_value.strip().split()
        if len(words) >= 2:
            code = (words[0][0] + words[1][:2]).upper()
        else:
            code = words[0][:3].upper()
        return code, customer_first_word
    else:
        print("[WARNING] Customer value is not a string.")
        return None
  

def check_required_files(master_folder: str, esn: str) -> dict[str, str]:
    results = {}
    esn_norm = normalize_string(str(esn))
 
    for key, (subfolder, keyword) in EXPECTED_FILES.items():
        folder_path = f"{master_folder}/{subfolder}".strip("/") + "/"
        all_files = list_files_in_s3(folder_path)
 
        filenames = [f.split("/")[-1] for f in all_files]
        keyword_norm = normalize_string(keyword)
 
        # 🔹 Apply ESN filter ONLY for Billing folder
        if subfolder.lower() == "billing":
            esn_filtered = [
                f for f in filenames
                if esn_norm in normalize_string(f)
            ]
        else:
            esn_filtered = filenames  # ⛔ no ESN filtering
 
        # 🔹 Apply keyword filtering
        matches = [
            f for f in esn_filtered
            if keyword_norm in normalize_string(f)
        ]
 
        selected_file = None
 
        if matches:
            if key == "internal":
                non_alter = [f for f in matches if "alter" not in f.lower()]
                selected_file = non_alter[0] if non_alter else matches[0]
            else:
                selected_file = matches[0]
 
            results[key] = f"{folder_path}{selected_file}"
        else:
            results[key] = None
            print(
                f"[WARNING] ❌ Missing: {keyword} "
                f"(ESN filter={'ON' if subfolder=='Billing' else 'OFF'}) "
                f"in {folder_path}"
            )
 
    return results

 


def load_excel_sheets(filepath: str) -> dict[str, pd.DataFrame]:
    """
    Load all sheets from an Excel file stored in S3 into a dict of DataFrames.
    Keys = sheet names, Values = DataFrames.
    """
    dfs = {}
 
    if not filepath.lower().endswith((".xls", ".xlsx")):
        print(f"[WARNING] ⚠️ Skipping non-Excel file: {filepath}")
        return dfs
 
    try:
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
 
        # ✅ Always fetch from S3
        file_bytes = read_excel_from_s3(filepath, return_bytes=True)
 
        # ✅ Load Excel from memory (no local disk needed)
        excel = pd.ExcelFile(BytesIO(file_bytes))
 
        if not excel.sheet_names:
            print(f"[WARNING] ⚠️ Excel file '{filepath}' has no sheets.")
            return dfs
 
        for sheet in excel.sheet_names:
            try:
                df_temp = excel.parse(sheet, header=None)
        
                # Detect header row dynamically
                header_row = None
                for i, row in df_temp.iterrows():
                    if row.notna().sum() >= 2:
                        header_row = i
                        break
        
                if header_row is not None:
                    dfs[sheet] = excel.parse(sheet, header=header_row)
                else:
                    dfs[sheet] = excel.parse(sheet)
        
            except Exception as e:
                print(f"[ERROR] ❗ Failed to load sheet '{sheet}' in '{filepath}': {e}")
 
    except Exception as e:
        print(f"[ERROR] ❗ Failed to open Excel file from S3 '{filepath}': {e}")
        
        
    # print("DFFFS",dfs)
 
    return dfs

# 892547
def user_input_file_checks(master_folder: str, esnno):
    engine_family, esn = get_engine_details(esnno)
 
    # 🔹 Pass ESN to checker
    file_dict = check_required_files(master_folder, esn)
 
    all_dataframes = {}
 
    for key, path in file_dict.items():
        if path:
            dfs = load_excel_sheets(path)
            all_dataframes[key] = dfs
        else:
            print(f"[INFO] Skipping loading for missing file: {EXPECTED_FILES[key]}")
 
    return {
        "engine_family": engine_family,
        "esn": esn,
        "file_paths": file_dict,
        "dataframes": all_dataframes
    }

# User Input
def get_engine_details(esnno) -> tuple[str, str]:
    engine_family = "CFM56" # RESTRICTING TO CFM56 ENGINE FAMILY
    esn = esnno #input("Enter ESN (Engine Serial Number): ").strip() # SELECTION OF THE ESN HAPPENS HERE = = = = =
    return engine_family.upper(), esn


def get_dataframe_by_file_and_sheet(all_dataframes: dict[str, dict[str, pd.DataFrame]],
                                    file_key: str,
                                    sheet_name_substring: str) -> pd.DataFrame:
    """
    Smarter sheet picker:
    1) exact name match (case/space/punct-insensitive)
    2) exact token match (same words, any order)
    3) whole-word subset (prefer fewest extra words)
    4) substring (old behavior, lowest penalty)
    5) fallback: first sheet
    """
 
    # --- helpers (local, so this works even if you don't have a global normalize_string) ---
    import re
    def _norm(s: str) -> str:
        # lower, replace non-alnum with spaces, collapse spaces
        s = re.sub(r"[\W_]+", " ", str(s).lower())
        return " ".join(s.split())
 
    def _tokens(s: str) -> list[str]:
        n = _norm(s)
        return n.split() if n else [] #Saini Shubham.xlsx: Saini, Shubham
 
    # --- basic checks ---
    if file_key not in all_dataframes:
        raise ValueError(f"[ERROR] ❌ File key '{file_key}' not loaded or does not exist in all_dataframes.")
 
    file_sheets = all_dataframes[file_key]
    if not file_sheets:
        raise ValueError(f"[ERROR] ❌ File '{file_key}' has no loaded sheets. Check if the Excel file is empty, corrupted, or unreadable.")
 
    target_norm = _norm(sheet_name_substring)
    target_tok  = _tokens(sheet_name_substring)
 
    # Fast path: if user gave nothing meaningful, fallback to first
    if not target_norm:
        first_sheet = next(iter(file_sheets), None) #billing request.xlsx --> billing
        if first_sheet:
            return file_sheets[first_sheet]
        raise ValueError(f"[ERROR] ❌ No sheets available to fallback in file '{file_key}'.")
 
    # Build candidates with scores
    exact_name = []
    exact_token = []   # same set of tokens
    subset = []        # target tokens ⊆ sheet tokens; track extra words count
    substr = []        # plain substring match on normalized
 
    for sheet in file_sheets.keys():
        s_norm = _norm(sheet)
        s_tok  = _tokens(sheet)
        s_tok_set = set(s_tok)
        tgt_set   = set(target_tok)
 
        if s_norm == target_norm: #"T&M Billing Timeliness" --> tM billing timeliness
            exact_name.append(sheet)
            continue
 
        if s_tok_set == tgt_set:
            # exactly the same words (order may differ)
            exact_token.append((sheet, len(s_tok)))
            continue
 
        if tgt_set.issubset(s_tok_set):
            # all target words appear as whole words in the sheet name
            extra = len(s_tok) - len(target_tok)
            subset.append((sheet, extra, len(s_tok)))
            continue
 
        if target_norm and target_norm in s_norm:
            # fallback, penalize by length difference
            penalty = abs(len(s_norm) - len(target_norm))
            substr.append((sheet, penalty, len(s_norm)))
 
    # Decide best match by priority and tie-breakers
    if exact_name:
        # exact normalized name match
        return file_sheets[exact_name[0]]
 
    if exact_token:
        # prefer the one with fewest total words (tighter)
        exact_token.sort(key=lambda x: (x[1], x[0]))
        return file_sheets[exact_token[0][0]]
 
    if subset:
        # prefer fewest extra words; tie-breaker: fewer total words, then name
        subset.sort(key=lambda x: (x[1], x[2], x[0]))
        return file_sheets[subset[0][0]]
 
    if substr:
        # prefer closest length; tie-breaker: shorter name, then name
        substr.sort(key=lambda x: (x[1], x[2], x[0]))
        return file_sheets[substr[0][0]]
 
    # Final fallback: first sheet
    fallback_sheet = next(iter(file_sheets), None)
    if fallback_sheet:
        return file_sheets[fallback_sheet]
 
    raise ValueError(f"[ERROR] ❌ No sheets available to fallback in file '{file_key}'.")


def find_first_data_row_in_df(df: pd.DataFrame, keyword: str) -> pd.DataFrame:
    if df.empty:
        print("[WARNING] Dataframe is empty. Defaulting to row 0.")
        return 0
    for idx, row in df.iterrows():
        if row.astype(str).str.lower().str.contains(keyword.lower(), na=False).any():
            new_df = df.iloc[idx:].reset_index(drop=True)
            new_df.columns = new_df.iloc[0]
            final_df = new_df[1:].reset_index(drop=True)
            return final_df
    print(f"[WARNING] Not found keyword '{keyword}'. Defaulting to row 0.")
    return df
 
# Extract Q_YEAR from timeliness tracker - timeliness sheet
def extract_q_year_from_timeliness(result: dict) -> tuple[str, datetime, str, str, str]:
    esn = result["esn"].strip().upper()
    engine_family = result["engine_family"].strip().upper()
 
    df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="timeliness",
        sheet_name_substring="timeliness"
    )
 
    # Normalize columns
    df.columns = [col.strip().lower() for col in df.columns]
 
    # Identify columns
    eng_col = next((col for col in df.columns if "engine" in col and "family" in col), None)
    esn_col = next((col for col in df.columns if "esn" in col), None)
    induction_col = next((col for col in df.columns if "induction" in col and "date" in col), None)
    g3_col = next((col for col in df.columns if "g3" in col and "close" in col), None)
    cust_col = next((col for col in df.columns if "bill to customer name" in col), None)
 
    if not all([eng_col, esn_col, induction_col, g3_col]):
        raise ValueError("Required columns not found in Timeliness Tracker sheet.")
 
    # --------------------------------------------------
    # Filter by Engine + ESN
    # --------------------------------------------------
    filtered = df[
        (df[eng_col].astype(str).str.strip().str.upper() == engine_family) &
        (df[esn_col].astype(str).str.strip().str.upper() == esn)
    ].copy()
 
    if filtered.empty:
        raise ValueError(f"No matching record found for ESN {esn} and Engine {engine_family}")
 
    # --------------------------------------------------
    # Convert induction column to datetime safely
    # --------------------------------------------------
    filtered[induction_col] = pd.to_datetime(filtered[induction_col], errors="coerce")
 
    # Drop rows where induction date is invalid
    filtered = filtered.dropna(subset=[induction_col])
 
    if filtered.empty:
        raise ValueError(f"No valid induction dates found for ESN {esn}")
 
    # --------------------------------------------------
    # Take row with latest induction date
    # --------------------------------------------------
    latest_row = filtered.loc[filtered[induction_col].idxmax()]
 
    induction_date = latest_row[induction_col]
 
    # Existing Q_YEAR logic (from induction)
    quarter = (induction_date.month - 1) // 3 + 1
    q_year = f"Q{quarter}_{induction_date.year}"
 
    
    # --------------------------------------------------
    # Get Contract Month from Term Sheet
    # --------------------------------------------------
    term_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="term sheet"
    )
    
    term_df.columns = [col.strip().lower() for col in term_df.columns]
    print(term_df.columns)
    
    contract_month_col = next((c for c in term_df.columns if "contract" in c and "month" in c), None)
    
    if not contract_month_col:
        raise ValueError("Contract Month column not found in Term Sheet")
    
    contract_month_raw = term_df[contract_month_col].dropna().iloc[0]

    import calendar
 
    if isinstance(contract_month_raw, str):
        contract_month = list(calendar.month_name).index(contract_month_raw.capitalize()) \
            if contract_month_raw.capitalize() in calendar.month_name else \
            list(calendar.month_abbr).index(contract_month_raw.capitalize())
    else:
        contract_month = int(contract_month_raw)


    # G3 Year logic
    g3_date = pd.to_datetime(latest_row[g3_col], errors="coerce")
    
    if pd.isna(g3_date):
        raise ValueError(f"G3 date missing for ESN {esn}")
    
    # --------------------------------------------------
    # Apply Contract Month Rule
    # --------------------------------------------------
    if g3_date.month >= contract_month:
        g3_date = pd.Timestamp(year=g3_date.year + 1, month=1, day=1)

    q_year_g3 = str(g3_date.year)
 
    # For customer logic use only the selected row
    customer_code, customer_first_word = get_customer_code_from_filtered_df(
        filtered.loc[[latest_row.name]]
    )
 
    print("Selected Latest Induction Row:")
    print("Induction Date:", induction_date)
    print("Q_YEAR (Induction):", q_year)
    print("Q_YEAR_G3 (Year Only):", q_year_g3)
 
    return q_year, induction_date, customer_code, customer_first_word, q_year_g3


def delete_files_in_folder(folder_path: str, pattern: str = "*", recursive: bool = False):
    """
    Delete files in a folder.
 
    Args:
        folder_path: Folder to clean.
        pattern: Glob pattern for files to delete (e.g., "*.tmp", "*").
        recursive: If True, also delete matching files in subfolders.
 
    Returns:
        (deleted, failed): lists of filenames and (filename, error) tuples.
    """
    p = Path(folder_path)
    if not p.is_dir():
        raise NotADirectoryError(f"Not a directory: {folder_path}")
 
    deleted, failed = [], []
    it = p.rglob(pattern) if recursive else p.glob(pattern)
 
    for path in it:
        if path.is_file():
            try:
                path.unlink()
                deleted.append(str(path))
            except Exception as e:
                failed.append((str(path), str(e)))
 
    return deleted, failed
 
 
def get_latest_clp_file(master_folder: str, q_year: str) -> str:
    """
    Now selects Parts Catalog file based only on YEAR (q_year_g3).
    No quarter logic. No induction date comparison.
    """
 
    # Build S3 folder path
    catalog_folder = f"{master_folder}/Catalog".strip("/")
    print("Catalog Path:", catalog_folder)
 
    # List all files under folder_path in S3
    all_files_s3 = list_files_in_s3(catalog_folder)
 
    # Extract filenames only
    all_files = [
        f[len(catalog_folder)+1:] if f.startswith(catalog_folder + "/") else f
        for f in all_files_s3
    ]
 
    print("All filenames:", all_files)
 
    EXTS = (".xlsx", ".xlsm", ".xls")
    target_year = q_year.strip()
 
    candidates = []
 
    for f in all_files:
        fl = f.lower()
 
        if not fl.endswith(EXTS):
            continue
 
        # Match pattern: "2024 - Parts Catalog.xlsx"
        if target_year in f and "parts" in fl and "catalog" in fl:
            candidates.append(f)
 
    if not candidates:
        raise FileNotFoundError(
            f"No Parts Catalog file found for year {q_year} in {catalog_folder}"
        )
 
    # If multiple matches exist, choose latest alphabetically
    chosen = sorted(candidates)[-1]
 
    print(f"[INFO] ✅ Selected Parts Catalog file: {chosen}")
 
    return f"{catalog_folder}/{chosen}"

# Merge columns in CLP
def create_pn_clp_lookup_column(clp_df: pd.DataFrame) -> pd.DataFrame:
    df = clp_df.copy()
    print("Original columns:", df.columns)
 
    df.columns = [col.strip().lower() for col in df.columns]
 
    # Step 1: Detect Part Number column
    part_col = next((col for col in df.columns if "part" in col and "number" in col), None)
 
    # Step 2: Detect CLP column (price column)
    clp_col = next((col for col in df.columns if "price" in col), None)
 
    if not part_col or not clp_col:
        raise ValueError("Could not detect both 'part number' and 'CLP' columns.")
 
    print(f"[INFO] Part Column Detected: {part_col}")
    print(f"[INFO] CLP Column Detected: {clp_col}")
 
    # ✅ Convert to numeric safely
    df[clp_col] = pd.to_numeric(df[clp_col], errors="coerce")
 
    # Optional: Remove rows where CLP is not numeric
    df = df.dropna(subset=[clp_col])
 
    # Keep only required columns
    df = df[[part_col, clp_col]]
 
    print("[INFO] Lookup dataframe created successfully")
 
    return df


# Merge columns in Billing request
def merge_billing_with_clp(result: dict, clp_df: pd.DataFrame, pipeline_type) -> pd.DataFrame:
 
    billing_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="billing",
        sheet_name_substring="Sheet1"
    )
 
    billing = billing_df.copy()
    clp = clp_df.copy()
 
    # Normalize columns
    billing.columns = [str(col).strip().lower() for col in billing.columns]
    clp.columns = [col.strip().lower() for col in clp.columns]
 
    # ---------------------------
    # Ensure pn_clp_lookup exists
    # ---------------------------
    if "price" in clp.columns:
        clp["pn_clp_lookup"] = pd.to_numeric(clp["price"], errors="coerce")
    else:
        clp["pn_clp_lookup"] = 0.0
 
    if pipeline_type.lower() == "vendor":
        clp["pn_clp_lookup"] = 0.0
 
    # Detect columns
    material_col = next((col for col in billing.columns if "material" in col or "part" in col), None)
    unit_price_col = next((col for col in billing.columns if col.strip() == "cost"), None)
    qty_col = next((c for c in billing.columns if "target" in c and "quantity" in c), None)
    part_col_clp = next((col for col in clp.columns if "part" in col or "desc" in col or "description" in col), None)
 
    if not material_col or not part_col_clp:
        raise ValueError("Required columns not found for merge.")
 
    billing[material_col] = billing[material_col].astype(str)
    clp[part_col_clp] = clp[part_col_clp].astype(str)
 
    merged_df1 = billing.merge(
        clp[[part_col_clp, "pn_clp_lookup"]],
        left_on=material_col,
        right_on=part_col_clp,
        how="left"
    )
 
    # ----------------------------------------
    # Guarantee pn_clp_lookup exists ALWAYS
    # ----------------------------------------
    if "pn_clp_lookup" not in merged_df1.columns:
        merged_df1["pn_clp_lookup"] = 0.0
 
    # Compute matl_clp_lookup
    merged_df1["matl_clp_lookup"] = merged_df1["pn_clp_lookup"]
 
    fallback = merged_df1.apply(
        lambda r: (r[unit_price_col] / r[qty_col])
        if pd.notna(r[qty_col]) and r[qty_col] not in [0, None]
        else r[unit_price_col],
        axis=1
    )
 
    merged_df1["matl_clp_lookup"] = merged_df1["matl_clp_lookup"].where(
        merged_df1["matl_clp_lookup"].notna() & merged_df1["matl_clp_lookup"].ne(0),
        fallback
    )
 
    # Final safeguard
    if "pn_clp_lookup" not in merged_df1.columns:
        merged_df1["pn_clp_lookup"] = 0.0
 
    return merged_df1

# Compute Extended Price & replace 'Blank' values with Price column
def compute_extended_price(billing_df: pd.DataFrame, pipeline_type) -> pd.DataFrame:
    df = billing_df.copy()
    df.columns = [col.strip().lower() for col in df.columns]
 
    # Detect necessary columns
    clp_lookup_col = next((col for col in df.columns if "matl_clp_lookup" in col), None)
    qty_col = next((col for col in df.columns if "target" in col and "quantity" in col), None)
    unit_price_col = next((col for col in df.columns if col.strip().lower() == "cost"), None)
    # unit_price_col = next((col for col in df.columns if "cost" in col), None)

    # if pipeline_type == "vendor":
    #     if not clp_lookup_col or not qty_col or not unit_price_col:
    #         raise ValueError("One or more required columns (matl_clp_lookup, target quantity, cost) are missing.")
    # else:
    if not clp_lookup_col or not qty_col or not unit_price_col:
        raise ValueError("One or more required columns (matl_clp_lookup, target quantity, cost) are missing.")

    # Compute extended price
    df["matl_clp_lookup"] = pd.to_numeric(df['matl_clp_lookup'], errors='coerce')
    df[qty_col] = pd.to_numeric(df[qty_col], errors='coerce')
    
    df["extended_price"] = df["matl_clp_lookup"] * df[qty_col]
    df["extended_price"] = pd.to_numeric(df.get("extended_price"), errors="coerce")

    # New logicccccccccccccccccccccc
    computed_from_cost = (
        (df[unit_price_col] / df[qty_col])
        if (unit_price_col and qty_col)
        else np.nan 
    )

    if unit_price_col and qty_col:
        df[unit_price_col] = pd.to_numeric(df[unit_price_col], errors="coerce")
        df[qty_col] = pd.to_numeric(df[qty_col], errors = "coerce")
        computed_from_cost = np.where(
            df[unit_price_col].notna() & df[qty_col].notna() & (df[qty_col] != 0), 
            df[unit_price_col] / df[qty_col], 
            np.nan
        )

    df["extended_price"] = np.where(
        df["extended_price"].notna() & (df["extended_price"] != 0),
        df["extended_price"],
        np.where(
            ~pd.isna(computed_from_cost),
            computed_from_cost,
            pd.to_numeric(df[unit_price_col], errors = "coerce")
        )
    )

    return df
 
# Extract LLP & NON-LLP Cost Categories from the Contract and flag them
def apply_cost_category_flags(result: dict, billing_df: pd.DataFrame) -> pd.DataFrame:
    """
    Assigns cost category flags (LLP_CC, NON_LLP_CC, CUST_FUR_MATL_CC, MATL_CC)
    to billing_df based on contract cost categories. Includes a second-pass override
    for stubborn categories like SCC-CONSUMABLE-AERO.
    """
 
    # Step 1: Load Cost Category sheet from Contract Summary
    contract_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="cost category"
    )

    contract_df.columns = [col.strip().lower() for col in contract_df.columns]
    # print("sheet_name_substring=cost category : ", contract_df.columns)
 
    # Step 2: Extract all cost category values into a list
    cc_col = next((col for col in contract_df.columns if "cost" in col and "category" in col), None)

    if not cc_col:
        raise ValueError("Could not locate 'Cost Category' column in contract summary.")

    cost_categories = contract_df[cc_col].dropna().astype(str).str.strip().tolist()
    print("cost categories",cost_categories)
    
 
    # --- Normalization helper ---
    def normalize_cc(cc: str) -> str:
        if cc is None:
            return ""
        return (
            str(cc)
            .replace("–", "-").replace("—", "-")   # normalize dashes
            .replace("\u00A0", " ")                # non-breaking space
            .strip()
        )
 
    # Step 3: Build sets
    LLP_CC = [
        cc for cc in cost_categories
        if any(tag in cc for tag in [
            "SCC-LLP-NEW-GE",
            "SCC-LLP-USED-LP",
            "SCC-LLP-USED-GE"
        ])
    ]

    NON_LLP_CC = [
        cc for cc in cost_categories
        if any(tag in cc for tag in [
            "SCC-MATL-NEW-LP",
            "SCC-MATL-NEW-GE",
            "SCC-MATL-NEW-LP-MOR",
            "SCC-MATL-NEW-GE-MOR",
            "SCC-MATL-NEW-SCRAPREP",
            "SCC-CONSUMABLE-AERO",
            "SCC-CONSUMABLE-NON-AERO"
        ])
    ]

    CUST_FUR_MATL_CC = [
        cc for cc in cost_categories
        if any(tag in cc for tag in ["SCC-MATL-CUSTOMER", "SCC-MATL-CUST-SCRAPREP"])
    ]

    MATL_CC = [
        cc for cc in cost_categories
        if any(tag in cc for tag in ["SCC-MATL-USED-LP", "SCC-MATL-MLIP", "SCC-MATL-USED-GE"]) # "SCC-LLP-USED-LP",
    ]
 
    LLP_CC_set = {normalize_cc(cc) for cc in LLP_CC}
    NON_LLP_CC_set = {normalize_cc(cc) for cc in NON_LLP_CC}
    CUST_FUR_MATL_CC_set = {normalize_cc(cc) for cc in CUST_FUR_MATL_CC}
    MATL_CC_set = {normalize_cc(cc) for cc in MATL_CC}
 
    # Step 4: Build dictionary lookup
    category_map = {}
    category_map.update({cc: "LLP_CC" for cc in LLP_CC_set})
    category_map.update({cc: "NON_LLP_CC" for cc in NON_LLP_CC_set})
    category_map.update({cc: "CUST_FUR_MATL_CC" for cc in CUST_FUR_MATL_CC_set})
    category_map.update({cc: "MATL_CC" for cc in MATL_CC_set})    

    # print("NON_LLP_CC_set : ", NON_LLP_CC_set)

    # Step 5: Prepare billing df
    df = billing_df.copy()
    df.columns = [col.strip().lower() for col in df.columns]
    cost_cat_col = next((col for col in df.columns if "cost" in col and "category" in col), None)

    if not cost_cat_col:
        raise ValueError("Could not find 'Cost Category' column in billing data.")
 
    # First pass assignment
    df["cc_flag"] = df[cost_cat_col].map(lambda x: category_map.get(normalize_cc(x), "UNKNOWN"))
 
    # --- Step 6: Second pass strict corrections ---
    df["_cc_norm"] = df[cost_cat_col].astype(str).map(normalize_cc)
 
    forced_map = {
        "SCC-CONSUMABLE-AERO": "NON_LLP_CC",
        "SCC-CONSUMABEL-AERO": "NON_LLP_CC",
    }
 
    df["cc_flag"] = df.apply(
        lambda row: forced_map.get(row["_cc_norm"], row["cc_flag"]),
        axis=1
    )
    
    df.drop(columns=["_cc_norm"], inplace=True)
 
    # Step 7: Filter out unknowns
    removed_unknown_cc_df = df[df["cc_flag"] != "UNKNOWN"].copy()
    return removed_unknown_cc_df #WLogicccccccc


# Assign Discounts for LLP & NON-LLP - Cost categories from the contract
def apply_discount_from_contract(result: dict, billing_df: pd.DataFrame) -> pd.DataFrame:
    # Step 1: Load Discount sheet
    discount_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="discount"
    )
 
    discount_df.columns = [col.strip().lower() for col in discount_df.columns]

    # print("sheet_name_substring=discount", discount_df.columns)
 
    # Step 2: Identify columns
    cost_cat_col = next((col for col in discount_df.columns if "cost" in col and "category" in col), None)
    discount_col = next((col for col in discount_df.columns if "discount" in col), None)
 
    if not cost_cat_col or not discount_col:
        raise ValueError("Could not detect required columns in Discount sheet.")
 
    # Step 3: Create a discount mapping dictionary
    discount_mapping = dict(
        zip(discount_df[cost_cat_col].astype(str).str.strip(), discount_df[discount_col])
    )
 
    # Step 4: Apply mapping to billing DataFrame based on actual Cost Category value
    df = billing_df.copy()
    df.columns = [col.strip().lower() for col in df.columns]
    billing_cost_col = next((col for col in df.columns if "cost" in col and "category" in col), None)
 
    if not billing_cost_col:
        raise ValueError("Could not find 'Cost Category' in billing data.")
 
    df["discount"] = df[billing_cost_col].apply(
        lambda x: discount_mapping.get(str(x).strip(), 0) # UNKNOWN
    )
 
    # print("[INFO] ✅ Discount values assigned from Contract Summary.")
    # df.to_excel("check_discounts.xlsx", index=False)
    return df

# Compute Discount Amount (Actual discount value)
def compute_discount_amount(df: pd.DataFrame) -> pd.DataFrame:
    billing = df.copy()
    billing.columns = [col.strip().lower() for col in billing.columns]
 
    # Detect relevant columns
    clp_lookup_col = next((col for col in billing.columns if "matl_clp_lookup" in col), None)
    qty_col = next((col for col in billing.columns if "target" in col and "quantity" in col), None)
    discount_rate_col = next((col for col in billing.columns if col == "discount"), None)
 
    if not clp_lookup_col or not qty_col or not discount_rate_col:
        raise ValueError("Required columns (matl_clp_lookup, qty, discount) not found.")
 
    # Step 2: Convert Discount % (if it's a string like "10%")
    def parse_discount(val):
        if isinstance(val, str) and "%" in val:
            return float(val.replace("%", "").strip()) / 100
        try:
            return float(val)
        except:
            return 0.0
 
    billing["discount_rate_numeric"] = billing[discount_rate_col].apply(parse_discount)
 
    # Step 3: Calculate Discount Amount
    billing["discount_amount"] = (
        billing["matl_clp_lookup"] * billing[qty_col] * billing["discount_rate_numeric"] #0.2, 0.02
    )
 
    # print("[INFO] ✅ Computed 'discount_amount' using MATL_CLP_LOOKUP * Target Quantity * Discount Rate.")
    return billing
 

def apply_handling_fees(result: dict, billing_df: pd.DataFrame) -> pd.DataFrame:
    q_year, induction_date, cust_name, customer_first_word,q_year_g3 = extract_q_year_from_timeliness(result)
 
    # 1) Load handling fee table (from contract)
    handling_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="handling"
    )

 
    # Normalize columns
    handling_df.columns = [col.strip().lower() for col in handling_df.columns]
    # print("sheet_name_substring=handling ", handling_df.columns)

    billing = billing_df.copy()
    billing.columns = [col.strip().lower() for col in billing.columns]
 
    # Detect columns
    cost_cat_col      = next((c for c in handling_df.columns if "cost" in c and "category" in c), None)
    fee_col           = next((c for c in handling_df.columns if "handling" in c and "fee" in c), None)
    max_part_col      = next((c for c in handling_df.columns if "max" in c and "per" in c and "part" in c), None)
    max_set_col       = next((c for c in handling_df.columns if "max" in c and "per" in c and "set" in c), None)
    billing_cost_cat_col = next((c for c in billing.columns     if "cost" in c and "category" in c), None)
 
    # Revised logic columns
    clp_col   = next((c for c in billing.columns if "matl_clp_lookup" in c), None)
    cost_col = next((c for c in billing.columns if c.strip().lower() == "cost"), None)
    #cost_col  = next((c for c in billing.columns if "cost" in c), None)  # optional, used only to backfill CLP
    qty_col   = next((c for c in billing.columns if "target" in c and "quantity" in c), None)
    part_keyword_col = next((c for c in billing.columns if "material" in c), None)
 
    if not all([cost_cat_col, fee_col, max_part_col, max_set_col,
                billing_cost_cat_col, clp_col, qty_col, part_keyword_col]):
        raise ValueError("Required columns not found for handling fee logic.")
 
    cols_to_remove = [fee_col, max_part_col, max_set_col]
    cols_to_remove = [c for c in cols_to_remove if c and c in billing.columns]

    if cols_to_remove:
        billing.drop(columns = cols_to_remove, inplace = True)
 
    # 2) Merge contract fee setup onto billing lines
    billing = billing.merge(
        handling_df[[cost_cat_col, fee_col, max_part_col, max_set_col]],
        left_on=billing_cost_cat_col,
        right_on=cost_cat_col,
        how='left'
    )
 
    # 3) Numeric coercion for safe math
    for c in [clp_col, qty_col, fee_col]:
        billing[c] = pd.to_numeric(billing[c], errors="coerce")
    if cost_col and cost_col in billing.columns:
        billing[cost_col] = pd.to_numeric(billing[cost_col], errors="coerce")
 
    # 4) Backfill CLP: if CLP is 0/NaN and we have cost & qty, set CLP = cost / qty
    if cost_col and cost_col in billing.columns:
        needs_fill  = billing[clp_col].isna() | (billing[clp_col] == 0)
        can_compute = billing[cost_col].notna() & billing[qty_col].notna() & (billing[qty_col] != 0)
        fill_mask   = needs_fill & can_compute
        billing.loc[fill_mask, clp_col] = (
            billing.loc[fill_mask, cost_col] / billing.loc[fill_mask, qty_col]
        )
 
    # 5) Compute handling fee = CLP * QTY * FEE%
    billing["handling_fee"] = (
        billing[clp_col].fillna(0) *
        billing[qty_col].fillna(0) *
        billing[fee_col].fillna(0)
    )
 
    # 6) Apply caps (from escalation)
    escalation_results = get_escalation_values(result, induction_date)
    cfe = escalation_results["cfe"]  # [max_per_part, max_per_set]
 
    # Cap per part (qty == 1)
    billing.loc[billing[qty_col] == 1, "handling_fee"] = billing.loc[
        billing[qty_col] == 1, "handling_fee"
    ].clip(upper=float(cfe[0]))
 
    # Cap per set (distribute proportionally if above cap)
    def cap_per_set(df, part_keyword_col, qty_col, max_per_set):
        df = df.copy()
        grouped = df.groupby(part_keyword_col)
        for name, group in grouped:
            total_qty = group[qty_col].sum() #160
            total_fee = group["handling_fee"].sum() #0
            if pd.notna(max_per_set) and total_fee > max_per_set and total_qty not in [0, None]:
                # 4567 & 0 > 4567 and 160 
                adjusted_fee = (group[qty_col] / total_qty * max_per_set).round(2)
                df.loc[group.index, "handling_fee"] = adjusted_fee
        return df
 
    handling_fee_final_df = cap_per_set(
        df=billing,
        part_keyword_col=part_keyword_col,
        qty_col=qty_col,
        max_per_set=float(cfe[1])
    )
    return handling_fee_final_df
  

# Applies SCC-LLP-USED-LLP logic using data from a template file.
def apply_scc_llp_logic(main_df: pd.DataFrame, used_llp_df: pd.DataFrame, cost_col) -> pd.DataFrame:
    # print("[INFO] Applying SCC-LLP-USED-LLP Logic...")
    
    # Normalize column names
    used_llp_df.columns = [col.strip().lower() for col in used_llp_df.columns]
    main_df.columns = [col.strip().lower() for col in main_df.columns]

    # Merge the template values with the main DataFrame
    enriched_df = main_df.merge(used_llp_df, left_on="material", right_on="part_number", how="left")
    
    # Compute the Pro-Rata factor
    enriched_df["pro_rata_factor"] = (
        enriched_df["total_remaining_cicles"].astype(float) / enriched_df["total_life_cicles"].astype(float)
    ) * (enriched_df["pro-rata %"].astype(float) / 100)
    
    # Apply the calculated Pro-Rata factor
    enriched_df.loc[enriched_df["cost category"].str.upper() == "SCC-LLP-USED-LP", "total_price"] = (
        enriched_df["pro_rata_factor"] * enriched_df["matl_clp_lookup"]
    ).round(2)

    # 3. Fallback: If total_price is still null or zero → use 'cost' if available
    if cost_col:
        enriched_df["total_price"] = enriched_df["total_price"].where(
            enriched_df["total_price"].notna() & (enriched_df['total_price'] !=0),
            enriched_df[cost_col]
        )
    return enriched_df


def apply_min_pricing_for_matl_used(billing: pd.DataFrame) -> pd.DataFrame:
    """
    CAT-1 / CAT-2 (MATL-USED) logic:
      - If Unit CLP exists and > 0:
            chosen = min(Cost, Qty * Unit CLP * 70%)
      - If Unit CLP missing or zero:
            chosen = Cost (NO comparison)
    """
 
    df = billing.copy()
    df.columns = [c.strip().lower() for c in df.columns]
 
    # --- SAME column detection as your original function ---
    cost_cat_col = next((c for c in df.columns if "cost" in c and "category" in c), None)
 
    qty_col = next(
        (c for c in df.columns if c in ("qty", "quantity") or ("target" in c and "quantity" in c)),
        None
    )
 
    cost_col = next(
        (c for c in df.columns if c.strip().lower() == "cost"),
        None
    )
 
    matl_clp_col = next((c for c in df.columns if "matl_clp_lookup" in c), None)
 
    if matl_clp_col is None:
        matl_clp_col = next((c for c in df.columns if "clp" in c and "unit" not in c), None)
 
    unit_clp_col = next(
        (c for c in df.columns if ("unit" in c and "clp" in c) or ("clp unit" in c)),
        None
    )
 
    if not all([cost_cat_col, qty_col, cost_col, matl_clp_col]):
        print("❌ Detected columns:")
        print("cost_cat_col :", cost_cat_col)
        print("qty_col      :", qty_col)
        print("cost_col     :", cost_col)
        print("matl_clp_col :", matl_clp_col)
        raise ValueError("Missing required columns")
 
    # --- normalize cost category (unchanged) ---
    def norm(s: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", str(s).lower())
 
    allowed = {
        "sccmatlusedge",
        "sccmatlusedlp",
        "matlusedge",
        "matlusedlp",
    }
 
    mask = df[cost_cat_col].apply(norm).isin(allowed)
    if not mask.any():
        return df
 
    # --- numeric extraction ---
    qty  = pd.to_numeric(df.loc[mask, qty_col], errors="coerce").fillna(0)
    cost = pd.to_numeric(df.loc[mask, cost_col], errors="coerce").fillna(0)
 
    if unit_clp_col and unit_clp_col in df.columns:
        unit_clp = pd.to_numeric(df.loc[mask, unit_clp_col], errors="coerce").fillna(0)
    else:
        unit_clp = pd.Series(0, index=df.loc[mask].index)
 
    # --- BUSINESS RULE FIX ---
    cap_total = qty * unit_clp * 0.70
 
    # If Unit CLP > 0 → compare
    # Else → directly take Cost
    chosen = np.where(
        unit_clp > 0,
        np.minimum(cost, cap_total),
        cost
    )
 
    # --- write back ---
    df.loc[mask, matl_clp_col] = chosen
 
    # --- optional audit ---
    df.loc[mask, "_cap_total_70"] = cap_total
    df.loc[mask, "_final_selected"] = chosen
 
    # --- print & save ---
    print("\n===== FINAL DF AFTER MATL-USED PRICING =====")
    print(df)
 
    output_file = "matl_used_pricing_final.xlsx"
    df.to_excel(output_file, index=False)
    print(f"\n✅ Saved output to {output_file}")
 
    return df



def apply_matl_new_lp_pricing(
    df: pd.DataFrame,
    discount_rate: float | None = None,  # if provided, recompute for CLP rows; else keep existing
    ) -> pd.DataFrame:

    """
    For rows where Cost Category ~ 'MATL-NEW-LP':
      - If CLP > 0  -> keep existing discount (or recompute if discount_rate provided)
      - If CLP <= 0 -> set discount to 0
    Leaves other rows unchanged.
    """

    out = df.copy()
    out.columns = [c.strip().lower() for c in out.columns]
 
    # detect columns flexibly
    cost_cat_col = next((c for c in out.columns if "cost" in c and "category" in c), None)
    disc_col     = next((c for c in out.columns if "discount" in c and "amount" in c), None) \
        or next((c for c in out.columns if c.strip() == "discount"), None)

    qty_col      = next((c for c in out.columns if c in ("qty","quantity") or ("target" in c and "quantity" in c)), None)

    # prefer explicit CLP lookup; fallback to any 'clp' column
    clp_col = next((c for c in out.columns if "matl_clp_lookup" in c), None) \
        or next((c for c in out.columns if "clp" in c and "unit" not in c), None)
 
    if not all([cost_cat_col, disc_col, clp_col]):
        raise ValueError("Missing required columns: cost category, discount(/_amount), and CLP lookup.")
 
    # mask: MATL-NEW-LP (lenient to separators/case)
    mnlp_mask = out[cost_cat_col].astype(str).str.contains(
        r'\bmatl[-_\s]*new[-_\s]*lp\b', flags=re.I, regex=True, na=False
    )

    if not mnlp_mask.any():
        return out
 
    # CLP present?
    clp_vals = pd.to_numeric(out.loc[mnlp_mask, clp_col], errors="coerce").fillna(0)
    clp_used = clp_vals.gt(0)
 
    # Optionally recompute discount for CLP rows
    if discount_rate is not None:
        if not qty_col:
            raise ValueError("qty/quantity column needed to recompute discount.")

        qty_vals = pd.to_numeric(out.loc[mnlp_mask, qty_col], errors="coerce").fillna(0)
        recomputed = (qty_vals * clp_vals * float(discount_rate)).fillna(0)
        out.loc[mnlp_mask & clp_used, disc_col] = recomputed.values
 
    # Always zero out discount where CLP not used
    out.loc[mnlp_mask & ~clp_used, disc_col] = 0
    return out
 
# Apply the above logi


def flag_discount(df: pd.DataFrame) -> pd.DataFrame:
    """
    Flags rows based on cost-category discount rules:
      1. If cost category contains 'MATL-NEW' and matl_clp_lookup == unit_clp → 'Applied Discount'
      2. If cost category contains 'MATL-NEW' and matl_clp_lookup == cost → 'No Discount'
      3. Else → blank
    """
    # Create new column with default blank
    df["discount_flag"] = ""
 
    # Condition: filter only MATL-NEW rows
    mask_matl = df["cost category"].astype(str).str.contains("MATL-NEW", case=False, na=False)

    # cost == clp --> not mat - apply discount and not mathcing dont apply discount
 
    # Rule 1 → Applied Discount    
    df.loc[
        mask_matl & (df["pn_clp_lookup"] == 0),
        "discount_flag"
    ] = "No Discount"
 
    # Rule 2 → No Discount
    df.loc[
        mask_matl & ((df["pn_clp_lookup"].isna() | df["pn_clp_lookup"].astype(str).str.strip() == "")),
        "discount_flag"
    ] = "No Discount"

    # Rule 3 → DIsocunt 2
    df.loc[
        mask_matl & ((df["discount_flag"].isna() | df["discount_flag"].astype(str).str.strip() == "")),
        "discount_flag"
    ]  = "Discount"
 
    return df

def is_united_airlines(series: pd.Series) -> pd.Series:
    """
    Returns True for vendor names like:
    'UNITED AIRLINES', 'United Airlines Inc', 'UNITED AIRLINES, LLC', etc.
    """
    return (
        series
        .astype(str)
        .str.strip()
        .str.contains(r"AVIANCA\s+AIRLINES\b", case=False, na=False)
    )
    

def compute_total_price(df: pd.DataFrame, cust_name, esn, q_year, result, customer_first_word, pipeline_type) -> pd.DataFrame:

    billing = df.copy()
    billing.columns = [c.strip().lower() for c in billing.columns]
 
    # --- detect columns (robustly) ---
    material_col   = next((c for c in billing.columns if "material" in c), None)
    description_col= next((c for c in billing.columns if "description" in c), None)
    ext_price_col  = next((c for c in billing.columns if "extended_price" in c), None)
    handling_col   = next((c for c in billing.columns if "handling_fee" in c), None)
    contract_handling_fee = next((c for c in billing.columns if "handling fee" in c), None)
    discount_col   = next((c for c in billing.columns if "discount_amount" in c), None)
    cost_col       = next((c for c in billing.columns if c.strip().lower() == "cost"), None) #next((c for c in billing.columns if c == "cost"), None)
    clp_col        = next((c for c in billing.columns if "matl_clp_lookup" in c), None)
    cost_cat_col   = next((c for c in billing.columns if "cost" in c and "category" in c), None)
    ata_long_col   = next((c for c in billing.columns if "ata_long" in c), None)
    qty_col        = next((c for c in billing.columns if "target quantity" in c), None)
    vendor_col     = next((c for c in billing.columns if "vendor description" in c), None)
    
    if pipeline_type == "vendor":
        if not all([material_col, description_col, ext_price_col, cost_cat_col]):
            raise ValueError("Missing required columns for total price computation.")
    else:
        if not all([material_col, description_col, ext_price_col, handling_col, discount_col, clp_col, cost_cat_col]):
            raise ValueError("Missing required columns for total price computation.")
        
    # normalize merge keys
    billing[material_col] = billing[material_col].astype(str).str.strip()
    # billing[ata_long_col] = billing[ata_long_col].astype(str).str.strip()
 
    # 1) Base total price for everyone
    if pipeline_type == "vendor":
        handling_col = "handling_fee"
        discount_col = "discount_amount"
        contract_handling_fee = "handling fee"
        billing[handling_col] = 0
        billing[discount_col] = 0
        billing[contract_handling_fee] = 0

    #     billing["total_price"] = (
    #         pd.to_numeric(billing[ext_price_col], errors="coerce").fillna(0)
    #         + pd.to_numeric(billing[handling_col], errors="coerce").fillna(0)
    #         - pd.to_numeric(billing[discount_col], errors="coerce").fillna(0)
    #     )
    # else:
    #     billing["total_price"] = (
    #         pd.to_numeric(billing[ext_price_col], errors="coerce").fillna(0)
    #         + pd.to_numeric(billing[handling_col], errors="coerce").fillna(0)
    #         - pd.to_numeric(billing[discount_col], errors="coerce").fillna(0)
    #     )
    # - - - -  >
    # For CFE Part - turn off (zero-out) extended price and disount (SCC-MATL-CUSTOMER, SCC-MATL-CUST-SCRAPREP)
    # "SCC-MATL-CUSTOMER", "SCC-MATL-CUST-SCRAPREP -- > CFE Handling fee will be applied when vendor is "UNITED AIRLINES" (vendor description)

    cfe_cost_cat = {"SCC-MATL-CUSTOMER", "SCC-MATL-CUST-SCRAPREP","SCC-LLP-CUST-SCRAPREP"}
    is_cfe = billing[cost_cat_col].astype(str).str.upper().isin(cfe_cost_cat)

    # vendor description == "UNITED AIRLINES"
    # is_united = (
    #     billing[vendor_col]
    #     .astype(str).str.strip().str.upper()
    #     .eq("UNITED AIRLINES")
    # )
    
    is_united = is_united_airlines(billing[vendor_col])

    mask_apply_fee = is_cfe & is_united # applying cfe handling fee
    mask_no_united = is_cfe & (~is_united) # set extended price, discount == 0 (and handling fee = 0)
        
    billing.loc[mask_no_united, [ext_price_col, discount_col]] = 0

    # For UNITED AIRLINES CFE rows -> Compute handling fee
    qty = pd.to_numeric(billing.loc[mask_apply_fee, qty_col], errors="coerce").fillna(0)
    clp = pd.to_numeric(billing.loc[mask_apply_fee, clp_col], errors="coerce")
    ext = pd.to_numeric(billing.loc[mask_apply_fee, ext_price_col], errors = "coerce").fillna(0)
    fee_rate = pd.to_numeric(
        billing.loc[mask_apply_fee, contract_handling_fee], errors = "coerce"
    ).fillna(0)

    # Fallback unit price when CLP is missing/zero
    unit_from_ext = (ext / qty).where(qty > 0,0)  #(billing.loc[cfe_parts_df, ext_price_col] / qty).where(qty.notna() & (qty != 0), 0)
    unit_basis = clp.where((clp.notna()) & (clp > 0), unit_from_ext).fillna(0)  #clp.where(clp.notna() & (clp > 0), unit_from_ext)

    billing.loc[mask_apply_fee, handling_col] =  (unit_basis * qty * fee_rate).round(2)

    billing.to_excel("hf_check_ual.xlsx", index=False)

    # Pick Min value from Cost (Billing-Req) & (Qty * CLP * 70%) --> SCC-MATL-USED-LP, SCC-MATL-USED-GE
    out_df = apply_min_pricing_for_matl_used(billing)
    
    # ✅ FINAL total_price calculation (single source of truth)
    out_df["total_price"] = (
        pd.to_numeric(out_df[ext_price_col], errors="coerce").fillna(0)
        + pd.to_numeric(out_df[handling_col], errors="coerce").fillna(0)
        - pd.to_numeric(out_df[discount_col], errors="coerce").fillna(0)
    )
    
    # ✅ FORCE zero for CFE + non-United Airlines (business rule)
    mask_no_united_out = (
    out_df[cost_cat_col].astype(str).str.upper().isin(cfe_cost_cat)
    & ~is_united_airlines(out_df[vendor_col])
    )
    
    out_df.loc[mask_no_united_out, "total_price"] = 0
    
    out_df.to_excel("after_applying_min_pricing.xlsx", index=False)

    # Start clean: normalize any legacy "(Blanks)" etc.
    out_df["discount_flag"] = out_df.get("discount_flag", "")
    out_df["discount_flag"] = out_df["discount_flag"].astype(str).replace({"(Blanks)": ""}).str.strip()
    
    # 1) Category mask (rows we care about)
    mask_matl = (
        out_df["cost category"].astype(str).str.contains("MATL-NEW", case=False, na=False) |
        out_df["cost category"].astype(str).str.contains("SCC-CONSUMABEL-NON-AERO", case=False, na=False)
    )
    
    # 2) Work off the original column but normalize it
    pn_raw = out_df.get("pn_clp_lookup")
    
    # Ensure column exists
    if pn_raw is None:
        raise KeyError("Column 'pn_clp_lookup' not found in 'out'.")
    
    # Normalize whitespace & common placeholders
    pn_str = pn_raw.astype(str).str.replace("\u00A0", " ", regex=False).str.strip()  # remove NBSP, trim
    blank_like = pn_str.eq("") | pn_str.str.casefold().isin({"nan", "none", "null", "(blanks)"})
    
    # Numeric view for zero check (non-numeric -> NaN)
    pn_num = pd.to_numeric(pn_raw, errors="coerce")
    is_zero = pn_num.eq(0)
    
    # 3) Initialize flag for the target rows, then assign in one shot
    out_df.loc[mask_matl, "discount_flag"] = np.where(
        blank_like[mask_matl] | is_zero[mask_matl],
        "No Discount",
        "Discount"
    )

    out_df.to_excel("cost_clp_min_value.xlsx", index=False)
    
    return out_df

# Compute total price 
def compute_vendor_total_price(billing_df: pd.DataFrame) -> pd.DataFrame:
    df = billing_df.copy()
    df.columns = [col.strip().lower() for col in df.columns]
 
    # Detect necessary columns
    qty_col = next((col for col in df.columns if "target" in col and "quantity" in col), None)
    unit_price_col = next((c for c in df.columns if c.strip().lower() == "cost"), None) #next((col for col in df.columns if "cost" in col), None)
    
 
    if not qty_col or not unit_price_col:
        raise ValueError("One or more required columns (target quantity, cost) are missing.")
 
    df[qty_col] = pd.to_numeric(df[qty_col], errors='coerce')
    df["total_price"] = pd.to_numeric(df.get("cost"), errors="coerce")

    # New logicccccccccccccccccccccc
    computed_from_cost = (
        (df[unit_price_col] / df[qty_col])
        if (unit_price_col and qty_col)
        else np.nan 
    )

    if unit_price_col and qty_col:
        df[unit_price_col] = pd.to_numeric(df[unit_price_col], errors="coerce")
        df[qty_col] = pd.to_numeric(df[qty_col], errors = "coerce")
        computed_from_cost = np.where(
            df[unit_price_col].notna() & df[qty_col].notna() & (df[qty_col] != 0), 
            df[unit_price_col] / df[qty_col], 
            np.nan
        )

    df["total_price"] = np.where(
        df["total_price"].notna() & (df["total_price"] != 0),
        df["total_price"],
        np.where(
            ~pd.isna(computed_from_cost),
            computed_from_cost,
            pd.to_numeric(df[unit_price_col], errors = "coerce")
        )
    )

    # print("[INFO] ✅ 'EXTENDED_PRICE_FINAL' computed using CLP or fallback Price.")
    return df


# Part keyword enrichment - Genpact part keywords 
def enrich_with_part_keywords(result: dict, main_df: pd.DataFrame, pipeline_type: str) -> pd.DataFrame:
    # print(f"[INFO] 🔍 Enriching Genpact Part Keywords for {pipeline_type} pipeline...")
 
    # Load the CC Part Number sheet
    cc_parts_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="cc part"
    )
            # Normalize column names
    cc_parts_df.columns = [col.strip().lower() for col in cc_parts_df.columns]
    # print("contract_cc-columns",cc_parts_df.columns)
    
    main_df = main_df.copy()
    print("contract_cc")
    main_df.to_excel("contract_cc_1.xlsx",index=False)
    main_df.columns = [col.strip().lower() for col in main_df.columns]
 
    part_num_col = next((col for col in cc_parts_df.columns if "part" in col and "num" in col), None)
    keyword_col = next((col for col in cc_parts_df.columns if "part" in col and "keyword" in col), None)
 
    if pipeline_type == "material":
        pn_col = next((col for col in main_df.columns if "material" in col), None)
        description_col = next((col for col in main_df.columns if "description" in col), None)

    elif pipeline_type == "repair":
        pn_col = next((col for col in main_df.columns if "inspected" in col and "part no" in col), None)
        description_col = next((col for col in main_df.columns if "csn description" in col and "description" in col), None)

    else:
        pn_col = next((col for col in main_df.columns if "material" in col), None)
        description_col = next((col for col in main_df.columns if "description" in col), None)

    if not all([part_num_col, keyword_col, pn_col, description_col]):
        raise ValueError("Required columns not found for keyword enrichment.")
 
    # Merge
    enriched_df = main_df.merge(
        cc_parts_df[[part_num_col, keyword_col]],
        left_on=pn_col,
        right_on=part_num_col,
        how='left'
    )
 
    # Step 4: Create 'GENPACT_PART_KEYWORDS' column
    enriched_df["genpact_part_keywords"] = enriched_df[keyword_col]
    enriched_df["genpact_part_keywords"] = enriched_df["genpact_part_keywords"].fillna(enriched_df[description_col])
 
    # print(f"[INFO] ✅ Genpact Part Keywords Enrichment completed for {pipeline_type}.")
    return enriched_df

# EIPC data refinement
def load_cleaned_eipc_dataframe(file_path: str, threshold: float = 0.5) -> pd.DataFrame:
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
    
    # Step 1: Read raw file with no headers
    raw_df = read_excel_from_s3(file_path, header=None)
 
    # Step 2: Identify the first valid header row
    header_row_idx = None
    for idx, row in raw_df.iterrows():
        if row.notna().mean() >= threshold:
            header_row_idx = idx
            break
 
    if header_row_idx is None:
        raise ValueError("Unable to find a valid header row in EIPC file.")
 
    # Step 3: Re-read using detected header
    df = read_excel_from_s3(file_path, header=header_row_idx)
    return df


def enrich_with_ata_long(result: dict, eipc_df: pd.DataFrame, main_df: pd.DataFrame, pipeline_type: str) -> pd.DataFrame:
    # print(f"[INFO] 🔍 Enriching ATA_LONG for {pipeline_type} pipeline...")
 
    eipc_df.columns = [col.strip().lower() for col in eipc_df.columns]
    main_df = main_df.copy()
    main_df.columns = [col.strip().lower() for col in main_df.columns]
 
    # Step 1: Detect required columns
    figure_col = next((col for col in eipc_df.columns if "figure" in col), None)
    part_num_col = next((col for col in eipc_df.columns if "part" in col and "number" in col), None)
 
    if pipeline_type in ["material", "vendor"]:
        mat_col = next((col for col in main_df.columns if "material" in col), None)
        ata_section_code = next((col for col in main_df.columns if "ata section code" in col), None)
        ata_chapter_col = next((col for col in main_df.columns if "ata chapter" in col), None)
        csn_col = next((col for col in main_df.columns if "component" in col and "csn" in col), None)
 
    elif pipeline_type in ["repair"]:
        mat_col = next((col for col in main_df.columns if "inspected part no" in col and "part" in col), None)
        ata_chapter_col = next((col for col in main_df.columns if "ata chapter" in col), None)
        csn_col = next((col for col in main_df.columns if "csn" in col), None)
        ata_section_code = None  # Not used in repair
 
    if not all([figure_col, part_num_col, mat_col, ata_chapter_col, csn_col]):
        raise ValueError("[ERROR] ❌ Required columns not found for ATA_LONG enrichment.")
 
    # Step 2: Create empty ATA_LONG column
    main_df["ata_long"] = pd.NA
 
    # Step 3: Extract 6-digit from CSN if ATA Chapter is 7200–7209
    valid_ata = main_df[ata_chapter_col].str.replace("-", "", regex=False).str[:4]
    valid_ata = pd.to_numeric(valid_ata, errors='coerce')
 
    main_df["valid_ata"] = valid_ata
 
    #Old logiccccc
    # ata_filter = valid_ata.between(7200, 7209, inclusive="both")
 
    # New logicccc
    # ata_filter = main_df["valid_ata"].between(7200, 7209, inclusive="both") | valid_ata.eq(7260)
    ata_filter = (
        main_df["valid_ata"].between(7200, 7209, inclusive="both")
        | main_df["valid_ata"].isin([7210,7220, 7230, 7240,7250,7260])
    )
 
    csn_six = main_df[csn_col].astype(str).str.extract(r'_(\d{6})')[0]
    main_df.loc[ata_filter & csn_six.notna(), "ata_long"] = csn_six[ata_filter]
 
    # Step 4: Fill with numeric ATA Chapter where ata_long is still missing#20-1
    numeric_ata_chapter = main_df[ata_chapter_col].str.replace("-", "").str[:6]
    main_df["ata_long"] = main_df["ata_long"].fillna(numeric_ata_chapter)
 
    # Step 5: Fill with ATA Section Code if provided (material/vendor only)
    if ata_section_code:
        main_df["ata_long"] = main_df["ata_long"].fillna(main_df[ata_section_code])
 
    # Step 6: Merge with EIPC for fallback
    eipc_df["figure_six"] = eipc_df[figure_col].astype(str).str.extract(r'(\d{6})')[0]
    merged_df = main_df.merge(
        eipc_df[[part_num_col, "figure_six"]],
        left_on=mat_col,
        right_on=part_num_col,
        how="left"
    )
 
    # Step 7: Fill missing ata_long from EIPC figure_six
    merged_df["ata_long"] = merged_df["ata_long"].fillna(merged_df["figure_six"])
    merged_df["formatted_ata_long"] = ( merged_df["ata_long"].astype("string").str.replace(r"^(\d{2})(\d{2})(\d{2})$", r"\1-\2-\3", regex=True))
    merged_df["ata_long_v1"]= merged_df["ata_long"].astype(str).str[:4]
    merged_df["ata_long_v1"] = merged_df["ata_long_v1"].astype(str).apply(lambda x: x.ljust(6, '0') if x.isdigit() and len(x) < 6 else x)
       
    # Step 8: Drop temp columns
    merged_df.drop(columns=["figure_six", part_num_col], inplace=True, errors="ignore")
   
    merged_df.drop_duplicates(inplace=True)
    # print(f"[INFO] ✅ ATA_LONG Enrichment completed for {pipeline_type}.")
    return merged_df


def assign_llp_flag(result: dict, billing_df: pd.DataFrame) -> pd.DataFrame:
 
    # ------------------------------------------------------------------
    # helpers
    # ------------------------------------------------------------------
    def _pick_col(cols, *need_tokens, regex=None, case_insensitive=True):
        if regex:
            flags = re.I if case_insensitive else 0
            for c in cols:
                if re.search(regex, str(c), flags=flags):
                    return c
            return None
 
        toks = [t.lower() if case_insensitive else t for t in need_tokens]
        for c in cols:
            name = str(c)
            chk = name.lower() if case_insensitive else name
            if all(t in chk for t in toks):
                return c
        return None
 
    def _to_num(s):
        return pd.to_numeric(s, errors="coerce")
 
    # order-independent keyword normalization
    def _normalize_keyword(s):
        if pd.isna(s):
            return s
        s = re.sub(r"[^a-z0-9 ]+", " ", str(s).lower())
        return " ".join(sorted(s.split()))
 
    # ------------------------------------------------------------------
    # Load & normalize
    # ------------------------------------------------------------------
    llp_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="llp"
    ).copy()
    llp_df.columns = [str(c).strip() for c in llp_df.columns]
 
    billing = billing_df.copy()
    billing.to_excel("contract_hand_1.xlsx",index=False)
    # cc
    billing.columns = [str(c).strip().lower() for c in billing.columns]
 
    # Billing columns
    material_col   = _pick_col(billing.columns, "material")
    cost_cat_col   = _pick_col(billing.columns, "cost category")
    matl_clp_col   = _pick_col(billing.columns, "matl_clp_lookup") or _pick_col(billing.columns, "clp")
    genpact_pk_col = (
        _pick_col(billing.columns, "genpact_part_keywords")
        or _pick_col(billing.columns, "genpact", "part", "keyword")
    )
 
    if not all([material_col, cost_cat_col, matl_clp_col]):
        raise ValueError("Missing required billing columns.")
 
    # LLP master part column
    llp_part_col = (
        _pick_col(llp_df.columns, "part", "number")
        or _pick_col(llp_df.columns, "part", "no")
    )
    if not llp_part_col:
        raise ValueError("Part Number column not found in LLP list.")
 
    # Normalize join keys
    billing[material_col] = billing[material_col].astype(str).str.strip().str.upper()
    llp_df[llp_part_col]  = llp_df[llp_part_col].astype(str).str.strip().str.upper()
 
    # ------------------------------------------------------------------
    # Step 1: LLP flag
    # ------------------------------------------------------------------
    flagged = billing.merge(
        llp_df[[llp_part_col]].drop_duplicates(),
        left_on=material_col,
        right_on=llp_part_col,
        how="left",
        indicator=True
    )
    flagged["llp_flag"] = np.where(flagged["_merge"].eq("both"), "Y", "N")
    flagged.drop(columns=["_merge", llp_part_col], inplace=True, errors="ignore")
 
    # ------------------------------------------------------------------
    # Step 2: LLP Pricing
    # ------------------------------------------------------------------
    llp_pricing = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="llp pricing"
    ).copy()
    llp_pricing.columns = [str(c).strip() for c in llp_pricing.columns]
 
    # Induction year
    _, induction_date, _, _,_= extract_q_year_from_timeliness(result)
    induction_year = str(induction_date.year)
 
    # Pricing columns
    pn_col = (
        _pick_col(llp_pricing.columns, "part", "number")
        or _pick_col(llp_pricing.columns, "part", "no")
    )
    kw_col = _pick_col(llp_pricing.columns, "keyword")
 
    # LLP-NEW → Adjusted Price (year-specific preferred)
    new_adj_col = (
        _pick_col(llp_pricing.columns, "adjusted", "price", induction_year)
        or _pick_col(llp_pricing.columns, "adjusted", "price")
    )
 
    # LLP-USED → Year (USD) ONLY (explicitly exclude Adjusted)
    used_year_usd_col = next(
        (
            c for c in llp_pricing.columns
            if induction_year in c
            and "usd" in c.lower()
            and "adjusted" not in c.lower()
        ),
        None
    )
 
    if not pn_col:
        raise ValueError("Part Number column not found in LLP Pricing.")
 
    # ------------------------------------------------------------------
    # Merge pricing by Part Number
    # ------------------------------------------------------------------
    flagged = flagged.merge(
        llp_pricing[[c for c in [pn_col, new_adj_col, used_year_usd_col] if c]],
        left_on=material_col,
        right_on=pn_col,
        how="left",
        suffixes=("", "_by_pn")
    )
 
    # ------------------------------------------------------------------
    # Keyword normalization & fallback merge
    # ------------------------------------------------------------------
    if genpact_pk_col and kw_col:
        flagged["_norm_kw"] = flagged[genpact_pk_col].apply(_normalize_keyword)
        llp_pricing["_norm_kw"] = llp_pricing[kw_col].apply(_normalize_keyword)
 
        flagged = flagged.merge(
            llp_pricing[[c for c in ["_norm_kw", new_adj_col, used_year_usd_col] if c]],
            left_on="_norm_kw",
            right_on="_norm_kw",
            how="left",
            suffixes=("", "_by_kw")
        )
 
    # Resolve LLP-NEW price (PN → Keyword)
    if new_adj_col:
        flagged["_new_adjusted_price"] = _to_num(
            flagged.get(new_adj_col)
        ).combine_first(
            _to_num(flagged.get(f"{new_adj_col}_by_kw"))
        )
    else:
        flagged["_new_adjusted_price"] = np.nan
 
    # Resolve LLP-USED price (PN → Keyword)
    if used_year_usd_col:
        flagged["_used_year_price"] = _to_num(
            flagged.get(used_year_usd_col)
        ).combine_first(
            _to_num(flagged.get(f"{used_year_usd_col}_by_kw"))
        )
    else:
        flagged["_used_year_price"] = np.nan
 
    # ------------------------------------------------------------------
    # Step 3: Apply pricing
    # ------------------------------------------------------------------
    cc_upper = flagged[cost_cat_col].astype(str).str.upper()
 
    is_new  = cc_upper.str.contains(r"\bLLP-NEW\b",  regex=True, na=False)
    is_used = cc_upper.str.contains(r"\bLLP-USED\b", regex=True, na=False)
 
    mask_new  = is_new  & flagged["_new_adjusted_price"].notna()
    mask_used = is_used & flagged["_used_year_price"].notna()
 
    flagged.loc[mask_new,  matl_clp_col] = flagged.loc[mask_new,  "_new_adjusted_price"]
    flagged.loc[mask_used, matl_clp_col] = flagged.loc[mask_used, "_used_year_price"]
 
    # ✅ SINGLE AUDIT PRINT
    print(
        f"LLP Pricing Applied | "
        f"NEW (Adjusted Price) = {mask_new.sum()} rows | "
        f"USED (Year USD) = {mask_used.sum()} rows"
    )
 
    # ------------------------------------------------------------------
    # Cleanup
    # ------------------------------------------------------------------
    drop_cols = [
        pn_col,
        kw_col,
        "_norm_kw",
        new_adj_col,
        used_year_usd_col,
        f"{new_adj_col}_by_kw" if new_adj_col else None,
        f"{used_year_usd_col}_by_kw" if used_year_usd_col else None,
        "_new_adjusted_price",
        "_used_year_price"
    ]
 
    flagged.drop(
        columns=[c for c in drop_cols if c in flagged.columns],
        inplace=True,
        errors="ignore"
    )
 
    flagged.drop_duplicates(inplace=True)
    flagged.to_excel("LLP_check.xlsx", index=False)
     
    return flagged

# ---------- small helpers ----------
def _pick_col(cols, *need_tokens, regex=None, case_insensitive=True):
    """Pick the first column whose name contains all tokens (or matches regex)."""
    if regex:
        flags = re.I if case_insensitive else 0
        for c in cols:
            if re.search(regex, str(c), flags=flags):
                return c
        return None
    toks = [t.lower() if case_insensitive else t for t in need_tokens]
    for c in cols:
        name = str(c)
        chk = name.lower() if case_insensitive else name
        if all(t in chk for t in toks):
            return c
    return None
 
def _to_num(s):
    return pd.to_numeric(s, errors="coerce")


def append_module_to_remarks(formatted_ata, module, remarks):
    """
    Append module to remarks with a space
    only if formatted_ata and module are different.
    """
    formatted_ata = str(formatted_ata).strip() if pd.notna(formatted_ata) else ""
    module = str(module).strip() if pd.notna(module) else ""
    remarks = str(remarks).strip() if pd.notna(remarks) else ""
 
    # If same or module empty → do nothing
    if formatted_ata == module or module == "":
        return remarks
 
    # Append naturally with space
    if remarks:
        return f"{remarks},{module}"
 
    # If remarks empty, just return module
    return module
  

def apply_llp_exclusion(llp_tagging_df: pd.DataFrame, result: dict,
                        esn: str | None = None, customer_first_word: str | None = None) -> pd.DataFrame:
    df = llp_tagging_df.copy()
    df.columns = [c.strip().lower() for c in df.columns]
 
    # ---- locate common columns ----
    material_col   = _pick_col(df.columns, "material")
    cost_cat_col   = _pick_col(df.columns, "cost", "category")
    clp_col        = _pick_col(df.columns, "matl_clp_lookup") or "matl_clp_lookup"
    ata_long_col   = _pick_col(df.columns, "ata_long")
    cost_col       = _pick_col(df.columns, "cost")  # optional
    remarks_col    = "remarks"
    genpact_pk_col = (_pick_col(df.columns, "genpact", "part", "keyword")
                      or _pick_col(df.columns, "genpact", "keyword")
                      or _pick_col(df.columns, "keyword"))
    llp_flag_col   = (_pick_col(df.columns, "llp", "flag")
                      or _pick_col(df.columns, "llp_flag")
                      or _pick_col(df.columns, "llpflag"))
 
    if not all([material_col, cost_cat_col]):
        raise ValueError("Missing required columns: material / cost category.")
 
    if remarks_col not in df.columns:
        df[remarks_col] = ""
 
    # ---- mark all LLP as exclusion (robust truthy set) ----
    is_llp_any = df[llp_flag_col].astype(str).str.strip().str.upper().isin({"Y","YES","TRUE","1"}) if llp_flag_col else pd.Series(False, index=df.index)
    df.loc[is_llp_any, remarks_col] = "Exclusion as LLP material"
 
    # ---- USED LLP mask ----
    used_llp_cc = {"SCC-LLP-USED-LP", "SCC-LLP-USED-GE"}
    mask_used_llp = is_llp_any & df[cost_cat_col].astype(str).str.upper().isin(used_llp_cc)
 
    # Default: if we never enter pro-rata path, we’ll return df
    billing_df_used_llp_df = None
 
    # ====== PRICING (by PN then Keyword) ONLY IF there are USED-LLP rows ======
    if mask_used_llp.any():
        q_year, induction_date, *_ = extract_q_year_from_timeliness(result)
        year_token = str(induction_date.year if induction_date is not None else (q_year.split("_")[-1] if "_" in q_year else q_year))
 
        llp_pr = get_dataframe_by_file_and_sheet(
            all_dataframes=result["dataframes"],
            file_key="contract",
            sheet_name_substring="llp pricing"
        ).copy()
        llp_pr.columns = [c.strip().lower() for c in llp_pr.columns]
 
        pn_col = _pick_col(llp_pr.columns, "part", "number") or _pick_col(llp_pr.columns, "part", "no")
        kw_col = _pick_col(llp_pr.columns, "keyword")
 
        # find '<YEAR> (USD)' or any col with year token
        year_col = None
        for c in llp_pr.columns:
            if re.search(rf"\b{re.escape(year_token)}\b", c, flags=re.I) and re.search(r"(usd|\$)", c, flags=re.I):
                year_col = c; break
        if not year_col:
            candidates = [c for c in llp_pr.columns if re.search(rf"\b{re.escape(year_token)}\b", c, flags=re.I)]
            year_col = candidates[0] if candidates else None
 
        if pn_col and year_col:
            map_pn = (
                llp_pr[[pn_col, year_col]]
                .dropna(subset=[pn_col])
                .assign(**{pn_col: lambda d: d[pn_col].astype(str).str.strip().str.upper()})
                .drop_duplicates(subset=[pn_col])
                .set_index(pn_col)[year_col]
            )
            keys_pn = df.loc[mask_used_llp, material_col].astype(str).str.strip().str.upper()
            val_pn = pd.to_numeric(keys_pn.map(map_pn), errors="coerce")
 
            if kw_col and genpact_pk_col:
                map_kw = (
                    llp_pr[[kw_col, year_col]]
                    .dropna(subset=[kw_col])
                    .assign(**{kw_col: lambda d: d[kw_col].astype(str).str.strip().str.upper()})
                    .drop_duplicates(subset=[kw_col])
                    .set_index(kw_col)[year_col]
                )
                keys_kw = df.loc[mask_used_llp, genpact_pk_col].astype(str).str.strip().str.upper()
                val_kw = pd.to_numeric(keys_kw.map(map_kw), errors="coerce")
                chosen_used_price = val_pn.combine_first(val_kw)
            else:
                chosen_used_price = val_pn
 
            if clp_col not in df.columns:
                df[clp_col] = np.nan
 
            has_price = chosen_used_price.notna()
            idx_masked = df.index[mask_used_llp]
            if has_price.any():
                df.loc[idx_masked[has_price.values], clp_col] = chosen_used_price[has_price].values
 
        # ====== PRO-RATA (only if file exists AND we have metadata to match) ======
        used_llp_fil_folder = r"SMBA_Avianca_Celma\Billing"
        # target_str = esn &"_" &"AD Status"
        target_str = esn + "_" + "AD Status"
        sheet_substr = "LLP On Report"
        out_path= "AD_Status_baked.xlsx" 
        print("target_str",target_str)
        used_llp_fil_path, sheet_name = find_file_and_sheet(used_llp_fil_folder, target_str, sheet_substr)
        print(f"File: {used_llp_fil_path}")
        print(f"Sheet: {sheet_name}")

        # file_exists = Path(used_llp_fil_path).exists()
        file_exists = s3_file_exists(used_llp_fil_path)

        # new logic ad status - finding it
        

        if file_exists and esn is not None and customer_first_word is not None:
            print("file_exists :", file_exists)

            print("used_llp_fil_path, out_path, target_sheet :", used_llp_fil_path, out_path, sheet_substr)

            bake_formulas_and_remove_hidden_sheets(used_llp_fil_path, "AD_Status_baked.xlsx", "LLP ON Report")
            
            used_llp_lrc = extract_tlc_lcr_used_llp(out_path, esn, customer_first_word).copy()
            
            print("saving extract Ad file")
            used_llp_lrc.to_excel("extract_ad_file.xlsx",index=False)
            used_llp_lrc.columns = [c.strip().lower() for c in used_llp_lrc.columns]
 
            used_llp_part_num_col = _pick_col(used_llp_lrc.columns, "part_number")
            # If there’s truly no data (empty or no PN col), SKIP pro-rata cleanly
            if used_llp_part_num_col and not used_llp_lrc.empty:
                used_llp_lrc[used_llp_part_num_col] = used_llp_lrc[used_llp_part_num_col].astype(str).str.strip().str.upper()
                used_llp_lrc = used_llp_lrc.drop_duplicates(subset=[used_llp_part_num_col], keep="first")
 
                # Merge (left) – result always exists; missing columns will be NaN
                billing_df_used_llp_df = df.merge(
                    used_llp_lrc, left_on=material_col, right_on=used_llp_part_num_col, how="left"
                )
 
                # Used LLP % (by ATA-4)
                used_llp_pct = get_dataframe_by_file_and_sheet(
                    all_dataframes=result["dataframes"], file_key="contract", sheet_name_substring="used llp price"
                )
                used_llp_pct = find_first_data_row_in_df(used_llp_pct, "atalong").copy()
                used_llp_pct.columns = [
                    (f"{float(c):g}" if isinstance(c, (int, float, np.integer, np.floating)) else str(c)).replace("\xa0"," ").strip().lower()
                    for c in used_llp_pct.columns
                ]
 
                used_pct_ata_col  = _pick_col(used_llp_pct.columns, "atalong")
                year_only         = (q_year.split("_")[-1] if "_" in q_year else q_year)[-4:]
                used_pct_year_col = _pick_col(used_llp_pct.columns, year_only)
 
                if used_pct_ata_col and used_pct_year_col and ata_long_col:
                    tmp_pct = used_llp_pct[[used_pct_ata_col, used_pct_year_col]].copy()
                    tmp_pct[used_pct_ata_col] = tmp_pct[used_pct_ata_col].astype(str).str.strip()
                    tmp_pct["ata_4_digit"] = tmp_pct[used_pct_ata_col].str.replace(r"\D", "", regex=True).str[:4]
                    tmp_pct = (tmp_pct.dropna(subset=["ata_4_digit"])
                                     .drop_duplicates(subset=["ata_4_digit"], keep="first")
                                     .rename(columns={used_pct_year_col: "used_llp_percent"}))
 
                    billing_df_used_llp_df["ata_4_digit"] = billing_df_used_llp_df[ata_long_col].astype(str).str.replace(r"\D", "", regex=True).str[:4]
                    billing_df_used_llp_df = billing_df_used_llp_df.merge(tmp_pct[["ata_4_digit","used_llp_percent"]], on="ata_4_digit", how="left")
 
                    # ---- SAFE column extraction: fall back to zeros if missing ----
                    def _col_or_zero(sframe: pd.DataFrame, *cands) -> pd.Series:
                        col = _pick_col(sframe.columns, *cands)
                        if col:
                            return _to_num(sframe[col])
                        return pd.Series(0, index=sframe.index, dtype="float64")
 
                    rem  = _col_or_zero(billing_df_used_llp_df, "total_remaining_cicles", "life cicles remaining",
                                        "life cycles remaining", "remaining cycles")
                    life = _col_or_zero(billing_df_used_llp_df, "total_life_cicles", "total life cicles",
                                        "total life cycles", "life cycles total")
                    pct  = _col_or_zero(billing_df_used_llp_df, "used_llp_percent", "pro-rata %", "pro rata %", "used llp %")
 
                    # pro-rata factor (avoid div-by-zero/NaN)
                    with np.errstate(divide="ignore", invalid="ignore"):
                        factor = (rem / life).replace([np.inf, -np.inf], 0).fillna(0) * (pct.fillna(0) / 100.0)
                    billing_df_used_llp_df["pro_rata_factor"] = factor
 
                    # total price for SCC-LLP-USED-LP only
                    cc_upper = billing_df_used_llp_df[cost_cat_col].astype(str).str.upper()
                    is_used_lp = cc_upper.eq("SCC-LLP-USED-LP")
                    billing_df_used_llp_df[clp_col] = _to_num(billing_df_used_llp_df.get(clp_col))
                    billing_df_used_llp_df["total_price_usedllp_calc"] = (billing_df_used_llp_df["pro_rata_factor"] * billing_df_used_llp_df[clp_col]).round(2)
                    billing_df_used_llp_df.loc[is_used_lp, "total_price"] = billing_df_used_llp_df.loc[is_used_lp, "total_price_usedllp_calc"]
 
                    if cost_col in df.columns:
                        billing_df_used_llp_df[cost_col] = _to_num(billing_df_used_llp_df[cost_col])
                        billing_df_used_llp_df["total_price"] = billing_df_used_llp_df["total_price"].where(
                            billing_df_used_llp_df["total_price"].notna() & (billing_df_used_llp_df["total_price"] != 0),
                            billing_df_used_llp_df[cost_col]
                        )
                # else: silently skip pro-rata if percentage sheet columns not present
 
    # ---- finalize safely ----
    out = billing_df_used_llp_df if billing_df_used_llp_df is not None else df
    out = out.drop_duplicates()
    out.to_excel("NEW_USED_LLP.xlsx", index=False)
    return out
 
    
# Rule-2: BLADE HPT & FAN BLADE - Inclusion 
FAN_BLADE_TERMS = ["fan", "blade"]
HPT_BLADE_TERMS = ["hpt", "blade"]
 
def normalize_phrase(text: str) -> str:
    """Normalize text for loose comparison (remove symbols, lowercase, collapse spaces)."""
    return re.sub(r'\s+', ' ', re.sub(r'[^a-zA-Z0-9 ]', ' ', str(text).lower())).strip()
 
def generate_pattern_permutations(terms: list[str]) -> set[str]:
    """Generate all permutations of given terms (e.g., fan-blade → blade-fan)"""
    return set([" ".join(p) for p in permutations(terms)])
 
# def apply_blade_inclusion(result: dict, main_df: pd.DataFrame, pipeline_type: str) -> pd.DataFrame:
#     # print(f"[INFO] 🔍 Applying Blade Inclusion for {pipeline_type} pipeline...")
 
#     # Step 1: Load the Inclusion sheet
#     inclusion_df = get_dataframe_by_file_and_sheet(
#         all_dataframes=result["dataframes"],
#         file_key="contract",
#         sheet_name_substring="inclusion"
#     )
#     inclusion_df.columns = [col.strip().lower() for col in inclusion_df.columns]
#     print("sheet_name_substring=inclusion : ", inclusion_df.columns)
#     main_df = main_df.copy()
    
#     main_df.columns = [col.strip().lower() for col in main_df.columns]
#     main_df.to_excel("main_df_apply_blade_inc.xlsx", index=False)
    
 
#     # Step 2: Detect relevant columns
#     part_desc_col = next((col for col in inclusion_df.columns if "description" in col), None)
#     if pipeline_type == "material":
#         part_keyword_col = "genpact_part_keywords"
#         remarks = "MATL UNKNOWN"
#     elif pipeline_type == "repair":
#         part_keyword_col = next((col for col in main_df.columns if "csn description" in col), None)
#         remarks = "REPAIR UNKNOWN"
#     elif pipeline_type == "vendor":
#         part_keyword_col = "genpact_part_keywords"
#         remarks = "VENDOR UNKNOWN"
    
#     # print("part_desc_col, part_keyword_col :", part_desc_col, part_keyword_col)
#     if not all([part_desc_col, part_keyword_col]):
#         raise ValueError("Required columns not found for Blade Inclusion.")
 
#     # Step 3: Normalize keyword columns
#     inclusion_df["norm_description"] = inclusion_df[part_desc_col].apply(normalize_phrase)
#     main_df["keyword_norm"] = main_df[part_keyword_col].apply(normalize_phrase)
 
#     # Step 4: Generate permutations for FAN & HPT blades
#     fan_blade_patterns = generate_pattern_permutations(FAN_BLADE_TERMS)
#     hpt_blade_patterns = generate_pattern_permutations(HPT_BLADE_TERMS)
 
#     # Step 5: Merge and check match
#     merged_df = main_df.merge(
#         inclusion_df[["norm_description"]],
#         left_on="keyword_norm",
#         right_on="norm_description",
#         how="left",
#         indicator=True
#     )


    
#     def identify_blade_type(text):
#         if not text:
#             return remarks
#         for p in generate_pattern_permutations(text.split()):
#             if p in fan_blade_patterns:
#                 return "Inclusion as Fan-Blade" #"Fan-Blade Inclusion"
#             if p in hpt_blade_patterns:
#                 return "Inclusion as Blade-HPT" # "HPT Blade Inclusion"
#         return remarks
 
#     # Step 6: Update only 'MATL UNKNOWN' remarks
#     mask = (merged_df["remarks"].str.upper() == remarks) & (merged_df["_merge"] == "both")
#     merged_df.loc[mask, "remarks"] = merged_df.loc[mask, "keyword_norm"].apply(identify_blade_type)
#     # print(f"[INFO] ✅ Blade Inclusion logic applied successfully for {pipeline_type}.")

#     remarks_expected = ["Inclusion as Fan-Blade", "Inclusion as Blade-HPT"]
#     blade_incl_df = merged_df[merged_df["remarks"].isin(remarks_expected)].copy()
#     blade_incl_df.drop_duplicates(inplace=True)
#     return merged_df

# def scrap_exclusion_calculation(df, result):
 
#     print("🔎 Starting Scrap Exclusion Calculation")
 
#     # -------------------------------------------------
#     # Load Required Sheets
#     # -------------------------------------------------
#     qpe_list = get_dataframe_by_file_and_sheet(
#         all_dataframes=result["dataframes"],
#         file_key="contract",
#         sheet_name_substring="qpe_list"
#     )
 
#     billing_df = get_dataframe_by_file_and_sheet(
#         all_dataframes=result["dataframes"],
#         file_key="billing",
#         sheet_name_substring="billing"
#     )
 
#     cc_part_num = get_dataframe_by_file_and_sheet(
#         all_dataframes=result["dataframes"],
#         file_key="contract",
#         sheet_name_substring="cc part number"
#     )
 
#     # Normalize column names
#     df.columns = [col.strip().lower() for col in df.columns]
#     billing_df.columns = [col.strip().lower() for col in billing_df.columns]
#     qpe_list.columns = [col.strip().lower() for col in qpe_list.columns]
#     cc_part_num.columns = [col.strip().lower() for col in cc_part_num.columns]
 
#     print("✅ Sheets Loaded Successfully")
 
#     # -------------------------------------------------
#     # STEP 0 — Create Clean Mapping (NO DUPLICATION)
#     # -------------------------------------------------
    
#     material_col = next((c for c in billing_df.columns if "material" in c), None)
#     part_num_col = next((c for c in cc_part_num.columns if "part_num" in c or "part number" in c), None)
#     part_keyword_col = next((c for c in cc_part_num.columns if "part_keyword" in c), None)
    
#     if not material_col or not part_num_col or not part_keyword_col:
#         raise ValueError("Required columns not found")
    
#     billing_df[material_col] = billing_df[material_col].astype(str)
#     cc_part_num[part_num_col] = cc_part_num[part_num_col].astype(str)
    
#     # Create unique mapping dictionary (IMPORTANT)
#     mapping_dict = (
#         cc_part_num
#         .drop_duplicates(subset=[part_num_col])
#         .set_index(part_num_col)[part_keyword_col]
#         .to_dict()
#     )
    
#     # Map instead of merge (NO row multiplication)
#     billing_df["genpact_part_keywords"] = billing_df[material_col].map(mapping_dict)
    
#     print("🔹 Keywords mapped using dictionary")
#     print("Missing Keywords:", billing_df["genpact_part_keywords"].isna().sum())
    
#     # -------------------------------------------------
#     # STEP 1 — Total Scrap Quantity (PURE AGGREGATION)
#     # -------------------------------------------------
    
#     billing_grouped = (
#         billing_df
#         .groupby("genpact_part_keywords", dropna=False)["target quantity"]
#         .sum()
#         .reset_index()
#         .rename(columns={"target quantity": "total_scrap_quantity"})
#     )
    
#     print("🔹 Billing grouped shape:", billing_grouped.shape)
    
#     # IMPORTANT: Ensure only one row per part
#     dup_check = billing_grouped["genpact_part_keywords"].duplicated().sum()
#     print("Duplicate keys in billing_grouped:", dup_check)
    
#     # -------------------------------------------------
#     # LEFT JOIN from df → billing_grouped
#     # -------------------------------------------------
    
#     df = df.merge(
#         billing_grouped,
#         on="genpact_part_keywords",
#         how="left",
#         validate="m:1"  # THIS WILL THROW ERROR IF DUPLICATE
#     )
    
#     df["total_scrap_quantity"] = df["total_scrap_quantity"].fillna(0)
    
#     print("🔹 Total Scrap Quantity Merged Safely")
 
#     # -------------------------------------------------
#     # STEP 2 — Merge QPE
#     # -------------------------------------------------
#     qpe_grouped = (
#         qpe_list[["part keyword", "qpe"]]
#         .rename(columns={
#             "part keyword": "genpact_part_keywords",
#             "qpe": "total_qpe"
#         })
#     )
 
#     df = df.merge(
#         qpe_grouped,
#         on="genpact_part_keywords",
#         how="left"
#     )
 
#     df["total_qpe"] = df["total_qpe"].replace(0, np.nan)
 
#     print("🔹 QPE Merged")
 
#     print("Missing QPE Count:", df["total_qpe"].isna().sum())
 
#     # -------------------------------------------------
#     # STEP 3 — Actual Scrap Rate
#     # -------------------------------------------------
#     df["actual_scrap_rate"] = (
#         df["total_scrap_quantity"] / df["total_qpe"]
#     )
 
#     print("🔹 Actual Scrap Rate Calculated")
 
#     # -------------------------------------------------
#     # STEP 4 — Compare scrap_cap vs actual rate
#     # scrap_cap is numeric (20, 40 etc)
#     # actual_scrap_rate is decimal (0.2)
#     # -------------------------------------------------
#     mask = df["scrap_exclusion"].str.lower() == "yes"
 
#     condition = mask & (
#         df["scrap_cap"] > df["actual_scrap_rate"] * 100
#     )
 
#     df.loc[condition, "scrap_exclusion"] = "No"
 
#     print("🔹 Scrap Cap Comparison Done")
#     print("Rows turned to No:", condition.sum())
 
#     # -------------------------------------------------
#     # STEP 5 — Excluded / Included Quantity
#     # -------------------------------------------------
#     df["allowed_quantity"] = df["total_qpe"] * (df["scrap_cap"] / 100)
 
#     df["excluded_quantity"] = (
#         df["total_scrap_quantity"] - df["allowed_quantity"]
#     )
 
#     df.loc[df["excluded_quantity"] < 0, "excluded_quantity"] = 0
 
#     df["included_quantity"] = (
#         df["total_scrap_quantity"] - df["excluded_quantity"]
#     )
 
#     print("🔹 Excluded & Included Quantities Calculated")
 
#     # -------------------------------------------------
#     # STEP 6 — Effective Price
#     # -------------------------------------------------
#     df["effective_price"] = df["matl_clp_lookup"]
 
#     df.loc[
#         (df["effective_price"].isna()) | (df["effective_price"] == 0),
#         "effective_price"
#     ] = df["cost"]
 
#     # If still zero → cancel exclusion
#     cancel_mask = (
#         (df["effective_price"].isna()) |
#         (df["effective_price"] == 0)
#     )
 
#     df.loc[cancel_mask, "scrap_exclusion"] = "No"
 
#     print("🔹 Effective Price Determined")
#     print("Rows cancelled due to price issue:", cancel_mask.sum())
 
#     # -------------------------------------------------
#     # STEP 7 — Allocate Excluded Price (Lowest First)
#     # -------------------------------------------------
#     def allocate_excluded(group):
 
#         group = group.sort_values("effective_price")
    
#         remaining_qty = group["excluded_quantity"].iloc[0]
#         total_excluded_price = 0
    
#         for _, row in group.iterrows():
    
#             if remaining_qty <= 0:
#                 break
    
#             # Row-level scrap quantity
#             row_scrap_qty = row["target quantity"]
    
#             take_qty = min(remaining_qty, row_scrap_qty)
    
#             total_excluded_price += take_qty * row["effective_price"]
    
#             remaining_qty -= take_qty
    
#         # discount = group["discount"].iloc[0] if "discount" in group else 0
#         discount = group["discount"].max() if "discount" in group.columns else 0
#         total_excluded_price *= (1 - discount)
    
#         group["excluded_price"] = total_excluded_price
    
#         return group
 
#     df = (
#         df.groupby("genpact_part_keywords", group_keys=False)
#           .apply(allocate_excluded)
#     )
 
#     df["excluded_price"] = df["excluded_price"].fillna(0)
 
#     print("🔹 Excluded Price Allocation Done")
 
#     # -------------------------------------------------
#     # STEP 8 — Included Price
#     # included = sum(matl_clp * scrap_qty) * (1-dis) - excluded_price
#     # -------------------------------------------------
#     def calculate_included_price(group):
 
#         remaining_qty = group["excluded_quantity"].iloc[0]
    
#         group = group.sort_values("effective_price")
    
#         included_value = 0
    
#         for _, row in group.iterrows():
    
#             row_scrap_qty = row["target quantity"]
    
#             # First allocate excluded portion
#             take_excluded = min(remaining_qty, row_scrap_qty)
#             remaining_qty -= take_excluded
    
#             # Remaining in that row is included
#             included_qty = row_scrap_qty - take_excluded
    
#             if included_qty > 0:
#                 included_value += included_qty * row["effective_price"]
    
#         # discount = group["discount"].iloc[0] if "discount" in group else 0
#         discount = group["discount"].max() if "discount" in group.columns else 0
#         included_value *= (1 - discount)
    
#         group["included_price"] = included_value
    
#         return group
    
    
#     df = (
#         df.groupby("genpact_part_keywords", group_keys=False)
#         .apply(calculate_included_price)
#     )
 
#     print("🔹 Included Price Calculated")
 
#     # -------------------------------------------------
#     # FINAL DEBUG SUMMARY
#     # -------------------------------------------------
#     print("📊 FINAL SUMMARY")
#     print("Total Rows:", len(df))
#     print("Scrap Yes:", (df["scrap_exclusion"].str.lower() == "yes").sum())
#     print("Scrap No:", (df["scrap_exclusion"].str.lower() == "no").sum())
#     # print("Total Excluded Price:", df["excluded_price"].sum())
#     # print("Total Included Price:", df["included_price"].sum())
 
#     print("✅ Scrap Exclusion Calculation Completed")
 
#     return df

def scrap_exclusion_calculation(df, result, pipeline):
 
 
    print("🔎 Starting Scrap Exclusion Calculation")
 
    # -------------------------------------------------
    # Load Required Sheets
    # -------------------------------------------------
    qpe_list = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="qpe_list"
    )

    if pipeline == 'repair':
        billing_df = get_dataframe_by_file_and_sheet(
            all_dataframes=result["dataframes"],
            file_key="internal",
            sheet_name_substring="Sheet1"
        )
 
    else:
        billing_df = get_dataframe_by_file_and_sheet(
            all_dataframes=result["dataframes"],
            file_key="billing",
            sheet_name_substring="billing"
        )
 
    cc_part_num = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="cc part number"
    )
    
    cost_category_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="cost category"
    )
    
    
 
    # Normalize columns
    df.columns = [col.strip().lower() for col in df.columns]
    billing_df.columns = [col.strip().lower() for col in billing_df.columns]
    qpe_list.columns = [col.strip().lower() for col in qpe_list.columns]
    cc_part_num.columns = [col.strip().lower() for col in cc_part_num.columns]
    cost_category_df.columns = [col.strip().lower() for col in cost_category_df.columns]
    
    print("✅ Sheets Loaded Successfully")
    
    
    # -------------------------------------------------
    # Extract Customer Furnished Material List
    # -------------------------------------------------
    customer_col = next(
        (c for c in cost_category_df.columns if "customer" in c and "material" in c),
        None
    )
    
    if not customer_col:
        raise ValueError("Customer furnished Material column not found")
    
    customer_priority_set = set(
        cost_category_df[customer_col]
        .dropna()
        .astype(str)
        .str.strip()
        .str.upper()
    )
    
    print("🔹 Customer Furnished Material Categories Loaded")
    print("Count:", len(customer_priority_set))
    
 
 
    # -------------------------------------------------
    # STEP 0 — Create Mapping (NO DUPLICATION)
    # -------------------------------------------------
    if pipeline == 'repair':
        material_col = next((c for c in billing_df.columns if "inspected" in c), None)
    else:
        material_col = next((c for c in billing_df.columns if "material" in c), None)
    part_num_col = next((c for c in cc_part_num.columns if "part_num" in c or "part number" in c), None)
    part_keyword_col = next((c for c in cc_part_num.columns if "part_keyword" in c), None)
 
    if not material_col or not part_num_col or not part_keyword_col:
        raise ValueError("Required columns not found")
 
    billing_df[material_col] = billing_df[material_col].astype(str)
    cc_part_num[part_num_col] = cc_part_num[part_num_col].astype(str)
 
    mapping_dict = (
        cc_part_num
        .drop_duplicates(subset=[part_num_col])
        .set_index(part_num_col)[part_keyword_col]
        .to_dict()
    )
 
    billing_df["genpact_part_keywords"] = billing_df[material_col].map(mapping_dict)
 
    print("🔹 Keywords mapped")
 
    # -------------------------------------------------
    # STEP 1 — Total Scrap Quantity
    # -------------------------------------------------
    if pipeline == "repair":
        quantity_col = "inspected quantity"
        billing_grouped = (
            billing_df
            .groupby("genpact_part_keywords", dropna=False)[quantity_col]
            .mean()
            .reset_index()
            .rename(columns={quantity_col: "total_scrap_quantity"})
        )
    else:
        quantity_col = "target quantity"
        billing_grouped = (
            billing_df
            .groupby("genpact_part_keywords", dropna=False)[quantity_col]
            .sum()
            .reset_index()
            .rename(columns={quantity_col: "total_scrap_quantity"})
        )
 
    df = df.merge(
        billing_grouped,
        on="genpact_part_keywords",
        how="left",
        validate="m:1"
    )
 
    df["total_scrap_quantity"] = df["total_scrap_quantity"].fillna(0)
 
    print("🔹 Total Scrap Quantity merged")
 
    # -------------------------------------------------
    # STEP 2 — Merge QPE
    # -------------------------------------------------
    qpe_grouped = (
        qpe_list[["part keyword", "qpe"]]
        .rename(columns={
            "part keyword": "genpact_part_keywords",
            "qpe": "total_qpe"
        })
    )
 
    df = df.merge(
        qpe_grouped,
        on="genpact_part_keywords",
        how="left"
    )
 
    df["total_qpe"] = df["total_qpe"].replace(0, np.nan)
 
    print("🔹 QPE merged")
 
    # -------------------------------------------------
    # STEP 3 — Actual Scrap Rate
    # -------------------------------------------------
    df["actual_scrap_rate"] = (
        df["total_scrap_quantity"] / df["total_qpe"]
    )
 
    print("🔹 Actual Scrap Rate calculated")
 
    # -------------------------------------------------
    # STEP 4 — Compare scrap_cap vs actual rate
    # -------------------------------------------------
    mask = df["scrap_exclusion"].str.lower() == "yes"
 
    condition = mask & (
        df["scrap_cap"] > df["actual_scrap_rate"] * 100
    )
 
    df.loc[condition, "scrap_exclusion"] = "No"
 
    print("🔹 Scrap Cap comparison done")
 
    # -------------------------------------------------
    # STEP 5 — Excluded / Included Quantity
    # -------------------------------------------------
    df["allowed_quantity"] = df["total_qpe"] * (df["scrap_cap"] / 100)
 
    df["excluded_quantity"] = (
        df["total_scrap_quantity"] - df["allowed_quantity"]
    )
 
    df.loc[df["excluded_quantity"] < 0, "excluded_quantity"] = 0
 
    df["included_quantity"] = (
        df["total_scrap_quantity"] - df["excluded_quantity"]
    )
 
    print("🔹 Excluded & Included quantities calculated")
 
    # -------------------------------------------------
    # STEP 6 — Effective Price
    # -------------------------------------------------
    if 'total_price_final' in df.columns:
        df["effective_price"] = df["total_price_final"]
    else:
        df["effective_price"] = df["matl_clp_lookup"]

    df.to_excel("merged_df_after_scrap_calculation_TEST.xlsx", index=False)
    
    
    if 'cost' in df.columns:
        df.loc[
            (df["effective_price"].isna()) | (df["effective_price"] == 0),
            "effective_price"
        ] = df["cost"]

    else:
        df.loc[
            (df["effective_price"].isna()) | (df["effective_price"] == 0),
            "effective_price"
        ] = df["amount"]
 
    print("🔹 Effective price determined")
 
    # -------------------------------------------------
    # STEP 7 & 8 — Row-Level Allocation + Totals
    # -------------------------------------------------
    def allocate_prices(group):
 
        group = group.copy()
    
        # -------------------------------------------------
        # PRIORITY LOGIC USING CUSTOMER FURNISHED LIST
        # -------------------------------------------------
        if 'cost category' not in group.columns:
            group["cost category"] = "REPAIR"
        group["priority_flag"] = (
            group["cost category"]
            .astype(str)
            .str.strip()
            .str.upper()
            .isin(customer_priority_set)
        ).astype(int)
    
        # Sort:
        # 1️⃣ Customer furnished first
        # 2️⃣ Then cheapest inside
        group = group.sort_values(
            by=["priority_flag", "effective_price"],
            ascending=[False, True]
        )
    
        remaining_excluded = group["excluded_quantity"].iloc[0]
    
        group["excluded_qty_indiv"] = 0.0
        group["included_qty_indiv"] = 0.0
        group["excluded_price_indiv"] = 0.0
        group["included_price_indiv"] = 0.0
    
        for idx, row in group.iterrows():
            
            if 'target quantity' in group.columns:
                row_scrap_qty = row["target quantity"]
            else:
                row_scrap_qty = row["inspected quantity"]
    
            take_excluded = min(remaining_excluded, row_scrap_qty)
            included_qty = row_scrap_qty - take_excluded
    
            group.at[idx, "excluded_qty_indiv"] = take_excluded
            group.at[idx, "included_qty_indiv"] = included_qty
    
            group.at[idx, "excluded_price_indiv"] = take_excluded * row["effective_price"]
            group.at[idx, "included_price_indiv"] = included_qty * row["effective_price"]
    
            remaining_excluded -= take_excluded
    
        # Apply discount
        discount = group["discount"].max() if "discount" in group.columns else 0
    
        group["excluded_price_indiv"] *= (1 - discount)
        group["included_price_indiv"] *= (1 - discount)
    
        # Totals per part
        group["excluded_price"] = group["excluded_price_indiv"].sum()
        group["included_price"] = group["included_price_indiv"].sum()
    
        group.drop(columns=["priority_flag"], inplace=True)
    
        return group
 
    df = (
        df.groupby("genpact_part_keywords", group_keys=False)
          .apply(allocate_prices)
    )
 
    print("🔹 Allocation completed")
 
    # -------------------------------------------------
    # FINAL DEBUG SUMMARY
    # -------------------------------------------------
    print("📊 FINAL SUMMARY")
    print("Total Rows:", len(df))
    print("Scrap Yes:", (df["scrap_exclusion"].str.lower() == "yes").sum())
    print("Scrap No:", (df["scrap_exclusion"].str.lower() == "no").sum())
 
    print("✅ Scrap Exclusion Calculation Completed")
 
    return df

def apply_blade_inclusion(result: dict,
                          main_df: pd.DataFrame,
                          pipeline_type: str,
                          workscope: str) -> pd.DataFrame:
 
    # -------------------------------------------------
    # 1️⃣ Load Inclusion Sheet
    # -------------------------------------------------
    inclusion_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="inclusion"
    )
    
    print("Inclusion_DF")
    inclusion_df.to_excel("inclusion_df.xlsx", index=False)
    
    
    inclusion_df = inclusion_df.copy()
    inclusion_df.columns = [col.strip().lower() for col in inclusion_df.columns]
    
 
    main_df = main_df.copy()
    main_df.columns = [col.strip().lower() for col in main_df.columns]
    
 
    pipeline_type = pipeline_type.strip().lower()
    workscope_norm = workscope.strip().lower()
 
    # -------------------------------------------------
    # 2️⃣ Define UNKNOWN remark
    # -------------------------------------------------
    if pipeline_type == "material":
        unknown_remark = "MATL UNKNOWN"
    elif pipeline_type == "repair":
        unknown_remark = "REPAIR UNKNOWN"
    elif pipeline_type == "vendor":
        unknown_remark = "VENDOR UNKNOWN"
    else:
        raise ValueError("Invalid pipeline_type")
 
    # -------------------------------------------------
    # 3️⃣ Detect Workscope Column
    # -------------------------------------------------
    workscope_col_index = None
    for idx, col in enumerate(inclusion_df.columns):
        if workscope_norm in col:
            workscope_col_index = idx
            break
 
    if workscope_col_index is None:
        raise ValueError(f"Workscope '{workscope}' not found")
 
    # -------------------------------------------------
    # 4️⃣ Select Target Column
    # -------------------------------------------------
    if pipeline_type in ["material"]:
        target_col = inclusion_df.columns[workscope_col_index]
    else:
        target_col = inclusion_df.columns[workscope_col_index + 1]
 
    # -------------------------------------------------
    # 5️⃣ Detect Part Description Column
    # -------------------------------------------------
    part_desc_col = next(
        (c for c in inclusion_df.columns
         if "nomenclature" in c or "type" in c or "part" in c),
        None
    )
 
    if not part_desc_col:
        raise ValueError("Part description column not found")
 
    # -------------------------------------------------
    # 6️⃣ Normalize Columns
    # -------------------------------------------------
    inclusion_df["norm_description"] = inclusion_df[part_desc_col].apply(normalize_phrase)
 
    if "genpact_part_keywords" not in main_df.columns:
        raise ValueError("genpact_part_keywords missing")
 
    main_df["keyword_norm"] = main_df["genpact_part_keywords"].apply(normalize_phrase)
 
    # -------------------------------------------------
    # 7️⃣ Merge (LEFT JOIN)
    # -------------------------------------------------
    merged_df = main_df.merge(
        inclusion_df[["norm_description", target_col]],
        left_on="keyword_norm",
        right_on="norm_description",
        how="left"
    )
 
    # Initialize scrap_exclusion
    merged_df["scrap_exclusion"] = "Unknown"
 
    # -------------------------------------------------
    # 8️⃣ Classification Logic
    # -------------------------------------------------
    def classify_row(row):
 
        current_remark = str(row.get("remarks", "")).upper()
 
        # If not pipeline UNKNOWN → do nothing
        if current_remark != unknown_remark:
            return row.get("remarks"), "Unknown"
 
        # If no match
        if pd.isna(row["norm_description"]):
            return unknown_remark, "Unknown"
 
        # Matched case
        status = str(row[target_col]).strip().lower()
 
        if status == "included":
            return "Inclusion", "No"
        else:
            return unknown_remark, "Yes"
 
    results = merged_df.apply(classify_row, axis=1)
 
    merged_df["remarks"] = [r[0] for r in results]
    merged_df["scrap_exclusion"] = [r[1] for r in results]
    
    
 
    merged_df.drop(columns=["norm_description"], inplace=True, errors="ignore")
    
    # -------------------------------------------------
    # 9️⃣ Scrap Cap Logic (Final Clean Version)
    # -------------------------------------------------
    
    def calculate_scrap_cap(row):
 
        status_raw = row.get(target_col)
    
        if pd.isna(status_raw):
            return None
    
        # If numeric (Excel decimal case like 0.2)
        if isinstance(status_raw, (int, float)):
            # If between 0 and 1 → treat as decimal percentage
            if 0 <= status_raw <= 1:
                return round(status_raw * 100, 2)
            else:
                return round(status_raw, 2)
    
        status_value = str(status_raw).strip().lower()
    
        # Excluded → 0
        if "excluded" in status_value:
            return 0
    
        # Included → 100
        if "included" in status_value:
            return 100
    
        # Extract percentage like 40%
        percent_match = re.search(r'(\d+)\s*%', status_value)
        if percent_match:
            return float(percent_match.group(1))
    
        # If pure numeric string like "40"
        if status_value.replace(".", "", 1).isdigit():
            num = float(status_value)
            if 0 <= num <= 1:
                return round(num * 100, 2)
            return num
    
        return None
    
    
    merged_df["scrap_cap"] = merged_df.apply(calculate_scrap_cap, axis=1)
    
    merged_df.to_excel("merged_df_scrap_cal.xlsx", index=False)
    
    
    merged_df = scrap_exclusion_calculation(merged_df,result, pipeline_type)
    
    # -------------------------------------------------
    # FINAL REMARK UPDATE BASED ON SCRAP
    # -------------------------------------------------
    merged_df.loc[
        merged_df["scrap_exclusion"].str.lower() == "yes",
        "remarks"
    ] = f"Exclusion as Scrap {pipeline_type}"
    
    merged_df.to_excel("merged_df_after_scrap_calculation.xlsx", index=False)
    
    return merged_df
 
# Rule-3: Missing on receipt Exclusion
def apply_missing_receipts_exclusion(df: pd.DataFrame) -> pd.DataFrame: 
    df = df.copy()
    print("saving missing receipt")
    df.to_excel("missing_rec_1.xlsx",index=False)
    df.columns = [col.strip().lower() for col in df.columns]

    # Step 1: Detect the relevant column
    irc_desc_col = next((col for col in df.columns if "irc description" in col.lower()), None)
    remarks_col = next((col for col in df.columns if "remarks" in col), None)

    if not all([irc_desc_col, remarks_col]):
        raise ValueError("Required column not found in DataFrame.")

    # Step 2: Apply exclusion logic
    condition_unkwn = df[remarks_col].str.strip().str.upper() == "MATL UNKNOWN"
    df_unkwn = df[condition_unkwn].copy() 

    variants = [
        "missing on receipt", "missing receipt", "receipt missing",
        "missing upon receipt", "receipt not found", "not received",
        "missing during receving"
    ]

    pattern = "|".join(variants)
    match_condition = df_unkwn[irc_desc_col].str.contains(pattern, case=False, na=False)

    df_unkwn.loc[match_condition, remarks_col] = "Exclusion as Missing on Receipt" 
    df_final = pd.concat([df[~condition_unkwn], df_unkwn], ignore_index=True)
    df_final.drop_duplicates(inplace=True)

    remarks_expected = ("Exclusion as Missing on Receipt",)
    mor_excl_df = df_final[df_final["remarks"].str.startswith(remarks_expected,na=False)].copy()
    mor_excl_df.drop_duplicates(inplace=True)    
    return df_final

# Rule-4: PMA/DER - Issue reason code (14, 35) - Exclusion
def apply_pma_der_exclusion(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [col.strip().lower() for col in df.columns]
    
    # Step 1: Detect the relevant column
    issue_reason_col = next((col for col in df.columns if "issue reason code" in col), None) #14, 35
    remarks_col = next((col for col in df.columns if "remarks" in col), None)
 
    if not all([issue_reason_col, remarks_col]):
        raise ValueError("Required column 'Issue Reason Code' not found in DataFrame.")
 
    # Step 2: Apply exclusion logic
    condition_unkwn = df[remarks_col].str.strip().str.upper() == "MATL UNKNOWN"
    df_unkwn = df[condition_unkwn].copy()

    condition_pma_der = df_unkwn[issue_reason_col].isin([14, 35])
    df_unkwn.loc[condition_pma_der, remarks_col] = "Exclusion as PMA & DER"     
    df_final = pd.concat([df[~condition_unkwn], df_unkwn], ignore_index=True)
    df_final.drop_duplicates(inplace=True)
    # print("[INFO] ✅ 'PMA/DER Codes Exclusion' applied successfully.")
    
    remarks_expected = ("Exclusion as PMA & DER",)
    pma_der_excl = df_final[df_final["remarks"].str.startswith(remarks_expected,na=False)].copy()
    pma_der_excl.drop_duplicates(inplace=True)    
    return df_final

def apply_sb_cap(df: pd.DataFrame, pipeline_type: str, sb_remarks_col: str, price_col: str):
    if pipeline_type == "material":
        total_included_price_col = 'total_mtl_sb_included_value'
    elif pipeline_type == "repair":
        total_included_price_col = 'total_repair_sb_included_value'
        
        
    # 1. Create mask for target rows
    mask = df[sb_remarks_col] == 'Inclusion as SB'
    
    # 2. Calculate the total for those specific rows
    total_sum = df.loc[mask, price_col].sum()
    
    # 3. Initialize the column with total_sum
    df.loc[mask, total_included_price_col] = total_sum
    
    # 4. If sum > 100k, update only the target rows
    if total_sum > 100000:
        df.loc[mask, sb_remarks_col] = "Exclusion as Inclusion as SB"
        
    return df

def apply_sb_logic(result: dict, billing_df: pd.DataFrame) -> pd.DataFrame:
    print("\n===== APPLY SB LOGIC START =====")
    # --------------------------------------------------
    # Step 1: Load SB data and extract SB codes
    # --------------------------------------------------
    sb_df1 = get_dataframe_by_file_and_sheet(result["dataframes"], "service bullet", "ZE14")
    sb_df2 = get_dataframe_by_file_and_sheet(result["dataframes"], "SB", "ZE57")
    
 
    print("ZE14 count:", len(sb_df1))
    print("ZE57 count:", len(sb_df2))
 
    sb_df2_slim = sb_df2[["SB Number", "Notification Status"]]
 
    sb_df = sb_df1.merge(sb_df2_slim, on="SB Number", how="left")
    sb_df.columns = [col.strip().lower() for col in sb_df.columns]
    print("After ZE14 + ZE57 merge:", len(sb_df))
 
    # --------------------------------------------------
    # Step 2: Extract valid SB codes
    # --------------------------------------------------
    def extract_valid_code(sb_value):
        if isinstance(sb_value, str) and sb_value.startswith("5B/"):
            after_slash = sb_value.split("5B/")[1]
            if len(after_slash) == 7:
                return after_slash
        return None
 
    sb_df["extracted_code"] = sb_df["sb number"].apply(extract_valid_code)
 
    print("Valid extracted_code count:", sb_df["extracted_code"].notna().sum())
 
    sb_df = sb_df[sb_df["extracted_code"].notna()].copy()
    print("After removing null extracted_code:", len(sb_df))
 
    all_sb_codes = sb_df["extracted_code"].unique()
    sb_codes = [
        c for c in all_sb_codes
        if len(str(c)) == 7 and str(c).startswith("7")
    ]
 
    print("Unique SB codes starting with 7:", len(sb_codes))
 
    # --------------------------------------------------
    # Step 3: CMP + category filtering
    # --------------------------------------------------
    sb_df["oem category"] = pd.to_numeric(sb_df["oem sb category"], errors="coerce")
    sb_df["amended category"] = pd.to_numeric(sb_df["amend sb category"], errors="coerce")
    
    sb_filtered = sb_df[
        (
            (sb_df["status"].str.strip().str.upper() == "CMP") |
            (sb_df["notification status"].str.strip().str.upper() == "CMP")
        ) &
        (sb_df["oem category"] >= 7) &
        (sb_df["amended category"] <= 6)
    ].copy()
    # sb_filtered.to_excel("sb_filtered.xlsx", index=False)
    print("After CMP + category filter:", len(sb_filtered))
 
    # --------------------------------------------------
    # Step 4: Load SB Contract mapping
    # --------------------------------------------------
    sb_contract_df = get_dataframe_by_file_and_sheet(
        result["dataframes"], "contract", "sb"
    )
    sb_contract_df.columns = [str(col).strip().lower() for col in sb_contract_df.columns]
 
    print("Contract SB sheet count:", len(sb_contract_df))
 
    sb_data_merged = sb_contract_df[
        sb_contract_df["sb"].isin(sb_codes)
    ].copy()
    
    print("Contract matched SB count:", len(sb_data_merged))
 
    # --------------------------------------------------
    # Step 5: Merge SB with contract (SB-level)
    # --------------------------------------------------
    sb_filtered_sb_incl_excl = sb_filtered.merge(
        sb_data_merged[
            ["sb", "mtl sb remarks", "category", "new part number", "sb description"]
        ],
        left_on="extracted_code",
        right_on="sb",
        how="left"
    )
 
    # IMPORTANT: create sb_x BEFORE using it
    sb_filtered_sb_incl_excl["sb_x"] = sb_filtered_sb_incl_excl["extracted_code"]
    print("After SB-contract merge:", len(sb_filtered_sb_incl_excl))
    print("sb_x present count:", sb_filtered_sb_incl_excl["sb_x"].notna().sum())
    
    ''' Filters the data to keep only one row per part number. If a part has multiple entries 
    across different Service Bulletins (SBs), it prioritizes the row with a status of 'CMP'.'''
    sb_filtered_sb_incl_excl['is_cmp'] = sb_filtered_sb_incl_excl['status'] == 'CMP'
    sb_filtered_sb_incl_excl = sb_filtered_sb_incl_excl.sort_values(['new part number', 'is_cmp'], ascending=[True, False])
    sb_filtered_sb_incl_excl = sb_filtered_sb_incl_excl.drop_duplicates(subset='new part number', keep='first').drop(columns='is_cmp')
    
    # sb_filtered_sb_incl_excl.to_excel("sb_filtered_sb_incl_excl.xlsx", index=False)
    
    # --------------------------------------------------
    # Step 6: Propagate sb_x back to sb_filtered
    # --------------------------------------------------
    sb_filtered = sb_filtered.merge(
        sb_filtered_sb_incl_excl[["extracted_code", "sb_x"]],
        on="extracted_code",
        how="left"
    )
    sb_filtered.to_excel("sb_filtered after sb_x propagation.xlsx", index=False)
    print("sb_filtered after sb_x propagation:", len(sb_filtered))
    print("sb_filtered sb_x non-null:", sb_filtered["sb_x"].notna().sum())
 
    # --------------------------------------------------
    # Step 7: Part-level merge with billing
    # --------------------------------------------------
    sb_filtered_sb_incl_excl_pn_pk = billing_df.merge(
        sb_filtered_sb_incl_excl,
        left_on="material",
        right_on="new part number",
        how="left"
    )
    sb_filtered_sb_incl_excl_pn_pk.to_excel("After billing merge.xlsx", index=False)
    print("After billing merge:", len(sb_filtered_sb_incl_excl_pn_pk))
 
    # --------------------------------------------------
    # Step 8: Default remarks
    # --------------------------------------------------
    sb_filtered_sb_incl_excl_pn_pk.loc[
        sb_filtered_sb_incl_excl_pn_pk["mtl sb remarks"].isna(),
        "mtl sb remarks"
    ] = "Exclusion as no matching SB"
 
    # --------------------------------------------------
    # Step 9: Final SB-applied rows
    # --------------------------------------------------
    apply_sb_logic_df = sb_filtered_sb_incl_excl_pn_pk[
        sb_filtered_sb_incl_excl_pn_pk["sb_x"].notna()
    ].copy()
 
    print("Rows with SB applied:", len(apply_sb_logic_df))
 
    # --------------------------------------------------
    # Step 10: Category exclusion
    # --------------------------------------------------
    apply_sb_logic_df = apply_sb_logic_df[
        apply_sb_logic_df["category"] != 9
    ].copy()
 
    print("After category != 9 filter:", len(apply_sb_logic_df))
    
    # --------------------------------------------------
    # Step 11: Apply Cap
    # --------------------------------------------------
    # apply_sb_logic_df = apply_sb_cap(apply_sb_logic_df, pipeline_type='material')
    # apply_sb_logic_df = apply_sb_cap(df= sb_filtered_sb_incl_excl_pn_pk, pipeline_type= "material", sb_remarks_col= "mtl sb remarks", price_col="matl_clp_lookup")
    apply_sb_logic_df.to_excel("service_bulletin_remarks.xlsx", index=False)
    
 
    print("===== APPLY SB LOGIC END =====\n")
    return apply_sb_logic_df, sb_filtered
def identify_creap_and_adders(result: dict, main_df: pd.DataFrame,):
   pass
    
def apply_workscope_inclusion_exclusion(
    result: dict,
    main_df: pd.DataFrame,
    pipeline_type: str, #material and repair and Vendor
    remarks: str,
    workscope : str = None
    ) -> pd.DataFrame:
    """
        Apply workscope inclusion/exclusion logic:
        - Merge main_df with sap_df to get Customer Workscope Level
        - Create ata4 for filtering
        - Apply rules for Inclusion as 72, WKS full, Exclusion as 0
        - Merge with WSPG for A/B coverage (for WKS=1,2)
        - Add 'covered_status' column (Covered, Not Covered, No WSPG match)
    """
    df = main_df.copy() #internal_repair_df
    df.columns = [c.strip().lower() for c in df.columns]
    workscope = workscope.lower()
    
    # ---------- Load WSPG ----------
    wspg = get_dataframe_by_file_and_sheet(
            all_dataframes=result["dataframes"],
            file_key="contract",
            sheet_name_substring="wspg incl excl"
        ).copy()
    
    # ---------- Load SAP ----------
    sap_df = get_dataframe_by_file_and_sheet(
            all_dataframes=result["dataframes"],
            file_key="sap",
            sheet_name_substring="ze14"
    ).copy()
    
    # ==========================================================
    # ---------- Load Workscope Level Contract Summary----------
    # ==========================================================
    wks_level_df = get_dataframe_by_file_and_sheet(
            all_dataframes=result["dataframes"], 
            file_key="contract", 
            sheet_name_substring="workscope level"
        )
    wks_level_df.columns = [str(c).strip().lower() for c in wks_level_df.columns]
    wks_level_df = wks_level_df[["wokscope defnitions", workscope]]
    
    # Convert empty strings to NaN, then drop all NaNs in that column
    wks_level_df = wks_level_df.replace('', np.nan).dropna(subset=[workscope])
    # Extracts 2 digits, a hyphen, and 2 digits from the start of the string
    wks_level_df['workscope_ata'] = wks_level_df['wokscope defnitions'].str.extract(r'^(\d{2})-(\d{2})').fillna('').add_prefix('').sum(axis=1)
    
    # Clean workscope
    wks_level_df[workscope] = pd.to_numeric(wks_level_df[workscope], errors='coerce')
    wks_level_df = wks_level_df.dropna(subset=[workscope])
    wks_level_df[workscope] = wks_level_df[workscope].astype(int)

    # Clean 'workscope_ata'
    wks_level_df['workscope_ata'] = pd.to_numeric(wks_level_df['workscope_ata'], errors='coerce')
    wks_level_df = wks_level_df.dropna(subset=['workscope_ata'])
    wks_level_df['workscope_ata'] = wks_level_df['workscope_ata'].astype(int)

    #    print(f"wks_level_df : {wks_level_df.shape} === {wks_level_df}")
        
        
    # --- Step 1: Normalize SAP ---
    sap_df.columns = [c.strip().lower() for c in sap_df.columns]
    sap_ata_col = next((c for c in sap_df.columns if "ata task reference" in c), None)
    sap_wks_col = next((c for c in sap_df.columns if "customer workscope level" in c), None)
    
    sap_df["ata_long"] = sap_df[sap_ata_col].fillna("").astype(str).str.extract(r"5B[_\- ]?(\d{6})", expand=False)
    sap_df = sap_df[["ata_long", sap_wks_col]].dropna().drop_duplicates()
    
    # --- Step 2: Normalize main_df ATA ---
    ata_chap_col = next((c for c in df.columns if "ata chapter" in c), None)
    ata_sec_col  = next((c for c in df.columns if "ata section" in c), None)
    csn_col      = next((c for c in df.columns if "component" in c and "csn" in c), None)
    def extract_ata(row):
        if ata_chap_col and pd.notna(row.get(ata_chap_col)):
            return re.sub(r"[^0-9]", "", str(row[ata_chap_col])).ljust(6, "0")[:6]
        elif ata_sec_col and pd.notna(row.get(ata_sec_col)):
            return re.sub(r"[^0-9]", "", str(row[ata_sec_col])).ljust(6, "0")[:6]
        elif csn_col and pd.notna(row.get(csn_col)):
            match = re.search(r"07[_\- ]?(\d{6})", str(row[csn_col]))
            if match:
                return match.group(1)
        return None
    
    # df["ata_long"] = df.apply(extract_ata, axis=1)
    
    # --- Step 3: Merge with SAP to get WKS level ---
    df = df.merge(sap_df.rename(columns={"ata_long": "_atalong6"}), left_on="ata_long_v1", right_on="_atalong6", how="left")
    df = df.rename(columns={sap_wks_col: "Customer_Workscope_Level"})
    
    # --- Step 4: Create ata4 ---
    #df["ata4"] = df["ata_long"].astype(str).str[:4]
    
    # --- Step 5: Normalize WSPG ---
    wspg.columns = [c.strip().lower() for c in wspg.columns]
    #    print("wspg.columns :", wspg.columns)
    
    wspg_ata_col = next((c for c in wspg.columns if "atalong" in c), None)
    wspg_a_col = "a"
    wspg_b_col = "b"
    
    wspg["atalong6"] = wspg[wspg_ata_col].astype(str).str.extract(r"(\d{6})", expand=False)
    wspg = wspg[["atalong6", wspg_a_col, wspg_b_col]].dropna(subset=["atalong6"])
    
    # --- Step 6: Merge WSPG ---
    #    df.to_excel("df.xlsx", index=False)
    #    wspg.to_excel("wspg.xlsx", index=False)
    
    df["ata4"] = df["ata_long"].astype(str).str.extract(r"(\d{4})", expand=False)
    wspg["ata4"] = wspg["atalong6"].astype(str).str.extract(r"(\d{4})", expand=False)
    
    df = df.merge(wspg, left_on="ata_long", right_on="atalong6", how="left")
    df.drop_duplicates(inplace=True)
    
    df.to_excel("merged_df_wspg.xlsx", index=False)
    
    df["wks_num"] = pd.to_numeric(
        df["Customer_Workscope_Level"],
        errors="coerce"
    )
    
    df["ata4"] = df["ata4_x"].astype("Int64")
    
    # ==========================================================
    # Merge the main df with the lookup table on the ATA column
    # ==========================================================
    df = df.merge(
        wks_level_df[['workscope_ata', workscope]].rename(columns={workscope: 'cs_wks_level'}), 
        left_on='ata4',
        right_on='workscope_ata',
        how='left'
        )
    
    # --- Step 7: Initialize covered_status ---
    df["covered_status"] = "No WSPG match"
    
    # --- Step 8: Apply Rules, but only if remarks == "MATL UNKNOWN" ---
    if pipeline_type == "material":
        target_mask = df["remarks"].astype(str).str.upper().eq("MATL UNKNOWN")
    elif pipeline_type == "vendor":
        target_mask = df["remarks"].astype(str).str.upper().eq("VEND UNKNOWN")
    else:
        target_mask = df["remarks"].astype(str).str.upper().eq("REPAIR UNKNOWN")
    
    
    remarks_pp = df["remarks"].astype(str).str.strip().str.upper()
    
    # Common inclusions to always include
    extra_inclusions = {"INCLUSION AS FAN-BLADE", "INCLUSION AS BLADE-HPT"}
    
    if pipeline_type == "material":
        base = {"MATL UNKNOWN"}
    elif pipeline_type == "vendor":
        base = {"VEND UNKNOWN"}
    else:
        base = {"REPAIR UNKNOWN"}
    
    # Combine base condition with extra inclusions
    target_mask_0 = remarks_pp.isin(base | extra_inclusions)
    
    
    # Inclusion as 72
    mask72 = target_mask & df["ata4"].between(7200, 7209, inclusive="both")#7200-7209, 7210,7220,7230,7240,7250,7260 then take CSN comment-
    df.loc[mask72, ["remarks", "covered_status"]] = ["Inclusion as 72", "Covered"]
    
    df.to_excel("apply_workscope_inclusion_exclusion_V1.xlsx", index=False)
    
    
    # WKS=3 → Inclusion full
    if pipeline_type == "material":
        #    mask_full = target_mask & df["ata4"].between(7221, 7263, inclusive="both") & (((df["Customer_Workscope_Level"] == 3) | df["final ata ws"] == 3|(df["Customer_Workscope_Level"] == "3")))      
        mask_full = (
            target_mask
            & df["ata4"].between(7221, 7263, inclusive="both")
            & (df["cs_wks_level"] == 3)
        )
    elif pipeline_type == "repair":
        #    mask_full = target_mask & df["ata4"].between(7221, 7263, inclusive="both") & ((df["Customer_Workscope_Level"] == "3") | (df["Customer_Workscope_Level"] == 3))
        mask_full = (
            target_mask
            & df["ata4"].between(7221, 7263, inclusive="both")
            & (df["cs_wks_level"] == "3")
        )    
    
    elif pipeline_type == "vendor":
        #    mask_full = target_mask & df["ata4"].between(7221, 7263, inclusive="both") & (((df["Customer_Workscope_Level"] == 3) | df["final ata ws"] == 3))
        mask_full = (
            target_mask
            & df["ata4"].between(7221, 7263, inclusive="both")
            & (df["cs_wks_level"] == 3)
        )
        
    
    df.loc[mask_full, ["remarks", "covered_status"]] = ["Inclusion as WKS is full", "Covered"]
    
    df.to_excel("apply_workscope_inclusion_exclusion_V2.xlsx", index=False)
    
    
    # WKS=0 → Exclusion
    if pipeline_type == "material":
        mask0 = target_mask_0 & df["ata4"].between(7221, 7263, inclusive="both") & ((df["final ata ws"] == "P110") | (df["final ata ws"] == "SP110") | (df["final ata ws"] == 0) | (df["cs_wks_level"] == 0))
    elif pipeline_type == "repair":
        mask0 = target_mask_0 & df["ata4"].between(7221, 7263, inclusive="both") & (df["cs_wks_level"] == 0)        
    elif pipeline_type == "vendor":
        mask0 = target_mask_0 & df["ata4"].between(7221, 7263, inclusive="both") & (((df["final ata ws"] == "P110") | (df["final ata ws"] == "SP110") | (df["final ata ws"] == 0)) | (df["cs_wks_level"] == 0))
    
    df.loc[mask0, ["remarks", "covered_status"]] = ["Exclusion as WKS = 0", "Not Covered"]
    
    
    # -----FOD-----
    if pipeline_type == "material":
        mask0 = target_mask_0 & df["ata4"].between(7221, 7263, inclusive="both") & (df["Customer_Workscope_Level"] == "FOD")
    elif pipeline_type == "repair":
        mask0 = target_mask_0 & df["ata4"].between(7221, 7263, inclusive="both") & (df["Customer_Workscope_Level"] == "FOD")      
    elif pipeline_type == "vendor":
        mask0 = target_mask_0 & df["ata4"].between(7221, 7263, inclusive="both") & (df["Customer_Workscope_Level"] == "FOD")
    
    df.loc[mask0, ["remarks", "covered_status"]] = ["Exclusion as FOD", "Not Covered"]
    #df.to_excel("check_remarks.xlsx", index=False)
    
    a_val = df[wspg_a_col].astype(str).str.strip().str.upper()
    b_val = df[wspg_b_col].astype(str).str.strip().str.upper()
    
    # WKS=1 → Check column A
    if pipeline_type == "material":
        mask1 = target_mask & df["ata4"].between(7221, 7263, inclusive="both") & (df["cs_wks_level"] == 1)
    elif pipeline_type == "repair":
        mask1 = target_mask & df["ata4"].between(7221, 7263, inclusive="both") & (df["cs_wks_level"] == 1)      
    elif pipeline_type == "vendor":
        mask1 = target_mask & df["ata4"].between(7221, 7263, inclusive="both") & (df["cs_wks_level"] == 1)
    
    #    df.loc[mask1 & (df[wspg_a_col].astype(str).str.strip().str.upper() == "X"), ["remarks", "covered_status"]] = ["Inclusion as WKS = 1", "Covered"]
    #    df.loc[mask1 & (df[wspg_a_col].isnull()), ["remarks", "covered_status"]] = ["Exclusion as WKS = 1", "Not Covered"]
    
    df.loc[mask1 & (a_val == "X"), ["remarks", "covered_status"]] = ["Inclusion as WKS = 1", "Covered"]
    df.loc[mask1 & (a_val != "X"), ["remarks", "covered_status"]] = ["Exclusion as WKS = 1", "Not Covered"]
    
    df.to_excel("apply_workscope_inclusion_exclusion_v2.1_mat.xlsx", index=False)
    
    # WKS=2 → Check column B
    if pipeline_type == "material":
        #    mask2 = target_mask & df["ata4"].between(7221, 7263, inclusive="both") & (df["Customer_Workscope_Level"] == 2)
        mask2 = (
            target_mask
            & df["ata4"].between(7221, 7263, inclusive="both")
            & (df["cs_wks_level"] == 2)
        )
        #    df.loc[mask2 & (df[wspg_b_col].astype(str).str.strip().str.upper() == "X"), ["remarks", "covered_status"]] = ["Inclusion as WKS = 2", "Covered"]
        #    df.loc[mask2 & (df[wspg_b_col].isnull()), ["remarks", "covered_status"]] = ["Exclusion as WKS = 2", "Not Covered"]
        df.loc[mask2 & (b_val == "X"), ["remarks", "covered_status"]] = ["Inclusion as WKS = 2", "Covered"]
        df.loc[mask2 & (b_val != "X"), ["remarks", "covered_status"]] = ["Exclusion as WKS = 2", "Not Covered"]
        df.to_excel("apply_workscope_inclusion_exclusion_v3_mat.xlsx", index=False)
    
            # df.to_excel("apply_workscope_inclusion_exclusion_V2.xlsx", index=False)
    
    elif pipeline_type == "repair":
        #    mask2 = target_mask & df["ata4"].between(7221, 7263, inclusive="both") & ((df["Customer_Workscope_Level"] == "2") | (df["Customer_Workscope_Level"] == 2))
        mask2 = (
            target_mask
            & df["ata4"].between(7221, 7263, inclusive="both")
            & (df["cs_wks_level"] == 2)
        )
        #    df.loc[mask2 & (df[wspg_b_col].astype(str).str.strip().str.upper() == "X"), ["remarks", "covered_status"]] = ["Inclusion as WKS = 2", "Covered"]
        #    df.loc[mask2 & (df[wspg_b_col].isnull()), ["remarks", "covered_status"]] = ["Exclusion as WKS = 2", "Not Covered"]
        df.loc[mask2 & (b_val == "X"), ["remarks", "covered_status"]] = ["Inclusion as WKS = 2", "Covered"]
        df.loc[mask2 & (b_val != "X"), ["remarks", "covered_status"]] = ["Exclusion as WKS = 2", "Not Covered"]  
        df.to_excel("rep_apply_workscope_inclusion_exclusion.xlsx", index=False)
    
    else:
        #    mask2 = target_mask & df["ata4"].between(7221, 7263, inclusive="both") & (df["Customer_Workscope_Level"] == 2 )|((df["Customer_Workscope_Level"] == "2"))
        mask2 = (
            target_mask
            & df["ata4"].between(7221, 7263, inclusive="both")
            & (df["cs_wks_level"]== 2)
        )
        #    df.loc[mask2 & (df[wspg_b_col].astype(str).str.strip().str.upper() == "X"), ["remarks", "covered_status"]] = ["Inclusion as WKS = 2", "Covered"]
        #    df.loc[mask2 & (df[wspg_b_col].isna()), ["remarks", "covered_status"]] = ["Exclusion as WKS = 2", "Not Covered"]
        
        df.loc[mask2 & (b_val == "X"), ["remarks", "covered_status"]] = ["Inclusion as WKS = 2", "Covered"]
        df.loc[mask2 & (b_val != "X"), ["remarks", "covered_status"]] = ["Exclusion as WKS = 2", "Not Covered"]
    
    # ==============================================
    # Identifying the Adder Creap
    # ==============================================
    adder_mask_2 = target_mask & df["ata4"].between(7261, 7263, inclusive="both") & (df['cs_wks_level'] == 1) & (df["wks_num"] == 2)
    df.loc[adder_mask_2, ['remarks']] = ['Creap 1 to 2']
    adder_mask_3 = target_mask & df["ata4"].between(7261, 7263, inclusive="both") & (df['cs_wks_level'] == 1) & (df["wks_num"] == 3)
    df.loc[adder_mask_3, ['remarks']] = ['Creap 1 to 3']

    df.to_excel("apply_workscope_inclusion_exclusion_v4_mat.xlsx", index=False)
    
    return df
 
# Repair Inclusion and Exclusion --- - -- - - - - -- - - - -- - #
MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12
}

def is_excel_xml_bytes(file_bytes: bytes) -> bool:
    """
    Detect Excel 2003 XML by inspecting the leading bytes.
    Returns True if bytes indicate Excel-XML (<?xml ... Workbook ...>).
    """
    if not file_bytes:
        return False
    head = file_bytes[:4096].lstrip()
    # Quick checks: starts with <?xml or contains Workbook or Office schema
    if head.startswith(b"<?xml") or b"<Workbook" in head[:2048] or b"urn:schemas-microsoft-com:office:spreadsheet" in head[:4096]:
        return True
    return False
 
def s3_is_excel_xml(s3_key: str) -> bool:
    """
    Fetch minimal bytes from S3 and detect Excel-2003-XML content.
    Uses read_excel_from_s3(key, return_bytes=True) from s3connect.
    """
    try:
        file_bytes = read_excel_from_s3(s3_key, return_bytes=True)
    except Exception:
        return False
    return is_excel_xml_bytes(file_bytes)
 
def _strip_prefix(key: str, prefix: str) -> str:
    """Return key minus prefix + slash if present."""
    p = prefix.rstrip("/") + "/"
    return key[len(p):] if key.startswith(p) else key

    
def get_latest_crd_file(master_folder: str, induction_date: datetime) -> str:
    """
    Return the S3 key to the latest CRD/Catalog file for the given induction_date.
    Preferred behavior:
      - Search in S3 under "{master_folder}/Catalog"
      - Find files with 'catalog' in the name and a matching year
      - Choose the latest month >= induction_month (or next year fallback)
      - If the selected file has a corresponding .xlsx already in S3, prefer that .xlsx key
    Returns the full S3 key (prefix + filename).
    Raises FileNotFoundError if no candidate is found.
    """
    crd_folder = f"{master_folder.rstrip('/')}/Catalog".lstrip("/")  # keep no leading slash
    print("[INFO] CRD folder (S3 prefix):", crd_folder)
 
    # List all S3 keys under the folder (list_files_in_s3 should return full keys)
    all_keys = list_files_in_s3(crd_folder)
    if not all_keys:
        raise FileNotFoundError(f"[ERROR] No files found under s3://{BUCKET_NAME}/{crd_folder}")
 
    # Build filename-only list and mapping to full key
    filenames = []
    key_by_filename = {}
    crd_prefix = crd_folder.rstrip("/")
    for k in all_keys:
        # normalize and extract filename part relative to prefix
        filename = _strip_prefix(k, crd_prefix)
        filenames.append(filename)
        key_by_filename[filename] = k
 
    print("[DEBUG] Found files count:", len(filenames))
 
    induction_year = induction_date.year
    induction_month = induction_date.month
 
    def collect_candidates_for_year(year: int) -> List[Tuple[int, str]]:
        candidates = []
        for fname in filenames:
            lname = fname.lower()
            # only consider common extensions and 'catalog' in name
            if not lname.endswith((".xls", ".xlsx", ".xml", ".xlsb")):
                continue
            if "catalog" not in lname:
                continue
            if "comp" not in lname:
                continue
            if str(year) not in lname:
                continue
 
            # Try to extract month token like 'catalog_aug' or 'catalog-aug' or '_Aug.'
            match = re.search(r'catalog[_\-\s]?([a-z]{3})', lname, re.IGNORECASE)
            month_num = 0
            if match:
                month_str = match.group(1).lower()
                month_num = MONTH_MAP.get(month_str, 0)
            else:
                # fallback: search for any 3-letter month anywhere
                m2 = re.search(r'(_|-|\b)(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)(_|-|\b)', lname, re.IGNORECASE)
                if m2:
                    month_num = MONTH_MAP.get(m2.group(2).lower(), 0)
 
            # accept if month found and >= induction_month
            # if month_num >= induction_month:
            if month_num >= 0:
                candidates.append((month_num, fname))
        return candidates
 
    # Primary search for induction year
    crd_candidates = collect_candidates_for_year(induction_year)
 
    # If none, check next year
    if not crd_candidates:
        next_year = induction_year + 1
        print(f"[WARN] No CRD candidates found for year {induction_year}. Searching {next_year}...")
        crd_candidates = collect_candidates_for_year(next_year)
 
    if not crd_candidates:
        raise FileNotFoundError(f"[ERROR] No valid CRD file found for induction year {induction_year} or next year.")
 
    # pick the candidate with the highest month number (latest)
    selected_filename = sorted(crd_candidates, key=lambda x: x[0], reverse=True)[0][1]
    print(f"[INFO] Candidate selected (filename): {selected_filename}")
 
    # Prefer an existing .xlsx version in S3 if present:
    base_name_no_ext = re.sub(r'\.[^.]+$', '', selected_filename)  # remove extension
    preferred_names = [
        f"{base_name_no_ext}.xlsx",
        f"{base_name_no_ext}.xls",   # fallback - same as original
        f"{base_name_no_ext}.xlsm",
        f"{base_name_no_ext}.xlsb",
        f"{base_name_no_ext}.xml"
    ]
 
    # search for an xlsx key in key_by_filename (which maps filename->full s3 key)
    for candidate_name in preferred_names:
        candidate_key = key_by_filename.get(candidate_name)
        if candidate_key:
            # prefer .xlsx if found
            if candidate_name.lower().endswith(".xlsx"):
                print(f"[INFO] Found existing converted xlsx in S3: {candidate_key}")
                return candidate_key
            # else keep note of the original if xlsx not present
    # If no .xlsx found, return the original selected file's full S3 key
    selected_full_key = key_by_filename[selected_filename]
    print(f"[INFO] Returning selected S3 key: {selected_full_key}")
    return selected_full_key


# def is_excel_xml_file(path: str) -> bool:
#     """Detect Excel 2003 XML by sniffing the file header, regardless of extension."""
#     try:

#         with open(path, "rb") as f:
#             head = f.read(4096)
#         text = head.decode("utf-8", errors="ignore")

#     except Exception:
#         return False

#     # Excel 2003 XML typically has these markers
#     return ("<?xml" in text and "urn:schemas-microsoft-com:office:spreadsheet" in text) or "<Workbook" in text
 
def read_crd_any_format(latest_crd_path: str, header_row: int = 0) -> pd.DataFrame:
    latest_crd_path = latest_crd_path.replace("\\", "/")
 
    base_no_ext, _ = os.path.splitext(latest_crd_path)
    xlsx_key = base_no_ext + ".xlsx"
    print(f"xlsx_key: {xlsx_key}")
    try:
        print("[INFO] Trying .xlsx:", xlsx_key)
        xlsx_bytes = read_excel_from_s3(xlsx_key, return_bytes=True)
        df = pd.read_excel(BytesIO(xlsx_bytes), header=header_row, engine="openpyxl")
        df.drop_duplicates(inplace=True)
        return df
    except Exception as e:
        print("[WARN] Could not read .xlsx:", e)
 
    try:
        print("[INFO] Falling back to original key:", latest_crd_path)
        bytes_data = read_excel_from_s3(latest_crd_path, return_bytes=True)
        df = pd.read_excel(BytesIO(bytes_data), header=header_row)
        df.drop_duplicates(inplace=True)
        return df
    except Exception as e:
        raise RuntimeError(
            f"Failed to read CRD from S3 ({latest_crd_path}). Error: {e}"
        )
        
def apply_qec_lru_logic(result: dict, main_df: pd.DataFrame, pipeline_type: str, workscope: str) -> pd.DataFrame:
    # --- normalize ---
    df = main_df.copy()
    df.to_excel("qec_df_first.xlsx",index=False)
    df.columns = [str(c).strip().lower() for c in df.columns]
 
    ata_col = next((c for c in df.columns if "ata_long" in c), None)
    ata_chapter_col = next((c for c in df.columns if "ata chapter" in c), None)
    ata_section_col = next((c for c in df.columns if "ata section code" in c), None)
    csn_col = next((c for c in df.columns if "component" in c and "csn" in c), None) or "csn"

    mat_col = next((c for c in df.columns if "material" in c or "inspected" in c), None)
    keyword_col = next((c for c in df.columns if "genpact_part_keyword" in c), None)
    
    description_org = next((c for c in df.columns if "csn description" in c or "description" in c), None)
   
   
    cc_parts_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="cc part number"
    )
   
    part_num_col_cc = next((col for col in cc_parts_df.columns if "part" in col and "num" in col), None)
    keyword_col_cc = next((col for col in cc_parts_df.columns if "part" in col and "keyword" in col), None)
   
    # Normalize
    cc_parts_df[part_num_col_cc] = cc_parts_df[part_num_col_cc].astype(str).str.strip().str.upper()
    cc_parts_df[keyword_col_cc] = cc_parts_df[keyword_col_cc].astype(str).str.strip().str.upper()
   
   
    # ✅ Create PN → Keyword mapping
    cc_parts_map = dict(zip(
        cc_parts_df[part_num_col_cc],
        cc_parts_df[keyword_col_cc]
    ))

    remarks_col = "remarks"
    if not all([ata_col, mat_col, keyword_col]):
        raise ValueError("❌ Required columns missing for ATA logic.")
 
    if remarks_col not in df.columns:
        df[remarks_col] = ""
 
    # helper numeric ATA
    # df["ata_numeric"] = df[ata_col].astype(str).str.extract(r"(\d{2,4})")[0].astype(float)
    chapter_clean = (df[ata_chapter_col].astype(str).str.replace("-", "", regex=False).str.strip())
    
    ata_from_chapter = chapter_clean.str[:4]
    
    # ---- Fallback from ATA SECTION ----
    if(ata_section_col):
        section_clean = (df[ata_section_col].astype(str).str.replace("-", "", regex=False).str.strip())
        ata_from_section = section_clean.str[:4]
    
    # combine chapter + fallback
    if ata_section_col:
        ata_combined = ata_from_chapter.where(
            ata_from_chapter.notna() & ata_from_chapter.str.isnumeric(),
            ata_from_section
        )

    else:
        ata_combined = ata_from_chapter
    
    df["ata_numeric"] = ata_combined.astype(str)
    
    # ---- Detect special ATA range ----
    ata_first4 = pd.to_numeric(df["ata_numeric"].str[:4], errors="coerce")
    mask_special = (
        ((ata_first4 >= 7200) & (ata_first4 <= 7209)) |
        ata_first4.isin([7210, 7220, 7230, 7240, 7250, 7260])
    )

    df["ata_numeric"] = pd.to_numeric(df["ata_numeric"], errors="coerce")
    
    # ---- Extract ATA from CSN ----
    csn_extract = (df[csn_col].astype(str).str.split("_").str[1].str[:4])
    csn_extract = pd.to_numeric(csn_extract, errors="coerce")
    df.loc[mask_special, "ata_numeric"] = csn_extract[mask_special]
    ata_col = "ata_numeric"
    #create the ata_numeric column as per the steps given and change the remarks if found in LRU file
 
    # work only on rows that are still UNKNOWN for this pipeline
    remark_value = {
        "material": "MATL UNKNOWN",
        "repair":   "REPAIR UNKNOWN",
        "vendor":   "VENDOR UNKNOWN"
    }.get(pipeline_type.lower(), "UNKNOWN")
 
    df_unknown = df[df[remarks_col].astype(str).str.strip().str.upper().eq(remark_value)].copy()
    if df_unknown.empty:
        return df  # nothing to change
 
    df_unknown["_orig_idx"] = df_unknown.index
 
    # ---------------- Rule 1: QEC ----------------
    #check the qec inclusion status for the workscope
    ws_df = get_dataframe_by_file_and_sheet(result["dataframes"], "contract", "workscope level")
    ws_df.columns = [str(c).strip().lower() for c in ws_df.columns]
    
    first_col = ws_df.columns[0]
 
    # Find row where first column == "Engine Accessories"
    engine_acc_row = ws_df[ws_df[first_col].astype(str).str.strip().str.upper() == "ENGINE ACCESSORIES"]
    
    if engine_acc_row.empty:
        raise ValueError("❌ 'Engine Accessories' row not found in Workscope Level sheet.")
    
    # Check if workscope column exists
    workscope = workscope.strip().lower()
    if workscope not in ws_df.columns:
        raise ValueError(f"❌ Workscope '{workscope}' not found in Workscope Level sheet.")
    
    status_value = str(engine_acc_row[workscope].iloc[0]).strip().lower()
    
    if(status_value == 'excluded'):
        df_unknown[remarks_col] = "Exclusion as Workscope Exclusion"
    
    # else:
    
    
    df_unknown.loc[df_unknown["ata_numeric"] <= 7199, remarks_col] = "Exclusion as QEC"
 
    # ---------------- Rule 2: WSPG ----------------
    wspg_df = get_dataframe_by_file_and_sheet(result["dataframes"], "contract", "wspg incl excl")
    wspg_df.columns = [str(c).strip().lower() for c in wspg_df.columns]
    wspg_df["ata_long"] = pd.to_numeric(wspg_df.iloc[:, 0].astype(str).str[:4], errors = "coerce")
    wspg_df["included"] = wspg_df.iloc[:, 1:].apply(
        lambda s: s.astype(str).str.upper().str.contains("X", na=False).any(), axis=1
    )
 
    df_unknown = df_unknown.merge(
        wspg_df[["ata_long", "included"]],
        left_on=ata_col, right_on="ata_long",
        how="left"
    )
     
    in_72xx = df_unknown["ata_numeric"].between(7221, 7263, inclusive="both")
    df_unknown.loc[in_72xx & (df_unknown["included"] == True), remarks_col] = "Inclusion as the WKS is full"
 
    # Only rows still unknown proceed to LRU/Schedule logic
    need_lru_logic = df_unknown[remarks_col].astype(str).str.strip().str.upper().eq(remark_value)
    work = df_unknown.loc[need_lru_logic].copy()
    if not work.empty:
        # ==== LRU ingestion (read ALL sources; flag per-source) ====
        def _read_lru(sheet_key, source_label):
            try:
                lru = get_dataframe_by_file_and_sheet(result["dataframes"], sheet_key,source_label).copy()
                # print("lru---",lru)
                lru.columns = [str(c).strip().lower() for c in lru.columns]
                # part/material + description columns (cover ze14/soft/ze57 variants)
                part_col = next((c for c in lru.columns if "material" in c or ("part" in c and "number" in c)), None)
                desc_col = next((c for c in lru.columns if "equipment description" in c or "description" in c), None)
 
                out = pd.DataFrame({
                    "material_norm": lru[part_col].astype(str).str.upper().str.strip() if part_col else pd.Series(dtype=str),
                    "desc_norm":     lru[desc_col].astype(str).str.upper().str.strip() if desc_col else pd.Series(dtype=str),
                })

                # ✅ NEW: map standardized keyword using cc_parts
                out["std_keyword"] = out["material_norm"].map(cc_parts_map)
                out["lru_source"] = source_label
                return out.dropna(how="all", subset=["material_norm","desc_norm"])
            except Exception as e:
                print(f"[WARN] LRU '{source_label}' not loaded: {e}")
                return pd.DataFrame(columns=["material_norm","desc_norm","lru_source"])
   
        lru_cat5 = _read_lru("lru_cat5", "ze14_cat5")
        lru_soft = _read_lru("lru_soft_time", "ze14_soft")
        lru_ze57 = _read_lru("lru_ze57", "ze57")
 
        # Build per-source sets (PN + desc) and a union
        mats_cat5 = set(lru_cat5["material_norm"].dropna().tolist())
        mats_soft = set(lru_soft["material_norm"].dropna().tolist())
        mats_ze57 = set(lru_ze57["material_norm"].dropna().tolist())
 
        desc_cat5 = set(lru_cat5["desc_norm"].dropna().tolist())
        desc_soft = set(lru_soft["desc_norm"].dropna().tolist())
        desc_ze57 = set(lru_ze57["desc_norm"].dropna().tolist())
        
        # Normalize working keys
        work["material_norm"] = work[mat_col].astype(str).str.upper().str.strip()
        work["desc_norm"]     = work[keyword_col].astype(str).str.upper().str.strip()
 
        # Per-source flags
        work["in_lru_ze14_cat5"] = work["material_norm"].isin(mats_cat5) | work["desc_norm"].isin(desc_cat5)
        work["in_lru_ze14_soft"] = work["material_norm"].isin(mats_soft) | work["desc_norm"].isin(desc_soft)
        work["in_lru_ze57"]      = work["material_norm"].isin(mats_ze57) | work["desc_norm"].isin(desc_ze57)
 
        # Any LRU
        work["in_lru"] = work[["in_lru_ze14_cat5","in_lru_ze14_soft","in_lru_ze57"]].any(axis=1)
        
        
        # print("work01---")
        # work.to_excel("work_qec_lru_01.xlsx", index=False)
        
        # ---------------- LRU MATCHING (STRICT RULE) ----------------
        
        TARGET_PN = "APS-73B-18BARD"
        TARGET_DESC = "INDICATOR, FUEL FILTER CLOGGING"
        
        # Normalize working keys
        # work["material_norm"] = work[mat_col].astype(str).str.upper().str.strip()
        # work["desc_norm"] = work[keyword_col].astype(str).str.upper().str.strip()

        work["material_norm"] = work[mat_col].astype(str).str.upper().str.strip()
 
        # 🔁 KEEP desc_norm (because clogging uses it)
        work["desc_norm"] = work[keyword_col].astype(str).str.upper().str.strip()
       
        # ✅ ADD actual description separately
        work["desc_main"] = work[description_org].astype(str).str.upper().str.strip()
       
        # ✅ Normalize keyword (same column, just clean)
        work["keyword_norm"] = work[keyword_col].astype(str).str.upper().str.strip()
        
        # EXTRA NORMALIZATION (DEBUG SAFETY)
        work["desc_norm"] = (
            work["desc_norm"]
            .str.replace(r"[^\w\s]", " ", regex=True)
            .str.replace(r"\s+", " ", regex=True)
            .str.strip()
        )
        
        # ================= DEBUG: SHOW ROWS WITH CLOGGING =================
        print("\n================ DEBUG: ROWS WITH 'CLOGGING' =================")
        
        # Boolean mask for clogging
        clogging_mask = work["desc_norm"].str.contains(
            r"\bCLOGGING\b", case=False, na=False
        )
        
        # Print count
        print(f"Total rows containing 'CLOGGING': {clogging_mask.sum()}")
        
        # Print actual rows (important columns only)
        if clogging_mask.any():
            clogging_rows = work.loc[
                clogging_mask,
                [mat_col, keyword_col, "material_norm", "desc_norm"]
            ]
            print("\n--- Rows detected with CLOGGING ---")
            # print(clogging_rows.head(20))
            # Also export for offline inspection
            clogging_rows.to_excel("debug_clogging_rows.xlsx", index=False)
        else:
            print("❌ NO rows detected with 'CLOGGING'")
        
        print("============================================================\n")
        
        # ================= DEBUG 1: Normalized input =================
        print("\n=== DEBUG 1: Normalized input values ===")
        # print(work[[mat_col, keyword_col, "material_norm", "desc_norm"]].head(10))
        
        # --- PN match (for everyone) ---
        work["pn_match"] = (
            work["material_norm"].isin(mats_cat5)
            | work["material_norm"].isin(mats_soft)
            | work["material_norm"].isin(mats_ze57)
        )
        
        # ================= DEBUG 2: PN match =================
        print("\n=== DEBUG 2: PN match ===")
        # print(work[["material_norm", "pn_match"]].head(10))
        
        # ================= DEBUG 3: CLOGGING keyword detection =================
        work["debug_contains_clogging"] = work["desc_norm"].str.contains(
            r"\bCLOGGING\b", case=False, na=False
        )
        
        print("\n=== DEBUG 3: CLOGGING keyword check ===")
        # print(work[["desc_norm", "debug_contains_clogging"]].head(10))
        
        # --- Identify the special target part (PN OR desc keyword) ---
        is_target_part = (
            (work["material_norm"] == TARGET_PN)
            | work["desc_norm"].str.contains(r"\bCLOGGING\b", case=False, na=False)
        )
        
        work["debug_is_target_part"] = is_target_part
        
        # ================= DEBUG 4: is_target_part evaluation =================
        print("\n=== DEBUG 4: is_target_part ===")
        # print(work[
        #     ["material_norm", "desc_norm", "pn_match",
        #     "debug_contains_clogging", "debug_is_target_part"]
        # ].head(10))
        
        # --- Description match (keyword-based, ONLY for target part) ---
        all_lru_desc = desc_cat5 | desc_soft | desc_ze57

        # ✅ NEW: collect all standardized keywords from LRU
        lru_keywords = set(
            lru_cat5["std_keyword"].dropna().tolist()
            + lru_soft["std_keyword"].dropna().tolist()
            + lru_ze57["std_keyword"].dropna().tolist()
        )
        
        # ---- CHECK IF LRU HAS ANY CLOGGING PART ----
        lru_has_clogging = any(
            isinstance(d, str) and "CLOGGING" in d
            for d in all_lru_desc
        )
        
        # print("DEBUG: LRU has CLOGGING =", lru_has_clogging)
        
        
        # Actual desc_match logic (UNCHANGED)
        # work["desc_match"] = (
        #     is_target_part
        # & work["desc_norm"].apply(
        #         lambda d: any(
        #             kw in d for kw in all_lru_desc
        #         ) if isinstance(d, str) else False
        #     )
        # )
        
        # ---- DESCRIPTION MATCH: CLOGGING ONLY ----
        work["desc_match"] = (
            work["desc_norm"].str.contains(r"\bCLOGGING\b", case=False, na=False)
        & lru_has_clogging
        )

        # -------------------------------
        # STEP 6: NORMAL MATCHING (NON-CLOGGING)
        # -------------------------------
        # PN match
        work["pn_match"] = (
            work["material_norm"].isin(mats_cat5)
            | work["material_norm"].isin(mats_soft)
            | work["material_norm"].isin(mats_ze57)
        )
        # Description match (using actual description)
        all_lru_desc = desc_cat5 | desc_soft | desc_ze57
        work["desc_match_main"] = work["desc_main"].isin(all_lru_desc)
        # Keyword match (standardized)
        work["keyword_match"] = work["keyword_norm"].isin(lru_keywords)
        
        # ================= DEBUG 6: desc_match =================
        # print("\n=== DEBUG 6: desc_match ===")
        # print(work[["desc_norm", "desc_match"]].head(10))
        
        # # --- Final LRU decision ---
        # work["in_lru"] = np.where(
        #     is_target_part,
        #     work["pn_match"], # | work["desc_match"],  # PN OR DESC for target
        #     work["pn_match"]                        # PN ONLY for others
        # )

        # ADD KEYWORD MATCH (NEW)
        # -------------------------------
        work["keyword_match"] = work["keyword_norm"].isin(lru_keywords)
       
        # -------------------------------
        # FINAL DECISION (WITH CLOGGING SAFE)
        # -------------------------------
        work["in_lru"] = np.where(
            is_target_part,
            work["pn_match"] | work["desc_match"],  # ✅ STEP 5 (clogging)
            np.where(
                work["pn_match"],
                True,
                np.where(
                    work["desc_match_main"],
                    True,
                    work["keyword_match"]
                )
            )
        )
        
        # ================= DEBUG 7: Final decision =================
        # print("\n=== DEBUG 7: Final LRU decision ===")
        # print(work[
        #     ["material_norm", "desc_norm",
        #     "pn_match", "debug_is_target_part",
        #     "desc_match", "in_lru"]
        # ].head(10))
        
        # --- DROP rows not found in LRU ---
        work = work[work["in_lru"]].copy()
        
        # print("\nwork--- (after LRU filtering)")
        # work.to_excel("work_qec_lru.xlsx", index=False)
 
        # Human-readable source list for debugging/auditing
        def _src(row):
            hits = []
            if row.get("in_lru_ze14_cat5"): hits.append("ze14_cat5")
            if row.get("in_lru_ze14_soft"): hits.append("ze14_soft")
            if row.get("in_lru_ze57"):      hits.append("ze57")
            return ", ".join(hits) if hits else ""
        work["lru_sources"] = work.apply(_src, axis=1)
 
        # # ==== Schedule (unchanged, but uses PN first, then keyword) ====
        # def _read_schedule(sheet_substring, tag_label):
        #     s = get_dataframe_by_file_and_sheet(result["dataframes"], "contract", sheet_substring).copy()
        #     s.columns = [str(c).strip().lower() for c in s.columns]
        #     s_mat  = next((c for c in s.columns if ("example" in c and "mpn" in c) or "material" in c), None)
        #     s_desc = next((c for c in s.columns if "part" in c and "des" in c), None)
        #     s_nom  = next((c for c in s.columns if "nomenclature" in c), None)
 
        #     out = pd.DataFrame()
        #     out["material_norm"] = s[s_mat].astype(str).str.upper().str.strip() if s_mat else pd.NA
        #     out["desc_norm"]     = s[s_desc].astype(str).str.upper().str.strip() if s_desc else pd.NA
        #     out["nomenclature"]  = s[s_nom] if s_nom in s.columns else pd.NA
        #     out["part_of"]       = tag_label
        #     return out
 
        # s2 = _read_schedule("schedule_2 parts", "Schedule_2")
        # s3 = _read_schedule("schedule_3 parts", "Schedule_3")
 
        # # PN-first match
        # s2_mats = set(s2["material_norm"].dropna().tolist())
        # s3_mats = set(s3["material_norm"].dropna().tolist())
 
        # work["in_s2_pn"] = work["material_norm"].isin(s2_mats)
        # work["in_s3_pn"] = work["material_norm"].isin(s3_mats)
 
        # # If PN not found, try keyword (description) match
        # s2_desc = set(s2["desc_norm"].dropna().tolist())
        # s3_desc = set(s3["desc_norm"].dropna().tolist())
 
        # work["in_s2_kw"] = (~work["in_s2_pn"]) & work["desc_norm"].isin(s2_desc)
        # work["in_s3_kw"] = (~work["in_s3_pn"]) & work["desc_norm"].isin(s3_desc)
 
        # work["in_s2"] = work["in_s2_pn"] | work["in_s2_kw"]
        # work["in_s3"] = work["in_s3_pn"] | work["in_s3_kw"]
 
        # # small audit helpers
        # work["schedule_part_of"] = np.select(
        #     [work["in_s2"], work["in_s3"]],
        #     ["Schedule_2", "Schedule_3"],
        #     default=""
        # )
        # work["schedule_match_source"] = np.select(
        #     [work["in_s2_pn"]|work["in_s3_pn"], work["in_s2_kw"]|work["in_s3_kw"]],
        #     ["pn","keyword"],
        #     default=""
        # )
 
        # ==== Remarks per rules ====
        crit_terms = ["OIL SCAVENGE", "SCAVENGE", "SCA"]
        is_critical = work["desc_norm"].fillna("").str.contains("|".join(crit_terms), case=False, regex=True)
        work.loc[is_critical, remarks_col] = "Inclusion as Critical Component"
 
        still_unknown = work[remarks_col].isna() | (work[remarks_col].astype(str).str.strip() == "") | work[remarks_col].astype(str).str.upper().eq(remark_value)
        #chaange---remarks-1/8/2026
        # if pipeline_type.lower() == "material":
        #     work.loc[still_unknown & work["in_lru"] & work["in_s2"], remarks_col] = "ACC/LRU Normal Inclusion"
        #     work.loc[still_unknown & work["in_lru"] & work["in_s3"], remarks_col] = "ACC/LRU Normal Inclusion"
        #     work.loc[still_unknown & ~(work["in_s2"] | work["in_s3"]), remarks_col] = "ACC/LRU Normal Inclusion"
        if pipeline_type.lower() == "material":
            work.loc[still_unknown & work["in_lru"], remarks_col] = "ACC/LRU Normal Exclusion"
            work.loc[still_unknown & ~(work["in_lru"]), remarks_col] = "ACC/LRU Normal Inclusion"
       
        elif pipeline_type.lower() in ("repair", "vendor"):
            work.loc[still_unknown & work["in_lru"], remarks_col] = "ACC/LRU Normal Exclusion"
        # elif pipeline_type.lower() in ("repair", "vendor"):
        #     work.loc[still_unknown & work["in_s2"], remarks_col] = "ACC/LRU schedule 2 Inclusion"
        #     work.loc[still_unknown & work["in_s3"], remarks_col] = "Exclusion as LRU schedule 3 parts"
        #     work.loc[still_unknown & ~(work["in_s2"] | work["in_s3"]), remarks_col] = "Non-LRU Inclusion"
 
        # write updated remarks + audit cols back to df_unknown
        cols_to_push = [remarks_col, "in_lru","in_lru_ze14_cat5","in_lru_ze14_soft","in_lru_ze57",
                        "lru_sources","in_s2","in_s3","in_s2_pn","in_s3_pn","in_s2_kw","in_s3_kw",
                        "schedule_part_of","schedule_match_source"]
        for c in cols_to_push:
            if c in work.columns:
                df_unknown.loc[work.index, c] = work.loc[work.index, c].values
 
    # finally, write df_unknown back into the main df (remarks & new flags only)
    for c in [remarks_col, "in_lru","in_lru_ze14_cat5","in_lru_ze14_soft","in_lru_ze57",
              "lru_sources","in_s2","in_s3","in_s2_pn","in_s3_pn","in_s2_kw","in_s3_kw",
              "schedule_part_of","schedule_match_source"]:
        if c in df_unknown.columns:
            df.loc[df_unknown["_orig_idx"], c] = df_unknown[c].values
 
    df.drop_duplicates(inplace=True)
    df.to_excel("qec_lru_logic_data.xlsx", index=False)
    
    return df

def convert_excel_xml_to_xlsx(xml_path: str) -> str:
    if not os.path.exists(xml_path):
        raise FileNotFoundError(f"[ERROR] ❌ File not found: {xml_path}")
 
    temp_folder = os.path.join(os.path.dirname(xml_path), "temp_converted_files")
    os.makedirs(temp_folder, exist_ok=True)
 
    base = os.path.splitext(os.path.basename(xml_path))[0]
    new_file_path = os.path.join(temp_folder, base + ".xlsx")
 
    namespaces = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}
 
    try:
        tree = etree.parse(xml_path)
        rows = tree.xpath("//ss:Row", namespaces=namespaces)
 
        data = []
        for row in rows:
            row_data = []
            for cell in row.xpath(".//ss:Cell", namespaces=namespaces):
                data_el = cell.xpath(".//ss:Data", namespaces=namespaces)
                row_data.append(data_el[0].text if data_el else None)
            data.append(row_data)
 
        max_columns = max((len(r) for r in data), default=0)
        padded = [r + [None]*(max_columns - len(r)) for r in data]
 
        df = pd.DataFrame(padded)
        with pd.ExcelWriter(new_file_path, engine="openpyxl") as w:
            df.to_excel(w, sheet_name="Sheet1", index=False)
        return new_file_path
    except Exception as e:
        print(f"[ERROR] ❌ Failed to convert Excel XML-based Excel: {e}")
        raise ValueError(f"[ERROR] ❌ Could not convert {xml_path} to .xlsx format.")


# New logicccccc
def get_component_repairs_discount_pct(result) -> float:
    """
    Find the 'Component Repairs' discount in the contract discount workbook.
    Returns a decimal (e.g., 0.15 for 15%). Defaults to 0 if not found.
    """
    possible_file_keys = ["contract", "discounts", "contract_discount", "contract discounts"]
    possible_sheet_keys = ["discount", "discounts", "contract", "price"]
 
    disc_df = None
    for fk in possible_file_keys:
        for sk in possible_sheet_keys:
            try:
                df_try = get_dataframe_by_file_and_sheet(
                    all_dataframes=result["dataframes"],
                    file_key=fk,
                    sheet_name_substring=sk
                )
                if df_try is not None and not df_try.empty:
                    disc_df = df_try.copy()
                    break
            except Exception:
                pass
        if disc_df is not None:
            break
 
    if disc_df is None or disc_df.empty:
        return 0.0
 
    disc_df.columns = [str(c).strip().lower() for c in disc_df.columns]
 
    # Strategy A: a column named like "component repairs", "comp repairs", etc., with numeric values
    comp_col = next(
        (c for c in disc_df.columns
         if ("component" in c or "comp" in c) and ("repair" in c or "repairs" in c)),
        None
    )
    if comp_col is not None:
        val = pd.to_numeric(disc_df[comp_col], errors="coerce").dropna()
        if not val.empty:
            pct = float(val.iloc[0])
            if pct > 1:  # assume given in % (e.g., 15)
                pct = pct / 100.0
            return max(0.0, min(1.0, pct))
 
    # Strategy B: look for a row describing component repairs + a generic 'discount' column
    desc_col = next((c for c in disc_df.columns if c in ["item", "category", "description", "type"]), None)
    discount_col = next((c for c in disc_df.columns if "discount" in c or "%" in c), None)
    if desc_col and discount_col:
        mask = disc_df[desc_col].astype(str).str.lower().str.contains("component") & \
               disc_df[desc_col].astype(str).str.lower().str.contains("repair")
        matches = pd.to_numeric(disc_df.loc[mask, discount_col], errors="coerce").dropna()
        if not matches.empty:
            pct = float(matches.iloc[0])
            if pct > 1:
                pct = pct / 100.0
            return max(0.0, min(1.0, pct))
 
    return 0.0


def extract_clean_sb(text: str) -> str | None:
    """
    Extracts the first SB code from a string and normalizes it to: 'SB NN-NNNN'.
    Rules:
      1) Only process rows that contain 'SB' (case-insensitive).
      2) Remove any '07/' substrings.
      3) Accept common variants like: 'SB 72-1011', 'SB72-1011', 'SB 721011', 'SB721011', 'SB 72 1011'.
    Returns 'SB NN-NNNN' or None.
    """
    if text is None:
        return None
 
    s = str(text)
    if "SB" not in s.upper():  # only rows that actually mention 'SB'
        return None
 
    # remove '07/' noise anywhere
    s_up = re.sub(r'07/', '', s, flags=re.IGNORECASE).upper()
 
    # Pattern A: 'SB' + 2 digits + optional sep + 3-4 digits (preferred)
    m = re.search(r'\bSB\b[\s\-]*([0-9]{2})[\s\-]?([0-9]{3,4})\b', s_up)
    if m:
        return f"SB {m.group(1)}-{m.group(2)}"
 
    # Pattern B: 'SB' + 5-6 digits → split as 2 + rest
    m = re.search(r'\bSB[\s\-]*([0-9]{5,6})\b', s_up)
    if m:
        digits = m.group(1)
        return f"SB {digits[:2]}-{digits[2:]}"
 
    # Pattern C: 'SB' + 2 digits + whitespace + 3-4 digits (e.g., 'SB 72 1011')
    m = re.search(r'\bSB\b[\s\-]*([0-9]{2})\s+([0-9]{3,4})\b', s_up)
    if m:
        return f"SB {m.group(1)}-{m.group(2)}"
 
    return None


def explode_child_repair_numbers(df, column='Child Repair Number'):
 
    pattern = r'''
        DISS
        | ASSY
        | CINSP
        | R\d+
        | SB\d+
        | RD\s.*?(?=(?:DISS|ASSY|CINSP|R\d+|SB\d+|RD\s|$))
    '''
    regex = re.compile(pattern, re.VERBOSE)
 
    def extract_codes(cell):
        if pd.isnull(cell):
            return []
        matches = regex.findall(str(cell))
        return [m.strip() for m in matches if m.strip()]
 
    df = df.copy()
 
    # 🔑 unique id per original CRD row
    df["_crd_row_id"] = df.index
 
    df["__split_codes__"] = df[column].apply(extract_codes)
 
    df_exploded = df.explode("__split_codes__")
    df_exploded = df_exploded.drop(columns=[column])
    df_exploded = df_exploded.rename(columns={"__split_codes__": column})
    df_exploded.to_excel("explode_child_repair.xlsx",index= False)
 
    return df_exploded.reset_index(drop=True)

def compute_repair_total_price(result: dict, repair_df: pd.DataFrame, new_crd_df: pd.DataFrame, induction_date: datetime) -> pd.DataFrame:
 
    # --- Normalize copies ---
    repair_df = repair_df.copy()
    new_crd_df = new_crd_df.copy()
    repair_df.columns = [str(col).strip().lower() for col in repair_df.columns]
    new_crd_df.columns = [str(col).strip().lower() for col in new_crd_df.columns]
 
    # --- Identify required columns ---
    code_col         = next((c for c in repair_df.columns if "repair code" in c), None)
    hours_col        = next((c for c in repair_df.columns if "repair hours" in c), None)
    repair_qty_col   = next((c for c in repair_df.columns if "inspected quantity" in c or c in ("qty","quantity")), None)
    repair_part_col  = next((c for c in repair_df.columns if "inspected part" in c or ("part" in c and "number" in c)), None)
    repair_desc_col  = next((c for c in repair_df.columns if "repair description" in c or ("repair" in c and "description" in c)), None)
 
 
    repair_num_col        = next((c for c in new_crd_df.columns if "repair" in c and "number" in c and "child" not in c), None)
    crd_part_num_col      = next((c for c in new_crd_df.columns if "part" in c and "number" in c), None)
    crd_child_repair_col  = next((c for c in new_crd_df.columns if "child repair number" in c and "number" in c), None)
    price_col             = next((c for c in new_crd_df.columns if c.strip().lower() == "price"), None)
    crd_rep_des             = next((c for c in new_crd_df.columns if c.strip().lower() == "repair description"), None)
    comments             = next((c for c in new_crd_df.columns if c.strip().lower() == "comments"), None)
    comp_desc_crd             = next((c for c in new_crd_df.columns if c.strip().lower() == "component description"), None)
   
    cc_parts_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="cc part number"
    )
   
    part_num_col = next((col for col in cc_parts_df.columns if "part" in col and "num" in col), None)
    keyword_col = next((col for col in cc_parts_df.columns if "part" in col and "keyword" in col), None)
   
    # Normalize
    cc_parts_df[part_num_col] = cc_parts_df[part_num_col].astype(str).str.strip().str.upper()
    cc_parts_df[keyword_col] = cc_parts_df[keyword_col].astype(str).str.strip().str.upper()
   
   
    if not all([code_col, hours_col]):
        raise ValueError("[ERROR] ❌ Required columns not found in Internal Repair file (need 'Repair Code' and 'Repair Hours').")
    if not all([crd_part_num_col, repair_num_col]) or price_col not in new_crd_df.columns:
        raise ValueError("[ERROR] ❌ Required columns not found in CRD file (need Part Number, Repair Number, Price).")
 
    # --- Extract short code from 'Repair Code' (part after '-') ---
   
    # repair_df["extracted_repair_code"] = repair_df[code_col].apply(
    #     lambda x: re.findall(r'-([A-Z0-9]+)', str(x))[0] if '-' in str(x) else None
    # )
   
    def normalize_repair_code(val):
 
        if "-" not in str(val):
            return None
   
        code = re.findall(r'-([A-Z0-9]+)', str(val).upper())
        if not code:
            return None
   
        code = code[0]
   
        # Apply special logic only for R codes
        if code.startswith("R"):
   
            # Remove trailing alphabet
            code = re.sub(r'[A-Z]$', '', code)
   
            digits = code[1:]
   
            # Pad digits to length 3
            if digits.isdigit() and len(digits) < 3:
                digits = digits.zfill(3)
   
            code = "R" + digits
   
        return code
   
   
    repair_df["extracted_repair_code"] = repair_df[code_col].apply(normalize_repair_code)
   
   
 
    # Extract SB Codes
    repair_df["sb_clean"] = repair_df[repair_desc_col].apply(extract_clean_sb)
 
    # Keep only rows where an SB was extracted
    sb_df = repair_df[repair_df["sb_clean"].notna()].copy()
 
    # --- Normalize join keys to uppercase/stripped ---
    # print("repair_df :", repair_df["extracted_repair_code"].unique())
    # repair_df = repair_df[~(repair_df["extracted_repair_code"].astype(str).str.startswith("S", na=False))]
    # print("repair_df :", repair_df["extracted_repair_code"].unique())
 
    repair_df["extracted_repair_code"] = repair_df["extracted_repair_code"].astype(str).str.strip().str.upper()
    repair_df[repair_part_col]         = repair_df[repair_part_col].astype(str).str.strip().str.upper()
    repair_df[repair_desc_col]         = repair_df[repair_desc_col].astype(str).str.strip().str.upper()
   
 
    new_crd_df[crd_part_num_col]       = new_crd_df[crd_part_num_col].astype(str).str.strip().str.upper()
    new_crd_df[repair_num_col]         = new_crd_df[repair_num_col].astype(str).str.strip().str.upper()
    new_crd_df[crd_rep_des]         = new_crd_df[crd_rep_des].astype(str).str.strip().str.upper()
    new_crd_df[comp_desc_crd]         = new_crd_df[comp_desc_crd].astype(str).str.strip().str.upper()
   
    # merge CRD with CC Part Number mapping
    new_crd_df = new_crd_df.merge(
        cc_parts_df[[part_num_col, keyword_col]],
        left_on=crd_part_num_col,
        right_on=part_num_col,
        how="left"
    )
    # normalized component description
    new_crd_df["normalized_component_description"] = new_crd_df[keyword_col]
   
   
   
    new_crd_df.rename(columns={crd_rep_des : "crd_rep_description"}, inplace=True)
   
   
   
    if crd_child_repair_col:
        new_crd_df[crd_child_repair_col] = new_crd_df[crd_child_repair_col].astype(str).str.strip().str.upper()
 
    # Explode Child Repair Numbers
    crd_child_num_explode_df = explode_child_repair_numbers(new_crd_df, crd_child_repair_col)
   
 
    # --- Step-1: merge on (Part, Repair Number) ---
    merged_df = repair_df.merge(
        crd_child_num_explode_df[[crd_part_num_col, repair_num_col, crd_child_repair_col, price_col,"crd_rep_description",comments,comp_desc_crd]],
        left_on=[repair_part_col, "extracted_repair_code"],
        right_on=[crd_part_num_col, repair_num_col],
        how="left",
        suffixes=("", "_step1")
    )
    merged_df.rename(columns={repair_num_col : "crd_repair_number"}, inplace=True)
    merged_df.to_excel("merged_df_repair_v1.xlsx", index=False)

    # ==============================
    # STEP 1.1: HANDLE MULTIPLE CRD MATCHES (DESCRIPTION FILTER)
    # ==============================
    
    print("\n===== STEP 1.1 : MULTI-MATCH DESCRIPTION FILTER =====")
    
    # Normalize descriptions
    merged_df["_rep_desc_norm"] = (
        merged_df["repair description"]
        .astype(str).str.strip().str.upper()
    )
    
    merged_df["_crd_desc_norm"] = (
        merged_df["crd_rep_description"]
        .astype(str).str.strip().str.upper()
    )
    
    group_cols = ["inspected part no.", "extracted_repair_code"]
    
    for (part, code), grp in merged_df.groupby(group_cols):
    
        if len(grp) == 1:
            continue  # nothing to fix
    
        # Identify matching rows
        match_mask = grp["_rep_desc_norm"] == grp["_crd_desc_norm"]
    
        match_count = match_mask.sum()
    
        if match_count == 1:
            # ✅ Only ONE correct row → zero others
            idx_to_zero = grp.loc[~match_mask].index
    
            print(f"Part: {part}, Code: {code} → Single match found, zeroing others")
    
            merged_df.loc[idx_to_zero, price_col] = 0
    
        elif match_count > 1:
            # ✅ Multiple matches → keep them, zero rest
            idx_to_zero = grp.loc[~match_mask].index
    
            print(f"Part: {part}, Code: {code} → Multiple matches, zeroing non-matches")
    
            merged_df.loc[idx_to_zero, price_col] = 0
    
        else:
            # ❗ No match → KEEP ALL
            print(f"Part: {part}, Code: {code} → No match, keeping all rows")
   
    # --- Step-2: for rows with no price, merge on (Part, Child Repair Number) ---
    need_child = merged_df[price_col].isna()
 
    if need_child.any() and crd_child_repair_col: # and crd_child_repair_col:
 
        child_map = crd_child_num_explode_df[[crd_part_num_col, crd_child_repair_col, price_col]].dropna(
            subset=[crd_child_repair_col]
        ).copy()
 
        child_map[crd_part_num_col] = child_map[crd_part_num_col].astype(str).str.strip().str.upper()
        child_map[crd_child_repair_col] = child_map[crd_child_repair_col].astype(str).str.strip().str.upper()
 
        # print("2: child_map :", child_map[crd_child_repair_col].unique())
 
        to_fix = merged_df.loc[need_child, [repair_part_col, "extracted_repair_code"]].copy()
        to_fix[repair_part_col] = to_fix[repair_part_col].astype(str).str.strip().str.upper()
        to_fix["extracted_repair_code"] = to_fix["extracted_repair_code"].astype(str).str.strip().str.upper()
        to_fix["_idx"] = to_fix.index
 
        retry = to_fix.merge(
            child_map.rename(columns={crd_child_repair_col: "_child_code"}),
            left_on=[repair_part_col, "extracted_repair_code"],
            right_on=[crd_part_num_col, "_child_code"],
            how="left"
        ).set_index("_idx")
 
        # print("1: child_map :", retry["_child_code"].unique())
 
        retry_price_and_child_rep_num = retry.groupby(level=0).agg(
            **{
                "_child_price": (price_col, "first"),
                "_child_code_out": ("_child_code", "first"),
            }
        )
 
        if crd_child_repair_col not in merged_df.columns:
            merged_df[crd_child_repair_col] = pd.NA
 
        # assign price and child code back by original indicies
        merged_df.loc[retry_price_and_child_rep_num.index, price_col] = pd.to_numeric(
            retry_price_and_child_rep_num["_child_price"], errors="coerce"
        ).values
        merged_df.loc[retry_price_and_child_rep_num.index, crd_child_repair_col] = retry_price_and_child_rep_num["_child_code_out"].values
 
 
    # Ensure numeric price
    merged_df[price_col] = pd.to_numeric(merged_df[price_col], errors="coerce")
   
    merged_df.to_excel("merged_df_repair_v2.xlsx", index=False)
   
    # ==============================
    # STEP-3 : COMMENT MATCH LOGIC
    # ==============================
   
    print("\n===== STEP 3 : COMMENT MATCHING STARTED =====")
   
    need_comment_match = merged_df[price_col].isna()
   
    print(f"Rows needing comment match: {need_comment_match.sum()}")
   
    if need_comment_match.any() and comments:
   
        comment_map = new_crd_df[[crd_part_num_col, price_col, comments]].copy()
   
        comment_map[crd_part_num_col] = comment_map[crd_part_num_col].astype(str).str.strip().str.upper()
        comment_map[comments] = comment_map[comments].astype(str).str.upper()
   
        print(f"Total CRD rows available for comment search: {len(comment_map)}")
   
        to_fix = merged_df.loc[need_comment_match, [repair_part_col, "extracted_repair_code"]].copy()
        to_fix["_idx"] = to_fix.index
   
        results = []
   
        for idx, row in to_fix.iterrows():
   
            part = str(row[repair_part_col]).strip().upper()
            code = str(row["extracted_repair_code"]).strip().upper()
   
            # print("\n-----------------------------------")
            # print(f"Checking Row Index : {idx}")
            # print(f"Part Number        : {part}")
            # print(f"Repair Code Search : {code}")
   
            matches = comment_map[
                (comment_map[crd_part_num_col] == part) &
                (comment_map[comments].str.contains(code, na=False))
            ]
   
            if not matches.empty:
   
                price_val = matches.iloc[0][price_col]
   
                print("✅ MATCH FOUND IN COMMENTS")
                print(f"Assigned Price : {price_val}")
   
                results.append((idx, price_val))
   
                # print("❌ No match found in comments")
   
        print("\n===== COMMENT MATCH SUMMARY =====")
        print(f"Total matches found : {len(results)}")
   
        for idx, price_val in results:
            merged_df.loc[idx, price_col] = pd.to_numeric(price_val, errors="coerce")
   
    print("===== STEP 3 : COMMENT MATCHING COMPLETED =====\n")
   
    # ==============================
    # STEP-4 : PN-INDEPENDENT MATCH
    # ==============================
   
    print("\n===== STEP 4 : PN INDEPENDENT MATCH STARTED =====")
   
    need_fallback = merged_df[price_col].isna()
   
    print("Rows needing fallback:", need_fallback.sum())
   
    if need_fallback.any():
   
        # normalize ATA
        merged_df["ata chapter"] = merged_df["ata chapter"].astype(str).str.strip().str.upper()
        new_crd_df["ata"] = new_crd_df["ata"].astype(str).str.strip().str.upper()
   
       
   
        merged_df["genpact_part_keywords"] = (
            merged_df["genpact_part_keywords"]
            .astype(str)
            .str.strip()
            .str.upper()
        )
   
        for idx in merged_df.loc[need_fallback].index:
   
            rep_code = merged_df.at[idx, "extracted_repair_code"]
            ata_val = merged_df.at[idx, "ata chapter"]
            keyword = merged_df.at[idx, "genpact_part_keywords"]
            rep_desc = merged_df.at[idx, "repair description"]
   
            # Step-1: repair code match
            crd_subset = new_crd_df[
                (new_crd_df[repair_num_col] == rep_code) |
                (new_crd_df[crd_child_repair_col] == rep_code) |
                (new_crd_df[comments].str.contains(rf"\b{rep_code}\b", na=False))
            ]
   
            if crd_subset.empty:
                continue
   
            # Step-2: ATA match
            crd_subset = crd_subset[crd_subset["ata"] == ata_val]
   
            if crd_subset.empty:
                continue
   
            # Step-3: component description match
            desc_match = crd_subset[
                # crd_subset["normalized_component_description"] == keyword
                crd_subset["normalized_component_description"].str.contains(
                    rf"\b{keyword}\b", na=False
                )
            ]
   
            if desc_match.empty:
                continue
   
            # Step-4: keyword similarity
            desc_match["similarity"] = desc_match["crd_rep_description"].apply(
                lambda x: fuzz.token_set_ratio(str(x), str(rep_desc))
            )
   
            desc_match = desc_match[desc_match["similarity"] >= 60]
   
            if desc_match.empty:
                continue
   
            # multiple prices → average
            prices = desc_match[price_col].astype(float).dropna()
 
            if prices.empty:
                continue
           
            # If all prices same → average (same value)
            if prices.nunique() == 1:
                price_val = prices.mean()
            else:
                # If different prices → take lowest
                price_val = prices.min()
           
            merged_df.at[idx, price_col] = price_val
           
            print("Matched row:", idx, "Price assigned:", price_val)
   
            print("Matched row:", idx, "Price assigned:", price_val)
   
    print("===== STEP 4 : PN INDEPENDENT MATCH COMPLETED =====\n")
   
   
    # Ensure blanks are handled properly
    merged_df["crd_repair_number"] = merged_df["crd_repair_number"].fillna("").astype(str).str.strip()
    merged_df["child repair number"] = merged_df["child repair number"].fillna("").astype(str).str.strip()
    merged_df["repair code group & code"] = merged_df["repair code group & code"].fillna("").astype(str)
    merged_df["repair description"] = (merged_df["repair description"].fillna("").astype(str).str.strip())
   
    merged_df.to_excel("merged_df_repair_v2_child_rep_num_v0.xlsx", index=False)
   
   
   
    def process_group(group):
 
        # rows where child repair logic applied
        eligible_mask = (
            (group["crd_repair_number"] == "") &
            (group["child repair number"] != "") &
            (group["extracted_repair_code"] == group["child repair number"])
        )
   
        eligible_rows = group[eligible_mask]
   
        if eligible_rows.empty:
            return group
   
        # group only by price
        for price, sub_group in eligible_rows.groupby("price"):
   
            if len(sub_group) <= 1:
                continue
   
            # combine repair codes
            combined_codes = (
                sub_group["repair code group & code"]
                .loc[sub_group["repair code group & code"] != ""]
                .unique()
            )
   
            combined_codes_str = ",".join(combined_codes)
   
            # assign merged repair code
            group.loc[sub_group.index, "repair code group & code"] = combined_codes_str
   
            # keep price only for first row
            first_idx = sub_group.index[0]
   
            group.loc[
                sub_group.index.difference([first_idx]),
                "price"
            ] = 0
   
        return group
   
   
    merged_df = (
        merged_df
        .groupby("inspected part no.", group_keys=False)
        .apply(process_group)
    )
 
    merged_df.to_excel("merged_df_repair_v2_child_rep_num.xlsx", index=False)
 
    # === Keep your downstream logic unchanged ===
    # Step 6: Assign direct price if available
    merged_df["total_price"] = merged_df[price_col]
 
    escalation_results  = get_escalation_values(result, induction_date)
    labor_amount        = escalation_results["labor_amount"]
    new_parts_non_llp   = escalation_results["new_parts_non_llp"]
    new_parts_llp       = escalation_results["new_parts_llp"]
    cfe                 = escalation_results["cfe"]
 
    merged_df["labor_amount"] = labor_amount
    merged_df["total_price_final"] = merged_df["total_price"].fillna(
        merged_df[hours_col] * float(labor_amount or 0)
    )
 
    # Extended = Qty * total_price_final (assumes qty column present/clean)
    merged_df["extended_price_final"] = merged_df[repair_qty_col] * merged_df["total_price_final"]
       
    # Keep non-zero totals, de-dup
    merged_df = merged_df[merged_df["total_price_final"] != 0].copy()
    merged_df.drop_duplicates(inplace=True)
    merged_df.to_excel("merged_df_repair_v3.xlsx", index=False)
   
   
    # Checking for Part Description if PN, Repair Code, matched then (New Scenario 23/1/2026)
   
    # ==============================
    # WORKING COPY
    # ==============================
    df = merged_df.copy()
   
    # ==============================
    # NORMALIZATION (strip + upper)
    # ==============================
    df["_rep_code_norm"] = (
        df["extracted_repair_code"]
        .astype(str).str.strip().str.upper()
    )
   
    df["_rep_desc_norm"] = (
        df["repair description"]
        .astype(str).str.strip().str.upper()
    )
   
    df["_crd_code_norm"] = (
        df["crd_repair_number"]
        .astype(str).str.strip().str.upper()
    )
   
    df["_crd_desc_norm"] = (
        df["crd_rep_description"]
        .astype(str).str.strip().str.upper()
    )
   
    # ==============================
    # FIND DUPLICATE PART NUMBERS
    # ==============================
    dup_pn_mask = df.duplicated(
        subset=["inspected part no."],
        keep=False
    )
   
    # print("\n===== DUPLICATE PART NUMBER CHECK =====")
    # print(f"Total rows        : {len(df)}")
    # print(f"Duplicate PN rows : {dup_pn_mask.sum()}")
   
    rows_to_drop = []
   
    # ==============================
    # PROCESS EACH DUPLICATE PN
    # ==============================
    for part_no, grp in df[dup_pn_mask].groupby("inspected part no."):
   
        # print("\n--------------------------------------------------")
        # print(f"Processing Part No : {part_no}")
        # print(f"Total Rows         : {len(grp)}")
   
        # ---- CRD MATCH CHECK ----
        crd_match_mask = (
            (grp["_rep_code_norm"] == grp["_crd_code_norm"]) &
            (grp["_rep_desc_norm"] == grp["_crd_desc_norm"])
        )
   
        crd_matched = grp[crd_match_mask]
   
        if crd_matched.empty:
            # print("❌ No CRD-matched row found → skipping")
            continue
   
        # ---- KEEP FIRST CRD MATCH ----
        keep_index = crd_matched.index[0]
        keep_row = df.loc[keep_index]
   
        # print("\n✅ CRD MATCH FOUND")
    #     print(f"""
    # Kept Row Index      : {keep_index}
    # Repair Code         : {keep_row['extracted_repair_code']}
    # Repair Description : {keep_row['repair description']}
    # """)
   
        # ---- NUMERIC VALIDATION (NEW) ----
        qty_col = "inspected quantity"
        rh_col  = "repair hours"
        toh_col = "total order hours"
   
        numeric_cols = [qty_col, rh_col, toh_col]
   
        # Compare numeric values with kept row
        numeric_match_mask = (
            (pd.to_numeric(grp[qty_col], errors="coerce").fillna(0)
                == pd.to_numeric(keep_row[qty_col], errors="coerce")) &
            (pd.to_numeric(grp[rh_col], errors="coerce").fillna(0)
                == pd.to_numeric(keep_row[rh_col], errors="coerce")) &
            (pd.to_numeric(grp[toh_col], errors="coerce").fillna(0)
                == pd.to_numeric(keep_row[toh_col], errors="coerce"))
        )
   
        if not numeric_match_mask.all():
            # print("⚠ Numeric mismatch detected → skipping deletion")
            continue
   
        print("✅ Numeric values match across duplicate rows")
   
        # ---- DELETE ALL OTHER SAME CODE + DESC ROWS ----
        delete_mask = (
            (grp["_rep_code_norm"] == keep_row["_rep_code_norm"]) &
            (grp["_rep_desc_norm"] == keep_row["_rep_desc_norm"]) &
            (grp.index != keep_index)
        )
   
        delete_rows = grp[delete_mask]
   
        for idx, row in delete_rows.iterrows():
    #         print("🗑 DELETING ROW")
    #         print(f"""
    # Row Index           : {idx}
    # Repair Code         : {row['extracted_repair_code']}
    # Repair Description : {row['repair description']}
    # Inspected Qty       : {row[qty_col]}
    # Repair Hours        : {row[rh_col]}
    # Total Order Hours   : {row[toh_col]}
    # """)
            rows_to_drop.append(idx)
   
    # ==============================
    # APPLY DELETION
    # ==============================
    before_rows = len(df)
   
    merged_df_final = df.drop(index=rows_to_drop).reset_index(drop=True)
   
    after_rows = len(merged_df_final)
   
    print("\n===== FINAL SUMMARY =====")
    print(f"Rows before cleanup : {before_rows}")
    print(f"Rows after cleanup  : {after_rows}")
    print(f"Rows removed        : {before_rows - after_rows}")
   
    # ==============================
    # CLEANUP HELPER COLUMNS
    # ==============================
    merged_df_final.drop(
        columns=[
            "_rep_code_norm",
            "_rep_desc_norm",
            "_crd_code_norm",
            "_crd_desc_norm"
        ],
        inplace=True,
        errors="ignore"
    )
   
    # ==============================
    # EXPORT
    # ==============================
    merged_df_final.to_excel("merged_df_repair_FINAL.xlsx", index=False)
   
    print("\n✅ File saved: merged_df_repair_FINAL.xlsx")
   
   
    # mmmmm
 
    return merged_df_final, labor_amount, new_parts_non_llp, new_parts_llp, cfe


def get_escalation_values(result: dict, induction_date: datetime) -> dict:
    # print("[INFO] 🔍 Fetching multiple values from Escalation File...")
    escalation_path = result["file_paths"]["escalation"]
    file_extension = os.path.splitext(escalation_path)[-1].lower()
    print(f"[DEBUG] Escalation file path: {escalation_path}, Extension: {file_extension}")
    # Load escalation sheet
    try:
        if file_extension == ".xls":
            escalation_df = read_excel_from_s3(escalation_path, sheet_name="Other Supplemental Prices", engine='xlrd', header=None)
        elif file_extension == ".xlsx":
            escalation_df = read_excel_from_s3(escalation_path, sheet_name="Other Supplemental Prices", engine='openpyxl', header=None)
        else:
            raise ValueError(f"[ERROR] ❌ Unsupported file format: {file_extension}")
    except Exception as e:
        raise ValueError(f"[ERROR] ❌ Could not read escalation file: {escalation_path} → {e}")
    # print("[INFO] ✅ Successfully loaded escalation sheet.")
    search_patterns = {
        "labor": [
            "over / above labor, including engineering support",
            "over / above labour, including engineering support",
            "over/above labor, including engineering support",
            "over/above labour, including engineering support",
            "over and above labor",
            "over and above labour",
            "over above labor",
            "over above labour",
            "over / above labor",
            "over / above labour"
        ],
        ""
        "new_parts_non_llp": ["new parts (non-llp)"],
        "new_parts_llp": ["new parts (llp)"],
        "cfe": ["customer furnished equipment"]
    }
    search_year = str(induction_date.year)
    print("induction_year",search_year)
    year_position = None
    # Locate year column position
    for i, row in escalation_df.iterrows():
        for j, val in enumerate(row):
            if str(search_year) in str(val):
                year_position = j
                break
        if year_position is not None:
            break
    if year_position is None:
        raise ValueError(f"[ERROR] ❌ Induction year {search_year} not found in Escalation sheet.")
    results = {
        "labor_amount": None,
        "new_parts_non_llp": [],
        "new_parts_llp": [],
        "cfe": []
    }
    # Helper to extract and append values
    def extract_values(pattern_list, target_list):
        count = 0
        for i, row in escalation_df.iterrows():
            row_str = row.astype(str).str.lower().str.replace(r'\s+', ' ', regex=True)
            for val in row_str:
                if any(pattern in val for pattern in pattern_list):
                    extracted = escalation_df.iat[i, year_position]
                    if not pd.isna(extracted):
                        cleaned = str(extracted).replace('$', '').replace(',', '')
                        try:
                            target_list.append(float(cleaned))
                            count += 1
                        except:
                            print(f"[WARNING] ⚠️ Skipping non-numeric value: {extracted}")
        # print(f"[INFO] ✅ Extracted {count} value(s) for pattern {pattern_list[0]}")
    # Extract labor (only one)
    for i, row in escalation_df.iterrows():
        row_str = row.astype(str).str.lower().str.replace(r'\s+', ' ', regex=True)
        for val in row_str:
            if any(pattern in val for pattern in search_patterns["labor"]):
                extracted = escalation_df.iat[i, year_position]
                if not pd.isna(extracted):
                    cleaned = str(extracted).replace('$', '').replace(',', '')
                    try:
                        results["labor_amount"] = float(cleaned)
                        # print(f"[INFO] ✅ Labor amount extracted: {results['labor_amount']}")
                    except:
                        print(f"[WARNING] ⚠️ Could not convert labor amount: {extracted}")
                break
        if results["labor_amount"] is not None:
            break
    # Extract other lists
    extract_values(search_patterns["new_parts_non_llp"], results["new_parts_non_llp"])
    extract_values(search_patterns["new_parts_llp"], results["new_parts_llp"])
    extract_values(search_patterns["cfe"], results["cfe"])
    return results

def apply_repair_sb_logic(result, labor_amount, induction_date, new_crd_df, sb_filtered, repair_blade_df):
    """
    Process internal repair records to extract S-codes and SB references, and enrich with contract SB info.
 
    Parameters:
    - internal_repair_df: DataFrame from internal repair sheet
    - sb_filtered_df: Pre-filtered SB data with 'sb_x' column
    - contract_sb_df: SB sheet from Contract Summary, must have 'SB', 'REPAIR SB Remarks', etc.
 
    Returns:
    - Merged/enriched DataFrame with SB remarks and part mappings
    """
    print("=======START apply_repair_sb_logic =======")
    internal_repair_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="internal",
        sheet_name_substring=""  # set your specific tab filter if needed
    ).copy()
    
    df = internal_repair_df.copy()
    df.columns = [col.strip().lower() for col in df.columns]
    # df.to_excel("repair_service_bulletin_irv1.xlsx", index=False)
    
    
    
    # Step 1: Filter only rows with "S" codes (4-character alphanumeric, starts with S)
    s_code_pattern = r"\bS[A-Z0-9]{3}\b"
    code_col = next((col for col in df.columns if "repair code group & code" in col), None)
    desc_col = next((col for col in df.columns if "repair description" in col), None)
    inspect_col = next((col for col in df.columns if "inspected part no." in col), None)
    hours_col = next((col for col in df.columns if "repair hours" in col), None)
    qty_col = next((col for col in df.columns if "inspected quantity" in col), None)
    ata_chap = next((col for col in df.columns if "ata chapter" in col), None)
    
    # Load the CC Part Number sheet
    cc_parts_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="cc part"
    )
    # Normalize column names
    cc_parts_df.columns = [col.strip().lower() for col in cc_parts_df.columns]
    print("contract_cc")
 
    part_num_col = next((col for col in cc_parts_df.columns if "part" in col and "num" in col), None)
    keyword_col = next((col for col in cc_parts_df.columns if "part" in col and "keyword" in col), None)
    
    # Normalize
    cc_parts_df[part_num_col] = cc_parts_df[part_num_col].astype(str).str.strip().str.upper()
    cc_parts_df[keyword_col] = cc_parts_df[keyword_col].astype(str).str.strip().str.upper()
    
    # Normalize internal repair PN
    df[inspect_col] = df[inspect_col].astype(str).str.strip().str.upper()
    
    # Merge standard description
    df = df.merge(
        cc_parts_df[[part_num_col, keyword_col]],
        left_on=inspect_col,
        right_on=part_num_col,
        how="left"
    )
    
    df.rename(columns={keyword_col: "std_description"}, inplace=True)
    

    if not code_col:
        raise ValueError("Could not find 'Repair Code Group & Code' column.")
 
    df = df[df[code_col].astype(str).str.contains(s_code_pattern, na=False, regex=True)].copy()
    # Step 2: Extract SB patterns like 72-1093, 73-2312, or malformed like 72-940
    def extract_sb_pattern(text):
        """
        Extracts first SB-like pattern such as 72-1093 or 72-940 (padded to 72-0940).
        """
        if not isinstance(text, str):
            return None
        # Look for patterns like 72-1093, 73-231, etc.
        matches = re.findall(r"7\d[-–](\d{3,4})", text)
        if matches:
            prefix_match = re.search(r"(7\d)[-–]", text)
            if prefix_match:
                prefix = prefix_match.group(1)
                suffix = matches[0].zfill(4)  # pad to 4 digits
                return f"{prefix}-{suffix}"
        return None
    df["Extracted SB"] = df[desc_col].apply(extract_sb_pattern)
    df.to_excel("repair_service_bulletin_irv2.xlsx", index=False)
 
    # Step 3: Drop rows where SB couldn't be extracted
    df = df[df["Extracted SB"].notna()].copy()
    # df.to_excel("df2.xlsx", index=False)
    sb_filtered.to_excel("repair_service_bulletin_sb_filtered.xlsx", index=False)
    
    # Step 4: Join with sb_filtered on Extracted SB vs sb_x
    
    if "sb_x" not in sb_filtered.columns:
        raise ValueError("'sb_filtered_df' must have a column named 'sb_x'")
    
    # ????????????????????????????????????????????????
    condition = sb_filtered['sb_x'].str.contains('72-0119', case=False, na=False)
   
    if condition.any():
        # Ensure columns exist
       # for col in ['Extracted SB', 'repair hours']:
            #if col not in sb_filtered.columns:
                #df[col] = None
 
        # Append one row at the end
        df = pd.concat([
            df,
            pd.DataFrame([{'Extracted SB': '72-0119', 'repair hours': 48, 'repair description': 'SB72-0119','repair code group & code':'SB72-0119'}])
        ], ignore_index=True)
 
    
    merged_df = df.merge(sb_filtered, left_on="Extracted SB", right_on="sb_x", how="left")
    merged_df.to_excel("sb_fil_intrep_SB.xlsx", index=False)
 
    # Step 5: Join with Contract SB sheet to get Repair SB Remarks and other fields
    contract_sb_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="sb"   # adjust if your tab is named differently
    ).copy()

    contract_sb_df.columns = [col.strip().lower() for col in contract_sb_df.columns]
    required_cols = ["sb", "repair sb remarks", "category", "new part number", "old part number", "sb description"]
 
    missing = [col for col in required_cols if col.lower() not in contract_sb_df.columns]
    if missing:
        raise ValueError(f"Missing expected columns in contract SB sheet: {missing}")
 
    contract_sb_df = contract_sb_df.rename(columns={"sb": "Extracted SB"})  # align column name for merge
 
    final_df = merged_df.merge(
        contract_sb_df[
            ["Extracted SB", "repair sb remarks", "category", "new part number", "old part number", "sb description"]
        ],
        on="Extracted SB",
        how="left"
    )
    

    pattern = r"\bS[A-Z0-9]{3}\b"
    final_df = final_df[final_df[code_col].astype(str).str.contains(pattern, regex=True, na=False)]

    # Extract canonical 'sb6' (six digits) from description
    # ---------- helpers: normalize/extract ----------
    def extract_sb6_from_text(text: str):
        """
        Return the 6-digit SB core as a string like '740004' or np.nan.
        Handles:
          - 'SB-74-0004', 'SB 74 0004', 'SB:74/0004', 'SB740004'
          - If 'SB' is missing (e.g., '74-0004') we still can extract (for Contract sheet).
        """
        s = str(text or "").upper()
 
        # Prefer matches that explicitly have 'SB' nearby
        m = re.search(r'SB\D*(\d{2})\D*(\d{4})', s)
        if m:
            return f"{m.group(1)}{m.group(2)}"
 
        # Fallback: 2 + 4 digits with separators (e.g., '72-0004')
        m = re.search(r'(\d{2})\D*(\d{4})', s)
        if m:
            return f"{m.group(1)}{m.group(2)}"
 
        return np.nan
    final_df["sb6"] = final_df[desc_col].apply(extract_sb6_from_text)

    print("\ninternal S Codes :", final_df.shape)
    print("\ninternal S Codes :", final_df[code_col].unique())
    
    
    print("final_df_cs")
    final_df.to_excel("repair_service_bulletin_remarks_cs.xlsx", index=False)
    

    # Keep only rows where SB was found
    sb_rows = final_df[final_df["sb6"].notna()].copy()
 
    # Normalize keys used for joins
    sb_rows[inspect_col] = sb_rows[inspect_col].astype(str).str.strip().str.upper()
    sb_rows["sb6"] = sb_rows["sb6"].astype(str)


    #  # ---------- Step 2: CRD (find price by (Part, SB)) ----------
    crd = new_crd_df.copy()
    crd.columns = [str(c).strip().lower() for c in crd.columns]
 
    crd_part_col      = next((c for c in crd.columns if "part" in c and "number" in c), None)
    crd_repair_col    = next((c for c in crd.columns if "repair" in c and "number" in c and "child" not in c), None)
    crd_child_col     = next((c for c in crd.columns if "child repair" in c and "number" in c), None)
    crd_price_col     = next((c for c in crd.columns if c.strip().lower() == "price"), None)
    ata_col_crd     = next((c for c in crd.columns if c.strip().lower() == "ata"), None)
    crd_desc_col     = next((c for c in crd.columns if c.strip().lower() == "component description"), None)
    
    if not all([crd_part_col, crd_repair_col, crd_price_col]):
        raise ValueError("CRD file missing required columns (part number / repair number / price).")
 
    # Normalize CRD keys
    crd[crd_part_col]   = crd[crd_part_col].astype(str).str.strip().str.upper()
    crd[crd_repair_col] = crd[crd_repair_col].astype(str).str.strip().str.upper()
    if crd_child_col:
        crd[crd_child_col] = crd[crd_child_col].astype(str).str.strip().str.upper()
        
    # Normalize ATA columns
    
    
    if ata_col_crd:
        crd[ata_col_crd] = crd[ata_col_crd].astype(str).str.strip().str.upper()
        
        
    # Map CRD part number to standardized description
    crd = crd.merge(
        cc_parts_df[[part_num_col, keyword_col]],
        left_on=crd_part_col,
        right_on=part_num_col,
        how="left"
    )
    
    crd.rename(columns={keyword_col: "std_description"}, inplace=True)
        
 
    # Build sb6 from CRD repair numbers (SB740004 → 740004)
    def crd_to_sb6(val):
        s = str(val or "").upper()
        m = re.search(r'SB\D*(\d{6})', s)
        if m:
            return m.group(1)
        # Also handle '74-0004' in CRD (rare)
        m = re.search(r'(\d{2})\D*(\d{4})', s)
        return f"{m.group(1)}{m.group(2)}" if m else np.nan
 
    crd["_sb6_repair"] = crd[crd_repair_col].apply(crd_to_sb6)
    if crd_child_col:
        crd["_sb6_child"] = crd[crd_child_col].apply(crd_to_sb6)
    
    crd.to_excel("repair_service_bulletin_remarks_crd.xlsx", index=False)
    
    # Step 2a: merge on (Part, sb6) to CRD repair number
    print("Step_2a")
    
    merged = final_df.merge(
        crd[[crd_part_col, "_sb6_repair", crd_price_col,ata_col_crd]],
        left_on=[inspect_col, "sb6"],
        right_on=[crd_part_col, "_sb6_repair"],
        how="left",
        suffixes=("", "_by_repair")
    )
    
    merged.to_excel("repair_service_bulletin_remarks_prev1.xlsx", index=False)
    
    # Step 2b: fallback to child repair number where price is missing
    if crd_child_col:
        print("Step_2b")
        child_merge = merged.merge(
            crd[[crd_part_col, "_sb6_child", crd_price_col]],
            left_on=[inspect_col, "sb6"],
            right_on=[crd_part_col, "_sb6_child"],
            how="left",
            suffixes=("", "_child")
        )
    
        # Fill price only where original is NaN
        merged[crd_price_col] = merged[crd_price_col].fillna(
            pd.to_numeric(child_merge[f"{crd_price_col}_child"], errors="coerce")
        )
    merged.to_excel("repair_service_bulletin_remarks_prev2.xlsx", index=False)
    # ---------- Step 2c: SB + ATA + Description (PN-independent) ----------
    if ata_chap and ata_col_crd:
    
        print("Step 2c: SB + ATA + Description pricing")
    
        # --- Normalize keys ---
        merged["sb6"] = merged["sb6"].astype(str).str.zfill(6)
        crd["_sb6_repair"] = crd["_sb6_repair"].astype(str).str.zfill(6)
    
        merged[ata_chap] = (
            merged[ata_chap]
            .astype(str)
            .str.strip()
            .str.upper()
        )
    
        crd[ata_col_crd] = (
            crd[ata_col_crd]
            .astype(str)
            .str.strip()
            .str.upper()
        )
    
        merged["std_description"] = (
            merged["std_description"]
            .astype(str)
            .str.strip()
            .str.upper()
        )
    
        crd["std_description"] = (
            crd["std_description"]
            .astype(str)
            .str.strip()
            .str.upper()
        )
    
        # --- Apply only where price is still missing ---
        need_ata = merged[crd_price_col].isna()
    
        for idx in merged.loc[need_ata].index:
    
            sb_val = merged.at[idx, "sb6"]
            ata_val = merged.at[idx, ata_chap]
            desc_val = merged.at[idx, "std_description"]
    
            # Step 1: Filter CRD by SB + ATA (PN ignored)
            crd_subset = crd[
                (crd["_sb6_repair"] == sb_val) &
                (crd[ata_col_crd] == ata_val)
            ]
    
            if crd_subset.empty:
                continue
    
            # Step 2: Additional filter using description
            desc_filtered = crd_subset[
                crd_subset["std_description"] == desc_val
            ]
    
            # If description matched → use those rows
            if not desc_filtered.empty:
                avg_price = desc_filtered[crd_price_col].astype(float).mean()
            else:
                # If no description match → fallback to SB+ATA only
                avg_price = crd_subset[crd_price_col].astype(float).mean()
    
            merged.at[idx, crd_price_col] = avg_price
    
    
    merged[crd_price_col] = pd.to_numeric(merged[crd_price_col], errors="coerce")
    
    merged.to_excel("repair_service_bulletin_remarks_prev3.xlsx", index=False)
 
    # ---------- Step 3: Compute totals (labor fallback) ----------
    # Honor passed-in labor_amount; if None, read from escalation
    if labor_amount is None:
        esc = get_escalation_values(result, induction_date)
        labor_amount = esc.get("labor_amount", 0.0)
 
    if hours_col:
        merged[hours_col] = pd.to_numeric(merged[hours_col], errors="coerce").fillna(0)
    else:
        merged["__repair_hours__"] = 0.0
        hours_col = "__repair_hours__"
 
    if qty_col:
        merged[qty_col] = pd.to_numeric(merged[qty_col], errors="coerce").fillna(1)
    else:
        merged["__qty__"] = 1.0
        qty_col = "__qty__"
 
    merged["sb_total_price"] = merged[crd_price_col]
    merged["total_price_final"] = merged["sb_total_price"].fillna(
        merged[hours_col] * float(labor_amount or 0)
    )
    merged["extended_price_final"] = merged[qty_col] * merged["total_price_final"]

    #     # ---------- Final tidy ----------
    for c in ("total_price_final", "extended_price_final", "sb_total_price"):
        if c in merged.columns:
            merged[c] = pd.to_numeric(merged[c], errors="coerce").round(2)
    merged.to_excel("repair_service_bulletin_remarks_prev4.xlsx", index=False)
    merged = merged[merged["total_price_final"].fillna(0) != 0].copy()
    merged.to_excel("repair_service_bulletin_remarks_prev5.xlsx", index=False)
    merged.drop_duplicates(inplace=True)
    # merged = apply_sb_cap(df= merged, pipeline_type = "repair", sb_remarks_col="repair sb remarks", price_col= "total_price_final")
    merged.to_excel("repair_service_bulletin_remarks_before_removing_scrap.xlsx", index=False)
    
    # Removing the Scrap Identified PN
    scrap_unknown = repair_blade_df.copy()
    scrap_unknown = scrap_unknown[['inspected part no.', 'genpact_part_keywords', 'remarks']]
    scrap_unknown = scrap_unknown[scrap_unknown['remarks'].isin(['Inclusion', 'Exclusion as Scrap repair'])]
    exclude_list = scrap_unknown['inspected part no.'].unique()
    merged = merged[~merged[inspect_col].isin(exclude_list)]
        
    merged.to_excel("repair_service_bulletin_remarks.xlsx", index=False)
    # rrrrr
    return merged

def extract_vendor_cost_category(result: dict) -> list:
    # print("[INFO] 🔍 Extracting Vendor Cost Category from Contract...")
 
    # Step 1: Load Cost Category sheet
    cost_category_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="cost category"
    )
 
    # Normalize column names
    cost_category_df.columns = [col.strip().lower() for col in cost_category_df.columns]
    # print("cost_category_df.columns :", cost_category_df.columns)
 
    # Step 2: Identify the relevant column
    vendor_col = next((col for col in cost_category_df.columns if "vendor cost category" in col), None)
    if not vendor_col:
        raise ValueError("[ERROR] ❌ 'Vendor Cost Category' column not found in Cost Category sheet.")
 
    # Step 3: Extract SUB-CON categories
    vendor_categories = cost_category_df[vendor_col].dropna().str.strip().tolist()
    # print(f"[INFO] ✅ Extracted Vendor Cost Categories: {vendor_categories}")
    return vendor_categories


def filter_billing_for_vendor(result: dict, vendor_categories: list) -> pd.DataFrame:
    # print("\n[INFO] 🔍 Filtering Billing Request for Vendor Cost Categories...")

    # Step 1: Load Billing Request DataFrame
    billing_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="billing",
        sheet_name_substring="billing"
    )

    # Normalize columns
    billing_df.columns = [col.strip().lower() for col in billing_df.columns]

    # Step 2: Identify the Cost Category column
    cost_cat_col = next((col for col in billing_df.columns if "cost category" in col), None)

    if not cost_cat_col:
        raise ValueError("[ERROR] ❌ 'Cost Category' column not found in Billing Request sheet.")

    # Step 3: Filter for Vendor Cost Categories
    filtered_billing = billing_df[billing_df[cost_cat_col].isin(vendor_categories)]

    return filtered_billing


def extract_workscope_pricing_for_year(df: pd.DataFrame, induction_year: str) -> pd.DataFrame:
    year_str = f"({induction_year} USD)"
    is_multiindex = isinstance(df.columns, pd.MultiIndex)
 
    col_positions = []
    if is_multiindex:
        for i, col in enumerate(df.columns):
            if year_str in str(col[0]):
                col_positions.append(i)
    else:
        for i, col in enumerate(df.columns):
            if year_str in str(col):
                col_positions.append(i)
 
    if not col_positions:
        raise ValueError(f"[ERROR] ❌ No columns found matching year '{year_str}'.")
 
    start_idx = col_positions[0]
    selected_indices = [start_idx, start_idx + 1, start_idx + 2, start_idx + 3]
 
    # === Always grab the first two columns (assuming identifier + description) ===
    first_two_cols = [df.columns[0], df.columns[1]]
 
    # === Add the year-based columns ===
    year_columns = []
    for idx in selected_indices:
        if idx < len(df.columns):
            year_columns.append(df.columns[idx])
 
    final_columns = first_two_cols + year_columns
    selected_df = df[final_columns].copy()
 
    # === Flatten MultiIndex if needed ===
    if is_multiindex:
        selected_df.columns = [
            f"{str(col[0]).strip()} | {str(col[1]).strip()}"
            for col in selected_df.columns
        ]
    return selected_df, year_str
   
 
def get_dataframe_and_filepath(
    all_dataframes: dict[str, dict[str, pd.DataFrame]],
    all_filepaths: dict[str, str],
    file_key: str,
    sheet_name_substring: str) -> tuple[pd.DataFrame, str]:
    if file_key not in all_dataframes:
        raise ValueError(f"[ERROR] ❌ File key '{file_key}' not loaded or does not exist in all_dataframes.")
 
    file_sheets = all_dataframes[file_key]
 
    if not file_sheets:
        raise ValueError(f"[ERROR] ❌ File '{file_key}' has no loaded sheets. Check if the Excel file is empty, corrupted, or unreadable.")
 
    target_norm = normalize_string(sheet_name_substring)
 
    # Try to find a sheet matching the substring
    matched_sheet = next(
        (sheet for sheet in file_sheets if target_norm in normalize_string(sheet)),
        None
    )
  
    # Get file path
    file_path = all_filepaths.get(file_key)
    if not file_path:
        raise ValueError(f"[ERROR] ❌ File path for '{file_key}' not found in all_filepaths dictionary.")
    return file_sheets[matched_sheet], file_path

 
def trim_and_clean_dataframe(df: pd.DataFrame, threshold: float = 0.8) -> pd.DataFrame:
    trimmed_df = df.copy()
 
    # Step 1: Find first continuous data row
    for idx, row in df.iterrows():
        non_null_ratio = row.notnull().sum() / len(row)
        if non_null_ratio >= threshold:
            trimmed_df = df.iloc[idx:].reset_index(drop=True)
            break
    else:
        print("[WARNING] ⚠️ No continuous data row found; returning original DataFrame.")
 
    # Step 2: Drop fully empty rows
    before_rows = trimmed_df.shape[0]
    trimmed_df = trimmed_df.dropna(how='all')
    after_rows = trimmed_df.shape[0]
 
    # Step 3: Drop fully empty columns
    before_cols = trimmed_df.shape[1]
    trimmed_df = trimmed_df.dropna(axis=1, how='all')
    after_cols = trimmed_df.shape[1]
    return trimmed_df

def compare_excel_sheets_contains(excel_path, target_sheet_names):
    # Load Excel
    xl = pd.ExcelFile(excel_path)
    file_sheets = xl.sheet_names
 
    # Normalize
    normalized_file_sheets = [normalize_string(name) for name in file_sheets]
    normalized_targets = [normalize_string(name) for name in target_sheet_names]
 
    # Compare with 'contains'
    matches = {}
    for target in normalized_targets:
        matched_in_file = [
            orig for orig, norm in zip(file_sheets, normalized_file_sheets)
            if target in norm
        ]
        matches[target] = matched_in_file if matched_in_file else None
    return matches

def write_multiple_dfs_to_sheet_with_gap(excel_path, target_sheet_name, dfs, gap_rows, start_row):
    wb = load_workbook(excel_path)
    file_sheets = wb.sheetnames
 
    # Normalize names
    normalized_file_sheets = [normalize_string(name) for name in file_sheets]
    normalized_target = normalize_string(target_sheet_name)
 
    # Find matching sheet
    matched_sheet = None
    for orig, norm in zip(file_sheets, normalized_file_sheets):
        if normalized_target in norm:
            matched_sheet = orig
            break

    if not matched_sheet:
        print(f"[❌] No matching sheet found for '{target_sheet_name}'.")
        return
 
    current_row = start_row
    summary_info = []
 
    for idx, df in enumerate(dfs):
        # Add blank first column
        df_with_blank = pd.concat([pd.Series([""] * len(df), name=""), df], axis=1)
        with pd.ExcelWriter(excel_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            df_with_blank.to_excel(
                writer,
                sheet_name=matched_sheet,
                startrow=current_row,
                index=False,
                header=True
            )
        df_rows = df.shape[0]
        df_cols = df.shape[1] + 1  # +1 for blank column
        col_start = 1  # column A (since we added a blank column)
        col_end = col_start + df_cols - 1
        row_end = current_row + df_rows + 1
        summary_info.append({
            'df_index': idx + 1,
            'row_start': current_row,
            'row_end': row_end,
            'num_rows': df_rows,
            'num_cols': df_cols,
            'col_start': col_start,
            'col_end': col_end
        })
        current_row += df_rows + gap_rows + 1  # +1 for header row   
    print(f"[✅] All DataFrames written to sheet: '{matched_sheet}' in '{excel_path}'")
    return summary_info


# def write_dfs_with_gaps(
#     file_paths,
#     output_file,
#     sheet_name,#="Out-of-Scope Material & Repair",
#     start_rows=[6, None, None],
#     row_gaps=5
#     ):
#     assert len(file_paths) <= 3, "Function expects exactly 3 Excel file paths"
 
#     dfs = [pd.read_excel(fp) for fp in file_paths]
#     metadata = []
 
#     # Load existing workbook or create a new one
#     try:
#         wb = load_workbook(output_file)
#     except FileNotFoundError:
#         wb = Workbook()
 
#     # Access or create the specified sheet
#     if sheet_name in wb.sheetnames:
#         ws = wb[sheet_name]
#     else:
#         ws = wb.create_sheet(title=sheet_name)
 
#     # Clear existing content in the target sheet
#     for row in ws.iter_rows():
#         for cell in row:
#             cell.value = None
 
#     current_row = start_rows[0] if start_rows[0] else 1
 
#     for idx, df in enumerate(dfs):
#         # Set up metadata
#         start_row = current_row
#         start_col = 2  # Column B (leave Column A blank)
#         end_row = start_row + len(df)
#         end_col = start_col + len(df.columns) - 1
#         num_rows = len(df)
#         num_cols = len(df.columns)
 
#         # Write DataFrame to sheet starting from column B
#         for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
#             for c_idx, value in enumerate(row, start=start_col):
#                 ws.cell(row=r_idx, column=c_idx, value=value)

#         metadata.append({
#             "df_index": idx + 1,
#             "start_row": start_row,
#             "end_row": end_row,
#             "start_col": start_col,
#             "end_col": end_col,
#             "num_rows": num_rows,
#             "num_cols": num_cols
#         })
 
#         current_row = end_row + row_gaps + 1
 
#     # Save final workbook
#     wb.save(output_file)
#     return metadata


def write_dfs_with_gaps(
    file_paths,
    output_file,
    sheet_name,
    start_rows=None,
    row_gaps=7,
):
    """
    Write 1 or 2 DataFrames from Excel files into a target sheet with row gaps.
 
    Args:
        file_paths (list): List of 1 or 2 Excel file paths.
        output_file (str): Path to save the final Excel file.
        sheet_name (str): Target sheet name.
        start_rows (list): List of starting rows (optional). If None, default to [6, None].
        row_gaps (int): Number of blank rows between DataFrames.
    Returns:
        metadata (list of dict): Metadata of written DataFrames.
    """
    assert len(file_paths) in [1, 2], "Function supports only 1 or 2 Excel files"
 
    dfs = [pd.read_excel(fp) for fp in file_paths]
    metadata = []
 
    # Default start_rows logic
    if start_rows is None:
        start_rows = [4] + [None] * (len(dfs) - 1)

    elif len(start_rows) < len(dfs):
        start_rows += [None] * (len(dfs) - len(start_rows))
        
    # Load or create workbook
    try:
        wb = load_workbook(output_file)
    except FileNotFoundError:
        wb = Workbook()
 
    # Access or create the sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
 
    # Clear existing content in the sheet
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
 
    # Initial row
    current_row = start_rows[0] if start_rows[0] else 1
 
    # Write each DataFrame
    for idx, df in enumerate(dfs):
        # Extract and drop 'Price From' Col
        price_from_values = []
        if 'Price From' in df.columns:
            price_from_values = df['Price From'].astype(str).tolist()
            df = df.drop(columns=['Price From'])
            
        start_row = current_row
        start_col = 2  # Column B
        # end_row = start_row + len(df)
        rows_written = len(list(dataframe_to_rows(df, index=False, header=True)))
        end_row = start_row + rows_written - 1
        end_col = start_col + len(df.columns) - 1
 
        # Write DataFrame with header
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                ws.cell(row=r_idx, column=c_idx, value=value)
 
        metadata.append({
            "df_index": idx + 1,
            "start_row": start_row,
            "end_row": end_row,
            "start_col": start_col,
            "end_col": end_col,
            "num_rows": len(df),
            "num_cols": len(df.columns),
            "price_from_data": price_from_values
        })
 
        # Prepare for next block
        current_row = end_row + row_gaps + 1
 
    wb.save(output_file)
    print("metadata----------",metadata)
    return metadata

def write_sb_df_with_gaps(
    file_paths,
    output_file,
    sheet_name,
    start_rows=None,
    row_gaps=7,
):
    """
    Write 1 or 2 DataFrames from Excel files into a target sheet with row gaps.
 
    Args:
        file_paths (list): List of 1 or 2 Excel file paths.
        output_file (str): Path to save the final Excel file.
        sheet_name (str): Target sheet name.
        start_rows (list): List of starting rows (optional). If None, default to [6, None].
        row_gaps (int): Number of blank rows between DataFrames.
    Returns:
        metadata (list of dict): Metadata of written DataFrames.
    """
    assert len(file_paths) in [1, 2], "Function supports only 1 or 2 Excel files"
 
    dfs = [pd.read_excel(fp) for fp in file_paths]
    metadata = []
 
    # Default start_rows logic
    if start_rows is None:
        start_rows = [9] + [None] * (len(dfs) - 1)

    elif len(start_rows) < len(dfs):
        start_rows += [None] * (len(dfs) - len(start_rows))
        
    # Load or create workbook
    try:
        wb = load_workbook(output_file)
    except FileNotFoundError:
        wb = Workbook()
 
    # Access or create the sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
 
    # Clear existing content in the sheet
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
 
    # Initial row
    current_row = start_rows[0] if start_rows[0] else 1
 
    # Write each DataFrame
    for idx, df in enumerate(dfs):
        start_row = current_row
        start_col = 2  # Column B
        end_row = start_row + len(df)
        end_col = start_col + len(df.columns) - 1
 
        # Write DataFrame with header
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                ws.cell(row=r_idx, column=c_idx, value=value)
 
        metadata.append({
            "df_index": idx + 1,
            "start_row": start_row,
            "end_row": end_row,
            "start_col": start_col,
            "end_col": end_col,
            "num_rows": len(df),
            "num_cols": len(df.columns)
        })
 
        # Prepare for next block
        current_row = end_row + row_gaps + 1
 
    wb.save(output_file)
    return metadata

def find_invoice_template_in_current_folder(keyword="draft ge celma ffp invoice template"):
    """Find a file in the current folder containing the keyword."""
    current_folder = Path.cwd()
    normalized_keyword = normalize_string(keyword)
 
    for file in current_folder.iterdir():
        if file.is_file() and normalized_keyword in normalize_string(file.name):
            print(f"[✅] Found file from current dir: {file.resolve()}")
            return file.resolve()
    
    #################
    print("File resolving done")
    raise FileNotFoundError(f"❌ No file found in '{current_folder}' containing '{keyword}'.")


def copy_excel_with_only_cover_and_remove_links(source_path, output_path=None):
    if not os.path.exists(source_path):
        raise FileNotFoundError(f"❌ File not found: {source_path}")
 
    # Step 1: Copy the file to avoid touching original
    if not output_path:
        base_name = os.path.splitext(os.path.basename(source_path))[0]
        output_path = f"{base_name}_COVER_Only.xlsx"
    shutil.copy2(source_path, output_path)
 
    # Step 2: Load the copied file
    wb = load_workbook(output_path, data_only=False)
 
    # Step 3: Find the COVER sheet
    cover_sheet = None
    for sheet_name in wb.sheetnames:
        if "cover" in sheet_name.lower():
            cover_sheet = sheet_name
            break
 
    if not cover_sheet:
        raise ValueError("❌ No 'COVER' sheet found in the workbook.")
 
    ws = wb[cover_sheet]
 
    # Step 4: Convert formulas to values in COVER
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                # Load current value (requires opening with data_only=True)
                formula_result = cell.value  # Keep formula reference in case needed
                try:
                    # Get evaluated value from original workbook
                    temp_wb = load_workbook(output_path, data_only=True)
                    evaluated_ws = temp_wb[cover_sheet]
                    cell.value = evaluated_ws[cell.coordinate].value
                except Exception:
                    cell.value = None  # fallback in case of error
 
    # Step 5: Delete other sheets
    for sheet in wb.sheetnames:
        if sheet != cover_sheet:
            wb.remove(wb[sheet])
 
    # Step 6: Save final file
    wb.save(output_path)
    print(f"[✅] Saved COVER-only file (with formulas converted to values): {output_path}")
    return output_path

def delete_empty_sheets(excel_path):
    print()
    wb = load_workbook(excel_path)
    sheets_to_delete = []
 
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        if all(row is None or all(cell.value in (None, "", " ") for cell in row) for row in ws.iter_rows()):
            sheets_to_delete.append(sheet)
 
    if not sheets_to_delete:
        print("[INFO] ✅ No empty sheets found.")
    else:
        for sheet in sheets_to_delete:
            wb.remove(wb[sheet])
            # print(f"[INFO] 🗑️ Deleted empty sheet: {sheet}")
 
        wb.save(excel_path)
        # print(f"[INFO] ✅ Final Invoice is ready!: {excel_path}")


def _normalize_wks(val):
    if pd.isna(val):
        return val
    s = str(val).strip()
    if s.lower() == 'a+':
        return 'A+'
    if s.lower() in {'a','b','c'}:
        return s.upper()
    try:
        f = float(s)
        if f in (1.0, 2.0, 3.0):
            return {1.0:'A', 2.0:'B', 3.0:'C'}[f]
    except:
        pass
    return s  # leave things like P110 untouched

def _force_text_for_headers(ws, header_row, data_start, data_end, headers):
    want = {h.strip().lower() for h in headers}
    for col in range(1, ws.max_column + 1):
        name = ws.cell(row=header_row, column=col).value
        if name and str(name).strip().lower() in want:
            for r in range(data_start, data_end + 1):
                ws.cell(row=r, column=col).number_format = '@'  # Text

def create_in_scope(result: dict, output_excel: str):
    print()
    q_year, induction_date, cust_name, customer_first_word,q_year_g3 = extract_q_year_from_timeliness(result) # United Airlines Inc. = UAI --> UAL
    induction_year = str(induction_date.year)

    wp_df, wp_file_path = get_dataframe_and_filepath(result["dataframes"], result["file_paths"], "contract", "workscope pricing")
    inscope_df, year_str = extract_workscope_pricing_for_year(wp_df, str(induction_year))

    wp2_df, wp_file_path2 = get_dataframe_and_filepath(result["dataframes"], result["file_paths"], "contract", "workscope pricing 2")
    inscope_df2, year_str = extract_workscope_pricing_for_year(wp2_df, str(induction_year))
    
    trimmed_inscope_df = trim_and_clean_dataframe(inscope_df, 0.8)
    trimmed_inscope_df2 = trim_and_clean_dataframe(inscope_df2, 0.8)

    temp_excel_path1, temp_excel_path2 = "temp_wprkscope_pricing.xlsx", "temp_wprkscope_pricing2.xlsx"
    trimmed_inscope_df.to_excel(temp_excel_path1, index=False)
    trimmed_inscope_df2.to_excel(temp_excel_path2, index=False)

    trimmed_inscope_df_2nd = pd.read_excel(temp_excel_path1, header = 1)
    trimmed_inscope_df2_2nd = pd.read_excel(temp_excel_path2, header = 1)

    trimmed_inscope_df_2nd.columns = [str(col[1]).strip().lower() if isinstance(col, tuple) else str(col).strip().lower() for col in trimmed_inscope_df_2nd.columns]
    trimmed_inscope_df2_2nd.columns = [str(col[1]).strip().lower() if isinstance(col, tuple) else str(col).strip().lower() for col in trimmed_inscope_df2_2nd.columns]
    
    # Normalize key columns
    trimmed_inscope_df_2nd['minor module'] = trimmed_inscope_df_2nd['minor module'].apply(lambda x: str(x).lower() if pd.notna(x) else x)
    trimmed_inscope_df2_2nd['module identifier'] = trimmed_inscope_df2_2nd['module identifier'].apply(lambda x: str(x).lower() if pd.notna(x) else x)

    # Detect price columns
    a_col = next((col for col in trimmed_inscope_df_2nd.columns if col == "a"), None)
    b_col = next((col for col in trimmed_inscope_df_2nd.columns if col == "b"), None)
    c_col = next((col for col in trimmed_inscope_df_2nd.columns if col == "c"), None)

    aa_col = next((col for col in trimmed_inscope_df2_2nd.columns if col == "a"), None)
    bb_col = next((col for col in trimmed_inscope_df2_2nd.columns if col == "b"), None)
    cc_col = next((col for col in trimmed_inscope_df2_2nd.columns if col == "c"), None)

    # Fill NaNs and clean formats
    for col in [a_col, "a+", b_col, c_col]:
        if col in inscope_df.columns:
            trimmed_inscope_df_2nd[col] = trimmed_inscope_df_2nd[col].fillna(0).astype(float)
    for col in [aa_col, bb_col, cc_col]:
        if col in trimmed_inscope_df2.columns:
            trimmed_inscope_df2_2nd[col] = trimmed_inscope_df2_2nd[col].replace({r'\$': '', r',': ''}, regex=True).fillna(0).astype(float)

    # Merge with SAP
    sap_df = get_dataframe_by_file_and_sheet(result["dataframes"], "sap", "ze14")
    sap_df.columns = [col.strip().lower() for col in sap_df.columns]

    sap_df['iin'] = sap_df['iin'].apply(lambda x: str(x).lower() if pd.notna(x) else x)

    minor_module = trimmed_inscope_df_2nd['minor module'].unique()
    modulex_sap_filtered = sap_df[sap_df['iin'].isin(minor_module)].copy()
    modulex_sap_filtered.rename(columns={"iin": "minor module"}, inplace=True)
    inscope_sap_merged_df = trimmed_inscope_df_2nd.merge(modulex_sap_filtered, on="minor module", how="left")

    inscope_sap_merged_df.columns = [col.strip().lower() for col in inscope_sap_merged_df.columns]
    
    inscope_sap_merged_df.rename(columns={'defined workscope level': 'initial wks', 'customer workscope level': 'actual wks'}, inplace=True)
    # inscope_sap_merged_df["initial wks"] = inscope_sap_merged_df["initial wks"].apply(_normalize_wks_value)
    # inscope_sap_merged_df["actual wks"] = inscope_sap_merged_df["actual wks"].apply(_normalize_wks_value)

    module_identifier = ['72x']
    moduleident_sap_filtered = sap_df[sap_df['iin'].isin(module_identifier)].copy()
    moduleident_sap_filtered.rename(columns={"iin": "module identifier"}, inplace=True)
    inscope_egx_df = trimmed_inscope_df2_2nd.merge(moduleident_sap_filtered, on="module identifier", how="left")
        
    inscope_egx_df.rename(columns={'defined workscope level': 'initial wks', 'customer workscope level': 'actual wks'}, inplace=True)
    # inscope_egx_df["initial wks"] = inscope_egx_df["initial wks"].apply(_normalize_wks_value)
    # inscope_egx_df["actual wks"] = inscope_egx_df["actual wks"].apply(_normalize_wks_value)

    inscope_egx_df['initial wks'] = 3
    inscope_egx_df['actual wks'] = 3

    # inscope_sap_merged_df['initial wks'] = inscope_sap_merged_df['initial wks'].replace({'a': 'A', 'b': 'B', 'c': 'C'})
    # inscope_sap_merged_df['actual wks'] = inscope_sap_merged_df['actual wks'].replace({'a': 'A', 'b': 'B', 'c': 'C'})

    inscope_egx_df['module identifier'] = inscope_egx_df['module identifier'].replace({'egx module': 'EGX Module', 'lru module': 'LRU Module'})
    # inscope_egx_df['initial wks'] = inscope_egx_df['initial wks'].replace({1: 'A', 2: 'B', 3: 'C', 'c':'C'})
    # inscope_egx_df['actual wks'] = inscope_egx_df['actual wks'].replace({1: 'A', 2: 'B', 3: 'C', 'c':'C'})

    for df in (inscope_sap_merged_df, inscope_egx_df):
        for col in ['initial wks','actual wks']:
            if col not in df.columns:
                df[col] = None
            df[col] = df[col].apply(_normalize_wks)
    # df['initial wks'] = df['initial wks'].apply(_normalize_wks)
    # df['actual wks'] = 3 #df['actual wks'].apply(_normalize_wks)

    wks_map = {1: 'A', 2: 'B', 3: 'C'}
    for df in [inscope_sap_merged_df, inscope_egx_df]:
        # df['initial wks'] = df.get('initial wks', 3)
        # df['actual wks'] = df.get('actual wks', 3)
        df['actual price'] = df.apply(
            lambda row: row.get(wks_map.get(int(row['actual wks']), None), 0) if pd.notna(row['actual wks']) and 
            str(row['actual wks']).isdigit() else 0,
            axis=1
        )
    #     df['initial wks'] = df['initial wks'].replace(wks_map)
    #     df['actual wks'] = df['actual wks'].replace(wks_map)

    if (os.path.exists(temp_excel_path1)) & (os.path.exists(temp_excel_path2)):
        os.remove(temp_excel_path1)
        os.remove(temp_excel_path2)
        # print("Temp file1 and 2 deleted")
    
    inscope_sap_merged_df.rename(columns = {
        'minor module' : 'Minor Module',
        'description_x' : 'Description',
        'a' : 'A',
        'a+':'A+',
        'b':'B',
        'c':'C',
        'initial wks':'Initial WKS',
        'actual wks':'Actual WKS',
        'actual price':'Actual Price'}, inplace=True)
    # print(inscope_sap_merged_df['Initial WKS'].unique())
    inscope_sap_merged_df = inscope_sap_merged_df[['Minor Module', 'Description', 'A', 'A+', 'B', 'C', 'Initial WKS', 'Actual WKS', 'Actual Price']]

    inscope_egx_df.rename(columns = {
        'module identifier' : 'Module Identifier', 
        'description_x' : 'Description',
        'a' : 'A',
        'b':'B', 
        'c':'C', 
        'initial wks':'Initial WKS', 
        'actual wks':'Actual WKS',
        'actual price':'Actual Price'
        }, inplace=True)
    # inscope_df2 = inscope_egx_df.copy()
    # print(inscope_egx_df['Initial WKS'].unique())
    inscope_egx_df = inscope_egx_df[['Module Identifier', 'Description', 'A', 'B', 'C', 'Initial WKS', 'Actual WKS', 'Actual Price']]
    # inscope_egx_df.to_excel("inscope_egx_df.xlsx", index=False)
    inscope_sheet_name = "MFP" #"In-Scope"

    # Change P110 to SP110 & Change 0 to SP110
    inscope_sap_merged_df["Initial WKS"] = inscope_sap_merged_df["Initial WKS"].replace({"P110": "SP110", "0" : "SP110", "0.0" : "SP110"})
    inscope_sap_merged_df["Actual WKS"] = inscope_sap_merged_df["Actual WKS"].replace({"0" : "SP110", "0.0" : "SP110"})
    # inscope_sap_merged_df.to_excel("inscope_sap_merged_df.xlsx", index=False)

    return inscope_sap_merged_df, inscope_egx_df, inscope_sheet_name, year_str


def create_module_columns(billing_df: pd.DataFrame) -> pd.DataFrame:
    """
    Build 'module' and 'refine_module' for each row by checking (in order):
      1) ATA Chapter
      2) ATA Section Code  (format to NN-NN-NN)
      3) Component CSN     (take 6 digits after '_', format to NN-NN-NN)
    Then: refine_module = middle two digits + 'x' (e.g., 72-21-00 -> 21x)
    """
    df = billing_df.copy()
    df.columns = [c.strip().lower() for c in df.columns]
 
    # Auto-detect likely column names
    chapter_col = next((c for c in df.columns if "ata" in c and "chapter" in c), None)
    section_col = next((c for c in df.columns if "ata" in c and "section" in c), None)
    csn_col     = next((c for c in df.columns if "component" in c and "csn" in c), None)
 
    def to_module_format(raw) -> str | None:
        """Normalize any 4/6 digit or already hyphenated value to NN-NN-NN."""
        if pd.isna(raw):
            return None
        s = str(raw).strip()
 
        # already NN-NN-NN?
        if re.fullmatch(r"\d{2}-\d{2}-\d{2}", s):
            return s
 
        # clean digits only
        digits = re.sub(r"\D", "", s)
        if len(digits) == 6:
            return f"{digits[0:2]}-{digits[2:4]}-{digits[4:6]}"
        if len(digits) == 4:
            d6 = digits + "00"
            return f"{d6[0:2]}-{d6[2:4]}-{d6[4:6]}"
        return None
 
    def module_from_csn(raw) -> str | None:
        """Extract 6 digits after first '_' and format to NN-NN-NN."""
        if pd.isna(raw):
            return None
        s = str(raw)
 
        # Prefer strictly '_(6 digits)'
        m = re.search(r"_(\d{6})", s)
        if not m:
            # fallback: find first '_' then any 6 consecutive digits after it
            pos = s.find("_")
            if pos != -1:
                tail = s[pos+1:]
                m = re.search(r"(\d{6})", tail)
        if m:
            d = m.group(1)
            return f"{d[0:2]}-{d[2:4]}-{d[4:6]}"
        return None
 
    def refine_from_module(mod) -> str | None:
        """Get middle two digits from NN-NN-NN and append 'x'."""
        if not isinstance(mod, str):
            return None
        m = re.fullmatch(r"\d{2}-(\d{2})-\d{2}", mod)
        return (m.group(1) + "x") if m else None
 
    # Build 'module' following your order of precedence
    modules = []
    for _, row in df.iterrows():
        mod = None
 
        # 1) ATA Chapter
        if chapter_col:
            mod = to_module_format(row.get(chapter_col))
 
        # 2) ATA Section Code (format with hyphens)
        if not mod and section_col:
            mod = to_module_format(row.get(section_col))
 
        # 3) Component CSN (extract 6 digits after '_' and format)
        if not mod and csn_col:
            mod = module_from_csn(row.get(csn_col))
 
        modules.append(mod)
 
    df["module"] = modules
    df["refine_module"] = df["module"].apply(refine_from_module)
    return df



def create_llp_replacement(result: dict, output_excel: str, cust_name: str):
 
    final_matl = pd.read_excel("final_summary_material.xlsx")
 
    # Create empty LIFE CYCLES COLUMNS
    required_cols = [
        'total_remaining_cicles',
        'total_life_cicles',
        'used_llp_percent'
    ]
    final_matl = final_matl.assign(
        **{c: pd.NA for c in required_cols if c not in final_matl.columns}
    )
 
    final_matl_with_mod = create_module_columns(final_matl)
    final_matl_with_mod.to_excel("revised_final_summary_material.xlsx", index=False)
 
    final_matl_df = pd.read_excel("revised_final_summary_material.xlsx")
 
    # =======================
    # NEW LLP
    # =======================
    new_llp_filt = ['SCC-LLP-NEW-GE']
    new_llp_df = final_matl_df[
        (final_matl_df['cost category'].isin(new_llp_filt)) &
        (final_matl_df['llp_flag'] == 'Y')
    ].copy()
 
    # 🔍 PRINT 1
    print(f"🆕 NEW LLP rows after filter: {new_llp_df.shape}")
 
    new_llp_df["Status"] = (
        new_llp_df["cost category"]
        .astype(str)
        .apply(
            lambda x: "NEW"
            if any(tag in x for tag in ["SCC-CONSUMABLE-AERO", "SCC-CONSUMABLE-NON-AERO"])
            else re.search(r"(?i)\b(NEW|USED)\b", x).group(1).upper()
            if re.search(r"(?i)\b(NEW|USED)\b", x)
            else None
        )
    )
 
    mask_aero = new_llp_df["cost category"].astype(str).str.contains("AERO", case=False, na=False)
    new_llp_df.loc[mask_aero, "Status"] = "NEW"
 
    new_llp_df['Material Type'] = new_llp_df['cost category'].str[4:]
    
    new_llp_df['remarks'] = new_llp_df.apply(
        lambda row: append_module_to_remarks(
            row['formatted_ata_long'],
            row['module'],
            row['remarks']
        ),
        axis=1
    )
 
    new_llp_df.rename(columns={
        'item (sd)': 'Item (SD)',
        'formatted_ata_long': 'ATA Chapter Code',
        'material': 'Part Number',
        'description': 'Part Description',
        'target quantity': 'Qty',
        'pn_clp_lookup': 'CLP',
        'matl_clp_lookup': 'Unit Price',
        'extended_price': 'Extended Price',
        'discount': 'Discount %',
        'discount_amount': 'Discount $',
        'handling fee': 'Handling Fee %',
        'handling_fee': 'Handling Fee Price',
        'total_price': 'Excluded [O&A] Total Price',
        'irc description': 'Replacement Remarks',
        'order': 'Service Order Equipment / MRB',
        'remarks': 'Comments'
    }, inplace=True)
 
    new_llp_df = new_llp_df[
        new_llp_df["Excluded [O&A] Total Price"] > 0
    ].drop_duplicates()
 
    new_llp_df = new_llp_df[
        [
            "Item (SD)", "ATA Chapter Code", "Material Type", "Part Number",
            "Part Description", "Qty", "CLP", "Unit Price", "Extended Price",
            "Discount %", "Discount $", "Handling Fee %", "Handling Fee Price",
            "Excluded [O&A] Total Price", "Replacement Remarks",
            "Comments", "Service Order Equipment / MRB"
        ]
    ]
 
    # 🔍 PRINT 2
    print(
        f"🆕 NEW LLP final | rows={new_llp_df.shape[0]} | "
        f"price nulls={new_llp_df['Excluded [O&A] Total Price'].isna().sum()}"
    )
 
    # =======================
    # USED LLP
    # =======================
    used_llp_cc = ['SCC-LLP-USED-LP', 'SCC-LLP-USED-GE']
    used_llp_df = final_matl_df[
        final_matl_df['cost category'].isin(used_llp_cc)
    ].copy()
 
    used_llp_df['Status'] = 'USED'
 
    # 🔍 PRINT 3
    print(f"♻️ USED LLP rows after filter: {used_llp_df.shape}")
 
    used_llp_df['Material Type'] = used_llp_df['cost category'].str[4:]
 
    used_llp_df['total_remaining_cicles'] = pd.to_numeric(
        used_llp_df['total_remaining_cicles'], errors='coerce'
    )
    used_llp_df['total_life_cicles'] = pd.to_numeric(
        used_llp_df['total_life_cicles'], errors='coerce'
    )
    used_llp_df['used_llp_percent'] = pd.to_numeric(
        used_llp_df['used_llp_percent'], errors='coerce'
    )
    used_llp_df['matl_clp_lookup'] = pd.to_numeric(
        used_llp_df['matl_clp_lookup'], errors='coerce'
    )
 
    used_llp_df['total_price'] = (
        (used_llp_df['total_remaining_cicles'] / used_llp_df['total_life_cicles']) *
        used_llp_df['used_llp_percent'] *
        used_llp_df['matl_clp_lookup']
    )
    
    used_llp_df['remarks'] = used_llp_df.apply(
        lambda row: append_module_to_remarks(
            row['formatted_ata_long'],
            row['module'],
            row['remarks']
        ),
        axis=1
    )
    
    
 
    used_llp_df.rename(columns={
        'item (sd)': 'Item (SD)',
        'formatted_ata_long': 'ATA Chapter Code',
        'material': 'Part Number',
        'description': 'Part Description',
        'target quantity': 'Qty',
        'pn_clp_lookup': 'CLP',
        'matl_clp_lookup': 'Unit Price',
        'extended_price': 'Extended Price',
        'discount': 'Discount %',
        'discount_amount': 'Discount $',
        'handling fee': 'Handling Fee %',
        'handling_fee': 'Handling Fee Price',
        'total_price': 'Excluded [O&A] Total Price',
        'irc description': 'Replacement Remarks',
        'order': 'Service Order Equipment / MRB',
        'remarks': 'Comments'
    }, inplace=True)
 
    used_llp_df = used_llp_df[
        used_llp_df["Excluded [O&A] Total Price"] > 0
    ].drop_duplicates()
 
    used_llp_df = used_llp_df[
        [
            "Item (SD)", "ATA Chapter Code", "Material Type", "Part Number",
            "Part Description", "Qty", "CLP", "Unit Price", "Extended Price",
            "Discount %", "Discount $", "Handling Fee %", "Handling Fee Price",
            "Excluded [O&A] Total Price", "Replacement Remarks",
            "Comments", "Service Order Equipment / MRB"
        ]
    ]
 
    used_llp_df.drop_duplicates(subset=["Part Number"], keep="first", inplace=True)
    used_llp_df['Unit Price'] = used_llp_df['Excluded [O&A] Total Price']
 
    # 🔍 PRINT 4
    print(
        f"♻️ USED LLP final | rows={used_llp_df.shape[0]} | "
        f"price min={used_llp_df['Excluded [O&A] Total Price'].min()} | "
        f"max={used_llp_df['Excluded [O&A] Total Price'].max()}"
    )
 
    # =======================
    # FINAL EXPORTS
    # =======================
    new_llp_df.to_excel("llp_new_table.xlsx", index=False)
    used_llp_df.to_excel("llp_used_table.xlsx", index=False)
 
    overall_llp_df = pd.concat([new_llp_df, used_llp_df])
    overall_llp_df.drop_duplicates(subset=["Item (SD)"], keep="first", inplace=True)
    overall_llp_df.to_excel("overall_llp_df.xlsx", index=False)

 
    return new_llp_df, used_llp_df, overall_llp_df



def _norm_series(s):
    return s.astype(str).str.strip().str.upper()
 
def _col(df, name_like):
    name_like = name_like.lower()
    for c in df.columns:
        if name_like in str(c).lower():
            return c
    return None
 
def _hyphenate_figure_to_ata_6(fig_value):
    if pd.isna(fig_value):
        return pd.NA
    digits = re.sub(r"\D", "", str(fig_value))[:6]
    digits = (digits + "000000")[:6]   # pad to 6 if short
    return f"{digits[0:2]}-{digits[2:4]}-{digits[4:6]}"
 

def annotate_material_with_schedule_and_eipc_from_store(
    result: dict,
    pipeline_type: str,
    material_file_key: str,
    material_sheet_substring: str,                 # e.g. "final"
    contract_file_key: str,                        # file key for the contract workbook
    schedule2_substring: str = "Schedule_2 parts",
    schedule3_substring: str = "Schedule_3 parts",
    eipc_file_key: str = "eipc",
    eipc_sheet_substring: str | None = None,
    ) -> pd.DataFrame:
 
    # ---------- helpers ----------
    def _norm_series(s: pd.Series) -> pd.Series:
        s = s.astype(str)
        s = s.replace({"\u00A0": " "}, regex=False)
        s = s.str.lower().str.strip()
        s = s.str.replace(r"\s+", " ", regex=True)
        s = s.where(~s.isin({"nan", "none", "null", ""}), np.nan)
        return s
 
    def _find_col(df: pd.DataFrame, *need_substrings, prefer: list[str] | None = None) -> str | None:
        cols = list(df.columns)
        lc   = [c.lower() for c in cols]
        hits = [cols[i] for i,c in enumerate(lc) if all(ns in c for ns in [s.lower() for s in need_substrings])]
        if not hits:
            return None
        if prefer:
            pref_l = [p.lower() for p in prefer]
            for h in hits:
                if any(p in h.lower() for p in pref_l):
                    return h
        return hits[0]
 
    def _hyphenate_figure_to_ata_6(val: str) -> str | None:
        if pd.isna(val):
            return None
        s = re.sub(r"\D", "", str(val))
        if not s:
            return None
        s = (s + "000000")[:6]
        return f"{s[0:2]}-{s[2:4]}-{s[4:6]}"
 
    # ---------- load inputs ----------
    # s2 = get_dataframe_by_file_and_sheet(
    #     all_dataframes=result["dataframes"],
    #     file_key=contract_file_key,
    #     sheet_name_substring=schedule2_substring
    # ).copy()
 
    # s3 = get_dataframe_by_file_and_sheet(
    #     all_dataframes=result["dataframes"],
    #     file_key=contract_file_key,
    #     sheet_name_substring=schedule3_substring
    # ).copy()
 
    eipc_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key=eipc_file_key,
        sheet_name_substring=(eipc_sheet_substring or "")
    ).copy()

    # s2.columns     = [c.strip() for c in s2.columns]
    # s3.columns     = [c.strip() for c in s3.columns]
    eipc_df.columns= [c.strip() for c in eipc_df.columns]

    # ---------- schedule columns ----------
    # Example MPN (PN) + Part Description + Nomenclature
    # s2_mpn = _find_col(s2, "example", "mpn") or _find_col(s2, "material")
    # s2_desc= _find_col(s2, "part", "description") or _find_col(s2, "description")
    # s2_nom = _find_col(s2, "nomenclature") or _find_col(s2, "description")
 
    # s3_mpn = _find_col(s3, "example", "mpn") or _find_col(s3, "material")
    # s3_desc= _find_col(s3, "part", "description") or _find_col(s3, "description")
    # s3_nom = _find_col(s3, "nomenclature") or _find_col(s3, "description")
    
    # print("s2_mpn, s2_desc, s2_nom, s3_mpn, s3_desc, s3_nom: ", s2_mpn, s2_desc, s2_nom, s3_mpn, s3_desc, s3_nom)
    # if not all([s2_mpn, s2_desc, s2_nom, s3_mpn, s3_desc, s3_nom]):
    #     raise ValueError("Schedules must have Example MPN, Part Description, and Nomenclature.")
 
    # Normalize schedule keys
    # s2["_mpn_norm"]  = _norm_series(s2[s2_mpn])
    # s2["_desc_norm"] = _norm_series(s2[s2_desc])
 
    # s3["_mpn_norm"]  = _norm_series(s3[s3_mpn])
    # s3["_desc_norm"] = _norm_series(s3[s3_desc])
 
    # ---------- build mappings ----------
    # PN → Nomenclature
    # map_s2_pn = dict(zip(s2["_mpn_norm"], s2[s2_nom]))
    # map_s3_pn = dict(zip(s3["_mpn_norm"], s3[s3_nom]))
    # # Description → Nomenclature
    # map_s2_desc = dict(zip(s2["_desc_norm"], s2[s2_nom]))
    # map_s3_desc = dict(zip(s3["_desc_norm"], s3[s3_nom]))

    # ---------- EIPC mapping (by PN) ----------
    fig_col = _find_col(eipc_df, "figure")
    pn_col  = _find_col(eipc_df, "part", "number") or _find_col(eipc_df, "pn")
    if not all([fig_col, pn_col]):
        raise ValueError("EIPC sheet must contain 'Figure' and 'Part Number' (or 'PN').")
 
    eipc_df["_pn_norm"] = _norm_series(eipc_df[pn_col])

    # --- Load dataframes from your store
    if pipeline_type in ['material', 'vendor']:
        annotate_df = pd.read_excel(f"{material_file_key}.xlsx")
  
        # ---------- normalize headers ----------
        annotate_df.columns = [c.strip() for c in annotate_df.columns]

        # ---------- material keys ----------
        # Material column
        mat_cols_lower = [c.lower() for c in annotate_df.columns]
        if "material" not in mat_cols_lower:
            raise ValueError("The material sheet does not contain a 'Material' column.")
        material_col = annotate_df.columns[mat_cols_lower.index("material")]
        annotate_df["_material_norm"] = _norm_series(annotate_df[material_col])
    
        # special_part_keywords column
        spk_col = _find_col(annotate_df, "special", "part", "keyword") \
                or _find_col(annotate_df, "special_part_keywords") \
                or _find_col(annotate_df, "genpact_part_keyword")
        if spk_col:
            annotate_df["_spk_norm"] = _norm_series(annotate_df[spk_col])
        else:
            annotate_df["_spk_norm"] = np.nan  # keeps logic simple
 
 
        # ---------- apply PN match first ----------
        # nom_s2_by_pn = annotate_df["_material_norm"].map(map_s2_pn)
        # nom_s3_by_pn = annotate_df["_material_norm"].map(map_s3_pn)
    
        # # ---------- fallback: keywords vs Part Description ----------
        # nom_s2_by_kw = annotate_df["_spk_norm"].map(map_s2_desc)
        # nom_s3_by_kw = annotate_df["_spk_norm"].map(map_s3_desc)
    
        # # Priority: PN S2 → PN S3 → KW S2 → KW S3
        # annotate_df["LRU Schedule Name"] = np.select(
        #     [
        #         nom_s2_by_pn.notna(),
        #         nom_s2_by_kw.notna(),
        #     ],
        #     [
        #         nom_s2_by_pn.astype(str) + " / Schedule 2",
        #         nom_s2_by_kw.astype(str) + " / Schedule 2",
        #     ],
        #     default=pd.NA,
        # )
        annotate_df["LRU Schedule Name"] = pd.NA
  
        tmp = annotate_df[["_material_norm"]].merge(
            eipc_df[["_pn_norm", fig_col]],
            left_on="_material_norm", right_on="_pn_norm",
            how="left", sort=False
        )
    
        last_fig = (
            tmp.dropna(subset=[fig_col], how="all")
            .groupby("_material_norm", as_index=False, sort=False)
            .tail(1)[["_material_norm", fig_col]]
        )
        fig_map = dict(zip(last_fig["_material_norm"], last_fig[fig_col]))
    
        annotate_df["_figure_raw"] = annotate_df["_material_norm"].map(fig_map)
        annotate_df["_figure_ata"] = annotate_df["_figure_raw"].apply(_hyphenate_figure_to_ata_6)
        annotate_df["EIPC Reference"] = annotate_df["_figure_ata"].apply(lambda x: f"EIPC {x}" if pd.notna(x) else pd.NA)
    
        # ---------- cleanup ----------
        annotate_df.drop(columns=[c for c in ["_material_norm", "_spk_norm", "_figure_raw", "_figure_ata"] if c in annotate_df.columns],
                    inplace=True, errors="ignore")
        annotate_df.drop_duplicates(inplace=True)
    
    else:
        
        annotate_df = pd.read_excel(f"{material_file_key}.xlsx")
  
        # ---------- normalize headers ----------
        annotate_df.columns = [c.strip() for c in annotate_df.columns]

        # ---------- material keys ----------
        # Material column
        rep_cols_lower = [c.lower() for c in annotate_df.columns]
        rep_pn_col = next((col for col in annotate_df.columns if "inspected part no" in col and "inspected part" in col), None)
        
        if rep_pn_col not in rep_cols_lower:
            raise ValueError("The inspected part no sheet does not contain a 'inspected part no.' column.")
        inspected_pn_col = annotate_df.columns[rep_cols_lower.index("inspected part no.")]
        annotate_df["_inspected_pn_norm"] = _norm_series(annotate_df[inspected_pn_col])
    
        # special_part_keywords column
        spk_col = _find_col(annotate_df, "csn description", "csn", "description") \
                or _find_col(annotate_df, "special_part_keywords") \
                or _find_col(annotate_df, "genpact_part_keyword")
        if spk_col:
            annotate_df["_spk_norm"] = _norm_series(annotate_df[spk_col])
        else:
            annotate_df["_spk_norm"] = np.nan  # keeps logic simple
 
 
        # ---------- apply PN match first ----------
        # nom_s2_by_pn = annotate_df["_inspected_pn_norm"].map(map_s2_pn)
        # nom_s3_by_pn = annotate_df["_inspected_pn_norm"].map(map_s3_pn)
    
        # # ---------- fallback: keywords vs Part Description ----------
        # nom_s2_by_kw = annotate_df["_spk_norm"].map(map_s2_desc)
        # nom_s3_by_kw = annotate_df["_spk_norm"].map(map_s3_desc)
    
        # Priority: PN S2 → PN S3 → KW S2 → KW S3
        # annotate_df["LRU Schedule Name"] = np.select(
        #     [
        #         nom_s3_by_pn.notna(),
        #         nom_s3_by_kw.notna(),
        #     ],
        #     [
        #         nom_s3_by_pn.astype(str) + " / Schedule 3",
        #         nom_s3_by_kw.astype(str) + " / Schedule 3",
        #     ],
        #     default=pd.NA,
        # )
        annotate_df["LRU Schedule Name"] = pd.NA
  
        tmp = annotate_df[["_inspected_pn_norm"]].merge(
            eipc_df[["_pn_norm", fig_col]],
            left_on="_inspected_pn_norm", right_on="_pn_norm",
            how="left", sort=False
        )
    
        last_fig = (
            tmp.dropna(subset=[fig_col], how="all")
            .groupby("_inspected_pn_norm", as_index=False, sort=False)
            .tail(1)[["_inspected_pn_norm", fig_col]]
        )
        fig_map = dict(zip(last_fig["_inspected_pn_norm"], last_fig[fig_col]))
    
        annotate_df["_figure_raw"] = annotate_df["_inspected_pn_norm"].map(fig_map)
        annotate_df["_figure_ata"] = annotate_df["_figure_raw"].apply(_hyphenate_figure_to_ata_6)
        annotate_df["EIPC Reference"] = annotate_df["_figure_ata"].apply(lambda x: f"EIPC {x}" if pd.notna(x) else pd.NA)
    
        # ---------- cleanup ----------
        annotate_df.drop(columns=[c for c in ["_inspected_pn_norm", "_spk_norm", "_figure_raw", "_figure_ata"] if c in annotate_df.columns],
                    inplace=True, errors="ignore")
        annotate_df.drop_duplicates(inplace=True)

    annotate_df.to_excel("check_PN_PK_QEC-LRU.xlsx", index=False)
    return annotate_df


def _pick_col(cols, *must_have, case_insensitive=True):
    for c in cols:
        s = str(c)
        hay = s.lower() if case_insensitive else s
        if all(tok.lower() in hay for tok in must_have):
            return c
    return None
 
def _normalize_text(s: pd.Series) -> pd.Series:
    return (s.astype(str)
              .str.upper()
              .str.replace(r"[^\w\s]", " ", regex=True)
              .str.replace(r"\s+", " ", regex=True)
              .str.strip())
 
def _anyorder_wholeword_pattern(keyword: str) -> re.Pattern | None:
    tokens = re.findall(r"\w+", str(keyword).upper())
    if not tokens:
        return None
    lookaheads = "".join(f"(?=.*\\b{re.escape(t)}\\b)" for t in tokens)
    return re.compile(lookaheads + ".*")

# ---------- helpers ----------
def pick_col(cols, *need):
    for c in cols:
        s = str(c).lower()
        if all(tok.lower() in s for tok in need):
            return c
    return None

def parse_symbol_driven_rule(rule_text: str):
    """
    Example:
    'Connector & Replacement OR Ring'
    →
    [
        ['CONNECTOR', 'REPLACEMENT'],
        ['RING']
    ]
    """
    if not rule_text or not isinstance(rule_text, str):
        return []
 
    rule_text = (
        rule_text.upper()
        .replace('"', '')
        .replace("'", '')
        .strip()
    )
 
    or_groups = []
 
    for or_part in re.split(r"\bOR\b", rule_text):
        and_terms = [
            term.strip()
            for term in or_part.split("&")
            if term.strip()
        ]
        if and_terms:
            or_groups.append(and_terms)
 
    return or_groups

def match_symbol_driven_rule(text: str, rule_groups: list) -> bool:
    """
    Returns True if:
    (AND group 1) OR (AND group 2) ...
    """
    if not rule_groups:
        return True  # No rule → no filtering
 
    text = str(text).upper()
 
    for and_group in rule_groups:
        if all(term in text for term in and_group):
            return True
 
    return False

def extract_allowed_harness_codes_from_excel(misc_excel_df):
    """
    Extract harness codes ONLY from Excel.
    Excel is the source of truth.
    """
    codes = set()
    col = "part description (internal repair)"
 
    for val in misc_excel_df[col].dropna().astype(str):
        text = val.upper()
 
        if "HARNESS" not in text:
            continue
 
        found = re.findall(r"\b[A-Z]+[-_/]?\d+\b", text)
        codes.update(found)
 
    return codes

def process_harness_repair(
    qec_lru_df_rep: pd.DataFrame,
    misc_excel_df: pd.DataFrame
    ) -> pd.DataFrame:
 
    df = qec_lru_df_rep.copy()
 
    # ---- Column detection ----
    csn_col = pick_col(df.columns, "csn", "description")
    repair_desc_col = pick_col(df.columns, "repair", "description")
    remarks_col = pick_col(df.columns, "remark")
 
    if not csn_col or not repair_desc_col or not remarks_col:
        print("[ERROR][REPAIR] Required columns missing")
        return df.iloc[0:0]
 
    # ---- Read Excel rules ----
    part_rule_text = misc_excel_df.loc[0, "part description (internal repair)"]
    repair_rule_text = misc_excel_df.loc[0, "repair description (internal repair)"]
 
    part_rule_groups = parse_symbol_driven_rule(part_rule_text)
    repair_rule_groups = parse_symbol_driven_rule(repair_rule_text)
 
    # ---- Step 1: Harness + dynamic code detection ----
    allowed_harness_codes = extract_allowed_harness_codes_from_excel(misc_excel_df)
 
    if allowed_harness_codes:
        pattern = "|".join(map(re.escape, allowed_harness_codes))
    
        mask_harness = (
            df[csn_col].astype(str).str.upper().str.contains("HARNESS", na=False) &
            df[csn_col].astype(str).str.upper().str.contains(pattern, na=False)
        )
    else:
        # Excel says only "HARNESS" → no code restriction
        mask_harness = df[csn_col].astype(str).str.upper().str.contains("HARNESS", na=False)
 
    df = df[mask_harness].copy()
 
    if df.empty:
        print("[WARN][REPAIR] No rows after harness filtering")
        return df
 
    # ---- Step 2: Symbol-driven repair description filter ----
    df = df[
        df[repair_desc_col].apply(
            lambda x: match_symbol_driven_rule(x, repair_rule_groups)
        )
    ].copy()
 
    if df.empty:
        print("[WARN][REPAIR] No rows after repair rule filtering")
        return df
 
    # ---- Step 3: Remarks ----
    df[remarks_col] = "HARNESS" 
 
    return df

def process_harness_material(
    qec_lru_df: pd.DataFrame,
    harness_repair_df: pd.DataFrame,
    misc_excel_df: pd.DataFrame
) -> pd.DataFrame:
 
    df = qec_lru_df.copy()
 
    material_desc_col = "description"
    order_col_mat = "order"
    order_col_rep = "order no"
    remarks_col = "remarks"
    csn_desc = "csn description"
 
    if harness_repair_df.empty:
        print("[WARN][MATERIAL] No repair rows → skipping material")
        return df.iloc[0:0]
 
    # ---- Step 1: Order matching ----
    valid_orders = (
        harness_repair_df[order_col_rep]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )
 
    df = df[df[order_col_mat].astype(str).isin(valid_orders)].copy()
 
    if df.empty:
        print("[WARN][MATERIAL] No material rows after order matching")
        return df
 
    # ---- Step 2: Read billing rule ----
    billing_rule_text = misc_excel_df.loc[0, "description(billing request)"]
    billing_rule_groups = parse_symbol_driven_rule(billing_rule_text)
 
    # ---- Step 3: Symbol-driven material filter ----
    df = df[
        df[material_desc_col].apply(
            lambda x: match_symbol_driven_rule(x, billing_rule_groups)
        )
    ].copy()
 
    if df.empty:
        print("[WARN][MATERIAL] No rows after billing rule filtering")
        return df
    
    # Build order → csn description mapping
    order_to_csn = (
        harness_repair_df
        .set_index(order_col_rep)[csn_desc]
        .astype(str)
        .to_dict()
    )
 
 
    # Apply mapped remarks
    df[remarks_col] = "SPAD " + df[order_col_mat].map(order_to_csn).fillna("")
 
    return df


def create_scrap_exclusion_sheet(result, output_excel):
    """
    Creates scrap exclusion sheets for material, vendor, and repair,
    and writes them into an Excel file (multiple sheets).
 
    Parameters:
        result (dict): currently unused
        output_excel (str): output Excel file path
    """
 
    try:
        # =========================
        # READ INPUT FILES
        # =========================
        matl_df = pd.read_excel('matl_apply_blade_inclusion.xlsx')
        vendor_df = pd.read_excel('vendor_blade_inclusion.xlsx')
        repair_df = pd.read_excel('repair_apply_blade_inclusion.xlsx')
 
        vendor_col = next((c for c in matl_df.columns if "vendor description" in c), None)
        cost_cat_col   = next((c for c in matl_df.columns if "cost" in c and "category" in c), None)
        cfe_cost_cat = {"SCC-MATL-CUSTOMER", "SCC-MATL-CUST-SCRAPREP","SCC-LLP-CUST-SCRAPREP"}
 
        # =========================
        # MATERIAL SHEET
        # =========================
        matl_filtered = matl_df[matl_df['scrap_exclusion'] == 'Yes'].copy()

        # =========================
        # CFE LOGIC OVERRIDE
        # =========================
        if 'bucket' in matl_filtered.columns:
            # cfe_mask = matl_filtered['bucket'].str.upper() == 'CFE'
            cfe_mask = matl_filtered[cost_cat_col].astype(str).str.upper().isin(cfe_cost_cat)

            # Keep only CFE rows where vendor_description contains 'avianca'
            avianca_mask = is_united_airlines(matl_filtered[vendor_col])#.str.contains('avianca', case=False, na=False)
            # Remove CFE rows that do NOT contain 'avianca'
            matl_filtered = matl_filtered[~cfe_mask | avianca_mask]
            # Override total price = handling fee for remaining valid CFE rows
            valid_cfe_mask = (matl_filtered['bucket'].str.upper() == 'CFE')
            matl_filtered.loc[valid_cfe_mask, 'total_price'] = matl_filtered.loc[valid_cfe_mask, 'handling_fee']
 
        matl_output = pd.DataFrame({
            'ATA Chapter Code': matl_filtered.get('ata chapter'),
            'Part Number': matl_filtered.get('material'),
            'Part Description': matl_filtered.get('description'),
            'QPE': matl_filtered.get('total_qpe'),
            '% Material Included under Exclusion Table': matl_filtered.get('scrap_cap'),
            'Qty Material Included under Exclusion Table': matl_filtered.get('allowed_quantity'),
            'Total Qty Replaced': matl_filtered.get('total_scrap_quantity'),
            'Material Type Qty New': matl_filtered.get('target quantity'),
            'Unit Price CLP$': matl_filtered.get('matl_clp_lookup'),
            'Handling Fee Price': matl_filtered.get('handling_fee'),
            'Total Price': matl_filtered.get('total_price'),
            'Included Price': matl_filtered.get('included_price'),
            'Excluded [O&A] Price': matl_filtered.get('excluded_price')
        })
 
        # Sum of Price
        matl_output['Sum of Price'] = (
            matl_output.groupby('Part Description')['Total Price']
            .transform('sum')
        )
        
        # Avoid division errors
        matl_output['Total Qty Replaced'] = matl_output['Total Qty Replaced'].replace(0, np.nan)
        
        matl_output['Included Price'] = np.where(
            matl_output['Total Qty Replaced'] > matl_output['Qty Material Included under Exclusion Table'],
            (matl_output['Qty Material Included under Exclusion Table'] / matl_output['Total Qty Replaced']) * matl_output['Sum of Price'],
            matl_output['Sum of Price']
        )
        
        matl_output['Included Price'] = matl_output['Included Price'].fillna(0)
        
        matl_output['Excluded [O&A] Price'] = (
            matl_output['Sum of Price'] - matl_output['Included Price']
        )
 
        matl_output = matl_output[
            [
                'ATA Chapter Code',
                'Part Number',
                'Part Description',
                'QPE',
                '% Material Included under Exclusion Table',
                'Qty Material Included under Exclusion Table',
                'Total Qty Replaced',
                'Material Type Qty New',
                'Unit Price CLP$',
                'Handling Fee Price',
                'Total Price',
                'Sum of Price',
                'Included Price',
                'Excluded [O&A] Price'
            ]
        ]
 
        # =========================
        # REPAIR SHEET (FINALIZED)
        # =========================
        repair_filtered = repair_df[repair_df['scrap_exclusion'] == 'Yes'].copy()
 
        repair_output = pd.DataFrame({
            'ATA Chapter Code': repair_filtered.get('ata chapter'),
            'Part Number': repair_filtered.get('inspected part no.'),
            'Part Description': repair_filtered.get('csn description'),
            'QPE': repair_filtered.get('total_qpe'),
            '% Repair Included under Exclusion Table': repair_filtered.get('scrap_cap'),
            'Qty Repair Included under Exclusion Table': repair_filtered.get('allowed_quantity'),
            'Total Qty Repaired': repair_filtered.get('total_scrap_quantity'),
            'Repair Price': repair_filtered.get('extended_price_final'),
            'Discount': repair_filtered.get('discount'),
            'Included Price': repair_filtered.get('included_price'),
            'Excluded [O&A] Price': repair_filtered.get('excluded_price'),
            'PO/Repair Tag': "",  # as requested
            'Repair Code / Service Level': repair_filtered.get('repair code group & code'),
            'Price From': repair_filtered.get('price from')
        })
 
        # Compute Total Price = Repair Price - Discount
        repair_output['Total Price'] = (
            repair_output['Repair Price'].fillna(0) -
            repair_output['Discount'].fillna(0)
        )
 
        # Sum of Price (grouped by Part Description)
        repair_output['Sum of Price'] = (
            repair_output.groupby('Part Description')['Total Price']
            .transform('sum')
        )
 
        # Reorder columns
        repair_output = repair_output[
            [
                'ATA Chapter Code',
                'Part Number',
                'Part Description',
                'QPE',
                '% Repair Included under Exclusion Table',
                'Qty Repair Included under Exclusion Table',
                'Total Qty Repaired',
                'Repair Price',
                'Discount',
                'Total Price',
                'Sum of Price',
                'Included Price',
                'Excluded [O&A] Price',
                'PO/Repair Tag',
                'Repair Code / Service Level',
                'Price From'
            ]
        ]
 
        # =========================
        # VENDOR SHEET (placeholder)
        # =========================
        vendor_filtered = vendor_df[vendor_df['scrap_exclusion'] == 'Yes'].copy()
        vendor_output = vendor_filtered.copy()
 
        # =========================
        # WRITE TO EXCEL
        # =========================
        matl_output.to_excel("scrap_table_matl.xlsx", index = False)
        repair_output.to_excel("scrap_table_repair.xlsx", index = False)
        # with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        #     matl_output.to_excel(writer, sheet_name='Material Scrap Exclusion', index=False)
        #     repair_output.to_excel(writer, sheet_name='Repair Scrap Exclusion', index=False)
        #     vendor_output.to_excel(writer, sheet_name='Vendor Scrap Exclusion', index=False)
 
        print(f"Scrap exclusion sheets created successfully: {output_excel}")

        scrap_mat_df = pd.read_excel("scrap_table_matl.xlsx")
        scrap_rep_df = pd.read_excel("scrap_table_repair.xlsx")
        
        scrap_mat_sum = scrap_mat_df["Excluded [O&A] Price"].sum()
        scrap_rep_sum = scrap_rep_df["Excluded [O&A] Price"].sum()

        # print(scrap_mat_sum)
        # print(scrap_rep_sum)
        
        if not scrap_mat_sum:
            scrap_mat_sum = 0
        
        if not scrap_rep_sum:
            scrap_rep_sum = 0

        return scrap_mat_sum, scrap_rep_sum
 
    except Exception as e:
        print(f"Error while creating scrap exclusion sheet: {str(e)}")



def create_contract_pricing_terms_file(result, ffp_price, ffp_year, output_path="contract_pricing_terms_sheet.xlsx"):
    ffp_year = str(ffp_year)
 
    # =========================
    # BASE DATA
    # =========================
    bucket = [
        "Fixed Price Labor",
        "O&A Labor",
        "New Parts (non-LLP)",
        "New Parts (LLP)",
        "Rotable Exchange",
        "Customer Furnished Equipment",
        "Component Repairs",
        "Accessory and QEC Repairs",
        "Other Sub-contracted Services",
        "Test Cell Usage Fees",
        "Used Serviceable LLP parts",
        "Used Serviceable non-LLP parts",
        "Transportation"
    ]
 
    material_type = [
        "Labor", "Labor", "New", "LLP New", "Rotable", "CFE",
        "CRD", "CRD", "PO", "Test", "Used", "LLP Used", "Transport"
    ]
 
    guideline = [
        "Per FPLS", "$ per Hour", "CLP", "CLP", "Incl in Rep price",
        "HF % of CLP or PO", "Engine CRD", "Accy CRD", "Vendor invoice",
        "", "X% of Rem. Cycles", "X% of CLP or PO", ""
    ]
 
    df = pd.DataFrame({
        "Bucket": bucket,
        "Material Type": material_type,
        "Guideline": guideline,
        "Amount": [None] * len(bucket),
        "Discount %": [None] * len(bucket),
        "Handling Fee %": [None] * len(bucket),
        "Handling Fee cap per part": [None] * len(bucket),
        "Handling Fee cap per set": [None] * len(bucket)
    })
 
    # =========================
    # 1. FIXED PRICE LABOR
    # =========================
    df.loc[df["Bucket"] == "Fixed Price Labor", "Amount"] = ffp_price
 
    # =========================
    # 2. O&A LABOR ($/Hour)
    # =========================
    try:
        labor_df = get_dataframe_by_file_and_sheet(
            all_dataframes=result["dataframes"],
            file_key="contract",
            sheet_name_substring="labor price"
        )
        labor_df.columns = labor_df.columns.astype(str).str.strip()
 
        # Normalize column names
        labor_df.columns = labor_df.columns.str.strip()
 
        # Find row containing "Labor" in component column
        component_col = [col for col in labor_df.columns if "component" in col.lower()][0]
        print(component_col)
 
        labor_row = labor_df[
            labor_df[component_col].astype(str).str.lower().str.contains("labor")
        ]
        print(labor_row)
 
        if not labor_row.empty and ffp_year in labor_df.columns:
            labor_value = labor_row.iloc[0][ffp_year]
            df.loc[df["Bucket"] == "O&A Labor", "Amount"] = labor_value
 
    except Exception as e:
        print(f"[WARN] Unable to fetch O&A Labor: {e}")
 
    # =========================
    # 3. TEST CELL USAGE FEES
    # =========================
    try:
        test_df = get_dataframe_by_file_and_sheet(
            all_dataframes=result["dataframes"],
            file_key="contract",
            sheet_name_substring="Test Cell Usage Price"
        )

        test_df.columns = test_df.columns.astype(str).str.strip()
 
        test_df.columns = test_df.columns.str.strip()
 
        # Find row containing "Test"
        first_col = [col for col in test_df.columns if "component" in col.lower()][0]
        print(first_col)
 
        test_row = test_df[
            test_df[first_col].astype(str).str.lower().str.contains("test")
        ]
        print(test_row)
 
        if not test_row.empty and ffp_year in test_df.columns:
            test_value = test_row.iloc[0][ffp_year]
            df.loc[df["Bucket"] == "Test Cell Usage Fees", "Amount"] = test_value
 
    except Exception as e:
        print(f"[WARN] Unable to fetch Test Cell Usage Fees: {e}")
 
    # =========================
    # WRITE TO EXCEL
    # =========================
    df.to_excel(output_path, index=False)
 
    print(f"[INFO] ✅ Contract Pricing Terms file created: {output_path}")
 
    return output_path



def create_qec_lru(result: dict, output_excel: str):
    # ---------- helpers ----------
    def _pick_col(cols, *need):
        for c in cols:
            s = str(c).lower()
            if all(tok.lower() in s for tok in need):
                return c
        return None
 
    def _norm_txt(s: pd.Series) -> pd.Series:
        return (s.astype(str)
                 .str.upper()
                 .str.replace(r"[^\w\s]", " ", regex=True)
                 .str.replace(r"\s+", " ", regex=True)
                 .str.strip())
 
    def _anyorder_pattern(keyword: str) -> re.Pattern | None:
        tokens = re.findall(r"\w+", str(keyword).upper())
        if not tokens:
            return None
        la = "".join(f"(?=.*\\b{re.escape(t)}\\b)" for t in tokens)
        return re.compile(la + ".*")
    
    def align_columns(df_new: pd.DataFrame, df_base: pd.DataFrame) -> pd.DataFrame:
        df_new = df_new.copy()
        for col in df_base.columns:
            if col not in df_new.columns:
                df_new[col] = None
        return df_new[df_base.columns]
 
    # ---------- build base df ----------
    updated_mat = annotate_material_with_schedule_and_eipc_from_store(
        result=result,
        material_file_key="final_summary_material",
        material_sheet_substring="",
        contract_file_key="contract",
        schedule2_substring="Schedule_2 parts",
        schedule3_substring="Schedule_3 parts",
        eipc_file_key="eipc",
        eipc_sheet_substring="",
        pipeline_type = "material"
    )
 
    # ---------- build base df ----------
    updated_rep = annotate_material_with_schedule_and_eipc_from_store(
        result=result,
        material_file_key="final_summary_repair",
        material_sheet_substring="",
        contract_file_key="contract",
        schedule2_substring="Schedule_2 parts",
        schedule3_substring="Schedule_3 parts",
        eipc_file_key="eipc",
        eipc_sheet_substring="",
        pipeline_type = "repair"
    )

    # ---------- build base df ----------
    updated_vend = annotate_material_with_schedule_and_eipc_from_store(
        result=result,
        material_file_key="final_summary_vendor",
        material_sheet_substring="",
        contract_file_key="contract",
        schedule2_substring="Schedule_2 parts",
        schedule3_substring="Schedule_3 parts",
        eipc_file_key="eipc",
        eipc_sheet_substring="",
        pipeline_type = "vendor"
    )

    final_matl_df = create_module_columns(updated_mat).copy()
    final_rep_df = create_module_columns(updated_rep).copy()
    final_vend_df = create_module_columns(updated_vend).copy()
 
    final_matl_df.to_excel("revised_final_summary_material.xlsx")
    final_matl_df = pd.read_excel("revised_final_summary_material.xlsx")
    final_matl_df.drop_duplicates(inplace=True)    
 
 
    final_rep_df.to_excel("revised_final_summary_repair.xlsx")
    final_rep_df = pd.read_excel("revised_final_summary_repair.xlsx")
    final_rep_df.drop_duplicates(inplace=True)    
 
    final_vend_df.to_excel("revised_final_summary_vendor.xlsx")
    final_vend_df = pd.read_excel("revised_final_summary_vendor.xlsx")
    final_vend_df.drop_duplicates(inplace=True)     
 
    qec_lru_df = final_matl_df.copy()
    qec_lru_df_rep = final_rep_df.copy()
    qec_lru_df_vend = final_vend_df.copy()
    
    # # Step 1: Repair processing
    # misc_excel_df = get_dataframe_by_file_and_sheet(result["dataframes"], "contract", "misc excl")
    # misc_excel_df.columns = [col.strip().lower() for col in misc_excel_df.columns]
    # print("misc excel col",misc_excel_df.columns)
    
    # print("misc exclusion",misc_excel_df)
    
    
    
    
    # # harness_repair_df = process_harness_repair(qec_lru_df_rep,misc_excel_df)
    
    # # Step 2: Material processing (dependent on repair)
    # # harness_material_df = process_harness_material(
    # #     qec_lru_df,
    # #     harness_repair_df,
    # #     misc_excel_df
    # # )
    
    # # Save Harness Repair Data
    # # harness_repair_df.to_excel(
    # #     "harness_repair_data.xlsx",
    # #     index=False
    # # )
    
    # # Save Harness Material Data
    # harness_material_df.to_excel(
    #     "harness_material_data.xlsx",
    #     index=False
    # )
    
    
    qec_lru_df.loc[qec_lru_df['pn_clp_lookup'].isna(), 'matl_clp_lookup'] = qec_lru_df['matl_clp_lookup']
    # Status from "cost category"
    if "cost category" in qec_lru_df.columns:
        qec_lru_df["Status"] = (
            qec_lru_df["cost category"].astype(str).apply(
                lambda x: "NEW" if any(tag in x for tag in ["SCC-CONSUMABLE-AERO", "SCC-CONSUMABLE-NON-AERO"])
                else (re.search(r"(?i)\b(NEW|USED)\b", x).group(1).upper() if re.search(r"(?i)\b(NEW|USED)\b", x) else None)
            )
        )
    else:
        qec_lru_df["Status"] = None
 
    # Status from "cost category"
    if "cost category" in qec_lru_df_rep.columns:
        qec_lru_df_rep["Status"] = (
            qec_lru_df_rep["cost category"].astype(str).apply(
                lambda x: "NEW" if any(tag in x for tag in ["SCC-CONSUMABLE-AERO", "SCC-CONSUMABLE-NON-AERO"])
                else (re.search(r"(?i)\b(NEW|USED)\b", x).group(1).upper() if re.search(r"(?i)\b(NEW|USED)\b", x) else None)
            )
        )
    else:
        qec_lru_df_rep["Status"] = None
 
    
    qec_lru_df_vend.loc[qec_lru_df_vend['pn_clp_lookup'].isna(), 'matl_clp_lookup'] = qec_lru_df_vend['matl_clp_lookup']
    # Status from "cost category"
    if "cost category" in qec_lru_df_vend.columns:
        qec_lru_df_vend["Status"] = (
            qec_lru_df_vend["cost category"].astype(str).apply(
                lambda x: "NEW" if any(tag in x for tag in ["SCC-CONSUMABLE-AERO", "SCC-CONSUMABLE-NON-AERO"])
                else (re.search(r"(?i)\b(NEW|USED)\b", x).group(1).upper() if re.search(r"(?i)\b(NEW|USED)\b", x) else None)
            )
        )
    else:
        qec_lru_df_vend["Status"] = None
 
    # Keep only LRU/QEC inclusions
    if "remarks" in qec_lru_df.columns:
        qec_lru_df = qec_lru_df[qec_lru_df["remarks"].astype(str).isin(["ACC/LRU Normal Exclusion", "Exclusion as QEC"])].copy()
        qec_lru_df.to_excel("qec_lru_df.xlsx", index = False)
 
    # Keep only LRU/QEC inclusions
    if "remarks" in qec_lru_df_rep.columns:
        qec_lru_df_rep = qec_lru_df_rep[qec_lru_df_rep["remarks"].isin(["ACC/LRU Normal Exclusion", "Exclusion as QEC"])].copy()
 
    # Keep only LRU/QEC inclusions
    if "remarks" in qec_lru_df_vend.columns:
        qec_lru_df_vend = qec_lru_df_vend[qec_lru_df_vend["remarks"].astype(str).isin(["ACC/LRU Normal Exclusion", "Exclusion as QEC"])].copy()
        qec_lru_df_vend.to_excel("qec_lru_df_vend.xlsx", index = False)
 
    # # Require non-blank LRU schedule name if present
    # lru_sched_col = _pick_col(qec_lru_df.columns, "lru", "schedule", "name")
    # if lru_sched_col:
    #     qec_lru_df = qec_lru_df[
    #         qec_lru_df[lru_sched_col].astype(str).str.strip().ne("") & qec_lru_df[lru_sched_col].notna()
    #     ].copy()
 
    # # Require non-blank LRU schedule name if present
    # lru_sched_col2 = _pick_col(qec_lru_df_rep.columns, "lru", "schedule", "name")
    # if lru_sched_col2:
    #     qec_lru_df_rep = qec_lru_df_rep[
    #         qec_lru_df_rep[lru_sched_col2].astype(str).str.strip().ne("") & qec_lru_df_rep[lru_sched_col2].notna()
    #     ].copy()
 
    # # ---------- load schedules ----------
    # s2_df = get_dataframe_by_file_and_sheet(
    #     all_dataframes=result["dataframes"], file_key="contract", sheet_name_substring="schedule_2 parts"
    # ).copy()
    # s3_df = get_dataframe_by_file_and_sheet(
    #     all_dataframes=result["dataframes"], file_key="contract", sheet_name_substring="schedule_3 parts"
    # ).copy()
 
    # # Detect columns
    # s2_mpn_col  = _pick_col(s2_df.columns, "example", "mpn")
    # s3_mpn_col  = _pick_col(s3_df.columns, "example", "mpn")
    # s2_desc_col = (_pick_col(s2_df.columns, "part", "description"))
 
    # s3_desc_col = (_pick_col(s3_df.columns, "part", "description"))
 
    # # Guard: if a schedule column is missing, create benign placeholders
    # if not s2_mpn_col: s2_df[s2_mpn_col := "_missing_mpn_"] = []
    # if not s3_mpn_col: s3_df[s3_mpn_col := "_missing_mpn_"] = []
    # if not s2_desc_col: s2_df[s2_desc_col := "_missing_desc_"] = ""
    # if not s3_desc_col: s3_df[s3_desc_col := "_missing_desc_"] = ""
 
    # Normalize keys/descriptions
    qec_lru_df["_pn_norm"] = _norm_txt(qec_lru_df.get("material", pd.Series("", index=qec_lru_df.index)))
    qec_lru_df_vend["_pn_norm"] = _norm_txt(qec_lru_df_vend.get("material", pd.Series("", index=qec_lru_df_vend.index)))
    qec_lru_df_rep["_pn_norm"] = _norm_txt(qec_lru_df_rep.get("inspected part no", pd.Series("", index=qec_lru_df_rep.index)))
 
    # s2_df["_mpn_norm"]     = _norm_txt(s2_df[s2_mpn_col])
    # s3_df["_mpn_norm"]     = _norm_txt(s3_df[s3_mpn_col])
    # s2_df["_desc_norm"]    = _norm_txt(s2_df[s2_desc_col])
    # s3_df["_desc_norm"]    = _norm_txt(s3_df[s3_desc_col])
 
    # ---------- Step 1: PN "join" (via normalized membership) ----------
    qec_lru_df["Part_of_Schedule"] = ""
    qec_lru_df_vend["Part_of_Schedule"] = ""
    # pn_hit_s2 = qec_lru_df["_pn_norm"].isin(set(s2_df["_mpn_norm"].dropna()))
    # pn_hit_s3 = qec_lru_df["_pn_norm"].isin(set(s3_df["_mpn_norm"].dropna()))
 
    qec_lru_df_rep["Part_of_Schedule"] = ""
    # pn_hit_s2_ = qec_lru_df_rep["_pn_norm"].isin(set(s2_df["_mpn_norm"].dropna()))
    # pn_hit_s3_ = qec_lru_df_rep["_pn_norm"].isin(set(s3_df["_mpn_norm"].dropna()))
 
    # Prefer Schedule_2 when both match
    # qec_lru_df.loc[pn_hit_s2, "Part_of_Schedule"] = "Schedule_2"
    # qec_lru_df.loc[~pn_hit_s2 & pn_hit_s3, "Part_of_Schedule"] = "Schedule_3"
 
    # Prefer Schedule_2 when both match
    # qec_lru_df_rep.loc[pn_hit_s2_, "Part_of_Schedule"] = "Schedule_2"
    # qec_lru_df_rep.loc[~pn_hit_s2_ & pn_hit_s3_, "Part_of_Schedule"] = "Schedule_3"
 
    # ---------- Step 3 & 4: Keyword path for PN-unmatched ----------
    # kw_col = _pick_col(qec_lru_df.columns, "genpact", "part", "keyword")  # "Genpact_Part_Keywords"
 
    # # ---------- Step 3 & 4: Keyword path for PN-unmatched ----------
    # kw_col_ = _pick_col(qec_lru_df_rep.columns, "genpact", "part", "keyword")  # "Genpact_Part_Keywords"
 
    # if kw_col:
    #     # Drop rows whose keyword has exactly 1 word (your "lastly" rule)
    #     word_count = (qec_lru_df[kw_col].fillna("").astype(str).str.findall(r"\w+").str.len())
    #     qec_lru_df = qec_lru_df[word_count.ne(1)].copy()
 
    #     need_kw = qec_lru_df["Part_of_Schedule"].eq("")
    #     sub = qec_lru_df.loc[need_kw, [kw_col]].copy()
    #     sub["_kw_norm"] = _norm_txt(sub[kw_col])
 
    #     unique_kws = [k for k in sub["_kw_norm"].dropna().unique().tolist() if k]
    #     kw_to_s2, kw_to_s3 = {}, {}
 
    #     for kw in unique_kws:
    #         pat = _anyorder_pattern(kw)
    #         if pat is None:
    #             kw_to_s2[kw] = False
    #             kw_to_s3[kw] = False
    #             continue
    #         kw_to_s2[kw] = s2_df["_desc_norm"].str.contains(pat, regex=True, na=False).any()
    #         kw_to_s3[kw] = s3_df["_desc_norm"].str.contains(pat, regex=True, na=False).any()
 
    #     # Assign schedule by keyword (prefer Schedule_2 if both)
    #     for idx, row in sub.iterrows():
    #         kw = row.get("_kw_norm", "")
    #         if not kw:
    #             continue
    #         s2hit = kw_to_s2.get(kw, False)
    #         s3hit = kw_to_s3.get(kw, False)
    #         if s2hit:
    #             qec_lru_df.at[idx, "Part_of_Schedule"] = "Schedule_2"
    #         elif s3hit:
    #             qec_lru_df.at[idx, "Part_of_Schedule"] = "Schedule_3"
    #         # else: remains blank (no schedule by keyword)
 
   
    # if kw_col_:
    #     # Drop rows whose keyword has exactly 1 word (your "lastly" rule)
    #     word_count = (qec_lru_df_rep[kw_col].fillna("").astype(str).str.findall(r"\w+").str.len())
    #     qec_lru_df_rep = qec_lru_df_rep[word_count.ne(1)].copy()
 
    #     need_kw = qec_lru_df_rep["Part_of_Schedule"].eq("")
    #     sub = qec_lru_df_rep.loc[need_kw, [kw_col]].copy()
    #     sub["_kw_norm"] = _norm_txt(sub[kw_col])
 
    #     unique_kws = [k for k in sub["_kw_norm"].dropna().unique().tolist() if k]
    #     kw_to_s2_, kw_to_s3_ = {}, {}
 
    #     for kw in unique_kws:
    #         pat = _anyorder_pattern(kw)
    #         if pat is None:
    #             kw_to_s2_[kw] = False
    #             kw_to_s3_[kw] = False
    #             continue
    #         kw_to_s2_[kw] = s2_df["_desc_norm"].str.contains(pat, regex=True, na=False).any()
    #         kw_to_s3_[kw] = s3_df["_desc_norm"].str.contains(pat, regex=True, na=False).any()
 
    #     # Assign schedule by keyword (prefer Schedule_2 if both)
    #     for idx, row in sub.iterrows():
    #         kw = row.get("_kw_norm", "")
    #         if not kw:
    #             continue
    #         s2hit = kw_to_s2_.get(kw, False)
    #         s3hit = kw_to_s3_.get(kw, False)
    #         if s2hit:
    #             qec_lru_df_rep.at[idx, "Part_of_Schedule"] = "Schedule_2"
    #         elif s3hit:
    #             qec_lru_df_rep.at[idx, "Part_of_Schedule"] = "Schedule_3"
    #         # else: remains blank (no schedule by keyword)
 
    # # Keep only rows matched to some schedule
    # final_qec_lru_df = qec_lru_df[qec_lru_df["Part_of_Schedule"].isin(["Schedule_2", "Schedule_3"])].copy()
 
    # # Keep only rows matched to some schedule
    # final_qec_lru_df_ = qec_lru_df_rep[qec_lru_df_rep["Part_of_Schedule"].isin(["Schedule_2", "Schedule_3"])].copy()
    
    # # --- Material ---
    # if not final_qec_lru_df.empty:
    #     print("[DEBUG] Scheduled material rows exist. Aligning harness material.")
    #     harness_material_aligned = align_columns(
    #         harness_material_df,
    #         final_qec_lru_df
    #     )
    #     final_qec_lru_df = pd.concat(
    #         [final_qec_lru_df, harness_material_aligned],
    #         ignore_index=True
    #     )
    
    #     # Apply schedule flag ONLY when dataframe is not empty
    #     final_qec_lru_df["Part_of_Schedule"] = (
    #         final_qec_lru_df.get(
    #             "Part_of_Schedule",
    #             pd.Series(index=final_qec_lru_df.index)
    #         ).fillna("HARNESS_MANUAL")
    #     )
    
    # else:
    #     print("[DEBUG] No scheduled material rows. Using harness material directly.")
    #     final_qec_lru_df = harness_material_df.copy()
    
    # # --- Repair ---
    # if not final_qec_lru_df_.empty:
    #     print("[DEBUG] Scheduled repair rows exist. Aligning harness repair.")
    #     harness_repair_aligned = align_columns(
    #         harness_repair_df,
    #         final_qec_lru_df_
    #     )
    #     final_qec_lru_df_ = pd.concat(
    #         [final_qec_lru_df_, harness_repair_aligned],
    #         ignore_index=True
    #     )
    
    #     # Apply schedule flag ONLY when dataframe is not empty
    #     final_qec_lru_df_["Part_of_Schedule"] = (
    #         final_qec_lru_df_.get(
    #             "Part_of_Schedule",
    #             pd.Series(index=final_qec_lru_df_.index)
    #         ).fillna("HARNESS_MANUAL")
    #     )
    
    # else:
    #     print("[DEBUG] No scheduled repair rows. Using harness repair directly.")
        # final_qec_lru_df_ = harness_repair_df.copy()

    final_qec_lru_df = qec_lru_df.copy()
    final_qec_lru_df_vend = qec_lru_df_vend.copy()
    final_qec_lru_df_vend = apply_handling_fees(result, final_qec_lru_df_vend)
    final_qec_lru_df_vend.to_excel("final_qec_lru_df_vend_TEST.xlsx", index = False)
    final_qec_lru_df_ = qec_lru_df_rep.copy()
    
    # final_qec_lru_df["Part_of_Schedule"] = final_qec_lru_df["Part_of_Schedule"].fillna("HARNESS_MANUAL")
    # final_qec_lru_df_["Part_of_Schedule"] = final_qec_lru_df_["Part_of_Schedule"].fillna("HARNESS_MANUAL")
    
 
    # # Discount fixed to 0
    # if "Discount" not in final_qec_lru_df.columns:
    #     final_qec_lru_df["Discount"] = 0
    # else:
    #     final_qec_lru_df["Discount"] = 0
 
    # # Discount fixed to 0
    # if "Discount" not in final_qec_lru_df_.columns:
    #     final_qec_lru_df_["Discount"] = 0
    # else:
    #     final_qec_lru_df_["Discount"] = 0
        
    # print("QEC_Mat",final_qec_lru_df.columns)
    # print("QEC_Rep",final_qec_lru_df_.columns)
    
 
    # Keep only priced rows if Total Price exists
    # final_qec_lru_df = final_qec_lru_df[final_qec_lru_df["total_price"] > 0].copy()
    # final_qec_lru_df_ = final_qec_lru_df_[final_qec_lru_df_["total_price"] > 0].copy()
    final_qec_lru_df_['discount_amount'] = final_qec_lru_df_['total_price_final'] * final_qec_lru_df_['discount']
    
    final_qec_lru_df_["order no"] = (
        final_qec_lru_df_["order no"].fillna("").astype(str).str.strip()
        + " - "
        + final_qec_lru_df_["repair description"].fillna("").astype(str).str.strip()
    )
    
    # New columns
    final_qec_lru_df['Material Type'] = final_qec_lru_df['cost category'].str[4:]
    final_qec_lru_df_vend['Material Type'] = final_qec_lru_df_vend['cost category'].str[4:]
    
    final_qec_lru_df['remarks'] = final_qec_lru_df.apply(
        lambda row: append_module_to_remarks(
            row['formatted_ata_long'],
            row['module'],
            row['remarks']
        ),
        axis=1
    )
    
    
    
    
    
    final_qec_lru_df.rename(columns={
        'item (sd)': 'Item (SD)',
        'formatted_ata_long': 'ATA Chapter Code',
        'material': 'Part Number',
        'description': 'Part Description',
        'target quantity': 'Qty',
        'pn_clp_lookup':'CLP',
        'matl_clp_lookup': 'Unit Price',
        'extended_price': 'Extended Price',
        'discount': 'Discount %',
        'discount_amount':'Discount $',
        'handling fee':'Handling Fee %',
        'handling_fee':'Handling Fee Price',
        'total_price': 'Excluded [O&A] Total Price',
        'irc description': 'Replacement Remarks',
        'remarks':'Comments',
        'order':'Service Order Equipment / MRB',
    }, inplace=True)
 

    final_qec_lru_df_vend['remarks'] = final_qec_lru_df_vend.apply(
        lambda row: append_module_to_remarks(
            row['formatted_ata_long'],
            row['module'],
            row['remarks']
        ),
        axis=1
    )

    final_qec_lru_df_vend.to_excel("final_qec_lru_df_vend_TEST_4.xlsx", index = False)
    
    final_qec_lru_df_vend.rename(columns={
        'item (sd)': 'Item (SD)',
        'formatted_ata_long': 'ATA Chapter Code',
        'material': 'Part Number',
        'description': 'Part Description',
        'target quantity': 'Qty',
        'pn_clp_lookup':'CLP',
        'matl_clp_lookup': 'Unit Price',
        'extended_price': 'Extended Price',
        'discount': 'Discount %',
        'discount_amount':'Discount $',
        'handling fee':'Handling Fee %',
        'handling_fee':'Handling Fee Price',
        'total_price': 'Excluded [O&A] Total Price',
        'irc description': 'Replacement Remarks',
        'remarks':'Comments',
        'order':'Service Order Equipment / MRB',
        'purchasing document' : 'Purchasing Document'
    }, inplace=True)


    final_qec_lru_df_vend.to_excel("final_qec_lru_df_vend_TEST_5.xlsx", index = False)
 
    # New columns
    # final_qec_lru_df_['Material Type'] = final_qec_lru_df['cost category'].str[4:]
    # final_qec_lru_df_.rename(columns={'Material Type':'Repair Source'}, inplace=True)
    
    final_qec_lru_df_['remarks'] = final_qec_lru_df_.apply(
        lambda row: append_module_to_remarks(
            row['formatted_ata_long'],
            row['module'],
            row['remarks']
        ),
        axis=1
    )
 
    final_qec_lru_df_.rename(columns={
        'br item': 'Item (SD)',
        'formatted_ata_long': 'ATA Chapter Code',
        'repair type': 'Repair Source',
        'inspected part no.' : 'Part Number',
        'csn description': 'Part Description',
        'inspected quantity' : 'Qty',
        'repair hours' : 'Hour / CRD',
        'total_price_final' : 'Unit Price',
        'extended_price_final': 'Extended Price',
        'discount': 'Discount %',
        'discount_amount':'Discount $',
        'total_price' : 'Excluded [O&A] Total Price',
        'remarks':'Comments',
        'order no':'PO/Repair Tag',
        'repair code group & code' : 'Repair Code / Service Level'
        }, inplace=True)
 
    # Column order (only keep those present)
    keep_cols = ["Item (SD)", "ATA Chapter Code", "Material Type", "Part Number", "Part Description", "Qty", "CLP", "Unit Price", "Extended Price","Discount %",
                 "Discount $", "Handling Fee %", "Handling Fee Price", "Excluded [O&A] Total Price", "Replacement Remarks",
                 'Comments', "Service Order Equipment / MRB", "Purchasing Document"
    ]
   
    keep_cols1 = [c for c in keep_cols if c in final_qec_lru_df.columns]
    final_qec_lru_df = final_qec_lru_df[keep_cols1]
    
    keep_cols2 = [c for c in keep_cols if c in final_qec_lru_df_vend.columns]
    final_qec_lru_df_vend = final_qec_lru_df_vend[keep_cols2]

    final_qec_lru_df_vend.to_excel("final_qec_lru_df_vend_TEST_3.xlsx", index = False)
 
    # Final cleaning and export
    keep_cols_ = [
        "Item (SD)", "ATA Chapter Code", "Repair Source", "Part Number", "Part Description", "Qty", "Hour / CRD", "Unit Price", "Extended Price",
        "Discount %", "Discount $", "Excluded [O&A] Total Price","Comments", "PO/Repair Tag",
        "Repair Code / Service Level"
    ]
 
    if final_qec_lru_df_.shape[0] == 0:
        final_qec_lru_df_ = pd.DataFrame(columns=keep_cols_)
    else:
        keep_cols_ = [c for c in keep_cols_ if c in final_qec_lru_df_.columns]
        final_qec_lru_df_ = final_qec_lru_df_[keep_cols_]
   
    # final_qec_lru_df.to_excel("qec_lru_data.xlsx", index=False)
    qec_lru_sheet_name = "QEC-LRU-ACC"
 
    final_qec_lru_df = final_qec_lru_df[["Item (SD)", "ATA Chapter Code", "Material Type", "Part Number", "Part Description", "Qty", "CLP", "Unit Price", "Extended Price","Discount %",
                 "Discount $", "Handling Fee %", "Handling Fee Price", "Excluded [O&A] Total Price", "Replacement Remarks",
                 'Comments', "Service Order Equipment / MRB" ]]

    # final_qec_lru_df_vend = final_qec_lru_df_vend[["Item (SD)", "ATA Chapter Code", "Material Type", "Part Number", "Part Description", "Qty", "CLP", "Unit Price", "Extended Price","Discount %",
    #             "Discount $", "Handling Fee %", "Handling Fee Price", "Excluded [O&A] Total Price", "Replacement Remarks",
    #             'Comments', "Service Order Equipment / MRB" ]]

    final_qec_lru_df_vend.to_excel("final_qec_lru_df_vend_TEST_2.xlsx", index = False)

    final_qec_lru_df_vend = final_qec_lru_df_vend[["Item (SD)", "ATA Chapter Code", "Material Type", "Part Number", "Part Description", "Qty", "CLP", "Unit Price", "Extended Price",
                "Discount $", "Handling Fee %", "Handling Fee Price", "Excluded [O&A] Total Price", "Replacement Remarks",
                'Comments', "Service Order Equipment / MRB", "Purchasing Document" ]]
 
   
    mask_zero = pd.to_numeric(final_qec_lru_df['CLP'], errors='coerce').fillna(0) == 0
    mask_zero_vend = pd.to_numeric(final_qec_lru_df_vend['CLP'], errors='coerce').fillna(0) == 0
    mask_blank = final_qec_lru_df['CLP'].astype(str).str.strip().eq('') | final_qec_lru_df['CLP'].isna()
    mask_blank_vend = final_qec_lru_df_vend['CLP'].astype(str).str.strip().eq('') | final_qec_lru_df_vend['CLP'].isna()
    mask = mask_zero | mask_blank
    mask_vend = mask_zero_vend | mask_blank_vend
 
    # Set discount to blank (empty string) wherever the condition holds
    final_qec_lru_df.loc[mask, 'Discount %'] = ""
    # final_qec_lru_df_vend.loc[mask, 'Discount %'] = ""
    
    price_cols = ["Excluded [O&A] Total Price", "Extended Price"]
 
    # Normalize blanks → 0
    final_qec_lru_df[price_cols] = (
        final_qec_lru_df[price_cols]
        .apply(
            lambda col: (
                col.fillna(0)
                .astype(str)
                .str.strip()
                .replace("", "0")
                .astype(float)
            )
        )
    )
    
    final_qec_lru_df = final_qec_lru_df[
        ~(
            (final_qec_lru_df["Excluded [O&A] Total Price"] == 0) &
            (final_qec_lru_df["Extended Price"] == 0)
        )
    ]
    final_qec_lru_df = final_qec_lru_df.drop_duplicates()
    final_qec_lru_df.to_excel("qec_mat.xlsx", index=False)

    final_qec_lru_df_vend[price_cols] = (
        final_qec_lru_df_vend[price_cols]
        .apply(
            lambda col: (
                col.fillna(0)
                .astype(str)
                .str.strip()
                .replace("", "0")
                .astype(float)
            )
        )
    )
    
    final_qec_lru_df_vend = final_qec_lru_df_vend[
        ~(
            (final_qec_lru_df_vend["Excluded [O&A] Total Price"] == 0) &
            (final_qec_lru_df_vend["Extended Price"] == 0)
        )
    ]
    final_qec_lru_df_vend = final_qec_lru_df_vend.drop_duplicates()
    final_qec_lru_df_vend.to_excel("qec_mat_vend.xlsx", index=False)


    # ---------------------------------------------------------
    # Append SUBCON vendor rows into QEC Repair
    # ---------------------------------------------------------
    
    # Step 1: Filter required vendor rows
    subcon_vendor_df = final_qec_lru_df_vend[
        final_qec_lru_df_vend["Material Type"].astype(str).str.lower().str.contains("subcon", na=False) &
        (final_qec_lru_df_vend["Comments"] == "Exclusion as QEC")
    ].copy()
    
    if not subcon_vendor_df.empty:
    
        # -------------------------------------------------
        # Step 2: Apply Handling Fees
        # -------------------------------------------------
        # subcon_vendor_df = apply_handling_fees(result, subcon_vendor_df)
    
        # EXPECTATION:
        # This function should return columns like:
        # - handling fee %
        # - handling fee price
    
        # If naming mismatch, normalize:
        subcon_vendor_df.rename(columns={
            'handling fee': 'Handling Fee %',
            'handling_fee': 'Handling Fee Price'
        }, inplace=True)
    
        # Safety (in case function naming differs)
        if "Handling Fee %" not in subcon_vendor_df.columns:
            subcon_vendor_df["Handling Fee %"] = ""
        if "Handling Fee Price" not in subcon_vendor_df.columns:
            subcon_vendor_df["Handling Fee Price"] = 0
    
        # -------------------------------------------------
        # Step 3: Transform to repair format
        # -------------------------------------------------
        subcon_vendor_df_renamed = subcon_vendor_df.rename(columns={
            "Material Type": "Repair Source",
            "Part Number": "Part Number",
            "Part Description": "Part Description",
            "Qty": "Qty",
            "Unit Price": "Unit Price",
            "Extended Price": "Extended Price",
            "Discount %": "Discount %",
            "Discount $": "Discount $",
            "Excluded [O&A] Total Price": "Excluded [O&A] Total Price",
            "Comments": "Comments",
            "Purchasing Document": "PO/Repair Tag"
        })
    
        # Add missing repair-specific columns
        # subcon_vendor_df_renamed["Item (SD)"] = ""
        subcon_vendor_df_renamed["ATA Chapter Code"] = subcon_vendor_df_renamed.get("ATA Chapter Code", "")
        subcon_vendor_df_renamed["Hour / CRD"] = ""
        subcon_vendor_df_renamed["Repair Code / Service Level"] = "REPAIR"
        # subcon_vendor_df_renamed["PO/Repair Tag"] = ""
    
        # -------------------------------------------------
        # Step 4: Ensure schema alignment
        # -------------------------------------------------
        repair_cols = [
            "Item (SD)", "ATA Chapter Code", "Repair Source", "Part Number",
            "Part Description", "Qty", "Hour / CRD", "Unit Price", "Extended Price",
            "Discount %", "Discount $","Handling Fee %", "Handling Fee Price", "Excluded [O&A] Total Price",
            "Comments", "PO/Repair Tag", "Repair Code / Service Level"
        ]
    
        for col in repair_cols:
            if col not in subcon_vendor_df_renamed.columns:
                subcon_vendor_df_renamed[col] = ""
    
        subcon_vendor_df_renamed = subcon_vendor_df_renamed[repair_cols]
    
        # -------------------------------------------------
        # Step 5: Append to repair dataframe
        # -------------------------------------------------
        final_qec_lru_df_ = pd.concat(
            [final_qec_lru_df_, subcon_vendor_df_renamed],
            ignore_index=True
        )
 
    final_qec_lru_df_ = final_qec_lru_df_[["Item (SD)", "ATA Chapter Code", "Repair Source", "Part Number", "Part Description", "Qty", "Hour / CRD", "Unit Price", "Extended Price",
        "Discount %", "Discount $", "Handling Fee %", "Handling Fee Price", "Excluded [O&A] Total Price", "Comments", "PO/Repair Tag", "Repair Code / Service Level"]]
    
    final_qec_lru_df_[price_cols] = (
        final_qec_lru_df_[price_cols]
        .apply(
            lambda col: (
                col.fillna(0)
                .astype(str)
                .str.strip()
                .replace("", "0")
                .astype(float)
            )
        )
    )
    
    final_qec_lru_df_ = final_qec_lru_df_[
        ~(
            (final_qec_lru_df_["Excluded [O&A] Total Price"] == 0) &
            (final_qec_lru_df_["Extended Price"] == 0)
        )
    ]
    
    final_qec_lru_df_ = final_qec_lru_df_[final_qec_lru_df_['Comments'] !='ACC/LRU Normal Exclusion']
    final_qec_lru_df_.drop_duplicates(inplace = True)
    final_qec_lru_df_.to_excel("qec_rep.xlsx", index=False)
 
    qec_mat_sum = final_qec_lru_df["Excluded [O&A] Total Price"].sum()
    qec_vend_sum = final_qec_lru_df_vend["Excluded [O&A] Total Price"].sum()
    qec_rep_sum = final_qec_lru_df_["Excluded [O&A] Total Price"].sum()
    return final_qec_lru_df, final_qec_lru_df_, qec_lru_sheet_name, qec_mat_sum, qec_rep_sum

def create_misc_data():
    # Define the data based on your provided table
    data = [
        ["Test Cell Charge", None, ""],
        ["Fuel & Oil", None, ""],
        ["Stand Repair", None, ""],
        ["CFE parts import freight", None, ""],
        ["Residue shipping", None, ""],
        ["Transportation", None, ""],
        ["Local Taxes", None, ""]
    ]
    
    # Define the column headers
    columns = ["Description", "Excluded [O&A] Price", "Comments"]
    
    # Create and return the DataFrame
    misc_df = pd.DataFrame(data, columns=columns)
    
    misc_df.to_excel("overall_misc_data.xlsx", index=False)
    misc_sheet_name = "Misc"
    return misc_df, misc_sheet_name 

def write_misc_df_with_gaps(
    file_paths,
    output_file,
    sheet_name,
    start_rows=None,
    row_gaps=3,
):
    """
    Write 1 or 2 DataFrames from Excel files into a target sheet with row gaps.
 
    Args:
        file_paths (list): List of 1 or 2 Excel file paths.
        output_file (str): Path to save the final Excel file.
        sheet_name (str): Target sheet name.
        start_rows (list): List of starting rows (optional). If None, default to [6, None].
        row_gaps (int): Number of blank rows between DataFrames.
    Returns:
        metadata (list of dict): Metadata of written DataFrames.
    """
    assert len(file_paths) in [1, 2], "Function supports only 1 or 2 Excel files"
 
    dfs = [pd.read_excel(fp) for fp in file_paths]
    metadata = []
 
    # Default start_rows logic
    if start_rows is None:
        start_rows = [5] + [None] * (len(dfs) - 1)

    elif len(start_rows) < len(dfs):
        start_rows += [None] * (len(dfs) - len(start_rows))
        
    # Load or create workbook
    try:
        wb = load_workbook(output_file)
    except FileNotFoundError:
        wb = Workbook()
 
    # Access or create the sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
 
    # Clear existing content in the sheet
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
 
    # Initial row
    current_row = start_rows[0] if start_rows[0] else 1
 
    # Write each DataFrame
    for idx, df in enumerate(dfs):
        start_row = current_row
        start_col = 1  # Column B
        end_row = start_row + len(df)
        end_col = start_col + len(df.columns) - 1
 
        # Write DataFrame with header
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                ws.cell(row=r_idx, column=c_idx, value=value)
 
        metadata.append({
            "df_index": idx + 1,
            "start_row": start_row,
            "end_row": end_row,
            "start_col": start_col,
            "end_col": end_col,
            "num_rows": len(df),
            "num_cols": len(df.columns)
        })
 
        # Prepare for next block
        current_row = end_row + row_gaps + 1
 
    wb.save(output_file)
    return metadata
# def fill_misc_sheet(invoice_path, misc_sheet_name, metadata_misc):
#     """
#     Fills the Misc sheet using metadata.
#     Extracts the DF from metadata, writes headers at Row 5, and data starting at Row 6.
#     """
#     # Extract metadata and the actual DataFrame object
#     num_rows_1 = metadata_misc[0]["num_rows"] 
#     misc_df = metadata_misc[0]["df_index"] # Assuming this contains the DF object
    
#     wb = load_workbook(invoice_path)
#     if misc_sheet_name not in wb.sheetnames:
#         wb.create_sheet(misc_sheet_name)
#     ws = wb[misc_sheet_name]

#     # === 1. Column & Row Formatting ===
#     ws.column_dimensions['A'].width = 45
#     ws.column_dimensions['B'].width = 25
#     ws.column_dimensions['C'].width = 35
#     ws.column_dimensions['D'].width = 20
    
#     for r in range(1, 20):
#         ws.row_dimensions[r].height = 22

#     # === 2. Styles ===
#     dark_blue_fill = PatternFill(start_color="050543", end_color="050543", fill_type="solid")
#     white_bold_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
#     black_font = Font(name="Arial", size=11)
#     border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), 
#                          top=Side(style='thin'), bottom=Side(style='thin'))
    
#     center_align = Alignment(horizontal="center", vertical="center")
#     left_align = Alignment(horizontal="left", vertical="center", indent=1)
#     right_align = Alignment(horizontal="right", vertical="center")

#     # === 3. Header & Logo ===
#     try:
#         img = Image('ge_new_logo.png')
#         img.height, img.width = 40, 100
#         ws.add_image(img, 'A1')
#     except: pass

#     ws["B1"].value = "GE Aerospace"
#     ws["B1"].font = Font(name='Arial', bold=True, size=16, color="050543")

#     # === 4. Grand Total Section (Row 1-2) ===
#     ws["D1"].value = "Excluded [O&A]\nPrice"
#     ws["D1"].fill = dark_blue_fill
#     ws["D1"].font = white_bold_font
#     ws["D1"].alignment = Alignment(horizontal="center", wrap_text=True)

#     ws["C2"].value = "Grand Total"
#     ws["C2"].fill = dark_blue_fill
#     ws["C2"].font = white_bold_font
#     ws["C2"].alignment = center_align
    
#     # Calculate range dynamically
#     last_data_row = 6 + num_rows_1 - 1
#     ws["D2"].value = f"=SUM(B6:B{last_data_row})"
#     ws["D2"].fill = dark_blue_fill
#     ws["D2"].font = white_bold_font
#     ws["D2"].alignment = center_align 
#     ws["D2"].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

#     # === 5. Main Table Headers (Row 5) ===
#     headers = ["Description", "Excluded [O&A] Price", "Comments"]
#     for i, text in enumerate(headers, start=1):
#         cell = ws.cell(row=5, column=i, value=text)
#         cell.fill = dark_blue_fill
#         cell.font = white_bold_font
#         cell.alignment = center_align
#         cell.border = border_thin

#     # === 6. Writing Table Data (Starts at A6) ===
#     # We iterate using .values to skip the original DF header
#     for i, row_values in enumerate(misc_df.values, start=6):
        
#         # Column A: Description (Index 0)
#         desc_cell = ws.cell(row=i, column=1, value=row_values[0])
#         desc_cell.alignment = left_align
        
#         # Column B: Price with TBD Logic (Index 1)
#         price_val = row_values[1]
#         price_cell = ws.cell(row=i, column=2)
        
#         if price_val is None or str(price_val).strip() in ["", "nan", "NaN", "None"]:
#             price_cell.value = "TBD"
#             price_cell.alignment = center_align
#         else:
#             price_cell.value = price_val
#             price_cell.alignment = right_align
#             price_cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

#         # Column C: Comments (Index 2)
#         comment_val = row_values[2] if len(row_values) > 2 else ""
#         ws.cell(row=i, column=3, value=comment_val).alignment = left_align

#         # Apply borders/font to the row for columns A, B, and C
#         for col_idx in range(1, 4):
#             cell = ws.cell(row=i, column=col_idx)
#             cell.border = border_thin
#             cell.font = black_font

#     ws.sheet_view.showGridLines = False
#     wb.save(invoice_path)
#     print(f"[INFO] ✅ Misc sheet populated via Metadata starting at A6.")

def fill_misc_sheet(invoice_path, misc_sheet_name, misc_df):
    """
    Fills the Misc sheet using a direct DataFrame object.
    Starts writing data at A6 and handles empty prices as 'TBD'.
    """
    wb = load_workbook(invoice_path)
    if misc_sheet_name not in wb.sheetnames:
        wb.create_sheet(misc_sheet_name)
    ws = wb[misc_sheet_name]

    # === 1. Column & Row Formatting ===
    ws.column_dimensions['A'].width = 45  # Description
    ws.column_dimensions['B'].width = 25  # Price
    ws.column_dimensions['C'].width = 35  # Comments
    ws.column_dimensions['D'].width = 20  # Grand Total Area
    
    # for r in range(1, 20): # Ensure enough rows have standard height
    #     ws.row_dimensions[r].height = 22

    # === 2. Styles ===
    dark_blue_fill = PatternFill(start_color="050543", end_color="050543", fill_type="solid")
    white_bold_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    black_font = Font(name="Arial", size=11)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", indent=1)
    right_align = Alignment(horizontal="right", vertical="center")
    white_side = Side(border_style="thin", color="FFFFFF")
    white_border = Border(left=white_side, right=white_side, top=white_side, bottom=white_side)

    # === 3. Header & Logo ===
    try:
        img = Image('ge_new_logo.png')
        img.height, img.width = 40, 100
        ws.add_image(img, 'A1')
    except: pass

    # ws["B1"].value = "GE Aerospace"
    # ws["B1"].font = Font(name='Arial', bold=True, size=16, color="050543")

    # === 4. Grand Total Section (Row 1-2) ===
    ws["D1"].value = "Excluded [O&A]\nPrice"
    ws["D1"].fill = dark_blue_fill
    ws["D1"].font = white_bold_font
    ws["D1"].alignment = center_align
    ws["D1"].border = white_border
    ws["D1"].alignment = Alignment(horizontal="center", wrap_text=True)

    ws["C2"].value = "Grand Total"
    ws["C2"].fill = dark_blue_fill
    ws["C2"].font = white_bold_font
    ws["C2"].border = white_border
    ws["C2"].alignment = center_align
    
    # Range is calculated dynamically based on misc_df length
    last_data_row = 6 + len(misc_df) - 1
    ws["D2"].value = f"=SUM(B6:B{last_data_row})"
    ws["D2"].fill = dark_blue_fill
    ws["D2"].font = white_bold_font
    ws["D2"].alignment = center_align 
    ws["D2"].border = white_border
    ws["D2"].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

    # === 5. Main Table Headers (Row 5) ===
    headers = ["Description", "Excluded [O&A] Price", "Comments"]
    for i, text in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=i, value=text)
        cell.fill = dark_blue_fill
        cell.font = white_bold_font
        cell.border = white_border
        cell.alignment = center_align

    # === 6. Writing Table Data (Starts at A6) ===
    # Use .values to automatically skip the DataFrame header
    for i, row_values in enumerate(misc_df.values, start=6):
        
        # Column A: Description
        desc_cell = ws.cell(row=i, column=1, value=row_values[0])
        desc_cell.alignment = left_align
        
        # Column B: Price with TBD Logic
        price_val = row_values[1]
        price_cell = ws.cell(row=i, column=2)
        
        if price_val is None or str(price_val).strip() in ["", "nan", "NaN"]:
            price_cell.value = "TBD"
            price_cell.alignment = center_align
        else:
            price_cell.value = price_val
            price_cell.alignment = right_align
            price_cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

        # Column C: Comments
        comment_val = row_values[2] if len(row_values) > 2 else ""
        ws.cell(row=i, column=3, value=comment_val).alignment = left_align

        # Apply borders/font to the row
        for col_idx in range(1, 4):
            cell = ws.cell(row=i, column=col_idx)
            cell.border = border_thin
            cell.font = black_font

    ws.sheet_view.showGridLines = False
    wb.save(invoice_path)
    print(f"[INFO] ✅ Misc sheet populated with TBD logic.")  
    
    
def create_fpla_and_labor_sheet(result: dict, year):
    year_col = str(year).strip().lower()
    
    files = ["final_summary_material.xlsx", "final_summary_repair.xlsx", "final_summary_vendor.xlsx"]
    cols = ["ata4", "cs_wks_level", "wks_num"]
    dataframes = []
    
    for file in files:
        try:
            # 1. Read only specific columns; throws ValueError if columns are missing
            df = pd.read_excel(file, usecols=cols)
            dataframes.append(df)
        except ValueError:
            print(f"Error: {file} is missing required columns: {cols}")
        except Exception as e:
            print(f"An unexpected error occurred reading {file}: {e}")

    try:
        # Combine and Filter
        combined_df = pd.concat(dataframes, ignore_index=True)
        
        # Condition: level is 1 AND num is 2 or 3
        mask = (combined_df["cs_wks_level"] == 1) & (combined_df["wks_num"].isin([2, 3]))
        
        # Remove duplicates and reset index for a clean final DF
        final_df = combined_df[mask].drop_duplicates().reset_index(drop=True)
        print(final_df)

    except Exception as e:
        print(f"Error during data processing: {e}")
        
    # ===================================
    # ===================================
    fpls_pricing_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"], 
        file_key="contract", 
        sheet_name_substring="fpls pricing"
    )
    print(fpls_pricing_df.columns)
    fpls_pricing_df.columns = [str(c).strip().lower() for c in fpls_pricing_df.columns]
    fpls_pricing_df = fpls_pricing_df[["esm chapter", "workscope", year_col]]
    fpls_pricing_df = fpls_pricing_df.rename(columns={year_col: "fpls_price"})
    print(f"FPLS Price DF Columns : {fpls_pricing_df.columns}")
    

    # This turns "EL 72-23-76" or "72-23" into "7223"
    def extract_ata4(text):
        if pd.isna(text): return None
        digits = "".join(re.findall(r'\d+', str(text)))
        return digits[:4] if len(digits) >= 4 else None

    # Create the temporary key in the other DataFrame
   
    fpls_pricing_df['ata4_key'] = fpls_pricing_df['esm chapter'].apply(extract_ata4)
    # Use .str.strip() instead of .strip()
    final_df['ata4'] = final_df['ata4'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()

    # Do the same for the pricing key to ensure they are identical strings
    fpls_pricing_df['ata4_key'] = fpls_pricing_df['ata4_key'].astype(str).str.strip()

    # Now the merge will run without the float/object error
    merged_df = pd.merge(
        final_df, 
        fpls_pricing_df, 
        left_on='ata4', 
        right_on='ata4_key', 
        how='left'
    )

    
    print(merged_df)
    
    
    fpls_columns = [
        "ATA", "Task", "Standard Price", "Minimum", "Performance", 
        "Full", "Contract Workscope", "Final Workscope", 
        "Excluded [O&A] Price", "Comments"
    ]
    
    fpls_df = pd.DataFrame(columns=fpls_columns)

    # Map the data from your merged_df
    fpls_df["ATA"] = merged_df["esm chapter"]
    fpls_df["Task"] = merged_df["workscope"]
    fpls_df["Standard Price"] = merged_df["fpls_price"]

    # Define the mapping for workscope descriptions
    wks_mapping = {1: "Minimum", 2: "Half", 3: "Full"}
    # Map the numeric levels to text (1 -> Minimum, etc.)
    fpls_df["Contract Workscope"] = merged_df["cs_wks_level"].map(wks_mapping)
    fpls_df["Final Workscope"] = merged_df["wks_num"].map(wks_mapping)

    # fpls_data = [
    #     ["72-00-00", "Engine Inspection", 5000, 4500, "High", "Yes", "Basic", "Full Overhaul", 1200.50, "Urgent repair needed"],
    #     ["72-30-00", "Compressor Wash", 1200, 1100, "Standard", "No", "Routine", "Routine", None, "Next due in 500 hrs"],
    #     ["73-10-00", "Fuel Pump Testing", 2500, 2200, "Optimal", "Yes", "Service Bulletin", "SB-442 Applied", 500.00, ""],
    #     ["75-20-00", "Anti-Ice Valve Replacement", 8000, 7500, "N/A", "Yes", "Ad-hoc", "Replaced", 0.00, "Part sourced from stock"],
    #     ["80-11-00", "Starter Motor Check", 900, 850, "Low", "No", "Minor", "Inspected", None, "TBD"]
    # ]

    # fpls_df = pd.DataFrame(fpls_data, columns=fpls_columns)
    # fpls_df = pd.DataFrame(columns=fpls_columns)
    print(fpls_df)
    
    labor_columns = [
        "ATA", 
        "Module Description", 
        "Hours", 
        "Excluded [O&A] Price", 
        "Comments"
    ]

    # labor_data = [
    #     ["72-31-00", "High Pressure Compressor", 45.5, 12500.00, "Routine overhaul"],
    #     ["72-41-00", "Combustion Section", 22.0, 8400.50, "Borescope inspection required"],
    #     ["72-51-00", "High Pressure Turbine", 60.0, np.nan, "TBD - Awaiting part inspection"],
    #     ["72-00-00", "Accessory Gearbox", 12.5, 3200.00, "Seal replacement"],
    #     ["72-30-00", "Fan Module", 18.0, 0.00, "Warranty coverage applies"]
    # ]

    # labor_df = pd.DataFrame(labor_data, columns=labor_columns)
    labor_df = pd.DataFrame(columns=labor_columns)
    
    fpls_df.to_excel("overall_fpls_data.xlsx", index=False)
    labor_df.to_excel("overall_labor_data.xlsx", index=False)
    
    fpls_labor_sheet_name = "FPLS & Labor"
    return fpls_df, labor_df, fpls_labor_sheet_name 

def create_fpla_and_labor_sheet(result: dict, year, scrap_mat_sum=0, scrap_rep_sum=0):
    year_col = str(year).strip().lower()
    
    files = ["final_summary_material.xlsx", "final_summary_repair.xlsx", "final_summary_vendor.xlsx"]
    cols = ["ata4", "cs_wks_level", "wks_num"]
    dataframes = []
    
    for file in files:
        try:
            df = pd.read_excel(file, usecols=cols)
            dataframes.append(df)
        except ValueError:
            print(f"Error: {file} is missing required columns: {cols}")
        except Exception as e:
            print(f"An unexpected error occurred reading {file}: {e}")

    try:
        combined_df = pd.concat(dataframes, ignore_index=True)
        
        mask = (combined_df["cs_wks_level"] == 1) & (combined_df["wks_num"].isin([2, 3]))
        final_df = combined_df[mask].drop_duplicates().reset_index(drop=True)

    except Exception as e:
        print(f"Error during data processing: {e}")
        
    # =========================
    # FPLS PRICING
    # =========================
    fpls_pricing_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"], 
        file_key="contract", 
        sheet_name_substring="fpls pricing"
    )

    fpls_pricing_df.columns = [str(c).strip().lower() for c in fpls_pricing_df.columns]
    fpls_pricing_df = fpls_pricing_df[["esm chapter", "workscope", year_col]]
    fpls_pricing_df = fpls_pricing_df.rename(columns={year_col: "fpls_price"})

    def extract_ata4(text):
        if pd.isna(text): return None
        digits = "".join(re.findall(r'\d+', str(text)))
        return digits[:4] if len(digits) >= 4 else None

    fpls_pricing_df['ata4_key'] = fpls_pricing_df['esm chapter'].apply(extract_ata4)

    final_df['ata4'] = final_df['ata4'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
    fpls_pricing_df['ata4_key'] = fpls_pricing_df['ata4_key'].astype(str).str.strip()

    merged_df = pd.merge(
        final_df, 
        fpls_pricing_df, 
        left_on='ata4', 
        right_on='ata4_key', 
        how='left'
    )

    # =========================
    # FPLS OUTPUT
    # =========================
    fpls_columns = [
        "ATA", "Task", "Standard Price", "Minimum", "Performance", 
        "Full", "Contract Workscope", "Final Workscope", 
        "Excluded [O&A] Price", "Comments"
    ]
    
    fpls_df = pd.DataFrame(columns=fpls_columns)

    fpls_df["ATA"] = merged_df["esm chapter"]
    fpls_df["Task"] = merged_df["workscope"]
    fpls_df["Standard Price"] = merged_df["fpls_price"]

    wks_mapping = {1: "Minimum", 2: "Half", 3: "Full"}
    fpls_df["Contract Workscope"] = merged_df["cs_wks_level"].map(wks_mapping)
    fpls_df["Final Workscope"] = merged_df["wks_num"].map(wks_mapping)

    # 🔥 NEW: Add scrap repair sum
    fpls_df["Excluded [O&A] Price"] = scrap_rep_sum

    # =========================
    # LABOR OUTPUT
    # =========================
    labor_columns = [
        "ATA", 
        "Module Description", 
        "Hours", 
        "Excluded [O&A] Price", 
        "Comments"
    ]

    labor_df = pd.DataFrame(columns=labor_columns)

    # 🔥 NEW: Add scrap material sum
    labor_df["Excluded [O&A] Price"] = [scrap_mat_sum]

    # =========================
    # WRITE OUTPUT
    # =========================
    fpls_df.to_excel("overall_fpls_data.xlsx", index=False)
    labor_df.to_excel("overall_labor_data.xlsx", index=False)

    return fpls_df, labor_df, "FPLS & Labor"

def fill_fpls_and_labor_sheet(invoice_path, fpls_labor_sheet_name, fpls_df, labor_df):
    """
    Fills the FPLS and Labor sheet. 
    Note: Column A is left empty. Data starts from Column B.
    """
    try:
        # --- 1. Load Workbook and Sheet ---
        if not os.path.exists(invoice_path):
            raise FileNotFoundError(f"File not found: {invoice_path}")
            
        wb = load_workbook(invoice_path)
        if fpls_labor_sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(fpls_labor_sheet_name)
        else:
            ws = wb[fpls_labor_sheet_name]

        # --- 2. Column Formatting (Starting from B) ---
        ws.column_dimensions['A'].width = 5   # Empty Column
        ws.column_dimensions['B'].width = 20  # ATA
        ws.column_dimensions['C'].width = 40  # Task
        ws.column_dimensions['D'].width = 15  # Standard Price
        ws.column_dimensions['E'].width = 25  # Minimum
        ws.column_dimensions['F'].width = 20  # Performance
        ws.column_dimensions['G'].width = 15  # Full
        ws.column_dimensions['H'].width = 25  # Contract Workscope
        ws.column_dimensions['I'].width = 25  # Final Workscope
        ws.column_dimensions['J'].width = 25  # Excluded [O&A] Price (Column 10)
        ws.column_dimensions['K'].width = 25  # Comments

        # --- 3. Styles ---
        dark_blue_fill = PatternFill(start_color="050543", end_color="050543", fill_type="solid")
        white_bold_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        black_font = Font(name="Arial", size=11)
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center", indent=1)
        right_align = Alignment(horizontal="right", vertical="center")
        white_side = Side(border_style="thin", color="FFFFFF")
        white_border = Border(left=white_side, right=white_side, top=white_side, bottom=white_side)
        curr_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

        # --- 4. Logo & Header ---
        try:
            if os.path.exists('ge_new_logo.png'):
                img = Image('ge_new_logo.png')
                img.height, img.width = 40, 100
                ws.add_image(img, 'B1') # Moved to B1
        except: pass

        # ws["B1"].value = "GE Aerospace"
        # ws["B1"].font = Font(name='Arial', bold=True, size=16, color="050543")

        # --- 5. Grand Total Header (Top Right) ---
        ws["K1"].value = "Excluded [O&A]\nPrice"
        ws["K1"].fill = dark_blue_fill
        ws["K1"].font = white_bold_font
        ws["K1"].border = white_border
        ws["K1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws["J2"].value = "Grand Total"
        ws["J2"].fill = dark_blue_fill
        ws["J2"].font = white_bold_font
        ws["J2"].border = white_border
        ws["J2"].alignment = center_align
        
        # --- 6. FPLS Table Headers (Row 4, Starting Col B/2) ---
        fpls_headers = [
            "ATA", "Task", "Standard Price", "Minimum", "Performance", 
            "Full", "Contract Workscope", "Final Workscope", 
            "Excluded [O&A] Price", "Comments"
        ]
        
        for i, text in enumerate(fpls_headers, start=2): # Start at 2 for Column B
            cell = ws.cell(row=4, column=i, value=text)
            cell.fill = dark_blue_fill
            cell.font = white_bold_font
            cell.alignment = center_align
            cell.border = white_border

        # --- 7. FPLS Data Writing (Starts at B6) ---
        # Data Loop
        for i, row_values in enumerate(fpls_df.values, start=6):
            for col_offset, value in enumerate(row_values, start=2): # Start at Column B
                cell = ws.cell(row=i, column=col_offset, value=value)
                # cell.border = border_thin
                cell.font = black_font
                cell.border = white_border
                
                # Excluded [O&A] Price logic (Column J / Index 10)
                if col_offset == 10: 
                    if pd.isna(value) or str(value).strip() in ["", "nan"]:
                        cell.value = "TBD"
                        cell.alignment = center_align
                    else:
                        cell.alignment = right_align
                        cell.number_format = curr_format
                elif col_offset in [4, 5]: # Price cols (D, E)
                    cell.alignment = right_align
                    cell.number_format = curr_format
                else:
                    cell.alignment = left_align

        # --- 8. FPLS Subtotal (Col J) ---
        last_fpls_row = 6 + len(fpls_df) - 1
        fpls_sub_row = last_fpls_row + 1
        
        ws.cell(row=fpls_sub_row, column=9, value="Sub Total").fill = dark_blue_fill
        ws.cell(row=fpls_sub_row, column=9).font = white_bold_font
        
        fpls_total_cell = ws.cell(row=fpls_sub_row, column=10) # Column J
        fpls_total_cell.value = f"=SUM(J6:J{last_fpls_row})"
        fpls_total_cell.fill = dark_blue_fill
        fpls_total_cell.font = white_bold_font
        fpls_total_cell.border = white_border
        fpls_total_cell.number_format = curr_format

        # --- 9. Labor Table (Starts 5 rows below FPLS Subtotal) ---
        labor_header_row = fpls_sub_row + 5
        ws.merge_cells(f'B{labor_header_row}:D{labor_header_row}') # Merge B:D
        lab_title = ws[f"B{labor_header_row}"]
        lab_title.value = "HOURLY LABOR"
        lab_title.fill = dark_blue_fill
        lab_title.font = white_bold_font
        lab_title.border = white_border
        lab_title.alignment = center_align

        # Labor Headers
        labor_cols_row = labor_header_row + 1
        labor_columns = ["ATA", "Module Description", "Hours", "Excluded [O&A] Price", "Comments"]
        for i, text in enumerate(labor_columns, start=2): # Start at 2 for Column B
            cell = ws.cell(row=labor_cols_row, column=i, value=text)
            cell.fill = dark_blue_fill
            cell.font = white_bold_font
            cell.alignment = center_align
            # cell.border = border_thin
            cell.border = white_border

        # --- 10. Labor Data Writing ---
        labor_data_start = labor_cols_row + 1
        for i, row_values in enumerate(labor_df.values, start=labor_data_start):
            for col_offset, value in enumerate(row_values, start=2): # Start Column B
                cell = ws.cell(row=i, column=col_offset, value=value)
                cell.border = border_thin
                cell.font = black_font
                
                if col_offset == 4: # Hours (Col D)
                    cell.alignment = right_align
                    cell.number_format = '0.00'
                elif col_offset == 5: # Labor Price (Col E)
                    if pd.isna(value) or str(value).strip() in ["", "nan"]:
                        cell.value = "TBD"
                        cell.alignment = center_align
                    else:
                        cell.alignment = right_align
                        cell.number_format = curr_format
                else:
                    cell.alignment = left_align

        # --- 11. Labor Subtotal & Grand Total Formula ---
        last_labor_row = labor_data_start + len(labor_df) - 1
        labor_sub_row = last_labor_row + 1
        
        ws.cell(row=labor_sub_row, column=4, value="Sub Total").fill = dark_blue_fill
        ws.cell(row=labor_sub_row, column=4).font = white_bold_font
        ws.cell(row=labor_sub_row, column=4).border = white_border
        
        labor_total_cell = ws.cell(row=labor_sub_row, column=5) # Column E
        labor_total_cell.value = f"=SUM(E{labor_data_start}:E{last_labor_row})"
        labor_total_cell.fill = dark_blue_fill
        labor_total_cell.font = white_bold_font
        labor_total_cell.border = white_border
        labor_total_cell.number_format = curr_format

        # FINAL GRAND TOTAL Formula (Sums both subtotal cells)
        ws["K2"].value = f"=J{fpls_sub_row} + E{labor_sub_row}"
        ws["K2"].fill = dark_blue_fill
        ws["K2"].font = white_bold_font
        ws["K2"].border = white_border
        ws["K2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # --- 12. Save ---
        ws.sheet_view.showGridLines = False
        wb.save(invoice_path)
        print(f"[SUCCESS] Sheet '{fpls_labor_sheet_name}' updated in {invoice_path}")

    except PermissionError:
        print(f"[ERROR] Permission Denied. Please close '{invoice_path}' and try again.")
    except Exception as e:
        print(f"[ERROR] An unexpected error occurred: {e}")

def fill_fpls_and_labor_sheet(invoice_path, sheet_name, metadata_fpls):

    wb = load_workbook(invoice_path)
    ws = wb[sheet_name]

    # =========================
    # STYLES
    # =========================
    dark_blue_fill = PatternFill(start_color="050543", end_color="050543", fill_type="solid")
    white_bold_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    normal_font = Font(name="Arial", size=11)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    curr_format = '_($* #,##0.00_);_($* (#,##0.00);_(@_)'

    # =========================
    # HELPER
    # =========================
    def get_col_map(header_row, sc, ec):
        col_map = {}
        for col in range(sc, ec + 1):
            val = ws.cell(row=header_row, column=col).value
            if val:
                col_map[val.strip()] = col
        return col_map

    # =========================
    # TABLE 1 → FPLS
    # =========================
    t1 = metadata_fpls[0]
    sr, er = t1["start_row"], t1["end_row"]
    sc, ec = t1["start_col"], t1["end_col"]

    # Header styling
    for col in range(sc, ec + 1):
        c = ws.cell(row=sr, column=col)
        c.fill = dark_blue_fill
        c.font = white_bold_font
        c.alignment = center
        c.border = border

    col_map = get_col_map(sr, sc, ec)

    # Data styling
    for r in range(sr + 1, er + 1):
        for col in range(sc, ec + 1):
            c = ws.cell(row=r, column=col)
            c.font = normal_font
            c.border = border

            if col == col_map.get("Excluded [O&A] Price"):
                if not c.value:
                    c.value = "TBD"
                    c.alignment = center
                else:
                    c.number_format = curr_format
                    c.alignment = right
            else:
                c.alignment = left

    # Subtotal
    sub_row = er + 1
    excl_col = col_map.get("Excluded [O&A] Price")

    ws.cell(row=sub_row, column=excl_col - 1, value="Sub Total").fill = dark_blue_fill
    ws.cell(row=sub_row, column=excl_col - 1).font = white_bold_font

    total_cell = ws.cell(row=sub_row, column=excl_col)
    total_cell.value = f"=SUM({get_column_letter(excl_col)}{sr+1}:{get_column_letter(excl_col)}{er})"
    total_cell.fill = dark_blue_fill
    total_cell.font = white_bold_font
    total_cell.number_format = curr_format

    # =========================
    # TABLE 2 → LABOR
    # =========================
    if len(metadata_fpls) > 1:

        t2 = metadata_fpls[1]
        sr2, er2 = t2["start_row"], t2["end_row"]
        sc2, ec2 = t2["start_col"], t2["end_col"]

        # Header styling
        for col in range(sc2, ec2 + 1):
            c = ws.cell(row=sr2, column=col)
            c.fill = dark_blue_fill
            c.font = white_bold_font
            c.alignment = center
            c.border = border

        col_map2 = get_col_map(sr2, sc2, ec2)

        # Data styling
        for r in range(sr2 + 1, er2 + 1):
            for col in range(sc2, ec2 + 1):
                c = ws.cell(row=r, column=col)
                c.font = normal_font
                c.border = border

                if col == col_map2.get("Excluded [O&A] Price"):
                    if not c.value:
                        c.value = "TBD"
                        c.alignment = center
                    else:
                        c.number_format = curr_format
                        c.alignment = right
                else:
                    c.alignment = left

        # Subtotal
        sub_row2 = er2 + 1
        excl_col2 = col_map2.get("Excluded [O&A] Price")

        ws.cell(row=sub_row2, column=excl_col2 - 1, value="Sub Total").fill = dark_blue_fill
        ws.cell(row=sub_row2, column=excl_col2 - 1).font = white_bold_font

        total_cell2 = ws.cell(row=sub_row2, column=excl_col2)
        total_cell2.value = f"=SUM({get_column_letter(excl_col2)}{sr2+1}:{get_column_letter(excl_col2)}{er2})"
        total_cell2.fill = dark_blue_fill
        total_cell2.font = white_bold_font
        total_cell2.number_format = curr_format

        # =========================
        # GRAND TOTAL
        # =========================
        ws["K2"] = f"={get_column_letter(excl_col)}{sub_row}+{get_column_letter(excl_col2)}{sub_row2}"
        ws["K2"].fill = dark_blue_fill
        ws["K2"].font = white_bold_font

    wb.save(invoice_path)
    print(f"[SUCCESS] FPLS & Labor formatted: {invoice_path}")

def create_pma_der(result: dict, output_excel: str):
    final_matl_df = pd.read_excel("revised_final_summary_material.xlsx")
    final_matl_df.drop_duplicates(subset=["item (sd)"], keep="first", inplace=True)

    pma_der_codes = [14, 35]
    pma_der_df = final_matl_df[final_matl_df['issue reason code'].isin(pma_der_codes) & (final_matl_df['net value']!=0)].copy()

    pma_der_df["Status"] = (
    pma_der_df["cost category"]
        .astype(str)
        .apply(
            lambda x: "NEW"
            if any(tag in x for tag in ["SCC-CONSUMABLE-AERO", "SCC-CONSUMABLE-NON-AERO"])
            else re.search(r"(?i)\b(NEW|USED)\b", x).group(1).upper()
            if re.search(r"(?i)\b(NEW|USED)\b", x)
            else None
        )
    )

    

    # pma_der_df['ata_long_module'] = pma_der_df['ata_long'].astype(str).str.zfill(6).str.replace(r'(.{2})(.{2})(.{2})', r'\1-\2-\3', regex=True)
    # pma_der_df.rename(columns={'refine_module': 'Module', 
    #                             'description': 'Part Name',
    #                             'material' : 'Part Number',
    #                             'target quantity' : 'Qty',
    #                             'matl_clp_lookup' : 'Unit Price',
    #                             'extended_price' : 'Extended Price',
    #                             'discount' : 'Discount',
    #                             'total_price' : 'Total Price',
    #                             'irc description': 'Removal Cause',
    #                             'remarks': 'Exclusion Remarks'}, inplace=True)
    # pma_der_df = pma_der_df[['Module', 'Part Name', 'Part Number', 'Qty', 'Unit Price',  'Extended Price', 'Discount', 'Total Price', 'Status', 'Removal Cause', 'Exclusion Remarks']]

    pma_der_df['Material Type'] = pma_der_df['cost category'].str[4:]
    
    pma_der_df['remarks'] = pma_der_df.apply(
        lambda row: append_module_to_remarks(
            row['formatted_ata_long'],
            row['module'],
            row['remarks']
        ),
        axis=1
    )
    
    
    pma_der_df.rename(columns={
        'item (sd)': 'Item (SD)',
        'formatted_ata_long': 'ATA Chapter Code',
        'material': 'Part Number',
        'description': 'Part Description',
        'target quantity': 'Qty',
        'pn_clp_lookup':'CLP',
        'matl_clp_lookup': 'Unit Price',
        'extended_price': 'Extended Price',
        'discount': 'Discount %',
        'discount_amount':'Discount $',
        'handling fee':'Handling Fee %',
        'handling_fee':'Handling Fee Price',
        'total_price': 'Excluded [O&A] Total Price',
        'irc description': 'Replacement Remarks',
        'order':'Service Order Equipment / MRB',
        'remarks':'Comments'
    }, inplace=True)
 
    # Final cleaning and export
    pma_der_df = pma_der_df[pma_der_df["Excluded [O&A] Total Price"] > 0].drop_duplicates()
    pma_der_df = pma_der_df[[
        "Item (SD)", "ATA Chapter Code", "Material Type", "Part Number", "Part Description", "Qty", "CLP", "Unit Price", "Extended Price",
        "Discount %", "Discount $", "Handling Fee %", "Handling Fee Price", "Excluded [O&A] Total Price", "Replacement Remarks",
        'Comments', "Service Order Equipment / MRB" 
    ]]

    # pma_der_sheet_name = "Out-of-Scope PMA & DER"
    pma_der_df.to_excel("pma_der_df.xlsx", index=False)
    return pma_der_df

def create_missing_on_receipts(result: dict, output_excel: str):
    final_matl_df = pd.read_excel("revised_final_summary_material.xlsx")

    mor_flag = ('Exclusion as Missing on Receipt',)
    mor_df = final_matl_df[final_matl_df['remarks'].str.startswith(mor_flag,na=False)].copy()
    
    mor_df["Status"] = (
    mor_df["cost category"]
        .astype(str)
        .apply(
            lambda x: "NEW"
            if any(tag in x for tag in ["SCC-CONSUMABLE-AERO", "SCC-CONSUMABLE-NON-AERO"])
            else re.search(r"(?i)\b(NEW|USED)\b", x).group(1).upper()
            if re.search(r"(?i)\b(NEW|USED)\b", x)
            else None
        )
    )

    # DO not Apply discounts when no CLP, otherise as it is
    mor_df.loc[mor_df["discount_flag"] == "No Discount", "discount"] = 0
    
    # Cost > 0,pn_clp_lookup is blank, discount = 0, discount_flag = "No Discount"
    mask_pn_clp_lookup_blank = (mor_df["cost"] > 0) & (mor_df["pn_clp_lookup"].astype(str).str.strip().isin(["", "nan", "NaN"]))
    mor_df.loc[mask_pn_clp_lookup_blank, "discount"] = 0
    mor_df.loc[mask_pn_clp_lookup_blank, "discount_flag"] = "No Discount"

    # mor_df.to_excel("checkkkkk.xlsx", index=False)
    # Cost > 0, pn_clp_lookup is > 0, discount = 0.12, disconut_falg = "Discount"
    mask_pn_clp_lookup_above_0 = (mor_df["cost"] > 0) & (mor_df["pn_clp_lookup"] > 0)
    mor_df.loc[mask_pn_clp_lookup_above_0, "discount"] = 0.05
    mor_df.loc[mask_pn_clp_lookup_above_0, "discount_flag"] = "Discount"

    # Discount = 0, disconut_falg = "No Discount"
    mask_discount_for_used_cc = (mor_df["Status"] == "USED")
    mor_df.loc[mask_discount_for_used_cc, "discount"] = 0 #
    mor_df.loc[mask_discount_for_used_cc, "discount_flag"] = "No Discount"

    # mor_df.rename(columns={'refine_module': 'Module', 
    #                             'description': 'Part Name',
    #                             'material' : 'Part Number',
    #                             'target quantity' : 'Qty',
    #                             'matl_clp_lookup' : 'Unit Price',
    #                             'extended_price' : 'Extended Price',
    #                             'discount' : 'Mark Up',
    #                             'total_price' : 'Total Price',
    #                             "remarks": "Exclusion Remarks"}, inplace=True)
    # mor_df = mor_df[mor_df['Total Price'] > 0].copy()
    # mor_df = mor_df[['Module', 'Part Name', 'Part Number', 'Qty', 'Unit Price',  'Extended Price', 'Mark Up', 'Total Price', 'Status', 'Exclusion Remarks']]
    # mor_sheet_name = "Out-of-Scope Missing Items"

    mor_df['Material Type'] = mor_df['cost category'].str[4:]
    
    mor_df['remarks'] = mor_df.apply(
        lambda row: append_module_to_remarks(
            row['formatted_ata_long'],
            row['module'],
            row['remarks']
        ),
        axis=1
    )
    mor_df.rename(columns={
        'item (sd)': 'Item (SD)',
        'formatted_ata_long': 'ATA Chapter Code',
        'material': 'Part Number',
        'description': 'Part Description',
        'target quantity': 'Qty',
        'pn_clp_lookup':'CLP',
        'matl_clp_lookup': 'Unit Price',
        'extended_price': 'Extended Price',
        'discount': 'Discount %',
        'discount_amount':'Discount $',
        'handling fee':'Handling Fee %',
        'handling_fee':'Handling Fee Price',
        'total_price': 'Excluded [O&A] Total Price',
        'irc description': 'Replacement Remarks',
        'order':'Service Order Equipment / MRB',
        'remarks':'Comments'
    }, inplace=True)
 
    # Final cleaning and export
    mor_df = mor_df[mor_df["Excluded [O&A] Total Price"] > 0].drop_duplicates()
    mor_df = mor_df[[
        "Item (SD)", "ATA Chapter Code", "Material Type", "Part Number", "Part Description", "Qty", "CLP", "Unit Price", "Extended Price",
        "Discount %", "Discount $", "Handling Fee %", "Handling Fee Price", "Excluded [O&A] Total Price", "Replacement Remarks",
        'Comments', "Service Order Equipment / MRB" 
    ]]
    mor_df.to_excel("mor_df.xlsx", index=False)
    return mor_df 


def apply_qpe_limits(
    result: dict,
    invoice_df: pd.DataFrame,
    qpe_file_key: str = "contract",
    qpe_sheet_substring: str = "qpe_list",
    text_cols: list[str] | None = None,
    qty_col_hint: str | None = None,
    highlight_col: str = "_qpe_highlight"
    ) -> pd.DataFrame:
    df = invoice_df.copy()
    df.columns = [str(c).strip() for c in df.columns]
 
    # --- helper: ensure Series even if duplicate headers exist ---
    def _as_series(dfx: pd.DataFrame, col: str) -> pd.Series:
        obj = dfx[col]
        if isinstance(obj, pd.DataFrame):
            # if duplicate column names, take the first occurrence
            obj = obj.iloc[:, 0]
        return obj
 
    # ---------- Detect text columns to search ----------
    if text_cols is None:
        candidates = []
        for needle in [
            "part name", "description",
            "genpact_part_keyword", "genpact_part_keywords",
            "csn description", "nomenclature"
        ]:
            hit = next((c for c in df.columns if needle == c.lower()), None)
            if hit:
                candidates.append(hit)
        if not candidates:
            candidates = [c for c in df.columns if re.search(r"(name|desc|nomenclature|keyword)", c, flags=re.I)]
        text_cols = candidates or [df.columns[0]]
 
    # Build a single lowercase search blob (always a Series)
    cols = text_cols or []
    if isinstance(cols, (str, int)):
        cols = [cols]
    cols = [c for c in cols if c in df.columns]
    if not cols: 
        cols = [df.columns[0]]

    search_blob = (
        df.loc[: , cols]
        .astype(str)
        .apply(lambda row: " ".join(map(str, row.values)), axis=1)
        .astype(str)
        .str.lower()
    )
 
    # ---------- Detect quantity column ----------
    if qty_col_hint and qty_col_hint in df.columns:
        qty_col = qty_col_hint
    else:
        qty_col = next(
            (c for c in df.columns
             if c.lower() in {"qty", "quantity", "billable qty", "billable quantity",
                              "inspected quantity", "target quantity"}
             or ("qty" in c.lower() or "quantity" in c.lower())),
            None
        )
    if qty_col is None:
        raise ValueError("Could not detect a quantity column. Provide `qty_col_hint` or ensure a qty column exists.")
    qty_series_all = pd.to_numeric(_as_series(df, qty_col), errors="coerce").fillna(0)
 
    # ---------- Load QPE list ----------
    qpe_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key=qpe_file_key,
        sheet_name_substring=qpe_sheet_substring
    ).copy()
    qpe_df.columns = [str(c).strip() for c in qpe_df.columns]
 
    kw_col = next((c for c in qpe_df.columns
                   if c.lower() in {"part keyword", "keyword", "part keywords"} or
                      ("keyword" in c.lower() and "part" in c.lower())), None)
    qpe_col = next((c for c in qpe_df.columns if c.lower() == "qpe"), None)
    if not kw_col or not qpe_col:
        raise ValueError("QPE_List must have columns: 'Part Keyword' and 'QPE'.")
 
    # Use _as_series to survive duplicate headers
    s_kw   = _as_series(qpe_df, kw_col)
    s_qpe  = _as_series(qpe_df, qpe_col)
 
    qpe_df = pd.DataFrame({
        "_kw_norm": s_kw.astype(str).str.strip().str.lower(),
        "_qpe":     pd.to_numeric(s_qpe, errors="coerce"),
    }).dropna(subset=["_kw_norm", "_qpe"])
 
    qpe_map = (
        qpe_df.groupby("_kw_norm", as_index=True)["_qpe"].max()
        .dropna()
        .to_dict()
    )
 
    # ---------- Helpers ----------
    def mask_contains_all_words(blob: pd.Series, keyword: str) -> pd.Series:
        tokens = [t for t in re.findall(r"[A-Za-z0-9]+", keyword.lower()) if t]
        if not tokens:
            return pd.Series(False, index=blob.index)
        m = pd.Series(True, index=blob.index)
        for t in tokens:
            pat = rf"(?<!\w){re.escape(t)}(?!\w)"
            m &= blob.str.contains(pat, regex=True, na=False)
        return m
 
    # ---------- Enforce ----------

    df[highlight_col] = False
    df["_qpe_keyword"] = ""

    for kw_norm, qpe in qpe_map.items():
        if qpe is None or qpe <= 0:
            continue

        # Rows whose search blob contains all words of the keyword
        m = mask_contains_all_words(search_blob, kw_norm)
        if not m.any():
            continue

        idx = df.index[m]
        qty_this = qty_series_all.loc[idx]  # numeric Series aligned to df

        # Always annotate matched rows with the keyword
        df.loc[idx, "_qpe_keyword"] = kw_norm

        # If any matched row exceeds QPE, cap only the row with the maximum quantity
        if (qty_this > qpe).any():
            idx_max = qty_this.idxmax()
            max_qty = float(qty_this.loc[idx_max])
            # Cap the retained max row's quantity at keyword QPE
            df.loc[idx_max, qty_col] = min(max_qty, float(qpe))
            # Highlight the capped row
            df.loc[idx_max, highlight_col] = True
 
 
 
    return df




def highlight_flagged_rows(xlsx_path: str, sheet_name: str, df_with_flags: pd.DataFrame, highlight_col="QPE_Highlight"):
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]
    # Assumes header on row 1 and df rows written in the same order starting at row 2
    fill = PatternFill(start_color="FFFDE9D9", end_color="FFFDE9D9", fill_type="solid")
    for excel_row, flag in enumerate(df_with_flags[highlight_col].tolist(), start=2):
        if flag:
            for cell in ws[excel_row]:
                cell.fill = fill
    wb.save(xlsx_path)


def enforce_qpe_by_keywords(
    lines_df: pd.DataFrame,
    qpe_df: pd.DataFrame,
    part_col: str = "Part Description",
    qty_col: str = "Qty",
    group_cols: list | None = None,          # e.g., ["ESN"] or ["Invoice No"]
    highlight_col: str = "QPE_Highlight",    # boolean flag column added
    ) -> pd.DataFrame:
    
    """
    Enforce QPE limits based on a QPE list with columns: 'Part Keyword', 'QPE'.
 
    Behavior:
      - For each keyword (optionally per group), compute cumulative Qty.
      - If total > QPE:
          * If cumsum == QPE at some row: drop rows after that point.
          * Else: keep all rows but mark them for highlight.
      - Adds a boolean column `highlight_col` and an optional 'QPE_Note' text column.
 
    Returns a new dataframe with rows potentially dropped and flags added.
    """
 
    if lines_df.empty:
        out = lines_df.copy()
        out[highlight_col] = False
        return out
 
    # Normalize part names for robust matching
    df = lines_df.copy()
    df["_norm_part"] = (
        df[part_col].astype(str)
        .str.upper()
        .str.replace(r"[^\w\s]", " ", regex=True)   # remove punctuation
        .str.replace(r"\s+", " ", regex=True)       # condense spaces
        .str.strip()
    )

    # Numeric Qty
    df["_qty_num"] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
 
    # Clean QPE list
    qdf = qpe_df.rename(columns=lambda c: str(c).strip().lower())
    qpe_df.to_excel("qpe_df.xlsx", index=False)

    # Find keyword column robustly
    if "part keyword" in qdf.columns:
        kw_col = "part keyword"
    elif "part_description" in qdf.columns:
        kw_col = "part_keyword"
    else:
        raise ValueError("QPE_List must have a 'Part Keyword' column.")
    if "qpe" not in qdf.columns:
        raise ValueError("QPE_List must have a 'QPE' column.")
 
    qdf["QPE"] = pd.to_numeric(qdf["qpe"], errors="coerce")
    qdf = qdf.dropna(subset=[kw_col, "QPE"])
    if qdf.empty:
        df[highlight_col] = False
        return df.drop(columns=["_norm_part", "_qty_num"])
 
    def _build_anyorder_pattern(keyword: str) -> re.Pattern | None:
        # Turn "fan blade" → tokens ["FAN","BLADE"] and require all tokens present in any order.
        tokens = re.findall(r"\w+", str(keyword).upper())
        if not tokens:
            return None
        lookaheads = "".join(f"(?=.*\\b{re.escape(t)}\\b)" for t in tokens)
        # Apply on normalized UPPER string, so no case flag needed
        return re.compile(lookaheads + ".*")
 
    to_drop: set[int] = set()
    highlight_notes: dict[int, list[str]] = {}
 
    # Process each QPE keyword
    for kw, qpe in qdf[[kw_col, "QPE"]].itertuples(index=False):
        pat = _build_anyorder_pattern(kw)
        if pat is None:
            continue
 
        mask_kw = df["_norm_part"].str.contains(pat, regex=True, na=False)
        if not mask_kw.any():
            continue
 
        # Operate globally or per group
        if group_cols:
            groups = df[mask_kw].groupby(group_cols, dropna=False).groups
            targets = [idx for _, idx in groups.items()]
        else:
            targets = [df.index[mask_kw]]
 
        for idx in targets:
            idx = list(idx)
            # qty = df.loc[idx, "_qty_num"]
            # csum = qty.cumsum()
            # total = float(csum.iloc[-1])
 
            # if total <= float(qpe):
            #     continue  # OK, nothing to do
 
            # # total > QPE: either trim on exact match or highlight if not exact
            # if (csum == qpe).any():
            #     exceed = csum > qpe
            #     to_drop.update(qty.index[exceed].tolist())
            # else:
            #     note = f"Total {int(total)} > QPE {int(qpe)} for '{kw}' (no exact match); review."
            #     for i in idx:
            #         highlight_notes.setdefault(i, []).append(note)
            
            qty = df.loc[idx, "_qty_num"]
 
            # Row-wise enforcement:
            # Qty <= QPE → safe
            # Qty > QPE  → stop and remove
            violations = qty[qty > float(qpe)]
            
            if violations.empty:
                continue  # all rows are safe
            
            # First row where Qty > QPE
            stop_idx = violations.index[0]
            
            # Drop this row and everything after it
            to_drop.update(i for i in idx if i >= stop_idx)            
 
    # Apply row deletions (trim)
    if to_drop:
        df = df.drop(index=list(to_drop))
 
    # Attach highlight flags/notes
    df[highlight_col] = False
    if highlight_notes:
        hit = list(highlight_notes.keys())
        # Some indices may have been dropped during trimming; keep those that remain
        hit = [i for i in hit if i in df.index]
        if hit:
            df.loc[hit, highlight_col] = True
            df["QPE_Note"] = ""
            notes_ser = df["QPE_Note"].copy()
            for i in hit:
                joined = " | ".join(dict.fromkeys(highlight_notes[i]))  # dedupe, keep order
                prev = notes_ser.at[i]
                notes_ser.at[i] = joined if not prev else f"{prev} | {joined}"
            df["QPE_Note"] = notes_ser
 
    # Cleanup helpers
    return df.drop(columns=[c for c in ["_norm_part", "_qty_num"] if c in df.columns])
    
def create_material_repair(result: dict, output_excel: str):
    # Load original material file
    mat_org_df = pd.read_excel("revised_final_summary_material.xlsx")
    mat_org_df.drop_duplicates(subset=["item (sd)"], keep="first", inplace=True)    
    
    mat_filt = ('Inclusion as WKS = 1', 'Inclusion as WKS = 2', 'Exclusion as WKS = 1', 'Exclusion as WKS = 2', 'Exclusion as WKS = 0','Exclusion as FOD')
    mat_updated_df = mat_org_df[mat_org_df['remarks'].str.startswith(mat_filt,na=False)]
    mat_updated_df.drop_duplicates(inplace=True)
    # mat_org_filt_df.to_excel("mat_org_df.xlsx", index=False)

    # blade_fan_rotor_pattern = r"(?=.*BLADE)(?=.*(?:FAN|HPT))" #r"(?=.*BLADE)(?=.*HPT)"
    # mat_updated_df = mat_org_filt_df[~mat_org_filt_df['genpact_part_keywords'].str.contains(blade_fan_rotor_pattern, regex=True, na=False)]    
    # mat_updated_df = mat_org_df.copy()
 
    # Load WSPG Incl/Excl
    wspg_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="wspg incl excl"
    )
    wspg_df.columns = [str(c).strip().lower().replace(" ", "_") for c in wspg_df.columns]
    wspg_df = wspg_df[["atalong", "a", "b"]]

    # Assign NEW/USED status
    mat_updated_df["Status"] = (
        mat_updated_df["cost category"]
        .astype(str)
        .apply(
            lambda x: "NEW"
            if any(tag in x for tag in ["SCC-CONSUMABLE-AERO", "SCC-CONSUMABLE-NON-AERO"])
            else re.search(r"(?i)\b(NEW|USED)\b", x).group(1).upper()
            if re.search(r"(?i)\b(NEW|USED)\b", x)
            else None
        )
    )
 
    # Discount Logic
    mat_updated_df.loc[mat_updated_df["discount_flag"] == "No Discount", "discount"] = 0
 
    mask_clp_blank = (
        (mat_updated_df["cost"] > 0) &
        (mat_updated_df["pn_clp_lookup"].astype(str).str.strip().isin(["", "nan", "NaN"]))
    )
    mat_updated_df.loc[mask_clp_blank, "discount"] = 0
    mat_updated_df.loc[mask_clp_blank, "discount_flag"] = "No Discount"
 
    mask_clp_above_0 = (
        (mat_updated_df["cost"] > 0) &
        (mat_updated_df["pn_clp_lookup"] > 0)
    )
    mat_updated_df.loc[mask_clp_above_0, "discount"] = 0.05
    mat_updated_df.loc[mask_clp_above_0, "discount_flag"] = "Discount"
 
    mask_used = mat_updated_df["Status"] == "USED"
    mat_updated_df.loc[mask_used, "discount"] = 0
    mat_updated_df.loc[mask_used, "discount_flag"] = "No Discount"
 
    mask_non_ual = (
        (~mat_updated_df["vendor description"].astype(str).str.strip().str.upper().eq("AVIANCA AIRLINES")) &
        mat_updated_df["cost category"].astype(str).str.upper().str.contains("CUSTOMER|CUST", na=False)
    )
    mat_updated_df.loc[mask_non_ual, ["matl_clp_lookup", "discount"]] = 0

    # Filter out atas that end with 0000
    mat_updated_df.rename(columns={'_atalong6_x':'_atalong6'}, inplace=True)
    # mat_updated_df.drop('_atalong6_y', axis=1, inplace=True)

    mat_updated_df['ata_long_clean'] = mat_updated_df['_atalong6'].astype(str).str.strip()
    mat_updated_df['ata_long_clean'] = mat_updated_df['ata_long_clean'].astype(str).str[:7]
    mat_updated_df = mat_updated_df[~mat_updated_df['ata_long_clean'].str.endswith("0000")]    
    # mat_updated_df.to_excel("mat_updated_df.xlsx", index=False)
    
    # Re-Check by Module column to see if remarks are correctly applied
    # Convert Module to MOD_ATA
    # mat_updated_df["MOD_ATA"] = mat_updated_df['module'].str.replace('-', '', regex=False)
    mat_updated_df["MOD_ATA"] = mat_updated_df['_atalong6']
    mat_updated_df['MOD_ATA'] = mat_updated_df['MOD_ATA'].astype(str).str[:6]
    
    # Conver MOD_ATA to MOD4
    mat_updated_df["MOD4"] = mat_updated_df["MOD_ATA"].str[:4]

    # Condition-1: MOD4 between 7200 to 7209
    # mask_cond1 = mat_updated_df["MOD4"].between(7200, 7209)
    # mat_updated_df.loc[mask_cond1, "remarks"] = "Inclusion as 72"

    # Checking Exclusions Remarks and Change to Inclusions if conditions does satisfies
    # For Condition-2 and Condition-3, define a helper function to process rows
    def check_exclusion_remarks(row):
        if row['remarks'] == "Exclusion as WKS = 1":
            if row['customer_workscope_level'] or row['final ata ws'] == 1:
                mod_ata = int(row['MOD_ATA'])

                # Find if MOD_ATA exists in wspg['atalong']
                wspg_df['atalong'] = wspg_df['atalong'].astype(int)
                wspg_row = wspg_df[wspg_df['atalong'] == mod_ata]
                
                if not wspg_row.empty:
                    val = wspg_row.iloc[0]['a']
                    # print("Printing val: exc to inc 1", val)
                    if val == "X":
                        return 'Inclusion as WKS = 1'
                    
        elif row['remarks'] == 'Exclusion as WKS = 2':
            if row['customer_workscope_level'] or row['final ata ws'] == 2:
                mod_ata = int(row['MOD_ATA'])
        
                # Find if MOD_ATA exists in wspg['atalong']
                wspg_df['atalong'] = wspg_df['atalong'].astype(int)
                wspg_row = wspg_df[wspg_df['atalong'] == mod_ata]
                
                if not wspg_row.empty:
                    val = wspg_row.iloc[0]['b']
                    # print("Printing val: exc to inc 2", val)
                    if val == "X":
                        return 'Inclusion as WKS = 2'

        return row['remarks'] 
    
    def is_blank(val):
        return pd.isna(val) or (isinstance(val, str) and val.strip() == "")
    
    def check_inclusion_remarks(row):
        if row['remarks'] == "Inclusion as WKS = 1":
            if row['customer_workscope_level'] or row['final ata ws'] == 1:
                mod_ata = int(row['MOD_ATA'])

                # Find if MOD_ATA exists in wspg['atalong']
                wspg_df['atalong'] = wspg_df['atalong'].astype(int)
                wspg_row = wspg_df[wspg_df['atalong'] == mod_ata]
                # print("Is blank for : ", wspg_row)

                if not wspg_row.empty:
                    val = wspg_row.iloc[0]['a']
                    if val != "X": #is_blank(val) or pd.isna(val):
                        # print("Is blank for : ", mod_ata)
                        return 'Exclusion as WKS = 1'
                    # print("Outside inside if (Incl to Ex 1): ", val)
                    
        elif row['remarks'] == 'Inclusion as WKS = 2':
            if row['customer_workscope_level'] or row['final ata ws'] == 2:
                mod_ata = str(row['MOD_ATA'])

                # Find if MOD_ATA exists in wspg['atalong']
                wspg_df['atalong'] = wspg_df['atalong'].astype(str)
                wspg_row = wspg_df[wspg_df['atalong'] == mod_ata]
                # print("Is blank for : ", wspg_row)

                if not wspg_row.empty:
                    val = wspg_row.iloc[0]['b']
                    if val != "X": #is_blank(val) or pd.isna(val):
                        # print("Inside: ", val)
                        return 'Exclusion as WKS = 2'
                    # print("Outside inside if: (Inc to Excl 2)", val)
                            
        return row['remarks'] 

    # mat_updated_df['remarks'] = mat_updated_df.apply(check_exclusion_remarks, axis=1)
    # mat_updated_df['remarks'] = mat_updated_df.apply(check_inclusion_remarks, axis=1)
    # mat_updated_df.to_excel("refined_mat.xlsx", index=False)

    # Rename columns for export
    mat_updated_df['Material Type'] = mat_updated_df['cost category'].str[4:]
    
    mat_updated_df['remarks'] = mat_updated_df.apply(
        lambda row: append_module_to_remarks(
            row['formatted_ata_long'],
            row['module'],
            row['remarks']
        ),
        axis=1
    )
    mat_updated_df.rename(columns={
        'item (sd)': 'Item (SD)',
        'formatted_ata_long': 'ATA Chapter Code',
        'material': 'Part Number',
        'description': 'Part Description',
        'target quantity': 'Qty',
        'pn_clp_lookup':'CLP',
        'matl_clp_lookup': 'Unit Price',
        'extended_price': 'Extended Price',
        'discount': 'Discount %',
        'discount_amount':'Discount $',
        'handling fee':'Handling Fee %',
        'handling_fee':'Handling Fee Price',
        'total_price': 'Excluded [O&A] Total Price',
        'irc description': 'Replacement Remarks',
        'order':'Service Order Equipment / MRB',
        'remarks':'Comments'
    }, inplace=True)
 
    # Final cleaning and export
    mat_updated_df = mat_updated_df[mat_updated_df["Excluded [O&A] Total Price"] > 0].drop_duplicates()
    mat_updated_df = mat_updated_df[[
        "Item (SD)", "ATA Chapter Code", "Material Type", "Part Number", "Part Description",	"Qty", "CLP", "Unit Price", "Extended Price",
        "Discount %", "Discount $", "Handling Fee %", "Handling Fee Price", "Excluded [O&A] Total Price", "Replacement Remarks",
        'Comments', "Service Order Equipment / MRB" 
    ]]
    
    mat_updated_df.to_excel("mat_table_before.xlsx")

    mat_filt = ('Exclusion as WKS = 1', 'Exclusion as WKS = 2', 'Exclusion as WKS = 0','Exclusion as FOD')
    mat_updated_df['Comments'] = mat_updated_df['Comments'].astype(str)
    mat_updated_df = mat_updated_df[mat_updated_df['Comments'].str.startswith(mat_filt,na=False)]
    mat_updated_df.drop_duplicates(inplace=True)
    

    # = = == = = = = = = = = = == = = = Repair df = = == = = = = = = = = = = = = = = 
    rep_df_ = pd.read_excel("final_summary_repair.xlsx")
    rep_df_with_mod = create_module_columns(rep_df_)

    rep_df_with_mod.to_excel("revised_repair_summary.xlsx")
    rep_dff = pd.read_excel("revised_repair_summary.xlsx")
    # rep_dff.drop_duplicates(subset=["extracted_repair_code"], keep="first", inplace=True)    


    rep_filt = ('Inclusion as WKS = 1', 'Inclusion as WKS = 2', 'Exclusion as WKS = 1', 'Exclusion as WKS = 2', 'Exclusion as WKS = 0','Exclusion as FOD')
    rep_df = rep_dff[rep_dff['remarks'].str.startswith(rep_filt,na=False)].copy()
    rep_df.drop_duplicates(inplace=True)

    rep_df.rename(columns={'_atalong6_x':'_atalong6'}, inplace=True)

    rep_df['ata_long_clean'] = rep_df['_atalong6'].astype(str).str.strip()
    rep_df['ata_long_clean'] = rep_df['ata_long_clean'].astype(str).str[:7]
    rep_df = rep_df[~rep_df['ata_long_clean'].str.endswith("0000")]    
    # rep_df.to_excel("rep_updated_df.xlsx", index=False)

    # Re-Check by Module column to see if remarks are correctly applied
    # Convert Module to MOD_ATA
    rep_df["MOD_ATA"] = rep_df['module'].str.replace('-', '', regex=False)
    rep_df['MOD_ATA'] = rep_df['MOD_ATA'].astype(str).str[:6]
    
    # Conver MOD_ATA to MOD4
    rep_df["MOD4"] = rep_df["MOD_ATA"].str[:4].astype(int)

    # Condition-1: MOD4 between 7200 to 7209
    mask_cond1 = rep_df["MOD4"].between(7200, 7209)
    rep_df.loc[mask_cond1, "remarks"] = "Inclusion as 72"

    # rep_df['remarks'] = rep_df.apply(check_exclusion_remarks, axis=1)
    # rep_df['remarks'] = rep_df.apply(check_inclusion_remarks, axis=1)
    # rep_df.to_excel("refined_rep.xlsx", index=False)
    
    # repair discount "Component Repairs"
    rep_df['discount_amount'] = rep_df['total_price_final'] * rep_df['discount']
    
    rep_df['remarks'] = rep_df.apply(
        lambda row: append_module_to_remarks(
            row['formatted_ata_long'],
            row['module'],
            row['remarks']
        ),
        axis=1
    )
    
    rep_df.rename(columns={
        'br item': 'Item (SD)',
        'formatted_ata_long': 'ATA Chapter Code',
        'inspected part no.' : 'Part Number',
        'csn description': 'Part Description',
        'inspected quantity' : 'Qty',
        'repair hours' : 'Hour / CRD',
        'total_price_final' : 'Unit Price',
        'extended_price_final': 'Extended Price',
        'discount': 'Discount %',
        'discount_amount':'Discount $',
        'total_price' : 'Excluded [O&A] Total Price',
        'order no':'PO/Repair Tag',
        'repair code group & code' : 'Repair Code / Service Level',
        'remarks':'Comments'}, inplace=True)

    # Final cleaning and export

    rep_df = rep_df[[
        "Item (SD)", "ATA Chapter Code", "Part Number", "Part Description",	"Qty", "Hour / CRD", "Unit Price", "Extended Price",
        "Discount %", "Discount $", "Excluded [O&A] Total Price", "Comments", "PO/Repair Tag",
        "Repair Code / Service Level", "MOD4", "repair description", "price"
    ]]

    rep_filt = ('Exclusion as WKS = 1', 'Exclusion as WKS = 2', 'Exclusion as WKS = 0','Exclusion as FOD')
    rep_df['Comments'] = rep_df['Comments'].astype(str)
    rep_dff = rep_df[(rep_df['Comments'].str.startswith(rep_filt,na=False))]
    rep_df = rep_dff[(rep_dff['MOD4'] >= 7221) & (rep_dff['MOD4'] <= 7263)]
    rep_df.drop_duplicates(inplace=True)
    
    # normalize a bit
    rep_df['Part Number'] = rep_df['Part Number'].astype(str).str.strip()
    rep_df["Repair Code / Service Level"] = rep_df["Repair Code / Service Level"].astype(str).str.strip()
    rep_df.to_excel("rep_df.xlsx", index=False)
    cc_parts_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="cc part"
    )
        # Normalize column names
    cc_parts_df.columns = [col.strip().lower() for col in cc_parts_df.columns]
    #main_df = main_df.copy()
    #m#ain_df.columns = [col.strip().lower() for col in main_df.columns]
 
    # Detect relevant columns
    part_num_col = next((col for col in cc_parts_df.columns if "part" in col and "num" in col), None)
    keyword_col = next((col for col in cc_parts_df.columns if "part" in col and "keyword" in col), None)
    pn_col = next((col for col in rep_df.columns if "Part" in col and "Number" in col), None)

    enriched_df =  rep_df.merge(
        cc_parts_df[[part_num_col, keyword_col]],
        left_on=pn_col,
        right_on=part_num_col,
        how='left'
    )
 
    # Step 4: Create 'GENPACT_PART_KEYWORDS' column
    enriched_df["genpact_part_keywords"] = enriched_df[keyword_col]
    #enriched_df["genpact_part_keywords"] = enriched_df["genpact_part_keywords"].fillna(enriched_df[description_col])
    enriched_df=enriched_df.drop_duplicates()
    enriched_df.to_excel("cehck_qpe.xlsx", index=False)
    
    rep_qpe_df = apply_qpe_limits(result, enriched_df)
    rep_qpe_df.to_excel("QPE_LImit.xlsx", index=False)

    # If module (ATA chapter) is 72-23-02
    # and Part Description is CASE-CONTAINMENT
    # and Repair Description contains ABRADABLE
    # then include the repair
    # BUT: skip rows where Comments == 'Exclusion as WKS = 0'
    
    mask_abradable = (
        rep_qpe_df["ATA Chapter Code"].str.lower().str.contains("72-23-02", na=False) &
        rep_qpe_df["Part Description"].str.lower().str.contains("case-containment", na=False) &
        rep_qpe_df["repair description"].str.lower().str.contains("abradable", na=False) &
        (rep_qpe_df["Comments"] != "Exclusion as WKS = 0")
    )
    
    if mask_abradable.any():
        rep_qpe_df.loc[mask_abradable, "Comments"] = "Inclusion"
    
    # Wherever the cost category is "SCC-LABOUR-REPAIR" as Part Number, there take Qty = 1
    rep_qpe_df.loc[(rep_qpe_df["Part Number"] == "SCC-LABOUR-REPAIR") & 
                       (rep_qpe_df["Qty"] == 0),
                       "Qty"] = 1
    
    # --- 1) Get Component Repair discount from contract (unchanged)
    comp_rep_contract_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="discount"
    )
    comp_rep_contract_df.columns = [col.strip().lower() for col in comp_rep_contract_df.columns]
    # print("comp_rep_contract_df.columns :", comp_rep_contract_df.columns)
    
    cost_cat_col = next((col for col in comp_rep_contract_df.columns if "cost" in col and "category" in col), None)
    comp_rep_disc_col = next((col for col in comp_rep_contract_df.columns if "discount" in col), None)
    
    mask_disc = (
        comp_rep_contract_df[cost_cat_col]
        .astype(str)
        .str.strip()
        .str.contains(r"\bcomponent repairs(s)?\b", flags=re.IGNORECASE, regex=True, na=False)
    )

    val = comp_rep_contract_df.loc[mask_disc, comp_rep_disc_col].dropna()

    component_repair_disc = float(val.iloc[0]) if not val.empty else 0.0
    
    # --- 2) Apply discounts
    # rep_qpe_df["discount"] = pd.to_numeric(rep_qpe_df.get("discount", 0), errors="coerce").fillna(0)
    if "discount" in rep_qpe_df.columns:
        rep_qpe_df["discount"] = pd.to_numeric(rep_qpe_df["discount"], errors="coerce").fillna(0)
    else:
        rep_qpe_df["discount"] = 0
    
    rep_qpe_df["discount"] = np.where(
        rep_qpe_df["price"].notna() & (rep_qpe_df["price"].astype(str).str.strip() != ""),
        component_repair_disc,
        0
    )
    
    # rep_qpe_df.to_excel("repair_summary_v1.xlsx",index=False)

    # Overwrite target quantity as 1 --> Whenever discount is 0
    rep_qpe_df.loc[rep_qpe_df["discount"] == 0, "Qty"] = 1
    
    # Alot Unit Price = 0 for Similar values of Module, Part Name, Part Number, Qty, Unit Price
    # subset_cols = ["ATA Chapter Code", "Part Description", "Part Number", "Qty", "Unit Price"]
    # mask = rep_qpe_df.duplicated(subset=subset_cols, keep="first")
    # rep_qpe_df.loc[mask, "Unit Price"] = 0

    rep_qpe_df.rename(columns={'discount' : 'Discount %'}, inplace=True)

    rep_qpe_df = rep_qpe_df[["Item (SD)", "ATA Chapter Code", "Part Number", "Part Description", "Qty", "Hour / CRD", "Unit Price", 
        "Extended Price", "Discount %", "Discount $", "Excluded [O&A] Total Price", "Comments", "PO/Repair Tag",
        "Repair Code / Service Level"
    ]]
    rep_qpe_df.to_excel("repair_summary_v2.xlsx",index=False)
    
    
    # = = = Vendor df = = = = = = =

    ###
    # Load original material file
    vend_data = pd.read_excel("final_summary_vendor.xlsx")
    vend_df_with_mod = create_module_columns(vend_data)

    vend_df_with_mod.to_excel("revised_vendor_summary.xlsx")
    ven_org_df = pd.read_excel("revised_vendor_summary.xlsx")
    # ven_org_df.drop_duplicates(subset=["item (sd)"], keep="first", inplace=True)

    ven_filt = ('Inclusion as WKS = 1', 'Inclusion as WKS = 2', 'Exclusion as WKS = 1', 'Exclusion as WKS = 2', 'Exclusion as WKS = 0','Exclusion as FOD')
    rep_df['Comments'] = rep_df['Comments'].astype(str)
    ven_updated_df = ven_org_df[ven_org_df['remarks'].str.startswith(ven_filt,na=False)]
    ven_updated_df.drop_duplicates(inplace=True)
    # ven_org_filt_df.to_excel("ven_org_df.xlsx", index=False)

    # blade_fan_rotor_pattern = r"(?=.*BLADE)(?=.*(?:FAN|HPT))" #r"(?=.*BLADE)(?=.*HPT)"
    # ven_updated_df = ven_org_filt_df[~ven_org_filt_df['genpact_part_keywords'].str.contains(blade_fan_rotor_pattern, regex=True, na=False)]    

    # Filter out atas that end with 0000
    ven_updated_df.rename(columns={'_atalong6_x':'_atalong6'}, inplace=True)
    # mat_updated_df.drop('_atalong6_y', axis=1, inplace=True)

    ven_updated_df['ata_long_clean'] = ven_updated_df['_atalong6'].astype(str).str.strip()
    ven_updated_df['ata_long_clean'] = ven_updated_df['ata_long_clean'].astype(str).str[:7]
    ven_updated_df = ven_updated_df[~ven_updated_df['ata_long_clean'].str.endswith("0000")]    
    # ven_updated_df.to_excel("ven_updated_df.xlsx", index=False)
    
    # Re-Check by Module column to see if remarks are correctly applied
    # Convert Module to MOD_ATA
    ven_updated_df["MOD_ATA"] = ven_updated_df['module'].str.replace('-', '', regex=False)
    ven_updated_df['MOD_ATA'] = ven_updated_df['MOD_ATA'].astype(str).str[:6]
    
    # Conver MOD_ATA to MOD4
    ven_updated_df["MOD4"] = ven_updated_df["MOD_ATA"].str[:4].astype(int)

    # Condition-1: MOD4 between 7200 to 7209
    mask_cond1 = ven_updated_df["MOD4"].between(7200, 7209)
    ven_updated_df.loc[mask_cond1, "remarks"] = "Inclusion as 72"

    ven_updated_df['remarks'] = ven_updated_df.apply(check_exclusion_remarks, axis=1)
    ven_updated_df['remarks'] = ven_updated_df.apply(check_inclusion_remarks, axis=1)
    # ven_updated_df.to_excel("refined_ven.xlsx", index=False)

    # # vend_df['ata_long_module'] = vend_df['ata_long'].astype(str).str.zfill(6).str.replace(r'(.{2})(.{2})(.{2})', r'\1-\2-\3', regex=True)    
    ven_updated_df['discount'] = 0
    ven_updated_df['Material Type'] = ven_updated_df['cost category'].str[4:]
    
    ven_updated_df['remarks'] = ven_updated_df.apply(
        lambda row: append_module_to_remarks(
            row['formatted_ata_long'],
            row['module'],
            row['remarks']
        ),
        axis=1
    )
    
    ven_updated_df.rename(columns={
        'item (sd)': 'Item (SD)',
        'formatted_ata_long': 'ATA Chapter Code',
        'material': 'Part Number',
        'description': 'Part Description',
        'target quantity': 'Qty',
        'unit clp':'CLP',
        'matl_clp_lookup': 'Vendor Invoice Amount',
        'extended_price': 'Extended Price',
        'discount': 'Discount %',
        'discount_amount':'Discount $',
        'handling fee':'Handling Fee %',
        'handling_fee':'Handling Fee Price',
        'total_price': 'Excluded [O&A] Total Price',
        'irc description': 'Replacement Remarks',
        'order':'PO #',
        'remarks':'Comments',
        'Material Type': 'Repair Source'}, inplace=True)
    
    # Final cleaning and export
    ven_updated_df = ven_updated_df[ven_updated_df["Excluded [O&A] Total Price"] > 0].drop_duplicates()
    ven_updated_df = ven_updated_df[[
        "Item (SD)", "ATA Chapter Code", "Repair Source", "Part Number", "Part Description", "Qty", "Vendor Invoice Amount", "Discount %","Discount $",
        "Handling Fee %", "Handling Fee Price", "Excluded [O&A] Total Price", 'Comments', "PO #" 
    ]]
    ven_updated_df["Vendor Invoice Amount"] = ven_updated_df["Vendor Invoice Amount"] * ven_updated_df["Qty"]
    ven_filt = ('Exclusion as WKS = 1', 'Exclusion as WKS = 2', 'Exclusion as WKS = 0','Exclusion as FOD')
    ven_updated_df['Comments'] = ven_updated_df['Comments'].astype(str)
    ven_updated_df = ven_updated_df[ven_updated_df['Comments'].str.startswith(ven_filt,na=False)]
    ven_updated_df.drop_duplicates(inplace=True)

    # QPE_List df
    qpe_list_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="qpe_list"
    )

    mat_qpe_result = enforce_qpe_by_keywords(
    mat_updated_df,
    qpe_list_df,
    part_col="Part Description",
    qty_col="Qty",
    group_cols=None,             # or ["ESN"] / ["Invoice No"] if you want per-group enforcement
    highlight_col="QPE_Highlight")

    mat_qpe_result["QPE_Highlight"] = mat_qpe_result["QPE_Highlight"].map({True: "REVIEW", False: ""})

    rep_qpe_result = enforce_qpe_by_keywords(
    rep_qpe_df,
    qpe_list_df,
    part_col="Part Description",
    qty_col="Qty",
    group_cols=None,             # or ["ESN"] / ["Invoice No"] if you want per-group enforcement
    highlight_col="QPE_Highlight")

    rep_qpe_result["QPE_Highlight"] = rep_qpe_result["QPE_Highlight"].map({True: "REVIEW", False: ""})
    rep_qpe_result.to_excel("internal_table01.xlsx", index=False)
    

    ven_qpe_result = enforce_qpe_by_keywords(
    ven_updated_df,
    qpe_list_df,
    part_col="Part Description",
    qty_col="Qty",
    group_cols=None,             # or ["ESN"] / ["Invoice No"] if you want per-group enforcement
    highlight_col="QPE_Highlight")

    ven_qpe_result["QPE_Highlight"] = ven_qpe_result["QPE_Highlight"].map({True: "REVIEW", False: ""})

    mat_qpe_result.drop_duplicates(inplace=True)
    mat_qpe_result['ATA Chapter Code'] = mat_qpe_result['ATA Chapter Code'].astype(str).str[:8]
    # mat_qpe_result.rename(columns={'QPE_Highlight':'Exclusion Remarks', 'Exclusion Remarks':'QPE_Highlight'}, inplace=True)
    
    ven_qpe_result.drop_duplicates(inplace=True)

    mat_sheet_name = "Material" 
    rep_sheet_name = "GE Repair"
    ven_sheet_name = "Sub Contracted Repair"

    mat_qpe_result['CLP'] = mat_qpe_result['CLP'].fillna(0)
    mat_qpe_result.drop_duplicates(inplace=True)
    rep_qpe_result.drop_duplicates(inplace=True)
    rep_qpe_result.to_excel("internal_table_02.xlsx", index=False)
    

    vendor_columns =  ["Item (SD)", "ATA Chapter Code", "Repair Source", "Part Number", "Part Description", "Qty", "Vendor Invoice Amount","Discount %", "Discount $", 
        "Handling Fee %", "Handling Fee Price", "Excluded [O&A] Total Price", 'Comments', "PO #" , "QPE_Highlight"]
    
    if ven_qpe_result.shape[0] == 0:
        ven_qpe_result = pd.DataFrame([["-"] * len(vendor_columns)], columns=vendor_columns)

    ven_qpe_result.drop_duplicates(inplace=True)

    mat_qpe_result["has_CUST"] = np.where(mat_qpe_result["Material Type"].astype(str).str.contains("CUST", case=False, na=False), 1, 0)

    rep_qpe_result = rep_qpe_result[["Item (SD)", "ATA Chapter Code", "Part Number", "Part Description", "Qty", "Hour / CRD", 
                                     "Unit Price", "Extended Price", "Discount %", "Discount $", "Excluded [O&A] Total Price",
                                     "Comments", "PO/Repair Tag", "Repair Code / Service Level", "QPE_Highlight"
    ]]
    rep_qpe_result.drop_duplicates(inplace=True)

    cols = list(rep_qpe_result.columns)
    target_col = 'Discount %'
    first_idx = cols.index(target_col)
    rep_qpe_result.columns = [f"{col}_{idx}" if col == target_col else col 
                              for idx, col in enumerate(cols)]
    rep_qpe_result = rep_qpe_result.drop(columns=[f"{target_col}_{first_idx}"])
    rep_qpe_result.columns = [target_col if col.startswith(target_col + "_") else col for col in rep_qpe_result.columns]                   
    
    mat_qpe_result.to_excel("mat_table.xlsx", index=False)
    ven_qpe_result.to_excel("external_table.xlsx", index=False)
    rep_qpe_result.to_excel("internal_table.xlsx", index=False)

    return mat_qpe_result, rep_qpe_result, ven_qpe_result, mat_sheet_name, rep_sheet_name, ven_sheet_name

def create_taxes(result:dict, output_excel: str):
    empty_df = pd.DataFrame(columns=['ESN', 'NF', 'ISS BRL', 'Exh Rate', 'ISS USD'])
    
    esn_list = [int(result['esn'])]
    esn = str(esn_list[0])
    new_row  = {'ESN': esn, 'NF': '12422-1', 'ISS BRL': "68303.06", 'Exh Rate':"5.7551", 'ISS USD': "11,868.27"}
    
    tax_df = pd.DataFrame([new_row])
    tax_final_df = pd.concat([empty_df, tax_df], ignore_index=True)

    taxes_sheet_name = "Taxes"
    return tax_final_df, taxes_sheet_name

def create_service_bulletin(result: dict, output_excel: str):
    fin_sb_ = pd.read_excel("service_bulletin_remarks.xlsx")
    
    # ========= SB CAP ========
    fin_sb_.drop_duplicates(
        subset=["sb_x","material",'item (sd)'],
        keep="first",
        inplace=True
    )
    # Create mask for target rows
    mask = fin_sb_['mtl sb remarks'] == 'Inclusion as SB'
    
    # Calculate the total for those specific rows
    total_sum = fin_sb_.loc[mask, 'matl_clp_lookup'].sum()
    print(f"Total Material Inclusion Sum : {total_sum}")
    
    # Initialize the column with total_sum
    fin_sb_.loc[mask, 'total_mtl_sb_included_value'] = total_sum
    
    # If sum > 100k, update only the target rows
    if total_sum > 100000:
        fin_sb_.loc[mask, 'mtl sb remarks'] = "Exclusion as MTL SB"
        
    fin_sb_.to_excel('service_bulletin_after_removing_duplicate.xlsx', index=False)
    
        
    # return sb_df
    
    # apply_sb_cap(df=sb_df, pipeline_type='Total Included Price', sb_remarks_col= 'material sb remarks', price_col="")
    
    fin_sb_df_with_mod = create_module_columns(fin_sb_)
 
 
    # Filter "remarks" column --> without previous rules-based remarks
    prev_remark = ["MATL UNKNOWN"]
    fin_sb_df_with_mod = fin_sb_df_with_mod[(fin_sb_df_with_mod['remarks'].isin(prev_remark))]   
 
    fin_sb_df_with_mod.to_excel("revised_sb_material.xlsx")
    fin_sb_dff = pd.read_excel("revised_sb_material.xlsx")
 
    fin_sb_filt = ['Exclusion as MTL SB']
    sb_df = fin_sb_dff[fin_sb_dff['mtl sb remarks'].isin(fin_sb_filt)].copy()
 
    # DO not Apply discounts when no CLP, otherise as it is
    # sb_df.loc[sb_df["discount_flag"] == "No Discount", "discount"] = 0
    # If discount is not allowed, zero out both discount % and discount amount
    sb_df.loc[
        sb_df["discount_flag"] == "No Discount",
        ["discount", "discount_amount"]
    ] = 0
 
    sb_df["Status"] = (
    sb_df["cost category"]
        .astype(str)
        .apply(
            lambda x: "NEW"
            if any(tag in x for tag in ["SCC-CONSUMABLE-AERO", "SCC-CONSUMABLE-NON-AERO"])
            else re.search(r"(?i)\b(NEW|USED)\b", x).group(1).upper()
            if re.search(r"(?i)\b(NEW|USED)\b", x)
            else None
        )
    )
    sb_df.to_excel("create_service_bulletin_2.xlsx", index=False)
    #23-12-2025
    # sb_df2=sb_df.copy().dropna(subset=['sb'])
    if 'sb_y' in sb_df.columns:
        sb_df2=sb_df.copy().dropna(subset=['sb_y'])
    else:
        sb_df2=sb_df.copy().dropna(subset=['sb'])
       
    # sb_df2['new_remark'] = sb_df2['mtl sb remarks'].astype(str) + sb_df2['sb'].astype(str)
    if 'sb_y' in sb_df.columns:
        sb_df2['new_remark'] = sb_df2['mtl sb remarks'].astype(str) + sb_df2['sb_y'].astype(str)
    else:
        sb_df2['new_remark'] = sb_df2['mtl sb remarks'].astype(str) + sb_df2['sb'].astype(str)
        
    sb_df2.rename(columns={
        'item (sd)': 'Item (SD)',
        'module': 'ATA Chapter Code',
        'cost category':'Material Type',
        'material': 'Part Number',
        'description_x': 'Part Description',
        'target quantity': 'Qty',
        'pn_clp_lookup':'CLP',
        'matl_clp_lookup': 'Unit Price',
        'extended_price': 'Extended Price',
        'discount': 'Discount %',
        'discount_amount':'Discount $',
        'handling fee':'Handling Fee %',
        'handling_fee':'Handling Fee Price',
        'total_price': 'Excluded [O&A] Total Price',
        'irc description': 'Replacement Remarks',
        'order':'Service Order Equipment / MRB',
        'new_remark':'Comments'
    }, inplace=True)
   
    print(sb_df2.columns)
    print("saving sb df2")
    sb_df2.to_excel("sb_df2.xlsx",index=False)
    sb_df2 = sb_df2[sb_df2["Excluded [O&A] Total Price"] > 0].drop_duplicates()
    
   
    sb_df2 = sb_df2[[
        "Item (SD)", "ATA Chapter Code", "Material Type", "Part Number", "Part Description",    "Qty", "CLP", "Unit Price", "Extended Price",
        "Discount %", "Discount $", "Handling Fee %", "Handling Fee Price", "Excluded [O&A] Total Price", "Replacement Remarks",
        'Comments', "Service Order Equipment / MRB"
    ]]
    sb_df2 ['QPE_Highlight'] = pd.NA
    sb_df2 ['has_CUST'] = pd.NA
 
    # Build a new comment_value column from OEM Category and UAL Code

    sb_df2.to_excel("Sb_test_23.xlsx", index=False)
    # … all your renaming and calculations …

    # Add the new column before writing sb_df out
    sb_df["Comments"] = (
        "OEM CAT " + sb_df["oem category"].fillna("").astype(int).astype(str) 
    )

    # Now export with the new column included
    sb_df.to_excel("Sb_before_concat_mat.xlsx", index=False)
 
    sb_df.rename(columns={'refine_module': 'Module',
                            'description_x': 'Part Name',
                            'material' : 'Part Number',
                            'target quantity' : 'Qty',
                            'matl_clp_lookup' : 'Unit Price',
                            'extended_price': 'Extended Price',
                            'discount' : 'Discount',
                            'item (sd)': 'Item (SD)',
                            'total_price' : 'Material', # 'SB Description' 'sb description': 'SB Description'
                            'irc description': 'Removal Cause', 
                            'sb_x' : 'SB', 
                            'sb description_x': 'Description', 
                            'category': 'SB Category',
                            }, inplace=True)
    sb_df['Description'] = sb_df['Description'].fillna(sb_df['Part Name'])
     
    # sb_df['Material'] = (sb_df['Extended Price']) - (sb_df['Extended Price'] * sb_df['Discount'])
    
    # sb_df['Material'] = np.where(
    # sb_df['Discount'] == 0,
    # sb_df['Unit Price'],
    # sb_df['Extended Price'] - (sb_df['Extended Price'] * sb_df['Discount'])
    # )
    
    sb_df['Material'] = np.where(
    (sb_df['Qty'] == 1) & (sb_df['Discount'] == 0),
    sb_df['Unit Price'],
    sb_df['Extended Price'] - (sb_df['Extended Price'] * sb_df['Discount'])
    )      
    
    sb_df['Labor'] = pd.NA
    sb_df["Repair"] = pd.NA 
 
    sb_df['Labor'] = sb_df['Labor'].fillna(0)
    # sb_df['Material'] = sb_df['Material'].fillna(0)
    sb_df['Repair'] = sb_df['Repair'].fillna(0)

    sb_df = sb_df[['Item (SD)','SB','Part Number', 'SB Category', 'Labor', 'Material', 'Repair', 'Comments']]
    # sb_df = sb_df[['Item (SD)','SB','Part Number', 'SB Category', 'Labor', 'Material', 'Repair', 'Comments', 'Material SB Remarks', 'Total Included Price']]
    
    
    sb_df.to_excel("Sb_before_concat_mat_1.xlsx", index=False)

    # Repair
    fin_rep_ = pd.read_excel("repair_service_bulletin_remarks.xlsx")
    # ========= SB CAP ========
    fin_rep_.drop_duplicates(
        subset=["sb_x","inspected part no.",'br item'],
        keep="first",
        inplace=True
    )
    # Create mask for target rows
    mask = fin_rep_['repair sb remarks'] == 'Inclusion as SB'
    
    # Calculate the total for those specific rows
    total_sum = fin_rep_.loc[mask, 'total_price_final'].sum()
    print(f"Total Repair Inclusion Sum : {total_sum}")
    
    # Initialize the column with total_sum
    fin_rep_.loc[mask, 'total_repair_sb_included_value'] = total_sum
    
    # If sum > 100k, update only the target rows
    if total_sum > 100000:
        fin_rep_.loc[mask, 'repair sb remarks'] = "Exclusion as REPAIR SB"
        
    fin_rep_.to_excel('service_bulletin_repair_after_removing_duplicate.xlsx', index=False)
        
    fin_rep_df_with_mod = create_module_columns(fin_rep_)
 
    fin_rep_df_with_mod.to_excel("revised_sb_repair.xlsx")
    fin_rep_dff = pd.read_excel("revised_sb_repair.xlsx")
 
    at_filters = ['Exclusion as REPAIR SB']
    rep_df = fin_rep_dff[fin_rep_dff['repair sb remarks'].isin(at_filters)].copy()
 
    # --- 1) Get Component Repair discount from contract (unchanged)
    comp_rep_contract_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="discount"
    )
    comp_rep_contract_df.columns = [col.strip().lower() for col in comp_rep_contract_df.columns]
    print("comp_rep_contract_df.columns :", comp_rep_contract_df.columns)
   
    cost_cat_col = next((col for col in comp_rep_contract_df.columns if "cost" in col and "category" in col), None)
    comp_rep_disc_col = next((col for col in comp_rep_contract_df.columns if "discount" in col), None)
    
    mask_disc = (
        comp_rep_contract_df[cost_cat_col]
        .astype(str)
        .str.strip()
        .str.contains(r"\bcomponent repairs(s)?\b", flags=re.IGNORECASE, regex=True, na=False)
    )
 
    val = comp_rep_contract_df.loc[mask_disc, comp_rep_disc_col].dropna()
 
    component_repair_disc = float(val.iloc[0]) if not val.empty else 0.0
   
    # --- 2) Apply discounts
    rep_df["discount"] = pd.to_numeric(rep_df.get("discount", 0), errors="coerce").fillna(0)
   
    rep_df["discount"] = np.where(
        rep_df["price"].notna() & (rep_df["price"].astype(str).str.strip() != ""),
        component_repair_disc,
        0
    )
 
    # Overwrite target quantity as 1 --> Whenever discount is 0
    rep_df.loc[rep_df["discount"] == 0, "inspected quantity"] = 1
 
    #23-12-2025
    sb_df2_r= rep_df.copy().dropna(subset=['sb_x'])
    sb_df2_r['discount_amount']= pd.NA
    sb_df2_r['new_remark'] = sb_df2_r['repair sb remarks'].astype(str) + sb_df2_r['sb_x'].astype(str)
    
    sb_df2_r.rename(columns={
        'br item': 'Item (SD)',
        'module': 'ATA Chapter Code',
        'inspected part no.' : 'Part Number',
        'csn description': 'Part Description',
        'inspected quantity' : 'Qty',
        'repair hours' : 'Hour / CRD',
        'total_price_final' : 'Unit Price',
        'extended_price_final': 'Extended Price',
        'discount': 'Discount %',
        'discount_amount':'Discount $',
        'sb_total_price' : 'Excluded [O&A] Total Price',
        'order no':'PO/Repair Tag',
        'repair code group & code' : 'Repair Code / Service Level',
        'new_remark':'Comments',
        }, inplace=True)
   
    #print(sb_df2_r.columns)
    #sb_df2_r = sb_df2_r[sb_df2_r["Excluded [O&A] Total Price"] > 0].drop_duplicates()
   
    sb_df2_r = sb_df2_r[[
        "Item (SD)", "ATA Chapter Code", "Part Number", "Part Description", "Qty", "Hour / CRD", "Unit Price", "Extended Price",
        "Discount %", "Discount $", "Excluded [O&A] Total Price", "Comments", "PO/Repair Tag",
        "Repair Code / Service Level"
    ]]
 
    sb_df2_r ['QPE_Highlight'] = pd.NA
    #sb_df2_r['has_CUST'] = pd.NA
 
    sb_df2_r.to_excel("Sb_test_23_r.xlsx", index=False)
    

    rep_df["Comments"] = (
        "OEM CAT " + rep_df["oem category"].apply(
            lambda x: str(int(x)) if pd.notnull(x) and str(x).strip() != "" and float(x).is_integer() else str(x) if pd.notnull(x) else ""
        ).astype(str)  
    )

    rep_df.to_excel("Sb_before_concat_rep.xlsx", index=False)
    
    # rep_df.drop_duplicates(inplace=True)
    
    rep_df.rename(columns={'refine_module': 'Module',
                                'csn description': 'Part Name',
                                'inspected part no.' : 'Part Number',
                                'inspected quantity' : 'Qty',
                                'br item': 'Item (SD)',
                                'total_price_final' : 'Unit Price',
                                'extended_price_final' : 'Extended Price',
                                'discount' : 'Discount',
                                'repair hours' : 'Hour / CRD',
                                'irc description': 'Removal Cause',
                                'repair code group & code' : 'Repair Code',
                                'repair description' : 'Service', 
                                'sb_x':'SB', 
                                'sb description_y':'Description', 
                                'category': 'SB Category',
                                }, inplace=True)
    # rep_df['Extended Price'] = rep_df['Unit Price'] * rep_df['Discount']
    # rep_df['Repair'] = (rep_df['Extended Price']) - (rep_df['Extended Price'] * rep_df['Discount'])  
    # rep_df['Repair'] = (rep_df['Unit Price'])        
    
    # rep_df['Repair'] = np.where(
    # rep_df['Discount'] == 0,
    # rep_df['Unit Price'],
    # rep_df['Extended Price'] - (rep_df['Extended Price'] * rep_df['Discount'])
    # )
    
    rep_df['Repair'] = np.where(
    (rep_df['Qty'] == 1) & (rep_df['Discount'] == 0),
    rep_df['Unit Price'],
    rep_df['Extended Price'] - (rep_df['Extended Price'] * rep_df['Discount'])
    )
    
    rep_df['Description'] = rep_df['Description'].fillna(rep_df['Part Name'])            
    # rep_df = rep_df[['SB', 'Description', 'Module', 'Part Name', 'Part Number', 'Qty', 'Unit Price',  'Extended Price', 'price from', 'Discount', 'Repair', 'Hour / CRD', 'Repair Code', 'Service']]
 
    # To have use concat we need to have common columns in material and repair
    rep_df['Labor'] = pd.NA
    rep_df["Material"] = pd.NA
    rep_df = rep_df[['Item (SD)','SB', 'Part Number','SB Category', 'Labor', 'Material', 'Repair', 'Comments']]
    rep_df['Labor'] = rep_df['Labor'].fillna(0)
    rep_df['Material'] = rep_df['Material'].fillna(0)
    rep_df['Repair'] = rep_df['Repair'].fillna(0)
    
    
    overall_sb = pd.concat([sb_df, rep_df])
 
    # Drop duplicates based on Part Number, keep the first occurrence
    overall_sb.drop_duplicates(
        subset=["SB","Part Number",'Item (SD)',"Comments"],
        keep="first",
        inplace=True
    )
    
 
    overall_sb.to_excel("Overall_sb.xlsx",index=False)
    # Mask for SB = 72-0119
    mask_720119 = overall_sb["SB"] == "72-0119"
    
    # Move Repair into Labor
    overall_sb.loc[mask_720119, "Labor"] = (
        overall_sb.loc[mask_720119, "Labor"] +
        overall_sb.loc[mask_720119, "Repair"]
    )
    
    # Set Repair to 0
    overall_sb.loc[mask_720119, "Repair"] = 0
    
    overall_sb_grouped = overall_sb.groupby(['SB', 'SB Category'], as_index=False).agg({
        'Material': 'sum',
        'Labor': 'sum',
        'Repair': 'sum',
        'Comments': 'first'
    })
 
    overall_sb_grouped.rename(columns={'SB':'SB Number'}, inplace=True)
    overall_sb_grouped['SB Total'] = overall_sb_grouped['Labor'] + overall_sb_grouped['Material'] + overall_sb_grouped['Repair']
 
    overall_sb_grouped = overall_sb_grouped[["SB Number", "SB Category", "Labor", "Material", "Repair", "SB Total", "Comments"]]
    overall_sb_grouped.to_excel("overall_sb_grouped.xlsx", index=False)
    sb_sheet_name = "SB O&A Summary"
    
   
    return overall_sb_grouped, sb_sheet_name


def apply_thick_outer_thin_inner_border(ws, start_row, end_row, start_col, end_col):
    thick = Side(style="thick", color="000000")
    thin = Side(style="thin", color="000000")
 
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
 
            # Determine each side's style
            top = thick if row == start_row else thin 
            bottom = thick if row == end_row else thin
            left = thick if col == start_col else thin
            right = thick if col == end_col else thin
 
            # Apply combined border
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

def auto_adjust_column_widths(ws):
    """Auto-adjust the width of each column based on the max length of its content."""
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)  # Get column letter (e.g., 'A', 'B')
        for cell in col:
            if cell.value:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        ws.column_dimensions[column_letter].width = max_length + 2

def format_row_as_currency(ws, row_num, start_col, end_col):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_num, column=col)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
            cell.alignment = Alignment(horizontal="right", vertical="center")
def get_adder_creep_data(result: dict, year):
    year_col = str(year).strip().lower()
    
    # Load files
    final_matl = pd.read_excel("final_summary_material.xlsx")
    final_rep = pd.read_excel("final_summary_repair.xlsx")
    final_vend = pd.read_excel("final_summary_vendor.xlsx")
    
    # Creep filter
    creep_values = ['Creap 1 to 2', 'Creap 1 to 3']
    
    # Extract unique 'ata4' values from all dataframes
    # We use a set to automatically handle uniqueness across all three sources
    unq_wks = set()
    for df in [final_matl, final_rep, final_vend]:
        filtered = df[df["remarks"].isin(creep_values)]
        if not filtered.empty:
            unq_wks.update(filtered['ata4'].astype(str).unique())
    
    if not unq_wks:
        return {}

    creep_price_df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="contract",
        sheet_name_substring="workscope creep adder"
    )
    
    # Convert all column names to string first, then strip and lower
    creep_price_df.columns = [str(col).strip().lower() for col in creep_price_df.columns]

    
    # Extract the 4-digit ATA (e.g., 72-89 -> 7289)
    extracted_ata = creep_price_df['ffp+price adders'].str.extract(r'^(\d{2})-(\d{2})').fillna('').sum(axis=1)
    
    # Create the module_key (last 2 digits + 'x')
    creep_price_df['module_key'] = extracted_ata.str[-2:] + 'x'

    # Filter rows based on the original 4-digit ATA matching your material list
    mask = extracted_ata.isin(unq_wks)
    filtered_df = creep_price_df[mask]

    # Return map of {module_key: price}
    return dict(zip(filtered_df['module_key'], filtered_df[year_col]))

    
    
    
def generate_cover_sheet(adder_creep_data: dict, output_path: str = "invoice_cover.xlsx", engine_workscope = "Engine Restoration"):
    wb = Workbook()
    ws = wb.active

    # Column width adjustments
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 21

    # Row height adjustments
    ws.row_dimensions[1].height = 33
    ws.row_dimensions[2].height = 13
    ws.row_dimensions[3].height = 17
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 17
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 17
    ws.row_dimensions[8].height = 72
    ws.row_dimensions[9].height = 17
    ws.row_dimensions[10].height = 72
    ws.row_dimensions[11].height = 17
    ws.row_dimensions[12].height = 18
    ws.row_dimensions[13].height = 17
    ws.row_dimensions[14].height = 17
    ws.row_dimensions[15].height = 8
    ws.row_dimensions[16].height = 17
    ws.row_dimensions[17].height = 7
    ws.row_dimensions[18].height = 17
    ws.row_dimensions[19].height = 18
    ws.row_dimensions[20].height = 18
    ws.row_dimensions[21].height = 18
    ws.row_dimensions[22].height = 18
    ws.row_dimensions[23].height = 18
    ws.row_dimensions[24].height = 18
    ws.row_dimensions[25].height = 18
    ws.row_dimensions[26].height = 18
    ws.row_dimensions[27].height = 17
    ws.row_dimensions[28].height = 17
    ws.row_dimensions[29].height = 7
    ws.row_dimensions[30].height = 17
    ws.row_dimensions[31].height = 18
    ws.row_dimensions[32].height = 18
    ws.row_dimensions[33].height = 18
    ws.row_dimensions[34].height = 18
    ws.row_dimensions[35].height = 18
    ws.row_dimensions[36].height = 18
    ws.row_dimensions[37].height = 17
    ws.row_dimensions[38].height = 17
    ws.row_dimensions[38].height = 17
    ws.row_dimensions[39].height = 17
    ws.row_dimensions[40].height = 17
    ws.row_dimensions[41].height = 17

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FF0000"
    ws.sheet_view.zoomScale = 90
    ws.title = "Summary"
 
    # === Font Styles ===
    blue_fill = PatternFill(start_color="050543", end_color="050543", fill_type="solid")
    white_bold_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    black_bold_font = Font(name="Arial", bold=True, color="000000", size=11)
    footer_note = Font(name="Arial", bold=True, color="000000", size=11)
    norma_black_font = Font(name="Arial", bold=False, color="000000", size=11)
    title_font = Font(name="Arial", bold=True, color="000000", size=22)

    # === Alignment Styles ===
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # === Border Styles ===
    no_border = Border()
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    black_colored_border = Side(style="thin", color="000000")
    white_colored_border = Side(style="thin", color="FFFFFF")
    thick_black_colored_border = Side(style="thick", color="000000")

    right_white_rest_black_border = Border(left=black_colored_border, top=black_colored_border, right=white_colored_border, bottom=black_colored_border)
    left_right_white_rest_black_border = Border(left=white_colored_border, top=black_colored_border, right=white_colored_border, bottom=black_colored_border)
    left_white_rest_black_border = Border(left=white_colored_border, top=black_colored_border, right=black_colored_border, bottom=black_colored_border)

    thick_black_border = Border(left=thick_black_colored_border, top=thick_black_colored_border, right=thick_black_colored_border, bottom=thick_black_colored_border)
    
    # === Insert GE Logo ===
    img = Image('ge_new_logo.png')
    img.height = int(0.56 * 96)
    img.width = int(2.32 * 96)
    ws.add_image(img, 'B1')

    # === Title Row ===
    ws.merge_cells("B1:H2")
    ws["B1"].border = no_border
    ws["B1"].value = "FINAL INVOICE"
    ws["B1"].font = title_font
    ws["B1"].alignment = center
 
    ws.merge_cells("B3:C3")
    ws["B3"].value = "Customer:"
    ws["B3"].fill = blue_fill 
    ws["B3"].font = white_bold_font
    ws["B3"].alignment = left
    for row in ws["B3:C3"]:
        for cell in row:
            cell.border = right_white_rest_black_border

    ws["D3"].value = "Contract #:"
    ws["D3"].fill = blue_fill 
    ws["D3"].font = white_bold_font
    ws["D3"].alignment = center
    ws["D3"].border = left_right_white_rest_black_border

    ws["E3"].value = "Engine Serial #:"
    ws["E3"].fill = blue_fill 
    ws["E3"].font = white_bold_font
    ws["E3"].alignment = center
    ws["E3"].border = left_right_white_rest_black_border

    ws.merge_cells("F3:H3")
    ws["F3"].value = "Notification #:"
    ws["F3"].fill = blue_fill 
    ws["F3"].font = white_bold_font
    ws["F3"].alignment = left
    for row in ws["F3:H3"]:
        for cell in row:
            cell.border = left_white_rest_black_border

    ws.merge_cells("B4:C4")
    ws["B4"].value = ""
    ws["B4"].font = norma_black_font
    ws["B4"].alignment = left
    for row in ws["B4:C4"]:
        for cell in row:
            cell.border = border

    ws["D4"].value = "" 
    ws["D4"].font = norma_black_font
    ws["D4"].alignment = center
    ws["D4"].border = border

    ws["E4"].value = "" 
    ws["E4"].font = norma_black_font
    ws["E4"].alignment = center
    ws["E4"].border = border

    ws.merge_cells("F4:H4")
    ws["F4"].value = "TBD"
    ws["F4"].font = norma_black_font
    ws["F4"].alignment = left
    for row in ws["F4:H4"]:
        for cell in row:
            cell.border = border

    ws.merge_cells("B5:D5")
    ws["B5"].value = "Customer Order #:"
    ws["B5"].fill = blue_fill 
    ws["B5"].font = white_bold_font
    ws["B5"].alignment = left
    for row in ws["B5:C5"]:
        for cell in row:
            cell.border = right_white_rest_black_border

    ws["E5"].value = "Invoice #" 
    ws["E5"].fill = blue_fill
    ws["E5"].font = white_bold_font
    ws["E5"].alignment = left
    ws["E5"].border = right_white_rest_black_border

    ws.merge_cells("F5:H5")
    ws["F5"].fill = blue_fill
    ws["F5"].value = "Invoice Date:"
    ws["F5"].font = white_bold_font
    ws["F5"].alignment = left
    for row in ws["F5:H5"]:
        for cell in row:
            cell.border = left_white_rest_black_border

    ws.merge_cells("B6:D6")
    ws["B6"].value = ""
    ws["B6"].font = norma_black_font
    ws["B6"].alignment = left
    for row in ws["B6:D6"]:
        for cell in row:
            cell.border = border

    ws["E6"].value = "" 
    ws["E6"].font = norma_black_font
    ws["E6"].alignment = left
    ws["E6"].border = border

    ws.merge_cells("F6:H6")
    ws["F6"].value = ""
    ws["F6"].font = norma_black_font
    ws["F6"].alignment = left
    for row in ws["F6:H6"]:
        for cell in row:
            cell.border = border

    ws.merge_cells("B7:D7")
    ws["B7"].value = "Work Performed At: "
    ws["B7"].fill = blue_fill 
    ws["B7"].font = white_bold_font
    ws["B7"].alignment = left
    for row in ws["B7:D7"]:
        for cell in row:
            cell.border = right_white_rest_black_border

    ws["E7"].value = "Foreign Affiliate of:" 
    ws["E7"].fill = blue_fill
    ws["E7"].font = white_bold_font
    ws["E7"].alignment = left
    ws["E7"].border = left_right_white_rest_black_border

    ws.merge_cells("F7:H7")
    ws["F7"].value = "Send payment to:"
    ws["F7"].fill = blue_fill 
    ws["F7"].font = white_bold_font
    ws["F7"].alignment = left
    for row in ws["F7:H7"]:
        for cell in row:
            cell.border = left_white_rest_black_border

    ws.merge_cells("B8:D8")
    ws["B8"].value = "" #f"GE Celma Ltda\n356 Rua Alice Herve\nPetropolis - RJ\n25669-900BR\nTel No. - 55 24 2233-4000"
    ws["B8"].font = norma_black_font
    ws["B8"].alignment = left
    for row in ws["B8:D8"]:
        for cell in row:
            cell.border = border

    ws["E8"].value = "" #f"GE Engine Services, LLC\n1 Neumann Way\nCincinnati , OH 45215"
    ws["E8"].font = norma_black_font
    ws["E8"].alignment = left
    ws["E8"].border = border

    ws.merge_cells("F8:H8")
    ws["F8"].value = "" #"GE Engine Services, LLC, C/o PNC BANK\nNATIONAL ASSOCIATION\nPITTSBURGH, PA\nS.W.I.F.T. : PNCCUS33 ABA # 043000096\nACCOUNT : 1010933861"
    ws["F8"].font = norma_black_font
    ws["F8"].alignment = left
    for row in ws["F8:H8"]:
        for cell in row:
            cell.border = border

    ws.merge_cells("B9:D9")
    ws["B9"].value = "Shipping Address:"
    ws["B9"].fill = blue_fill 
    ws["B9"].font = white_bold_font
    ws["B9"].alignment = left
    for row in ws["B9:D9"]:
        for cell in row:
            cell.border = left_white_rest_black_border

    ws.merge_cells("E9:H9")
    ws["E9"].value = "Billing Address:"
    ws["E9"].fill = blue_fill 
    ws["E9"].font = white_bold_font
    ws["E9"].alignment = left
    for row in ws["E9:H9"]:
        for cell in row:
            cell.border = left_white_rest_black_border

    ws.merge_cells("B10:D10")
    ws["B10"].value = ""
    ws["B10"].font = norma_black_font
    ws["B10"].alignment = left
    for row in ws["B10:D10"]:
        for cell in row:
            cell.border = border

    ws.merge_cells("E10:H10")
    ws["E10"].value = ""
    ws["E10"].font = norma_black_font
    ws["E10"].alignment = left
    for row in ws["E10:H10"]:
        for cell in row:
            cell.border = border

    ws["B11"].value = "Shop Visit Induction Date:"
    ws["B11"].fill = blue_fill
    ws["B11"].font = white_bold_font
    ws["B11"].alignment = left
    ws["B11"].border = right_white_rest_black_border

    ws["C11"].value = "Shipping Reference #:"
    ws["C11"].fill = blue_fill
    ws["C11"].font = white_bold_font
    ws["C11"].alignment = left
    ws["C11"].border = left_right_white_rest_black_border

    ws["D11"].value = "Shipped Date:"
    ws["D11"].fill = blue_fill
    ws["D11"].font = white_bold_font
    ws["D11"].alignment = left
    ws["D11"].border = left_right_white_rest_black_border

    ws.merge_cells("E11:F11")
    ws["E11"].value = "Payment Terms:"
    ws["E11"].fill = blue_fill 
    ws["E11"].font = white_bold_font
    ws["E11"].alignment = left
    for row in ws["E11:F11"]:
        for cell in row:
            cell.border = left_right_white_rest_black_border

    ws.merge_cells("G11:H11")
    ws["G11"].value = "Payment Due Date:"
    ws["G11"].fill = blue_fill 
    ws["G11"].font = white_bold_font
    ws["G11"].alignment = left
    for row in ws["G11:H11"]:
        for cell in row:
            cell.border = left_white_rest_black_border

    ws["B12"].value = "TBD"
    ws["B12"].font = norma_black_font
    ws["B12"].alignment = left
    ws["B12"].border = border

    ws["C12"].value = "TBD"
    ws["C12"].font = norma_black_font
    ws["C12"].alignment = left
    ws["C12"].border = border

    ws["D12"].value = "TBD"
    ws["D12"].font = norma_black_font
    ws["D12"].alignment = left
    ws["D12"].border = border

    ws.merge_cells("E12:F12")
    ws["E12"].value = "TBD"
    ws["E12"].font = norma_black_font
    ws["E12"].alignment = left
    for row in ws["E12:F12"]:
        for cell in row:
            cell.border = border

    ws.merge_cells("G12:H12")
    ws["G12"].value = "TBD"
    ws["G12"].font = norma_black_font
    ws["G12"].alignment = left
    for row in ws["G12:H12"]:
        for cell in row:
            cell.border = border

    # Leave row 13 empty
    ws.append([])
 
    # === Section 2: Final Cost Breakdown ===
    ws.merge_cells("B14:H14")
    ws["B14"].value = "Final Cost Breakdown"
    ws["B14"].fill = blue_fill
    ws["B14"].font = white_bold_font
    ws["B14"].alignment = center
    for row in ws["GB14:H14"]:
        for cell in row:
            cell.border = border

    # Leave row 15 empty
    ws.append([])

    ws.merge_cells("B16:G16")
    ws["B16"].value = f"FFP Price for workscope {engine_workscope}"
    ws["B16"].fill = blue_fill
    ws["B16"].font = white_bold_font
    ws["B16"].alignment = left
    for row in ws["B16:G16"]:
        for cell in row:
            cell.border = right_white_rest_black_border

    ws["H16"].value = "='Contract Pricing Terms'!E6"
    ws["H16"].fill = blue_fill
    ws["H16"].font = white_bold_font
    ws["H16"].number_format =  '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    ws.merge_cells("B17:H17")
    ws["B17"].value = ""
    for row in ws["B17:H17"]:
        for cell in row:
            cell.border = border

    ws.merge_cells("B18:G18")
    ws["B18"].value = "Over and Above (exclusions)"
    ws["B18"].fill = blue_fill
    ws["B18"].font = white_bold_font
    ws["B18"].alignment = left
    for row in ws["B18:G18"]:
        for cell in row:
            cell.border = right_white_rest_black_border

    ws["H18"].value = ""
    ws["H18"].fill = blue_fill
    ws["H18"].font = white_bold_font
    ws["H18"].border = left_white_rest_black_border

    ws.merge_cells("B19:G19")
    ws["B19"].value = "FPLS & Labor"
    ws["B19"].font = norma_black_font
    ws["B19"].alignment = left
    for row in ws["B19:G19"]:
        for cell in row:
            cell.border = border

    ws["H19"].value = ""
    ws["H19"].font = norma_black_font
    ws["H19"].border = border
    ws["H19"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    ws.merge_cells("B20:G20")
    ws["B20"].value = "Material"
    ws["B20"].font = norma_black_font
    ws["B20"].alignment = left
    for row in ws["B20:G20"]:
        for cell in row:
            cell.border = border

    ws["H20"].value = ""
    ws["H20"].font = norma_black_font
    ws["H20"].border = border
    ws["H20"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    ws.merge_cells("B21:G21")
    ws["B21"].value = "GE Repair"
    ws["B21"].font = norma_black_font
    ws["B21"].alignment = left
    for row in ws["B21:G21"]:
        for cell in row:
            cell.border = border

    ws["H21"].value = ""
    ws["H21"].font = norma_black_font
    ws["H21"].border = border
    ws["H21"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    ws.merge_cells("B22:G22")
    ws["B22"].value = "Sub Contracted Repair"
    ws["B22"].font = norma_black_font
    ws["B22"].alignment = left
    for row in ws["B22:G22"]:
        for cell in row:
            cell.border = border

    ws["H22"].value = ""
    ws["H22"].font = norma_black_font
    ws["H22"].border = border
    ws["H22"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    ws.merge_cells("B23:G23")
    ws["B23"].value = "QEC - LRU Repair"
    ws["B23"].font = norma_black_font
    ws["B23"].alignment = left
    for row in ws["B23:G23"]:
        for cell in row:
            cell.border = border

    ws["H23"].value = ""
    ws["H23"].font = norma_black_font
    ws["H23"].border = border
    ws["H23"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    ws.merge_cells("B24:G24")
    ws["B24"].value = "QEC - LRU Material"
    ws["B24"].font = norma_black_font
    ws["B24"].alignment = left
    for row in ws["B24:G24"]:
        for cell in row:
            cell.border = border

    ws["H24"].value = ""
    ws["H24"].font = norma_black_font
    ws["H24"].border = border
    ws["H24"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    ws.merge_cells("B25:G25")
    ws["B25"].value = "Misc (test cell, storage, etc)"
    ws["B25"].font = norma_black_font
    ws["B25"].alignment = left
    for row in ws["B25:G25"]:
        for cell in row:
            cell.border = border

    ws["H25"].value = ""
    ws["H25"].font = norma_black_font
    ws["H25"].border = border
    ws["H25"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    # Scrap Repair
    ws.merge_cells("B26:G26")
    ws["B26"].value = "Scrap - Repair"
    ws["B26"].font = norma_black_font
    ws["B26"].alignment = left
    for row in ws["B26:G26"]:
        for cell in row:
            cell.border = border
 
    ws["H26"].value = ""
    ws["H26"].font = norma_black_font
    ws["H26"].border = border
    ws["H26"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    
    # Scrap Material
    ws.merge_cells("B27:G27")
    ws["B27"].value = "Scrap - Material"
    ws["B27"].font = norma_black_font
    ws["B27"].alignment = left
    for row in ws["B27:G27"]:
        for cell in row:
            cell.border = border
    
    ws["H27"].value = ""
    ws["H27"].font = norma_black_font
    ws["H27"].border = border
    ws["H27"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    ws.merge_cells("B28:G28")
    ws["B28"].value = "Over and Above (exclusions) Total"
    ws["B28"].fill = blue_fill
    ws["B28"].font = white_bold_font
    ws["B28"].alignment = left
    for row in ws["B28:G28"]:
        for cell in row:
            cell.border = right_white_rest_black_border

    ws["H28"].value = "=SUM(H19:H27)"
    ws["H28"].fill = blue_fill
    ws["H28"].font = white_bold_font
    ws["H28"].border = left_white_rest_black_border
    ws["H28"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    ws.merge_cells("B29:H30")
    ws["B29"].value = ""
    for row in ws["B29:H30"]:
        for cell in row:
            cell.border = border

    ws.merge_cells("B31:G31")
    ws["B31"].value = "Other"
    ws["B31"].fill = blue_fill
    ws["B31"].font = white_bold_font
    ws["B31"].alignment = left
    for row in ws["B31:G31"]:
        for cell in row:
            cell.border = right_white_rest_black_border

    ws["H31"].value = ""
    ws["H31"].fill = blue_fill
    ws["H31"].border = left_white_rest_black_border

    ws.merge_cells("B32:G32")
    ws["B32"].value = "TAT Penalty"
    ws["B32"].font = norma_black_font
    ws["B32"].alignment = left
    for row in ws["B32:G32"]:
        for cell in row:
            cell.border = border

    ws["H32"].value = ""
    ws["H32"].font = norma_black_font
    ws["H32"].border = border
    ws["H32"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    ws.merge_cells("B33:G33")
    ws["B33"].value = "EGT margin"
    ws["B33"].font = norma_black_font
    ws["B33"].alignment = left
    for row in ws["B33:G33"]:
        for cell in row:
            cell.border = border

    ws["H33"].value = ""
    ws["H33"].font = norma_black_font
    ws["H33"].border = border
    ws["H33"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    
    ws.merge_cells("B34:G34")
    ws["B34"].value = "LLP Service Credit (Escalated per schedule 3)"
    ws["B34"].font = norma_black_font
    ws["B34"].alignment = left
    for row in ws["B34:G34"]:
        for cell in row:
            cell.border = border
          
    ws["H34"].value = ""
    ws["H34"].font = norma_black_font
    ws["H34"].border = border
    ws["H34"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    
    row_idx = 35

    # Loop through dynamic creep data
    for module, price in adder_creep_data.items():
        print(module, price)
        ws.merge_cells(f"B{row_idx}:G{row_idx}")
        ws[f"B{row_idx}"].value = f"FFP Workscope Upgrade - {module} Escalation"
        ws[f"B{row_idx}"].font = norma_black_font
        ws[f"H{row_idx}"].value = price
        ws[f"B{row_idx}"].alignment = left
        ws[f"H{row_idx}"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
        ws[f"H{row_idx}"].font = norma_black_font
        for row in ws[f"B{row_idx}:G{row_idx}"]:
            for cell in row:
                cell.border = border
        ws[f"H{row_idx}"].border = border
        row_idx += 1  # Move to the next row for the next module or the next section  

    # Transportation Credit
    ws.merge_cells(f"B{row_idx}:G{row_idx}")
    ws[f"B{row_idx}"].value = "Transportation Credit"
    ws[f"B{row_idx}"].font = norma_black_font
    ws[f"B{row_idx}"].alignment = left
    for row in ws[f"B{row_idx}:G{row_idx}"]:
        for cell in row:
            cell.border = border
            
    ws[f"H{row_idx}"].value = ""
    ws[f"H{row_idx}"].border = border
    ws[f"H{row_idx}"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    row_idx += 1

    # Progress Invoice
    ws.merge_cells(f"B{row_idx}:G{row_idx}")
    ws[f"B{row_idx}"].value = "Transportation Credit"
    ws[f"B{row_idx}"].font = norma_black_font
    ws[f"B{row_idx}"].alignment = left
    for row in ws[f"B{row_idx}:G{row_idx}"]:
        for cell in row:
            cell.border = border
    ws[f"H{row_idx}"].value = ""
    ws[f"H{row_idx}"].border = border
    ws[f"H{row_idx}"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    row_idx += 1

    # Initial Invoice
    ws.merge_cells(f"B{row_idx}:G{row_idx}")
    ws[f"B{row_idx}"].value = "Progress Invoice"
    ws[f"B{row_idx}"].font = norma_black_font
    ws[f"B{row_idx}"].alignment = left
    for row in ws[f"B{row_idx}:G{row_idx}"]:
        for cell in row:
            cell.border = border
    ws[f"H{row_idx}"].value = ""
    ws[f"H{row_idx}"].border = border
    ws[f"H{row_idx}"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    row_idx += 1

    # Other Total (Summing from B31 to the row just before this one)
    ws.merge_cells(f"B{row_idx}:G{row_idx}")
    ws[f"B{row_idx}"].value = "Other Total"
    ws[f"B{row_idx}"].fill = blue_fill
    ws[f"B{row_idx}"].font = white_bold_font
    ws[f"B{row_idx}"].alignment = left
    for row in ws[f"B{row_idx}:G{row_idx}"]:
        for cell in row:
            cell.border = border
   
    ws[f"H{row_idx}"].value = f"=SUM(H31:H{row_idx-1})"
    ws[f"H{row_idx}"].fill = blue_fill
    ws[f"H{row_idx}"].font = white_bold_font
    ws[f"H{row_idx}"].border = left_white_rest_black_border
    ws[f"H{row_idx}"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    other_total_row = row_idx
    row_idx += 1

    # Spacer Row
    ws.merge_cells(f"B{row_idx}:G{row_idx}")
    for row in ws[f"B{row_idx}:G{row_idx}"]:
        for cell in row:
            cell.border = border
    ws[f"H{row_idx}"].border = border
    row_idx += 1

    # Grand Total
    ws.merge_cells(f"B{row_idx}:G{row_idx}")
    ws[f"B{row_idx}"].value = "Grand Total"
    ws[f"B{row_idx}"].fill = blue_fill
    ws[f"B{row_idx}"].font = white_bold_font
    ws[f"B{row_idx}"].alignment = left
    for row in ws[f"B{row_idx}:G{row_idx}"]:
        for cell in row:
            cell.border = right_white_rest_black_border

    ws[f"H{row_idx}"].value = f"=H16+H28+H{other_total_row}"
    ws[f"H{row_idx}"].fill = blue_fill
    ws[f"H{row_idx}"].font = white_bold_font
    ws[f"H{row_idx}"].border = left_white_rest_black_border
    ws[f"H{row_idx}"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    row_idx += 2 # Leave one empty row

    # Disclaimer
    ws.merge_cells(f"B{row_idx}:H{row_idx}")
    ws[f"B{row_idx}"].value = "***All Values are in US Dollars***"
    ws[f"B{row_idx}"].font = black_bold_font
    ws[f"B{row_idx}"].alignment = center
    for row in ws[f"B{row_idx}:H{row_idx}"]:
        for cell in row:
            cell.border = thick_black_border

    wb.save(output_path)
    # print(f"[INFO] ✅ COVER Sheet invoice saved at: {output_path}")
    
    


def apply_actual_price_formulas_table(ws, start_row, end_row, start_col, end_col, flag):
    # Define columns by flag
    if flag == 1:
        col_map = { "A": None, "A+": None, "B": None, "C": None, "Initial WKS": None, "Actual WKS": None, "Actual Price": None }
    elif flag == 0:
        col_map = { "A": None, "B": None, "C": None, "Initial WKS": None, "Actual WKS": None, "Actual Price": None }
    else:
        raise ValueError(f"❌ Invalid flag value provided: {flag}")
 
    # Normalize keys for matching
    col_map_lower = { key.lower().strip(): key for key in col_map.keys() }
 
    # Find matching columns
    for col in range(start_col, end_col + 1):
        header_value = ws.cell(row=start_row, column=col).value
        if header_value:
            header_clean = str(header_value).strip().lower()
            if header_clean in col_map_lower:
                mapped_key = col_map_lower[header_clean]
                col_map[mapped_key] = get_column_letter(col)
                print(f"[MAPPED] '{mapped_key}' → {col_map[mapped_key]}")
 
    # Validate found columns
    missing = [key for key, val in col_map.items() if val is None]
    if missing:
        raise ValueError(f"❌ Could not find these required columns: {missing}")
 
    # print(f"[INFO] Final column mapping: {col_map}")
 
    # Apply formula row by row
    for row in range(start_row + 1, end_row + 1):  # skip header row
        if flag == 1:
            formula = (
                f'=IF({col_map["Actual WKS"]}{row}="A",{col_map["A"]}{row},'
                f'IF({col_map["Actual WKS"]}{row}="A+",{col_map["A+"]}{row},'
                f'IF({col_map["Actual WKS"]}{row}="B",{col_map["B"]}{row},'
                f'IF({col_map["Actual WKS"]}{row}="C",{col_map["C"]}{row},""))))'
            )
        elif flag == 0:
            formula = (
                f'=IF({col_map["Actual WKS"]}{row}="A",{col_map["A"]}{row},'
                f'IF({col_map["Actual WKS"]}{row}="B",{col_map["B"]}{row},'
                f'IF({col_map["Actual WKS"]}{row}="C",{col_map["C"]}{row},""))))'
            )
        ws[f'{col_map["Actual Price"]}{row}'] = formula
 
    print(f"[✅] Formulas applied to Actual Price column from rows {start_row + 1} to {end_row}")


def add_actual_price_total(ws, start_row, start_col, end_row, sum_col):
    # Find the Actual Price column
    actual_price_col = None

    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=start_row, column=col).value  # <=== use 'col' here!
        if header and str(header).strip().lower() == sum_col:
            actual_price_col = get_column_letter(col)
            break
 
    if actual_price_col is None:
        raise ValueError("❌ 'Actual Price' column not found.")
 
    # Define the total row (one after last data row)
    total_row = end_row + 1
 
    # Create the SUM formula
    sum_formula = f"=SUM({actual_price_col}{start_row + 1}:{actual_price_col}{end_row})"
    
    total_sum = 0
    #for row in range(start_row + 1, end_row + 1):
    #    cell_value = ws[f"{actual_price_col}{row}"].value
    #    if cell_value is not None:
    #        total_sum += cell_value

    #print(total_sum)
    # Write the formula into the first empty cell below the Actual Price column
    target_cell = ws[f"{actual_price_col}{total_row}"]
    target_cell.value = sum_formula
    target_cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    total_sum=ws["J25"].value
    print(total_sum)
 
    # Styling
    target_cell.font = Font(bold=True, color="FFFFFF", size=11)
    target_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    target_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    target_cell.alignment = Alignment(horizontal="center", vertical="center")
    # print(f"[✅] Total formula written in {actual_price_col}{total_row}: {sum_formula}")
    return actual_price_col, total_row


def fill_inscope_sheet(invoice_path, inscope_sheet_name, inscope_df_info, year_str):

    
    #df_index = inscope_df_info["df_index"]
    start_row = inscope_df_info["row_start"]
    end_row = inscope_df_info["row_end"]
    num_rows = inscope_df_info["num_rows"]
    num_cols = inscope_df_info["num_cols"]
    start_col = inscope_df_info["col_start"]
    end_col = inscope_df_info["col_end"]

    #df_index_1, df_index_2 = inscope_df_info[0]["df_index"], inscope_df_info[1]["df_index"]

    #start_row_1, start_row_2 = inscope_df_info[0]["row_start"], inscope_df_info[1]["row_start"]
    #end_row_1, end_row_2 = inscope_df_info[0]["row_end"], inscope_df_info[1]["row_end"]
    
    #num_rows_1, num_rows_2 = inscope_df_info[0]["num_rows"], inscope_df_info[1]["num_rows"]
    #num_cols_1, num_cols_2 = inscope_df_info[0]["num_cols"], inscope_df_info[1]["num_cols"]
    
    #tart_col_1, start_col_2 = inscope_df_info[0]["col_start"], inscope_df_info[1]["col_start"]
    #end_col_1, end_col_2 = inscope_df_info[0]["col_end"], inscope_df_info[1]["col_end"]
    
    wb = load_workbook(invoice_path)
    ws = wb[inscope_sheet_name]

    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
   
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "8DC858"
 
    # Define styles
    title_bold_black_font = Font(bold=True, color="000000", size=12)
    norma_black_font = Font(bold=False, color="000000", size=10)
    white_bold_font = Font(bold=True, color="FFFFFF", size=11)
    blue_fill = PatternFill(start_color="050543", end_color="050543", fill_type="solid") 
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    no_border = Border() 
    top_bottom_border = Border(top=Side(style='thin'), bottom=Side(style='thin')) 
    top_bottom_left_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))
    top_bottom_right_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin')) 
    right_border = Border(right = Side(style='thin'))
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
 
    # Insert logo
    img = Image('ge_new_logo.png')
    img.height = int(0.54 * 96)
    img.width = int(2.31 * 96)
    ws.add_image(img, 'B1')
     
    # # # Table-1 Values
    for row in ws.iter_rows(min_row=start_row+2, max_row=end_row, min_col = 2, max_col=num_cols):
        for cell in row:
            cell.font = norma_black_font
            cell.alignment = center
            cell.border = border
            cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    _force_text_for_headers(ws, header_row=start_row+1, data_start=start_row+2, data_end=end_row, headers=["Initial WKS", "Actual WKS"])    
    # Actual Price Formula - 1                      5+1 = 6                23                   1                        10
    apply_actual_price_formulas_table(ws, start_row=start_row+1, end_row=end_row, start_col=start_col, end_col=end_col, flag=1) 
    
    # Sub-total 1
    actual_price_col_1, row_no_1 = add_actual_price_total(ws, start_row=start_row+1, start_col=2, end_row=end_row, sum_col="actual price") 
            
    # # # Table-2 Values
    """"
        #     #for row in ws.iter_rows(min_row=start_row_2+2, max_row=end_row_2, min_col = 2, max_col=num_cols_2):
        for cell in row:
            cell.font = norma_black_font
            cell.alignment = center
            cell.border = border
            cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'          
    
    _force_text_for_headers(ws, header_row=start_row_2+1, data_start=start_row_2+2, data_end=end_row_2, headers=["Initial WKS", "Actual WKS"])    

    # Sub-total 2
    actual_price_col_2, row_no_2 = add_actual_price_total(ws, start_row=start_row_2+1, start_col=2, end_row=end_row_2, sum_col="actual price")
    """
    # Set header values
    ws["E2"].value = "United Pricing Workbook"
    ws["E2"].alignment = Alignment(horizontal="center", vertical="center")#, wrap_text=False)
    ws["E2"].font = title_bold_black_font

    ws["I2"].value = "Total"
    ws["I2"].fill = blue_fill
    ws["I2"].font = white_bold_font
    ws["I2"].alignment = center
    ws["I2"].border = top_bottom_left_border
     
    # Section headers - 1
    #ws.merge_cells("B5:C5")
    #ws["B5"].value = "Minor Modules"
    #ws["B5"].fill = blue_fill
    #ws["B5"].font = white_bold_font
    #ws["B5"].alignment = center
    #ws["B5"].border = border

    ws.merge_cells("D4:G4")
    ws["D5"].value = f"Repair Level FFP {year_str}"
    ws["D5"].fill = blue_fill
    ws["D5"].font = white_bold_font
    ws["D5"].alignment = center
    ws["D5"].border = border

    ws["B5"].value = "Minor Module"
    ws["B5"].fill = blue_fill
    ws["B5"].font = white_bold_font
    ws["B5"].alignment = center
    ws["B5"].border = border

    ws["C5"].value = "Description"
    ws["C5"].fill = blue_fill
    ws["C5"].font = white_bold_font
    ws["C5"].alignment = center
    ws["C5"].border = border

    ws["D5"].value = "A"
    ws["D5"].fill = blue_fill
    ws["D5"].font = white_bold_font
    ws["D5"].alignment = center
    ws["D5"].border = border

    ws["E5"].value = "A+"
    ws["E5"].fill = blue_fill
    ws["E5"].font = white_bold_font
    ws["E5"].alignment = center
    ws["E5"].border = border

    ws["F5"].value = "B"
    ws["F5"].fill = blue_fill
    ws["F5"].font = white_bold_font
    ws["F5"].alignment = center
    ws["F5"].border = border

    ws["G5"].value = "C"
    ws["G5"].fill = blue_fill
    ws["G5"].font = white_bold_font
    ws["G5"].alignment = center
    ws["G5"].border = border

    ws["H5"].value = "Initial WKS"
    ws["H5"].fill = blue_fill
    ws["H5"].font = white_bold_font
    ws["H5"].alignment = center
    ws["H5"].border = border

    ws["I5"].value = "Actual WKS"
    ws["I5"].fill = blue_fill
    ws["I5"].font = white_bold_font
    ws["I5"].alignment = center
    ws["I5"].border = border

    ws["J5"].value = "Actual Price"
    ws["J5"].fill = blue_fill
    ws["J5"].font = white_bold_font
    ws["J5"].alignment = center
    ws["J5"].border = border

    #ws.merge_cells("B26:C26")
    #ws["B26"].value = "Minor Modules"
    #ws["B26"].fill = blue_fill
    #ws["B26"].font = white_bold_font
    #ws["B26"].alignment = center
    #ws["B26"].border = border

    #ws.merge_cells("D26:F26")
    #ws["D26"].value = f"Repair Level FFP {year_str}"
    #ws["D26"].fill = blue_fill
    #ws["D26"].font = white_bold_font
    #ws["D26"].alignment = center
    #ws["D26"].border = border

    #ws["B27"].value = "Module Identifier"
    #ws["B27"].fill = blue_fill
    #ws["B27"].font = white_bold_font
    #ws["B27"].alignment = center
    #ws["B27"].border = border

    #ws["C27"].value = "Description"
    #ws["C27"].fill = blue_fill
    #ws["C27"].font = white_bold_font
    #ws["C27"].alignment = center
    #ws["C27"].border = border

    #ws["D27"].value = "A"
    #ws["D27"].fill = blue_fill
    #ws["D27"].font = white_bold_font
    #ws["D27"].alignment = center
    #ws["D27"].border = border

    #ws["E27"].value = "B"
    #ws["E27"].fill = blue_fill
    #ws["E27"].font = white_bold_font
    #ws["E27"].alignment = center
    #ws["E27"].border = border

    #ws["F27"].value = "C"
    #ws["F27"].fill = blue_fill
    #ws["F27"].font = white_bold_font
    #ws["F27"].alignment = center
    #ws["F27"].border = border

    #w#s["G27"].value = "Initial WKS"
    #ws["G27"].fill = blue_fill
    #ws["G27"].font = white_bold_font
    #ws["G27"].alignment = center
    #ws["G27"].border = border

    #ws["H27"].value = "Actual WKS"
    #ws["H27"].fill = blue_fill
    #ws["H27"].font = white_bold_font
    #w#s["H27"].alignment = center
    #ws["H27"].border = border

    #ws["I27"].value = "Actual Price"
    #ws["I27"].fill = blue_fill
    #ws["I27"].font = white_bold_font
    #ws["I27"].alignment = center
    #ws["I27"].border = border

    
    #ws["I28"] = '=IF(H28="A",D28,(IF(H28="B",E28,(IF(H28="C",F28,0)))))'
    #ws["I28"].font = norma_black_font
    #ws["I28"].alignment = center
    #ws["I28"].border = border

    #ws["I29"] = '=IF(H29="A",D29,(IF(H29="B",E29,(IF(H29="C",F29,0)))))'
    #ws["I29"].font = norma_black_font
    #ws["I29"].alignment = center
    #ws["I29"].border = border

    sub_total_1 = f"{actual_price_col_1}{row_no_1}"
    #sub_total_2 = f"{actual_price_col_1}{row_no_1}"

    ws["J2"] = '=J25'
    ws["J2"].number_format = '$#,##0.00' 
    ws["J2"].fill = blue_fill
    ws["J2"].font = white_bold_font
    ws["J2"].alignment = Alignment(horizontal="right", vertical="center")
    ws["J2"].border = top_bottom_right_border

    wb.save(invoice_path)
    # print(f"[INFO] ✅ In-Scope sheet filled and saved to {invoice_path}")


def is_blank_row(ws, row_idx):
    """Check if a row is blank."""
    return all(
        (cell.value is None or str(cell.value).strip() == '')
        for cell in ws[row_idx]
    )


def delete_empty_rows_and_return_metadata(ws, row_ranges):
 
    def is_blank_row(ws, row_idx):
        return all(
            (cell.value is None or str(cell.value).strip() == '')
            for cell in ws[row_idx]
        )
 
    rows_to_delete = []
    metadata = []
 
    for i, (start_row, end_row) in enumerate(row_ranges, start=1):
        data_rows = []
 
        # Collect only non-blank rows in the range
        for row in range(start_row, end_row + 1):
            if not is_blank_row(ws, row):
                data_rows.append(row)
            else:
                rows_to_delete.append(row)  # queue for deletion
 
        if not data_rows:
            # No data found in range
            metadata.append({
                'df_index': i,
                'start_row': None,
                'end_row': None,
                'num_rows': 0,
                'num_cols': 0,
                'start_col': None,
                'end_col': None
            })
            continue
 
        actual_start_row = min(data_rows)
        actual_end_row = max(data_rows)
 
        # Determine min/max columns
        actual_start_col, actual_end_col = None, None
        num_cols = 0
 
        for row in data_rows:
            row_cells = ws[row]
            non_blank_indices = [
                cell.column for cell in row_cells
                if cell.value is not None and str(cell.value).strip() != ''
            ]
            if non_blank_indices:
                min_col = min(non_blank_indices)
                max_col = max(non_blank_indices)
                if actual_start_col is None or min_col < actual_start_col:
                    actual_start_col = min_col
                if actual_end_col is None or max_col > actual_end_col:
                    actual_end_col = max_col
                if max_col - min_col + 1 > num_cols:
                    num_cols = max_col - min_col + 1
 
        num_rows = actual_end_row - actual_start_row + 1
 
        metadata.append({
            'df_index': i,
            'start_row': actual_start_row,
            'end_row': actual_end_row,
            'num_rows': num_rows,
            'num_cols': num_cols,
            'start_col': actual_start_col,
            'end_col': actual_end_col
        })
 
    # Delete all blank rows in reverse order
    for row_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_idx)
 
    return metadata

def apply_arithmetic_formula(
    ws,
    start_row,
    start_col,
    end_col,
    formula_str,
    input_col_headers,
    mode='horizontal',
    output_col_name=None,
    output_cell=None):

    # Step 1: Extract header mapping
    header_map = {}
    for col in range(start_col, end_col + 1):
        cell_val = ws.cell(row=start_row, column=col).value
        if cell_val:
            normalized = str(cell_val).strip().lower().replace(" ", "_")
            # print(normalized)
            header_map[normalized] = col
            print(header_map)
 
    # Step 2: Validate required columns
    missing = [hdr for hdr in input_col_headers if hdr.lower() not in header_map]

    if missing:
        raise ValueError(f"Missing required columns: {missing}")
 
    # Step 3: Row-wise Excel formula (horizontal mode)
    if mode == 'horizontal':
        output_col_index = header_map.get(output_col_name.lower())
        if output_col_index is None:
            output_col_index = end_col + 1
            ws.cell(row=start_row, column=output_col_index, value=output_col_name)
 
        # Determine number format
        col_name = output_col_name.lower()
        if any(key in col_name for key in ['qty']):
            # print("Updating int format")             
            number_format = '0'       
        elif any(key in col_name for key in ['discount %', 'handling fee %', 'mark up']):             
            # print("Updating Perce format")             
            number_format = '0.00%'         
        elif any(key in col_name for key in ['total price', 'labor', 'clp', 'extended price', 'unit price', 'handling fee price', 'vendor invoice amount', 'material', 'repair', 'sb total']):             
            # print("Updating Dollar format")             
            number_format = '$#,##0.00'          
        else:             
            number_format = None  # Leave default
 
        row = start_row + 1
        while ws.cell(row=row, column=start_col).value is not None:
            try:
                # Construct Excel formula dynamically
                excel_formula = formula_str
                for hdr in input_col_headers:
                    col_idx = header_map[hdr.lower()]
                    col_letter = get_column_letter(col_idx)
                    excel_formula = excel_formula.replace(hdr.lower(), f"{col_letter}{row}")
 
                # Write Excel formula to cell
                target_cell = ws.cell(row=row, column=output_col_index)
                target_cell.value = f"={excel_formula}"
                if number_format:
                    target_cell.number_format = number_format
 
            except Exception as e:
                ws.cell(row=row, column=output_col_index, value=f"#ERR: {e}")
            row += 1
 
    # Step 4: Column-wise (vertical mode)
    elif mode == 'vertical':
        target_col_idx = header_map[input_col_headers[0].lower()]
        values = []
        row = start_row + 1
        while ws.cell(row=row, column=target_col_idx).value is not None:
            val = ws.cell(row=row, column=target_col_idx).value
            try:
                values.append(float(val))
            except:
                pass
            row += 1
 
        try:
            result = eval(formula_str, {}, {input_col_headers[0].lower(): values})
            ws[output_cell] = result
        except Exception as e:
            if isinstance(output_cell, str):
                ws[output_cell] = f"ERR: {e}"
            else:
                ws.cell(row=output_cell.row, column=output_cell.column, value=f"#ERR: {e}")
 
    else:
        raise ValueError("Mode must be either 'horizontal' or 'vertical'")

def apply_number_format_by_column_name(ws, header_row, start_row, end_row, target_col_name): 
    # Normalize and find the column index
    target_col_index = None
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and str(val).strip().lower() == target_col_name.strip().lower():
            target_col_index = col
            break
 
    if not target_col_index:
        raise ValueError(f"Column '{target_col_name}' not found in header row {header_row}")
 
    name = target_col_name.lower()

    # Determine format type
    if any(keyword in name for keyword in ['price', 'cost', 'fee', 'handling fee price', 'unit price', 'clp unit price', 'uai unit price', 'clp', 'discount $', 'excluded [o&a] total price', 'labor', 'material', 'repair', 'sb total']):
        number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    elif any(keyword in name for keyword in ["discount %", 'rate', 'tax', 'pro-rata %', 'mark up', "handling fee %"]):
        number_format = '0%_);(0%);"-"??_);@'
    elif any(keyword in name for keyword in ['po/repair tag', 'qty', 'has cust', 'quantity',  'sb category', 'total life cicles', 'life cicles remaining', "item (sd)", "service order equipment / mrb", "po #"]):
        number_format = '0'
    elif any(keyword in name for keyword in ["hour / crd"]):
        number_format = '0.0'

    else:
        number_format = None  # Skip if unrecognized
 
    # Apply formatting
    if number_format:
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=target_col_index)
 
            # Handle percentage normalization
            if 'discount' in name or 'handling fee %' in name or 'tax' in name or 'pro-rata' in name:
                if isinstance(cell.value, (int, float)):
                    if cell.value > 1:
                        cell.value = cell.value / 100
                elif isinstance(cell.value, str) and '%' in cell.value:
                    try:
                        cell.value = float(cell.value.strip('%')) / 100
                    except:
                        pass  # Leave unchanged if conversion fails
 
            cell.number_format = number_format


# ------------------ Reusable Helpers ------------------ #
def set_style(cell, font=None, fill=None, align=None, border=None, number_format=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border
    if number_format: cell.number_format = number_format
 
def style_range(ws, start_row, end_row, start_col, end_col, font=None, fill=None, align=None, border=None, number_format=None):
    """Apply styles to a rectangular range"""
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            set_style(cell, font, fill, align, border, number_format)
 
def insert_logo(ws, position="A1", path="ge_logo.png"):
    img = Image(path)
    img.height, img.width = int(0.54 * 96), int(0.67 * 96)
    ws.add_image(img, position)
 
def insert_subtotal(ws, row, col_idx, start_row, end_row, font, fill, border, number_format):
    """Insert a SUM formula in subtotal cell"""
    col_letter = get_column_letter(col_idx)
    cell = ws.cell(row=row, column=col_idx)
    cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
    set_style(cell, font=font, fill=fill, align=Alignment(horizontal="center"), border=border, number_format=number_format)
    return cell
 
# ------------------ Main Function ------------------ #
def fill_llp_replacement_sheet(invoice_path, sheet_name, llp_info, year_str):
    print("llp_info :", llp_info)

    wb = load_workbook(invoice_path)
    ws = wb[sheet_name]
 
    # ---- Setup Sheet ----
    col_widths = [9, 12, 36, 18, 7, 12, 9, 12, 12, 15, 12, 9, 15]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "8DC858"
 
    # ---- Styles ----
    title_font = Font(bold=True, color="000000", size=12)
    normal_font = Font(bold=False, color="000000", size=10)
    white_bold = Font(bold=True, color="FFFFFF", size=11)
    blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
 
    # ---- Logo at Top ----
    insert_logo(ws, "A1")

    # Formats
    formats = {
        "Qty": '0',
        "Total Life Cicles": '0',
        "Life Cicles Remaining": '0',
        "Discount": '0%',
        "Pro-rata %": '0%',
    }
 
    # ---- Table 1 ----
    t1 = llp_info[0]
    start_row_1, end_row_1 = t1["start_row"], t1["end_row"]
    start_col_1, end_col_1 = t1["start_col"], t1["end_col"]
 
    style_range(ws, start_row_1, start_row_1, 2, t1["num_cols"]+1,
                fill=blue_fill, font=white_bold, align=Alignment(horizontal="center", vertical="center"),
                border=border, number_format='_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)')


    # Loop through columns and apply number format based on header
    data_row_start = start_row_1 + 1
    data_row_end = end_row_1
    
    for col in range(2, t1["num_cols"] + 2):  # +2 because range end is exclusive and index starts from 1
        header_cell = ws.cell(row=data_row_start - 1, column=col)
        header_value = header_cell.value.strip() if header_cell.value else ""
    
        # Determine number format
        number_format = formats.get(header_value, '_($* #,##0_);_($* (#,##0);_(@_)')  # Default to currency
    
        # Apply styling per cell in the column
        for row in range(data_row_start, data_row_end + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = normal_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            cell.number_format = number_format

    # Row-wise total_price
    apply_arithmetic_formula(
        ws=ws,
        start_row=start_row_1,
        start_col=start_col_1,
        end_col=end_col_1,
        formula_str="ual_unit_price - (ual_unit_price * discount)",
        input_col_headers=["ual_unit_price", "discount"],
        mode="horizontal",
        output_col_name="total_price"
    )
 
    # Column-wise sum for Table 1
    sub_total_1_cell = insert_subtotal(ws, end_row_1+1, 8, start_row_1+1, end_row_1,
                                       font=white_bold, fill=blue_fill, border=border,
                                       number_format='_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)')
 
    # ---- Table 2 ----
    t2 = llp_info[1]
    start_row_2, end_row_2 = t2["start_row"], t2["end_row"]
    start_col_2, end_col_2 = t2["start_col"], t2["end_col"]
    num_rows_2 = t2["num_rows"]
 
    if start_row_2 == end_row_2:  # Empty case
        fill_empty_df_headers(ws, start_row_2-1, start_col_2,
                              pd.DataFrame(columns=["Module", "Part Name", "Part Number", "Qty",
                                                    "CLP Unit Price", "Discount", "Total Life Cicles",
                                                    "Life Cicles Remaining", "Pro-rata %", "Total Price",
                                                    "Status", "Removal Cause", "Exclusion Remarks"]),
                              white_bold, blue_fill, Alignment(horizontal="center"), border,
                              number_format_cols={"qty": "0", "clp_unit_price": "_($* #,##0_);_($* (#,##0);_(@_)",
                                                  "discount": "0%", "total_life_cicles": "0",
                                                  "life_cicles_remaining": "0", "pro-rata_%": "0%",
                                                  "total_price": "_($* #,##0_);_($* (#,##0);_(@_)"})
        sub_total_2_cell = ws.cell(row=start_row_2+1, column=start_col_2+9)
        set_style(sub_total_2_cell, font=white_bold, fill=blue_fill,
                  align=Alignment(horizontal="center"), border=border,
                  number_format='_($* #,##0_);_($* (#,##0);_(@_)')
        sub_total_2_cell.value = 0
 
    elif num_rows_2 == 1:  # One row 
        header_row = start_row_2
        data_row_start = end_row_2
        data_row_end = end_row_2
 
        # Style headers and data
        style_range(ws, header_row, header_row, 2, t2["num_cols"]+1,
                    font=white_bold, fill=blue_fill, align=Alignment(horizontal="center"), border=border)

        # Loop through columns and apply number format based on header
        for col in range(2, t2["num_cols"] + 2):  # +2 because range end is exclusive and index starts from 1
            header_cell = ws.cell(row=data_row_start - 1, column=col)
            header_value = header_cell.value.strip() if header_cell.value else ""
        
            # Determine number format
            number_format = formats.get(header_value, '_($* #,##0_);_($* (#,##0);_(@_)')  # Default to currency
        
            # Apply styling per cell in the column
            for row in range(data_row_start, data_row_end + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = normal_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                cell.number_format = number_format


        # Row-wise formula
        apply_arithmetic_formula(
            ws=ws,
            start_row=header_row,
            start_col=start_col_2,
            end_col=end_col_2,
            formula_str="(life_cicles_remaining / total_life_cicles) * (pro-rata_%) * (clp_unit_price)",
            input_col_headers=["clp_unit_price", "total_life_cicles", "life_cicles_remaining", "pro-rata_%"],
            mode="horizontal",
            output_col_name="total_price"
        )
 
        # Column-wise subtotal
        sub_total_2_cell = insert_subtotal(ws, data_row_end+1, 11, data_row_start, data_row_end,
                                           font=white_bold, fill=blue_fill, border=border,
                                           number_format='_($* #,##0_);_($* (#,##0);_(@_)')

    
    else:  # multiple rows handled uniformly
        header_row = start_row_2-2
        data_row_start = start_row_2-1
        data_row_end = end_row_2-1 if num_rows_2 > 3 else start_row_2-1
 
        # Style headers and data
        style_range(ws, header_row, header_row, 2, t2["num_cols"]+1,
                    font=white_bold, fill=blue_fill, align=Alignment(horizontal="center"), border=border)

        # Loop through columns and apply number format based on header
        for col in range(2, t2["num_cols"] + 2):  # +2 because range end is exclusive and index starts from 1
            header_cell = ws.cell(row=data_row_start - 1, column=col)
            header_value = header_cell.value.strip() if header_cell.value else ""
        
            # Determine number format
            number_format = formats.get(header_value, '_($* #,##0_);_($* (#,##0);_(@_)')  # Default to currency
        
            # Apply styling per cell in the column
            for row in range(data_row_start, data_row_end + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = normal_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                cell.number_format = number_format


        # Row-wise formula
        apply_arithmetic_formula(
            ws=ws,
            start_row=header_row,
            start_col=start_col_2,
            end_col=end_col_2,
            formula_str="(life_cicles_remaining / total_life_cicles) * (pro-rata_%) * (clp_unit_price)",
            input_col_headers=["clp_unit_price", "total_life_cicles", "life_cicles_remaining", "pro-rata_%"],
            mode="horizontal",
            output_col_name="total_price"
        )
 
        # Column-wise subtotal
        sub_total_2_cell = insert_subtotal(ws, data_row_end+1, 11, data_row_start, data_row_end,
                                           font=white_bold, fill=blue_fill, border=border,
                                           number_format='_($* #,##0_);_($* (#,##0);_(@_)')
 
    # ---- Titles and Grand Total ----
    ws.cell(row=start_row_2-3, column=5, value="Out-of-Scope Used LLPs Replacement").font = title_font
    ws["E3"].value = "Material"
    ws["E3"].font = title_font
    ws["E3"].alignment = Alignment(horizontal="center")
 
    ws["I2"].value, ws["I2"].font = "Grand Total", white_bold
    ws["I2"].fill, ws["I2"].alignment = blue_fill, Alignment(horizontal="center")
    ws["I2"].border = Border(top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"))
 
    insert_logo(ws, "A22")  # Second logo
 
    ws["J2"].value = f"=SUM({sub_total_1_cell.coordinate},{sub_total_2_cell.coordinate})"
    set_style(ws["J2"], font=white_bold, fill=blue_fill,
              align=Alignment(horizontal="right"), border=Border(top=Side(style="thin"),
              bottom=Side(style="thin"), right=Side(style="thin")),
              number_format='_($* #,##0_);_($* (#,##0);_(@_)')
 
    wb.save(invoice_path)
    print(f"[✅] All DataFrames written to sheet: '{sheet_name}' in '{invoice_path}'")
    


def fill_qec_lru_replacement_sheet(invoice_path, new_qec_name_rep, metadata_qec):
    # print("metadata_qec :", metadata_qec)

    wb = load_workbook(invoice_path)
    ws = wb[new_qec_name_rep]
 
     # === Insert GE Logo ===
    img = Image('ge_new_logo.png')
    img.height = int(0.56 * 96)
    img.width = int(2.32 * 96)
    ws.add_image(img, 'B1')

    # ---------- column widths, styles ----------
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 21
    ws.column_dimensions['D'].width = 21
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 33
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 11
    ws.column_dimensions['I'].width = 17
    ws.column_dimensions['J'].width = 17
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 28
    ws.column_dimensions['M'].width = 28
    ws.column_dimensions['N'].width = 22
    ws.column_dimensions['O'].width = 30
    ws.column_dimensions['P'].width = 29
    ws.column_dimensions['Q'].width = 27
    ws.column_dimensions['R'].width = 34
 
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "8DC858"
    ws.sheet_view.zoomScale = 80
 
    # === Font Styles ===
    title_bold_black_font = Font(name='Arial', bold=True, color="000000", size=14)
    blue_fill = PatternFill(start_color="050543", end_color="050543", fill_type="solid")
    white_bold_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    norma_black_font = Font(name="Arial", bold=False, color="000000", size=11)

    # === Alignment Styles ===
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # === Border Styles ===
    no_border = Border()
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    black_colored_border = Side(style="thin", color="000000")
    white_colored_border = Side(style="thin", color="FFFFFF")
    thick_black_colored_border = Side(style="thick", color="000000")

    right_white_rest_black_border = Border(left=black_colored_border, top=black_colored_border, right=white_colored_border, bottom=black_colored_border)
    left_right_white_rest_black_border = Border(left=white_colored_border, top=black_colored_border, right=white_colored_border, bottom=black_colored_border)
    left_white_rest_black_border = Border(left=white_colored_border, top=black_colored_border, right=black_colored_border, bottom=black_colored_border)
    top_bottom_left_white_rest_black_border = Border(left=white_colored_border, top=white_colored_border, right=black_colored_border, bottom=white_colored_border)

    # Formats
    formats = {
        "Qty": '0',
        'Item (SD)': '0',
        "MRB": '0',
        "Discount %": '0%',
        "Handling Fee %": "0%"
    }
 
    # ---- Table 1 ----
    t1 = metadata_qec[0]
    start_row_1, end_row_1 = t1["start_row"], t1["end_row"]
    start_col_1, end_col_1 = t1["start_col"], t1["end_col"]
 
    style_range(ws, start_row_1, start_row_1, 2, t1["num_cols"]+1,
                fill=blue_fill, font=white_bold_font, align=Alignment(horizontal="center", vertical="center"),
                border=border, number_format='_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)')

    # Loop through columns and apply number format based on header
    data_row_start = start_row_1 + 1
    data_row_end = end_row_1
    
    for col in range(2, t1["num_cols"] + 2):  # +2 because range end is exclusive and index starts from 1
        header_cell = ws.cell(row=data_row_start - 1, column=col)
        header_value = header_cell.value.strip() if header_cell.value else ""
    
        # Determine number format
        number_format = formats.get(header_value, '_($* #,##0_);_($* (#,##0);_(@_)')  # Default to currency
    
        # Apply styling per cell in the column
        for row in range(data_row_start, data_row_end + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = norma_black_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            cell.number_format = number_format

    # formulas
    apply_arithmetic_formula(
        ws=ws, start_row=start_row_1, start_col=start_col_1, end_col=end_col_1,
        formula_str='=unit_price*qty',
        input_col_headers=['unit_price','qty'],
        mode='horizontal', output_col_name='extended_price'
    )

    apply_arithmetic_formula(
        ws=ws, start_row=start_row_1, start_col=start_col_1, end_col=end_col_1,
        formula_str='(extended_price) * (discount_%)',
        input_col_headers=['extended_price','discount_%'],
        mode='horizontal', output_col_name='discount_$'
    )

    apply_arithmetic_formula(
        ws=ws, start_row=start_row_1, start_col=start_col_1, end_col=end_col_1,
        formula_str='(clp) * (qty) * (handling_fee_%)',
        input_col_headers=['clp','qty','handling_fee_%'],
        mode='horizontal', output_col_name='handling_fee_price'
    )

    apply_arithmetic_formula(
        ws=ws, start_row=start_row_1, start_col=start_col_1, end_col=end_col_1,
        formula_str="=(extended_price) - (discount_$) + (handling_fee_price)",
        input_col_headers=['extended_price','discount_$', 'handling_fee_price'],
        mode='horizontal', output_col_name='excluded_[o&a]_total_price'
    )

    col_index_of_total_price = 15
    col_letter = get_column_letter(col_index_of_total_price)
    output_cell = f"{col_letter}{end_row_1+1}"

    apply_arithmetic_formula(
        ws=ws,
        start_row=start_row_1,
        start_col=15,
        end_col=15,
        formula_str="sum(excluded_[o&a]_total_price)",
        input_col_headers=['excluded_[o&a]_total_price'],
        mode='vertical',
        output_cell=output_cell
    )

    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Item (SD)")     
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Service Order Equipment / MRB")     
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Qty")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="CLP")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Unit Price")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Extended Price")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Discount %")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Discount $")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Handling Fee %")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Handling Fee Price")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Excluded [O&A] Total Price")
    
    sub_total_1_row_idx = end_row_1 + 1
    sub_total_1_col_idx = 15
    col_letter_sub_total_1 = get_column_letter(sub_total_1_col_idx)
    sub_total_1_cell = ws.cell(row=sub_total_1_row_idx, column=sub_total_1_col_idx)

    # Create the SUM formula
    sum_formula = f"=SUM({col_letter_sub_total_1}{start_row_1 + 1}:{col_letter_sub_total_1}{end_row_1})"
    sub_total_1_cell.value = sum_formula
    sub_total_1_cell.font = white_bold_font 
    sub_total_1_cell.fill = blue_fill   
    sub_total_1_cell.alignment = center
    sub_total_1_cell.border = border
    sub_total_1_cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'   

    # ---- Table 2 ----
    # t2 = metadata_qec[1]
    # start_row_2, end_row_2 = t2["start_row"], t2["end_row"]
    # start_col_2, end_col_2 = t2["start_col"], t2["end_col"]
    # num_rows_2 = t2["num_rows"]
    
    t2 = metadata_qec[1]
    start_row_2 = t2["start_row"] 
    end_row_2 = t2["end_row"] 
    start_col_2 = t2["start_col"]
    end_col_2 = t2["end_col"]
    num_rows_2 = t2["num_rows"]

    if start_row_2 == end_row_2:  # Empty case
        print("Con1 FOR Rep")
        fill_empty_df_headers(ws, start_row_2-1, start_col_2,
        pd.DataFrame(columns=["Item (SD)", "ATA Chapter Code", "Repair Source", "Part Number", "Part Description", "Qty", "Hour / CRD", "Unit Price", "Extended Price",
        "Discount %", "Discount $", "Excluded [O&A] Total Price", "Comments", "PO/Repair Tag", "Repair Code / Service Level"]),
                              white_bold_font, blue_fill, Alignment(horizontal="center"), border,
                              number_format_cols={"item_(SD)":"0", "qty": "0", "clp": "_($* #,##0_);_($* (#,##0);_(@_)",
                                                  "unit_price": "_($* #,##0_);_($* (#,##0);_(@_)",
                                                  "extended_price": "_($* #,##0_);_($* (#,##0);_(@_)",
                                                  "discount_%": "0%",
                                                  "discount_$": "_($* #,##0_);_($* (#,##0);_(@_)",
                                                  "handling_fee_%": "0%",
                                                  "handling_fee_price": "_($* #,##0_);_($* (#,##0);_(@_)",
                                                  "life_cicles_remaining": "0", "pro-rata_%": "0%",
                                                  "excluded_[o&a]_total_price": "_($* #,##0_);_($* (#,##0);_(@_)",
                                                  "mrb": "0"})
        sub_total_2_cell = ws.cell(row=start_row_2+1, column=start_col_2+11)
        set_style(sub_total_2_cell, font=white_bold_font, fill=blue_fill,
                  align=Alignment(horizontal="center"), border=border,
                  number_format='_($* #,##0_);_($* (#,##0);_(@_)')
        sub_total_2_cell.value = 0
 
    elif num_rows_2 == 1:  # One row 
        header_row = start_row_2
        data_row_start = start_row_2 + 1
        data_row_end = end_row_2
 
        
        # ✅ Style ONLY the existing header
        style_range(
            ws,
            header_row,
            header_row,
            start_col_2,
            end_col_2,
            font=white_bold_font,
            fill=blue_fill,
            align=Alignment(horizontal="center"),
            border=border
        ) 

        # Loop through columns and apply number format based on header
        for col in range(start_col_2, end_col_2 + 1):  # +2 because range end is exclusive and index starts from 1
            header_cell = ws.cell(row=header_row, column=col)
            header_value = header_cell.value.strip() if header_cell.value else ""
        
            # Determine number format
            number_format = formats.get(header_value, '_($* #,##0_);_($* (#,##0);_(@_)')  # Default to currency
        
            # Apply styling per cell in the column
            for row in range(data_row_start, data_row_end + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = norma_black_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                cell.number_format = number_format


        # formulas
        apply_arithmetic_formula(
            ws=ws, start_row=start_row_2, start_col=start_col_2, end_col=end_col_2,
            formula_str='=unit_price*qty',
            input_col_headers=['unit_price','qty'],
            mode='horizontal', output_col_name='extended_price'
        )

        apply_arithmetic_formula(
            ws=ws, start_row=start_row_2, start_col=start_col_2, end_col=end_col_2,
            formula_str='(extended_price) * (discount_%)',
            input_col_headers=['extended_price','discount_%'],
            mode='horizontal', output_col_name='discount_$'
        )
        
        apply_arithmetic_formula(
            ws=ws, start_row=start_row_2, start_col=start_col_2, end_col=end_col_2,
            # formula_str='=IF(discount_$="",(extended_price),(extended_price)-(discount_$))',
            formula_str='=(extended_price) - (discount_$) + (handling_fee_price)',
            input_col_headers=['extended_price','discount_$', 'handling_fee_price'],
            mode='horizontal', output_col_name='excluded_[o&a]_total_price'
        )
        

        col_index_of_total_price = 15
        col_letter = get_column_letter(col_index_of_total_price)
        output_cell = f"{col_letter}{end_row_2+1}"

        apply_arithmetic_formula(
            ws=ws,
            start_row=start_row_2,
            start_col=15,
            end_col=15,
            formula_str="sum(excluded_[o&a]_total_price)",
            input_col_headers=['excluded_[o&a]_total_price'],
            mode='vertical',
            output_cell=output_cell
        )

        apply_number_format_by_column_name(ws, header_row=start_row_2, start_row=start_row_2+1, end_row=end_row_2, target_col_name="Item (SD)")
        apply_number_format_by_column_name(ws, header_row=start_row_2, start_row=start_row_2+1, end_row=end_row_2, target_col_name="PO/Repair Tag")
        apply_number_format_by_column_name(ws, header_row=start_row_2, start_row=start_row_2+1, end_row=end_row_2, target_col_name="Qty")
        apply_number_format_by_column_name(ws, header_row=start_row_2, start_row=start_row_2+1, end_row=end_row_2, target_col_name="Hour / CRD")
        apply_number_format_by_column_name(ws, header_row=start_row_2, start_row=start_row_2+1, end_row=end_row_2, target_col_name="Unit Price")
        apply_number_format_by_column_name(ws, header_row=start_row_2, start_row=start_row_2+1, end_row=end_row_2, target_col_name="Extended Price")
        apply_number_format_by_column_name(ws, header_row=start_row_2, start_row=start_row_2+1, end_row=end_row_2, target_col_name="Discount %")
        apply_number_format_by_column_name(ws, header_row=start_row_2, start_row=start_row_2+1, end_row=end_row_2, target_col_name="Discount $")
        apply_number_format_by_column_name(ws, header_row=start_row_2, start_row=start_row_2+1, end_row=end_row_2, target_col_name="Excluded [O&A] Total Price")
        
        sub_total_2_row_idx = end_row_2 + 1
        sub_total_2_col_idx = 15   # only if Excluded [O&A] is column 13 in Repair
        col_letter_sub_total_2 = get_column_letter(sub_total_2_col_idx)
        
        sub_total_2_cell = ws.cell(row=sub_total_2_row_idx, column=sub_total_2_col_idx)
        
        sum_formula = f"=SUM({col_letter_sub_total_2}{start_row_2 + 1}:{col_letter_sub_total_2}{end_row_2})"
        sub_total_2_cell.value = sum_formula
        
        sub_total_2_cell.font = white_bold_font
        sub_total_2_cell.fill = blue_fill
        sub_total_2_cell.alignment = center
        sub_total_2_cell.border = border
        sub_total_2_cell.number_format = '_($* #,##0_);_($* (#,##0);_(@_)'     

    else:  # multiple rows handled uniformly
        print("Con3 FOR Rep") 
        
        header_row = start_row_2
        data_row_start = start_row_2 + 1
        data_row_end = end_row_2
 
        
        # ✅ Style ONLY the existing header
        style_range(
            ws,
            header_row,
            header_row,
            start_col_2,
            end_col_2,
            font=white_bold_font,
            fill=blue_fill,
            align=Alignment(horizontal="center"),
            border=border
        ) 

        # Loop through columns and apply number format based on header
        for col in range(start_col_2, end_col_2 + 1):  # +2 because range end is exclusive and index starts from 1
            header_cell = ws.cell(row=header_row, column=col)
            header_value = header_cell.value.strip() if header_cell.value else ""
        
            # Determine number format
            number_format = formats.get(header_value, '_($* #,##0_);_($* (#,##0);_(@_)')  # Default to currency
        
            # Apply styling per cell in the column
            for row in range(data_row_start, data_row_end + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = norma_black_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                cell.number_format = number_format


        # formulas
        apply_arithmetic_formula(
            ws=ws, start_row=start_row_2, start_col=start_col_2, end_col=end_col_2,
            formula_str='=unit_price*qty',
            input_col_headers=['unit_price','qty'],
            mode='horizontal', output_col_name='extended_price'
        )

        apply_arithmetic_formula(
            ws=ws, start_row=start_row_2, start_col=start_col_2, end_col=end_col_2,
            formula_str='(extended_price) * (discount_%)',
            input_col_headers=['extended_price','discount_%'],
            mode='horizontal', output_col_name='discount_$'
        )
        
        apply_arithmetic_formula(
            ws=ws, start_row=start_row_2, start_col=start_col_2, end_col=end_col_2,
            # formula_str='=IF(discount_$="",(extended_price),(extended_price)-(discount_$))',
            # input_col_headers=['extended_price','discount_$'],
            formula_str='=(extended_price) - (discount_$) + (handling_fee_price)',
            input_col_headers=['extended_price','discount_$', 'handling_fee_price'],
            mode='horizontal', output_col_name='excluded_[o&a]_total_price'
        )
        

        col_index_of_total_price = 15
        col_letter = get_column_letter(col_index_of_total_price)
        output_cell = f"{col_letter}{end_row_2+1}"

        apply_arithmetic_formula(
            ws=ws,
            start_row=start_row_2,
            start_col=15,
            end_col=15,
            formula_str="sum(excluded_[o&a]_total_price)",
            input_col_headers=['excluded_[o&a]_total_price'],
            mode='vertical',
            output_cell=output_cell
        )

        apply_number_format_by_column_name(ws, header_row=start_row_2, start_row=start_row_2+1, end_row=end_row_2, target_col_name="Item (SD)")
        apply_number_format_by_column_name(ws, header_row=start_row_2, start_row=start_row_2+1, end_row=end_row_2, target_col_name="PO/Repair Tag")
        apply_number_format_by_column_name(ws, header_row=start_row_2, start_row=start_row_2+1, end_row=end_row_2, target_col_name="Qty")
        apply_number_format_by_column_name(ws, header_row=start_row_2, start_row=start_row_2+1, end_row=end_row_2, target_col_name="Hour / CRD")
        apply_number_format_by_column_name(ws, header_row=start_row_2, start_row=start_row_2+1, end_row=end_row_2, target_col_name="Unit Price")
        apply_number_format_by_column_name(ws, header_row=start_row_2, start_row=start_row_2+1, end_row=end_row_2, target_col_name="Extended Price")
        apply_number_format_by_column_name(ws, header_row=start_row_2, start_row=start_row_2+1, end_row=end_row_2, target_col_name="Discount %")
        apply_number_format_by_column_name(ws, header_row=start_row_2, start_row=start_row_2+1, end_row=end_row_2, target_col_name="Discount $")
        apply_number_format_by_column_name(ws, header_row=start_row_2, start_row=start_row_2+1, end_row=end_row_2, target_col_name="Excluded [O&A] Total Price")
        
        sub_total_2_row_idx = end_row_2 + 1
        sub_total_2_col_idx = 15   # only if Excluded [O&A] is column 13 in Repair
        col_letter_sub_total_2 = get_column_letter(sub_total_2_col_idx)
        
        sub_total_2_cell = ws.cell(row=sub_total_2_row_idx, column=sub_total_2_col_idx)
        
        sum_formula = f"=SUM({col_letter_sub_total_2}{start_row_2 + 1}:{col_letter_sub_total_2}{end_row_2})"
        sub_total_2_cell.value = sum_formula
        
        sub_total_2_cell.font = white_bold_font
        sub_total_2_cell.fill = blue_fill
        sub_total_2_cell.alignment = center
        sub_total_2_cell.border = border
        sub_total_2_cell.number_format = '_($* #,##0_);_($* (#,##0);_(@_)'  
 
    # ---- Titles and Grand Total ----
    # ws.cell(row=start_row_1-1, column=5, value="Material").font = title_bold_black_font
    ws["E3"].value = "Material"
    ws["E3"].font = title_bold_black_font
    ws["E3"].alignment = center
 
    ws.cell(row=start_row_2-2, column=5, value="Repair").font = title_bold_black_font
    ws.cell(row=start_row_2-2, column=5).alignment = center

    ws["I3"].value, ws["I3"].font = "Grand Total", white_bold_font
    ws["I3"].fill, ws["I3"].alignment = blue_fill, Alignment(horizontal="center")
    ws["I3"].border = Border(top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"))

    ws["J1"].value = f"={sub_total_1_cell.coordinate}"
    ws["J1"].fill = blue_fill
    ws["J1"].font = white_bold_font
    ws["J1"].alignment = left
    ws["J1"].border = top_bottom_left_white_rest_black_border
    ws["J1"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'   

    ws["J2"].value = f"={sub_total_2_cell.coordinate}"
    ws["J2"].fill = blue_fill
    ws["J2"].font = white_bold_font
    ws["J2"].alignment = left
    ws["J2"].border = top_bottom_left_white_rest_black_border
    ws["J2"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'   

    ws["J3"].value = f"=SUM({sub_total_1_cell.coordinate},{sub_total_2_cell.coordinate})"
    set_style(ws["J3"], font=white_bold_font, fill=blue_fill,
              align=Alignment(horizontal="right"), border=Border(top=Side(style="thin"),
              bottom=Side(style="thin"), right=Side(style="thin")),
              number_format='_($* #,##0_);_($* (#,##0);_(@_)')

    wb.save(invoice_path)
    print(f"[INFO] ✅ QEC-LRU-ACC sheet filled and saved to {invoice_path}")




def merge_same_value_cells(ws, start_row, end_row, col_map, target_columns, key_column_name):
    from openpyxl.utils import get_column_letter
 
    key_col = col_map[key_column_name]
 
    current_start = start_row + 1
    prev_value = ws.cell(row=current_start, column=key_col).value
 
    for row in range(start_row + 2, end_row + 2):  # +2 to include last group
        curr_value = ws.cell(row=row, column=key_col).value if row <= end_row else None
 
        if curr_value != prev_value:
            group_end = row - 1
 
            # Merge only if more than 1 row
            if group_end > current_start:
                for col_name in target_columns:
                    col_idx = col_map[col_name]
                    col_letter = get_column_letter(col_idx)
 
                    merge_range = f"{col_letter}{current_start}:{col_letter}{group_end}"
                    ws.merge_cells(merge_range)
 
                    # Center align merged cell
                    cell = ws.cell(row=current_start, column=col_idx)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
 
            # Move to next group
            current_start = row
            prev_value = curr_value


def fill_scrap_exclusion_sheet(invoice_path, scrap_sheet_name, metadata_scrap):
    # from openpyxl import load_workbook
    # from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    # from openpyxl.drawing.image import Image
    # from openpyxl.utils import get_column_letter
 
    wb = load_workbook(invoice_path)
    ws = wb[scrap_sheet_name]
 
    # === Logo ===
    img = Image('ge_new_logo.png')
    img.height = int(0.56 * 96)
    img.width = int(2.32 * 96)
    ws.add_image(img, 'B1')
 
    # =========================
    # COLUMN WIDTHS (CUSTOM FOR SCRAP)
    # =========================
    widths = [
        2, 12, 22, 24, 12, 20, 22, 18, 18, 18, 18, 20, 22, 22, 26, 30
    ]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
 
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "8DC858"
    ws.sheet_view.zoomScale = 80
 
    # =========================
    # STYLES
    # =========================
    title_font = Font(name='Arial', bold=True, color="000000", size=14)
    blue_fill = PatternFill(start_color="050543", end_color="050543", fill_type="solid")
    white_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    normal_font = Font(name="Arial", size=11)
 
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
 
    currency_fmt = '_($* #,##0_);_($* (#,##0);_(@_)'
    percent_fmt = '0%'
 
    # =========================
    # HELPER: GET COL INDEX
    # =========================
    def get_col_map(header_row, start_col, end_col):
        col_map = {}
        for col in range(start_col, end_col + 1):
            val = ws.cell(row=header_row, column=col).value
            if val:
                col_map[val.strip()] = col
        return col_map
 
    # =========================
    # TABLE 1 → MATERIAL
    # =========================
    t1 = metadata_scrap[0]
    sr, er = t1["start_row"], t1["end_row"]
    sc, ec = t1["start_col"], t1["end_col"]
 
    # Header styling
    for col in range(sc, ec + 1):
        cell = ws.cell(row=sr, column=col)
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = border
 
    col_map = get_col_map(sr, sc, ec)
 
    # Data styling
    for row in range(sr + 1, er + 1):
        for col in range(sc, ec + 1):
            c = ws.cell(row=row, column=col)
            c.font = normal_font
            c.border = border
            c.alignment = center
 
    # -------------------------
    # FORMULAS (MATERIAL)
    # -------------------------
    for r in range(sr + 1, er + 1):
        QPE = get_column_letter(col_map['QPE'])
        pct = get_column_letter(col_map['% Material Included under Exclusion Table'])
        qty_inc = get_column_letter(col_map['Qty Material Included under Exclusion Table'])
        total_qty = get_column_letter(col_map['Total Qty Replaced'])
        qty_new = get_column_letter(col_map['Material Type Qty New'])
        unit_price = get_column_letter(col_map['Unit Price CLP$'])
        handling = get_column_letter(col_map['Handling Fee Price'])
        total_price = get_column_letter(col_map['Total Price'])
        sum_price = get_column_letter(col_map['Sum of Price'])
        included = get_column_letter(col_map['Included Price'])
        excluded = get_column_letter(col_map['Excluded [O&A] Price'])
        desc = get_column_letter(col_map['Part Description'])
 
        # Qty Included
        ws[f"{qty_inc}{r}"] = f"={QPE}{r}*{pct}{r}"
 
        # Total Price
        # ws[f"{total_price}{r}"] = f"=({qty_new}{r}*{unit_price}{r})+{handling}{r}-(({qty_new}{r}*{unit_price}{r})*5%)"
        ws[f"{total_price}{r}"] = f"=IF(OR({handling}{r}=0,ISBLANK({handling}{r})),({qty_new}{r}*{unit_price}{r})+{handling}{r}-(({qty_new}{r}*{unit_price}{r})*5%),{handling}{r})"
 
        # Sum of Price (SUMIF)
        ws[f"{sum_price}{r}"] = (
            f"=SUMIF({desc}{sr+1}:{desc}{er},{desc}{r},{total_price}{sr+1}:{total_price}{er})"
        )
 
        # Included Price
        ws[f"{included}{r}"] = (
            f"=IF({total_qty}{r}>{qty_inc}{r},"
            f"({qty_inc}{r}/{total_qty}{r})*{sum_price}{r},"
            f"{sum_price}{r})"
        )
 
        # Excluded Price
        ws[f"{excluded}{r}"] = f"={sum_price}{r}-{included}{r}"
    
    merge_same_value_cells(
        ws=ws,
        start_row=sr,
        end_row=er,
        col_map=col_map,
        key_column_name='Part Description',
        target_columns=[
            
            'QPE',
            '% Material Included under Exclusion Table',
            'Qty Material Included under Exclusion Table',
            'Total Qty Replaced',
            'Sum of Price',
            'Included Price',
            'Excluded [O&A] Price'
        ]
    )

    # # Subtotal
    # sub1_row = er + 1
    # total_price_col_letter = get_column_letter(col_map['Total Price'])
    # ws[f"{total_price_col_letter}{sub1_row}"] = f"=SUM({total_price_col_letter}{sr+1}:{total_price_col_letter}{er})"
    sub1_row = er + 1
 
    sum_price_col = get_column_letter(col_map['Sum of Price'])
    included_col = get_column_letter(col_map['Included Price'])
    excluded_col = get_column_letter(col_map['Excluded [O&A] Price'])
    
    # Label
    ws[f"{sum_price_col}{sub1_row}"] = "Sub Total"
    
    # Values
    ws[f"{included_col}{sub1_row}"] = f"=SUM({included_col}{sr+1}:{included_col}{er})"
    ws[f"{excluded_col}{sub1_row}"] = f"=SUM({excluded_col}{sr+1}:{excluded_col}{er})"
    
    # Style (same as header)
    for col in [col_map['Sum of Price'], col_map['Included Price'], col_map['Excluded [O&A] Price']]:
        cell = ws.cell(row=sub1_row, column=col)
        cell.font = white_font
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = border

    # Apply percentage format
    pct_col_idx = col_map['% Material Included under Exclusion Table']
 
    for r in range(sr + 1, er + 1):
        cell = ws.cell(row=r, column=pct_col_idx)
        if cell.value is not None:
            try:
                cell.value = cell.value / 100
            except:
                pass
        cell.number_format = '0%'
    
 
    # =========================
    # TABLE 2 → REPAIR
    # =========================
    t2 = metadata_scrap[1]
    sr2, er2 = t2["start_row"], t2["end_row"]
    sc2, ec2 = t2["start_col"], t2["end_col"]
    price_from_data = t2.get("price_from_data", [])
 
    for col in range(sc2, ec2 + 1):
        c = ws.cell(row=sr2, column=col)
        c.fill = blue_fill
        c.font = white_font
        c.alignment = center
        c.border = border
 
    col_map2 = get_col_map(sr2, sc2, ec2)
 
    for r in range(sr2 + 1, er2 + 1):
        data_idx = r - (sr2 + 1)
        for col in range(sc2, ec2 + 1):
            c = ws.cell(row=r, column=col)
            c.font = normal_font
            c.border = border
            c.alignment = center
 
        QPE = get_column_letter(col_map2['QPE'])
        pct = get_column_letter(col_map2['% Repair Included under Exclusion Table'])
        qty_inc = get_column_letter(col_map2['Qty Repair Included under Exclusion Table'])
        total_qty = get_column_letter(col_map2['Total Qty Repaired'])
        repair_price = get_column_letter(col_map2['Repair Price'])
        discount = get_column_letter(col_map2['Discount'])
        total_price = get_column_letter(col_map2['Total Price'])
        sum_price = get_column_letter(col_map2['Sum of Price'])
        included = get_column_letter(col_map2['Included Price'])
        excluded = get_column_letter(col_map2['Excluded [O&A] Price'])
        desc = get_column_letter(col_map2['Part Description'])
 
        # Qty Included
        ws[f"{qty_inc}{r}"] = f"={QPE}{r}*{pct}{r}"
 
        total_qty_lookup = f"LOOKUP(2,1/({total_qty}{sr2+1}:{total_qty}{r}<>\"\"),{total_qty}{sr2+1}:{total_qty}{r})"
 
        # Discount
        row_price_from = price_from_data[data_idx] if data_idx < len(price_from_data) else ""
        
        if row_price_from == "CRD":
            ws[f"{discount}{r}"] = f"={repair_price}{r}*{total_qty_lookup}*20%"
        else:
            ws[f"{discount}{r}"] = 0
    
        # Total Price
        ws[f"{total_price}{r}"] = (
            f"=({repair_price}{r}*{total_qty_lookup})-{discount}{r}"
        )
 
        # Sum of Price
        ws[f"{sum_price}{r}"] = (
            f"=SUMIF({desc}{sr2+1}:{desc}{er2},{desc}{r},{total_price}{sr2+1}:{total_price}{er2})"
        )
 
        # Included Price
        ws[f"{included}{r}"] = (
            f"=IF({total_qty}{r}>{qty_inc}{r},"
            f"({qty_inc}{r}/{total_qty}{r})*{sum_price}{r},"
            f"{sum_price}{r})"
        )
 
        # Excluded Price
        ws[f"{excluded}{r}"] = f"={sum_price}{r}-{included}{r}"
    
    merge_same_value_cells(
        ws=ws,
        start_row=sr2,
        end_row=er2,
        col_map=col_map2,
        key_column_name='Part Description',
        target_columns=[
            'QPE',
            '% Repair Included under Exclusion Table',
            'Qty Repair Included under Exclusion Table',
            'Total Qty Repaired',
            'Sum of Price',
            'Included Price',
            'Excluded [O&A] Price'
        ]
    )

    # sub2_row = er2 + 1
    # total_price_col_letter2 = get_column_letter(col_map2['Total Price'])
    # ws[f"{total_price_col_letter2}{sub2_row}"] = f"=SUM({total_price_col_letter2}{sr2+1}:{total_price_col_letter2}{er2})"
    sub2_row = er2 + 1
 
    sum_price_col2 = get_column_letter(col_map2['Sum of Price'])
    included_col2 = get_column_letter(col_map2['Included Price'])
    excluded_col2 = get_column_letter(col_map2['Excluded [O&A] Price'])
    
    ws[f"{sum_price_col2}{sub2_row}"] = "Sub Total"
    
    ws[f"{included_col2}{sub2_row}"] = f"=SUM({included_col2}{sr2+1}:{included_col2}{er2})"
    ws[f"{excluded_col2}{sub2_row}"] = f"=SUM({excluded_col2}{sr2+1}:{excluded_col2}{er2})"
    
    for col in [col_map2['Sum of Price'], col_map2['Included Price'], col_map2['Excluded [O&A] Price']]:
        cell = ws.cell(row=sub2_row, column=col)
        cell.font = white_font
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = border

    # Apply percentage format
    pct_col_idx = col_map2['% Repair Included under Exclusion Table']
 
    for r in range(sr2 + 1, er2 + 1):
        cell = ws.cell(row=r, column=pct_col_idx)
        if cell.value is not None:
            try:
                cell.value = cell.value / 100
            except:
                pass
        cell.number_format = '0%'
 
    # =========================
    # TITLES + GRAND TOTAL
    # =========================
    ws["E3"] = "Material"
    ws["E3"].font = title_font
    ws["E3"].alignment = center
 
    ws.cell(row=sr2 - 2, column=5, value="Repair").font = title_font
 
    ws["I3"] = "Grand Total"
    ws["I3"].font = white_font
    ws["I3"].fill = blue_fill
    ws["I3"].alignment = center
 
    ws["J1"] = f"={excluded_col}{sub1_row}"
    ws["J1"].font = white_font
    ws["J1"].fill = blue_fill
    ws["J1"].alignment = center

    ws["J2"] = f"={excluded_col2}{sub2_row}"
    ws["J2"].font = white_font
    ws["J2"].fill = blue_fill
    ws["J2"].alignment = center

    ws["J3"] = f"=SUM(J1,J2)"
    ws["J3"].font = white_font
    ws["J3"].fill = blue_fill
    ws["J3"].alignment = center
 
    wb.save(invoice_path)
    print(f"[INFO] ✅ Scrap sheet created: {invoice_path}")


def fill_contract_pricing_terms_sheet(invoice_path, sheet_name, metadata, escalation_year = '2025'):
    wb = load_workbook(invoice_path)
    ws = wb[sheet_name]
 
    # === Logo ===
    img = Image('ge_new_logo.png')
    img.height = int(0.56 * 96)
    img.width = int(2.32 * 96)
    ws.add_image(img, 'B1')
 
    # === Column Widths (custom for this sheet) ===
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 35   # Bucket
    ws.column_dimensions['C'].width = 20   # Material Type
    ws.column_dimensions['D'].width = 40   # Guideline
    ws.column_dimensions['E'].width = 18   # Amount
    ws.column_dimensions['F'].width = 14   # Discount %
    ws.column_dimensions['G'].width = 16   # Handling Fee %
    ws.column_dimensions['H'].width = 28   # Handling Fee cap per part
    ws.column_dimensions['I'].width = 30   # Handling Fee cap per set
 
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "8DC858"
    ws.sheet_view.zoomScale = 85
 
    # === Styles ===
    title_font = Font(name='Arial', bold=True, color="000000", size=14)
    blue_fill = PatternFill(start_color="050543", end_color="050543", fill_type="solid")
    white_bold_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    normal_font = Font(name="Arial", size=11)
 
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left = Alignment(horizontal="left", vertical="center", wrap_text=False)
 
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
 
    # === Table Metadata ===
    t1 = metadata[0]
    start_row = t1["start_row"]
    end_row = t1["end_row"]
    start_col = t1["start_col"]
    end_col = t1["end_col"]
 
    header_row = start_row
    data_row_start = start_row + 1
    data_row_end = end_row


    # === Insert Escalation Year Row ===
    ws.insert_rows(start_row)
    
    # Update references after row shift
    header_row = start_row + 1
    data_row_start = header_row + 1
    data_row_end = end_row + 1  # shift due to inserted row
    
    num_cols = end_col - start_col + 1
    
    # Merge from first column to second last column
    ws.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=start_row,
        end_column=end_col - 1
    )
    
    # === Escalation Label Cell ===
    esc_cell = ws.cell(row=start_row, column=start_col)
    esc_cell.value = "Escalation Year"
    esc_cell.font = white_bold_font
    esc_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dark blue fill (same as header)
    esc_cell.fill = blue_fill
    esc_cell.border = border
    
    # === Year Value Cell (last column) ===
    year_cell = ws.cell(row=start_row, column=end_col)
    year_cell.value = escalation_year   # <-- pass this as parameter
    year_cell.font = white_bold_font
    year_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Light blue fill (#00B0F0)
    light_blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    year_cell.fill = light_blue_fill
    year_cell.border = border


 
    # === Header Styling ===
    style_range(
        ws,
        header_row,
        header_row,
        start_col,
        end_col,
        font=white_bold_font,
        fill=blue_fill,
        align=center,
        border=border
    )
 
    # === Column Formats Mapping ===
    formats = {
        "Amount": '_($* #,##0.00_);_($* (#,##0.00);_(@_)',
        "Discount %": '0%',
        "Handling Fee %": '0%',
        "Handling Fee cap per part": '_($* #,##0_);_($* (#,##0);_(@_)',
        "Handling Fee cap per set": '_($* #,##0_);_($* (#,##0);_(@_)'
    }
 
    # === Apply Styling + Formats ===
    for col in range(start_col, end_col + 1):
        header_val = ws.cell(row=header_row, column=col).value
        header_val = header_val.strip() if header_val else ""
 
        number_format = formats.get(header_val, None)
 
        for row in range(data_row_start, data_row_end + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = normal_font
            cell.alignment = left if header_val == "Guideline" else center
            cell.border = border
 
            if number_format:
                cell.number_format = number_format
 
    # === Title ===
    # ws["E3"].value = "Contract Pricing Terms"
    # ws["E3"].font = title_font
    # ws["E3"].alignment = center
 
    wb.save(invoice_path)
    print(f"[INFO] ✅ Contract Pricing Terms sheet created: {invoice_path}")



def fill_pma_der_sheet(invoice_path, pma_der_sheet_name, pma_der_info_refined, year_str):

    df_index_1 = pma_der_info_refined[0]["df_index"]

    start_row_1 = pma_der_info_refined[0]["start_row"]
    end_row_1 = pma_der_info_refined[0]["end_row"]
    
    num_rows_1 = pma_der_info_refined[0]["num_rows"]
    num_cols_1 = pma_der_info_refined[0]["num_cols"]
    
    start_col_1 = pma_der_info_refined[0]["start_col"]
    end_col_1 = pma_der_info_refined[0]["end_col"]
    
    wb = load_workbook(invoice_path)
    ws = wb[pma_der_sheet_name]

    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 7
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 9
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 28
    ws.column_dimensions['L'].width = 21

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "8DC858"


    # Define styles
    title_bold_black_font = Font(bold=True, color="000000", size=12)
    norma_black_font = Font(bold=False, color="000000", size=10)
    white_bold_font = Font(bold=True, color="FFFFFF", size=11)
    blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid") 
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    no_border = Border() 
    top_bottom_border = Border(top=Side(style='thin'), bottom=Side(style='thin')) 
    top_bottom_left_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))
    top_bottom_right_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin')) 
    right_border = Border(right = Side(style='thin'))
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
 
    # Insert logo
    img = Image('ge_logo.png')
    img.height = int(0.54 * 96)
    img.width = int(0.67 * 96)
    ws.add_image(img, 'A1')

    # # # Table-1 Values
    dollar_cols = ["total price"]
    percent_cols = ["discount"]
    int_cols = ['qty']

    for row in ws.iter_rows(min_row=start_row_1+1, max_row=end_row_1-1, min_col = 2, max_col=num_cols_1+1):
        for cell in row:
            cell.font = norma_black_font
            cell.alignment = center
            cell.border = border
            cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    
    
    # Sub-total 1
    apply_arithmetic_formula(
        ws=ws,
        start_row=start_row_1,
        start_col=start_col_1,
        end_col=end_col_1,
        formula_str='unit_price * qty',
        input_col_headers=['unit_price', 'qty'],
        mode='horizontal',
        output_col_name='extended_price'
    )

    apply_arithmetic_formula(
        ws=ws,
        start_row=start_row_1,
        start_col=start_col_1,
        end_col=end_col_1,
        formula_str='(extended_price) - (extended_price * discount)',
        input_col_headers=['extended_price', 'discount'],
        mode='horizontal',
        output_col_name='total_price'
    )

    col_index_of_total_price = 9
    col_letter = get_column_letter(col_index_of_total_price)
    output_cell = f"{col_letter}{end_row_1}"

    apply_arithmetic_formula(
        ws=ws,
        start_row=start_row_1,
        start_col=9,
        end_col=9,
        formula_str='sum(total_price)',
        input_col_headers=['total_price'],
        mode='vertical',
        output_cell=output_cell
    )
     
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1-1, target_col_name="Qty")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1-1, target_col_name="Discount")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1-1, target_col_name="Unit Price")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1-1, target_col_name="Extended Price")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1-1, target_col_name="Total Price")
    
    sub_total_1_row_idx = end_row_1
    sub_total_1_col_idx = 9
    col_letter_sub_total_1 = get_column_letter(sub_total_1_col_idx)
    sub_total_1_cell = ws.cell(row=sub_total_1_row_idx, column=sub_total_1_col_idx)

    # Create the SUM formula
    sum_formula = f"=SUM({col_letter_sub_total_1}{start_row_1 + 1}:{col_letter_sub_total_1}{end_row_1-1})"
    sub_total_1_cell.value = sum_formula
    sub_total_1_cell.font = white_bold_font 
    sub_total_1_cell.fill = blue_fill   
    sub_total_1_cell.alignment = center
    sub_total_1_cell.border = border
    sub_total_1_cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'   

    # # Set header values
    ws["E2"].value = "Out-of-Scope PMA & DER"
    ws["E2"].alignment = Alignment(horizontal="center", vertical="center")#, wrap_text=False)
    ws["E2"].font = title_bold_black_font

    ws["I2"].value = "Grand Total"
    ws["I2"].fill = blue_fill
    ws["I2"].font = white_bold_font
    ws["I2"].alignment = center
    ws["I2"].border = top_bottom_left_border
     
    # Section headers - 1
    ws["B5"].value = "Module"
    ws["B5"].fill = blue_fill
    ws["B5"].font = white_bold_font
    ws["B5"].alignment = center
    ws["B5"].border = border

    ws["C5"].value = "Part Name"
    ws["C5"].fill = blue_fill
    ws["C5"].font = white_bold_font
    ws["C5"].alignment = center
    ws["C5"].border = border

    ws["D5"].value = "Part Number"
    ws["D5"].fill = blue_fill
    ws["D5"].font = white_bold_font
    ws["D5"].alignment = center
    ws["D5"].border = border

    ws["E5"].value = "Qty"
    ws["E5"].fill = blue_fill
    ws["E5"].font = white_bold_font
    ws["E5"].alignment = center
    ws["E5"].border = border

    ws["F5"].value = "Unit Price"
    ws["F5"].fill = blue_fill
    ws["F5"].font = white_bold_font
    ws["F5"].alignment = center
    ws["F5"].border = border

    ws["G5"].value = "Extended Price"
    ws["G5"].fill = blue_fill
    ws["G5"].font = white_bold_font
    ws["G5"].alignment = center
    ws["G5"].border = border

    ws["H5"].value = "Discount"
    ws["H5"].fill = blue_fill
    ws["H5"].font = white_bold_font
    ws["H5"].alignment = center
    ws["H5"].border = border

    ws["I5"].value = "Total Price"  
    ws["I5"].fill = blue_fill
    ws["I5"].font = white_bold_font
    ws["I5"].alignment = center
    ws["I5"].border = border

    ws["J5"].value = "Status"
    ws["J5"].fill = blue_fill
    ws["J5"].font = white_bold_font
    ws["J5"].alignment = center
    ws["J5"].border = border

    ws["K5"].value = "Removal Cause"
    ws["K5"].fill = blue_fill
    ws["K5"].font = white_bold_font
    ws["K5"].alignment = center
    ws["K5"].border = border

    ws["L5"].value = "Exclusion Remarks"
    ws["L5"].fill = blue_fill
    ws["L5"].font = white_bold_font
    ws["L5"].alignment = center
    ws["L5"].border = border

    ws["J2"].value = f"=({sub_total_1_cell.coordinate})"
    ws["J2"].fill = blue_fill
    ws["J2"].font = white_bold_font
    ws["J2"].alignment = Alignment(horizontal="right", vertical="center")
    ws["J2"].border = top_bottom_right_border
    ws["J2"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)' 
 
    wb.save(invoice_path)


def fill_mor_sheet(invoice_path, mor_sheet_name, mor_info_refined, year_str):

    df_index_1 = mor_info_refined[0]["df_index"]

    start_row_1 = mor_info_refined[0]["start_row"]
    end_row_1 = mor_info_refined[0]["end_row"]
    
    num_rows_1 = mor_info_refined[0]["num_rows"]
    num_cols_1 = mor_info_refined[0]["num_cols"]
    
    start_col_1 = mor_info_refined[0]["start_col"]
    end_col_1 = mor_info_refined[0]["end_col"]
    
    wb = load_workbook(invoice_path)
    ws = wb[mor_sheet_name]

    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 7
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 9
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 27

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "8DC858"


    # Define styles
    title_bold_black_font = Font(bold=True, color="000000", size=12)
    norma_black_font = Font(bold=False, color="000000", size=10)
    white_bold_font = Font(bold=True, color="FFFFFF", size=11)
    blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid") 
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    no_border = Border() 
    top_bottom_border = Border(top=Side(style='thin'), bottom=Side(style='thin')) 
    top_bottom_left_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))
    top_bottom_right_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin')) 
    right_border = Border(right = Side(style='thin'))
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
 
    # Insert logo
    img = Image('ge_logo.png')
    img.height = int(0.54 * 96)
    img.width = int(0.67 * 96)
    ws.add_image(img, 'A1')

    # # # Table-1 Values
    dollar_cols = ["total price"]
    percent_cols = ["mark up"]
    int_cols = ['qty']

    for row in ws.iter_rows(min_row=start_row_1+1, max_row=end_row_1-1, min_col = 2, max_col=num_cols_1+1):
        for cell in row:
            cell.font = norma_black_font
            cell.alignment = center
            cell.border = border
            cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    
    
    # Sub-total 1
    apply_arithmetic_formula(
        ws=ws,
        start_row=start_row_1,
        start_col=start_col_1,
        end_col=end_col_1,
        formula_str='unit_price * qty',
        input_col_headers=['unit_price', 'qty'],
        mode='horizontal',
        output_col_name='extended_price'
    )

    apply_arithmetic_formula(
        ws=ws,
        start_row=start_row_1,
        start_col=start_col_1,
        end_col=end_col_1,
        formula_str='(extended_price) - (extended_price * mark_up)',
        input_col_headers=['extended_price', 'mark_up'],
        mode='horizontal',
        output_col_name='total_price'
    )

    col_index_of_total_price = 9
    col_letter = get_column_letter(col_index_of_total_price)
    output_cell = f"{col_letter}{end_row_1}"

    apply_arithmetic_formula(
        ws=ws,
        start_row=start_row_1,
        start_col=9,
        end_col=9,
        formula_str='sum(total_price)',
        input_col_headers=['total_price'],
        mode='vertical',
        output_cell=output_cell
    )
     
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1-1, target_col_name="Qty")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1-1, target_col_name="Mark Up")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1-1, target_col_name="Unit Price")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1-1, target_col_name="Extended Price")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1-1, target_col_name="Total Price")
    
    sub_total_1_row_idx = end_row_1
    sub_total_1_col_idx = 9
    col_letter_sub_total_1 = get_column_letter(sub_total_1_col_idx)
    sub_total_1_cell = ws.cell(row=sub_total_1_row_idx, column=sub_total_1_col_idx)

    # Create the SUM formula
    sum_formula = f"=SUM({col_letter_sub_total_1}{start_row_1 + 1}:{col_letter_sub_total_1}{end_row_1-1})"
    sub_total_1_cell.value = sum_formula
    sub_total_1_cell.font = white_bold_font 
    sub_total_1_cell.fill = blue_fill   
    sub_total_1_cell.alignment = center
    sub_total_1_cell.border = border
    sub_total_1_cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'   

    # # Set header values
    ws["E2"].value = "Out-of-Scope Missing Items"
    ws["E2"].alignment = Alignment(horizontal="center", vertical="center")#, wrap_text=False)
    ws["E2"].font = title_bold_black_font

    ws["I2"].value = "Grand Total"
    ws["I2"].fill = blue_fill
    ws["I2"].font = white_bold_font
    ws["I2"].alignment = center
    ws["I2"].border = top_bottom_left_border
     
    # Section headers - 1
    ws["B5"].value = "Module"
    ws["B5"].fill = blue_fill
    ws["B5"].font = white_bold_font
    ws["B5"].alignment = center
    ws["B5"].border = border

    ws["C5"].value = "Part Name"
    ws["C5"].fill = blue_fill
    ws["C5"].font = white_bold_font
    ws["C5"].alignment = center
    ws["C5"].border = border

    ws["D5"].value = "Part Number"
    ws["D5"].fill = blue_fill
    ws["D5"].font = white_bold_font
    ws["D5"].alignment = center
    ws["D5"].border = border

    ws["E5"].value = "Qty"
    ws["E5"].fill = blue_fill
    ws["E5"].font = white_bold_font
    ws["E5"].alignment = center
    ws["E5"].border = border

    ws["F5"].value = "Unit Price"
    ws["F5"].fill = blue_fill
    ws["F5"].font = white_bold_font
    ws["F5"].alignment = center
    ws["F5"].border = border

    ws["G5"].value = "Extended Price"
    ws["G5"].fill = blue_fill
    ws["G5"].font = white_bold_font
    ws["G5"].alignment = center
    ws["G5"].border = border

    ws["H5"].value = "Mark Up"
    ws["H5"].fill = blue_fill
    ws["H5"].font = white_bold_font
    ws["H5"].alignment = center
    ws["H5"].border = border

    ws["I5"].value = "Total Price"  
    ws["I5"].fill = blue_fill
    ws["I5"].font = white_bold_font
    ws["I5"].alignment = center
    ws["I5"].border = border

    ws["J5"].value = "Status"
    ws["J5"].fill = blue_fill
    ws["J5"].font = white_bold_font
    ws["J5"].alignment = center
    ws["J5"].border = border

    ws["K5"].value = "Exclusion Remarks"
    ws["K5"].fill = blue_fill
    ws["K5"].font = white_bold_font
    ws["K5"].alignment = center
    ws["K5"].border = border

    ws["J2"].value = f"=({sub_total_1_cell.coordinate})"
    ws["J2"].fill = blue_fill
    ws["J2"].font = white_bold_font
    ws["J2"].alignment = Alignment(horizontal="right", vertical="center")
    ws["J2"].border = top_bottom_right_border
    ws["J2"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)' 
 
    wb.save(invoice_path)
    # print(f"[INFO] ✅ In-Scope sheet filled and saved to {invoice_path}")


def fill_empty_df_headers(ws, start_row, start_col, df, white_bold_font, blue_fill, center, border, number_format_cols=None):
    if number_format_cols is None:
        number_format_cols = {}
 
    # Write headers
    for col_idx, col_name in enumerate(df.columns, start=start_col):
        header_cell = ws.cell(row=start_row, column=col_idx)
        header_cell.value = col_name
        header_cell.font = white_bold_font
        header_cell.fill = blue_fill
        header_cell.alignment = center
        header_cell.border = border
 
    # Write one empty data row with formatting
    data_row_idx = start_row + 1
    for col_idx, col_name in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=data_row_idx, column=col_idx)
        cell.value = ""  # intentionally blank
        cell.font = Font(bold=False, color="000000", size=10)
        cell.alignment = center
        cell.border = border
 
        # Apply specific number format
        if col_name.lower() in number_format_cols:
            cell.number_format = number_format_cols[col_name.lower()]
        else:
            cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'


def format_specific_row_header(ws, row_idx, start_col, end_col, font, fill, alignment, border):

    for col_idx in range(start_col, end_col+1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = font 
        cell.fill = fill   
        cell.alignment = alignment
        cell.border = border
    
def fill_mat_sheet(invoice_path, mat_sheet_name, mat_info_refined):
    # print("mat_info_refined :", mat_info_refined)
    df_index_1 = mat_info_refined[0]["df_index"]

    start_row_1 = mat_info_refined[0]["start_row"]
    end_row_1 = mat_info_refined[0]["end_row"]
    
    num_rows_1 = mat_info_refined[0]["num_rows"]
    num_cols_1 = mat_info_refined[0]["num_cols"]
    
    start_col_1 = mat_info_refined[0]["start_col"]
    end_col_1 = mat_info_refined[0]["end_col"]
    
    wb = load_workbook(invoice_path)
    ws = wb[mat_sheet_name]
 
     # === Insert GE Logo ===
    img = Image('ge_new_logo.png')
    img.height = int(0.56 * 96)
    img.width = int(2.32 * 96)
    ws.add_image(img, 'B1')

    # ---------- column widths, styles ----------
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 26
    ws.column_dimensions['E'].width = 19
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 17
    ws.column_dimensions['I'].width = 17
    ws.column_dimensions['J'].width = 17
    ws.column_dimensions['K'].width = 11
    ws.column_dimensions['L'].width = 17
    ws.column_dimensions['M'].width = 10
    ws.column_dimensions['N'].width = 17
    ws.column_dimensions['O'].width = 21
    ws.column_dimensions['P'].width = 36
    ws.column_dimensions['Q'].width = 31
    ws.column_dimensions['R'].width = 17
    ws.column_dimensions['S'].width = 17
 
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "8DC858"
    ws.sheet_view.zoomScale = 80
 
    # === Font Styles ===
    title_bold_black_font = Font(name='Arial', bold=True, color="000000", size=14)
    blue_fill = PatternFill(start_color="050543", end_color="050543", fill_type="solid")
    white_bold_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    norma_black_font = Font(name="Arial", bold=False, color="000000", size=11)

    # === Alignment Styles ===
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # === Border Styles ===
    no_border = Border()
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    black_colored_border = Side(style="thin", color="000000")
    white_colored_border = Side(style="thin", color="FFFFFF")
    thick_black_colored_border = Side(style="thick", color="000000")

    right_white_rest_black_border = Border(left=black_colored_border, top=black_colored_border, right=white_colored_border, bottom=black_colored_border)
    left_right_white_rest_black_border = Border(left=white_colored_border, top=black_colored_border, right=white_colored_border, bottom=black_colored_border)
    left_white_rest_black_border = Border(left=white_colored_border, top=black_colored_border, right=black_colored_border, bottom=black_colored_border)
    top_bottom_left_white_rest_black_border = Border(left=white_colored_border, top=white_colored_border, right=black_colored_border, bottom=white_colored_border)

    # # # Table-1 Values
    for row in ws.iter_rows(min_row=start_row_1+1, max_row=end_row_1, min_col = 2, max_col=num_cols_1+1):
        for cell in row:
            cell.font = norma_black_font
            cell.alignment = center
            cell.border = border
            cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    for row in range(start_row_1+1, end_row_1+1):
        ws.row_dimensions[row].height = 16

    # formulas
    apply_arithmetic_formula(
        ws=ws, start_row=start_row_1, start_col=start_col_1, end_col=end_col_1,
        formula_str='=unit_price*qty',
        input_col_headers=['unit_price','qty'],
        mode='horizontal', output_col_name='extended_price'
    )

    apply_arithmetic_formula(
        ws=ws, start_row=start_row_1, start_col=start_col_1, end_col=end_col_1,
        formula_str='(extended_price)*(discount_%)',
        input_col_headers=['extended_price','discount_%'],
        mode='horizontal', output_col_name='discount_$'
    )

    apply_arithmetic_formula(
        ws=ws, start_row=start_row_1, start_col=start_col_1, end_col=end_col_1,
        formula_str='(clp)*(qty)*(handling_fee_%)',
        input_col_headers=['clp','qty','handling_fee_%'],
        mode='horizontal', output_col_name='handling_fee_price'
    )

    apply_arithmetic_formula(
        ws=ws, start_row=start_row_1, start_col=start_col_1, end_col=end_col_1,
        # formula_str="=(extended_price) - (discount_$) + (handling_fee_price)",
        formula_str="IF(has_cust=1,handling_fee_price,extended_price-discount_$+handling_fee_price)",
        input_col_headers=['has_cust', 'extended_price','discount_$', 'handling_fee_price'],
        mode='horizontal', output_col_name='excluded_[o&a]_total_price'
    )

    col_index_of_total_price = 15
    col_letter = get_column_letter(col_index_of_total_price)
    output_cell = f"{col_letter}{end_row_1+1}"

    apply_arithmetic_formula(
        ws=ws,
        start_row=start_row_1,
        start_col=15,
        end_col=15,
        formula_str="sum(excluded_[o&a]_total_price)",
        input_col_headers=['excluded_[o&a]_total_price'],
        mode='vertical',
        output_cell=output_cell
    )

    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Item (SD)")     
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Service Order Equipment / MRB")     
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Qty")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="CLP")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Unit Price")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Extended Price")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Discount %")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Discount $")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Handling Fee %")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Handling Fee Price")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Excluded [O&A] Total Price")

    sub_total_1_row_idx = end_row_1 + 1
    sub_total_1_col_idx = 15
    col_letter_sub_total_1 = get_column_letter(sub_total_1_col_idx)
    sub_total_1_cell = ws.cell(row=sub_total_1_row_idx, column=sub_total_1_col_idx)

    # Create the SUM formula
    sum_formula = f"=SUM({col_letter_sub_total_1}{start_row_1 + 1}:{col_letter_sub_total_1}{end_row_1})"
    sub_total_1_cell.value = sum_formula
    sub_total_1_cell.font = white_bold_font 
    sub_total_1_cell.fill = blue_fill   
    sub_total_1_cell.alignment = center
    sub_total_1_cell.border = border
    sub_total_1_cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'   

    for col in ws.iter_cols(min_col=14, max_col=14, min_row=5, max_row=ws.max_row):
        for cell in col:
            cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'   

    ws["O1"].value = "Excluded [O&A] Price"
    ws["O1"].fill = blue_fill
    ws["O1"].font = white_bold_font
    ws["O1"].alignment = center
    ws["O1"].border = border

    # Set header values
    ws["H2"].value = "Material"
    ws["H2"].alignment = Alignment(horizontal="center", vertical="center")#, wrap_text=False)
    ws["H2"].font = title_bold_black_font

    ws["N2"].value = "Grand Total"
    ws["N2"].fill = blue_fill
    ws["N2"].font = white_bold_font
    ws["N2"].alignment = center
    ws["N2"].border = right_white_rest_black_border
     
    # Section headers - 1
    ws["B4"].value = "Item (SD)"
    ws["B4"].fill = blue_fill
    ws["B4"].font = white_bold_font
    ws["B4"].alignment = center
    ws["B4"].border = right_white_rest_black_border

    ws["C4"].value = "ATA Chapter Code"
    ws["C4"].fill = blue_fill
    ws["C4"].font = white_bold_font
    ws["C4"].alignment = center
    ws["C4"].border = left_right_white_rest_black_border

    ws["D4"].value = "Material Type"
    ws["D4"].fill = blue_fill
    ws["D4"].font = white_bold_font
    ws["D4"].alignment = center
    ws["D4"].border = left_right_white_rest_black_border

    ws["E4"].value = "Part Number"
    ws["E4"].fill = blue_fill
    ws["E4"].font = white_bold_font
    ws["E4"].alignment = center
    ws["E4"].border = left_right_white_rest_black_border

    ws["F4"].value = "Part Description"
    ws["F4"].fill = blue_fill
    ws["F4"].font = white_bold_font
    ws["F4"].alignment = center
    ws["F4"].border = left_right_white_rest_black_border

    ws["G4"].value = "Qty"
    ws["G4"].fill = blue_fill
    ws["G4"].font = white_bold_font
    ws["G4"].alignment = center
    ws["G4"].border = left_right_white_rest_black_border

    ws["H4"].value = "CLP"
    ws["H4"].fill = blue_fill
    ws["H4"].font = white_bold_font
    ws["H4"].alignment = center
    ws["H4"].border = left_right_white_rest_black_border

    ws["I4"].value = "Unit Price"
    ws["I4"].fill = blue_fill
    ws["I4"].font = white_bold_font
    ws["I4"].alignment = center
    ws["I4"].border = left_right_white_rest_black_border

    ws["J4"].value = "Extended Price"
    ws["J4"].fill = blue_fill
    ws["J4"].font = white_bold_font
    ws["J4"].alignment = center
    ws["J4"].border = left_right_white_rest_black_border

    ws["K4"].value = "Discount %"
    ws["K4"].fill = blue_fill
    ws["K4"].font = white_bold_font
    ws["K4"].alignment = center
    ws["K4"].border = left_right_white_rest_black_border

    ws["L4"].value = "Discount $"
    ws["L4"].fill = blue_fill
    ws["L4"].font = white_bold_font
    ws["L4"].alignment = center
    ws["L4"].border = left_right_white_rest_black_border

    ws["M4"].value = "Handling Fee %"
    ws["M4"].fill = blue_fill
    ws["M4"].font = white_bold_font
    ws["M4"].alignment = center
    ws["M4"].border = left_right_white_rest_black_border

    ws["N4"].value = "Handling Fee Price"
    ws["N4"].fill = blue_fill
    ws["N4"].font = white_bold_font
    ws["N4"].alignment = center
    ws["N4"].border = left_right_white_rest_black_border

    ws["O4"].value = "Excluded [O&A] Total Price"
    ws["O4"].fill = blue_fill
    ws["O4"].font = white_bold_font
    ws["O4"].alignment = center
    ws["O4"].border = left_right_white_rest_black_border

    ws["P4"].value = "Replacement Remarks"
    ws["P4"].fill = blue_fill
    ws["P4"].font = white_bold_font
    ws["P4"].alignment = center
    ws["P4"].border = left_right_white_rest_black_border

    ws["Q4"].value = "Comments"
    ws["Q4"].fill = blue_fill
    ws["Q4"].font = white_bold_font
    ws["Q4"].alignment = center
    ws["Q4"].border = left_right_white_rest_black_border

    ws["R4"].value = "Service Order Equipment / MRB"
    ws["R4"].fill = blue_fill
    ws["R4"].font = white_bold_font
    ws["R4"].alignment = center
    ws["R4"].border = left_white_rest_black_border

    ws["S4"].value = "QPE_Highlight"
    ws["S4"].fill = blue_fill
    ws["S4"].font = white_bold_font
    ws["S4"].alignment = center
    ws["S4"].border = left_white_rest_black_border

    ws["T4"].value = "Has_CUST"
    ws["T4"].fill = blue_fill
    ws["T4"].font = white_bold_font
    ws["T4"].alignment = center
    ws["T4"].border = left_white_rest_black_border

    ws["O2"].value = f"={sub_total_1_cell.coordinate}"
    ws["O2"].fill = blue_fill
    ws["O2"].font = white_bold_font
    ws["O2"].alignment = Alignment(horizontal="right", vertical="center")
    ws["O2"].border = top_bottom_left_white_rest_black_border
    ws["O2"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)' 

    wb.save(invoice_path)
    print(f"[INFO] ✅ Material sheet filled and saved to {invoice_path}")

    # sub_total_cell = sub_total_1_cell.coordinate
    # o2_value = ws[sub_total_cell].value
    # print(f"[RESULT] O2 manually resolved to: {o2_value}")
    # return o2_value
    wb_data = load_workbook(invoice_path, data_only=True)
    ws_data = wb_data[mat_sheet_name]
    o2_value = ws_data["O2"].value
 
    print(f"[INFO] O2 evaluated value: {o2_value}")
    return o2_value

def fill_rep_sheet(invoice_path, repair_sheet_name, rep_info_refined):
    # print("rep_file_sheet")
    # invoice_path.to_excel("rep_file_sheet.xlsx",index=False)
    df_index_1 = rep_info_refined[0]["df_index"]

    start_row_1 = rep_info_refined[0]["start_row"]
    end_row_1 = rep_info_refined[0]["end_row"]
    
    num_rows_1 = rep_info_refined[0]["num_rows"]
    num_cols_1 = rep_info_refined[0]["num_cols"]
    
    start_col_1 = rep_info_refined[0]["start_col"]
    end_col_1 = rep_info_refined[0]["end_col"]
    
    wb = load_workbook(invoice_path)
    ws = wb[repair_sheet_name]
 
     # === Insert GE Logo ===
    img = Image('ge_new_logo.png')
    img.height = int(0.56 * 96)
    img.width = int(2.32 * 96)
    ws.add_image(img, 'B1')

    # ---------- column widths, styles ----------
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 32
    ws.column_dimensions['E'].width = 32
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 17
    ws.column_dimensions['I'].width = 17
    ws.column_dimensions['J'].width = 17
    ws.column_dimensions['K'].width = 17
    ws.column_dimensions['L'].width = 17
    ws.column_dimensions['M'].width = 22
    ws.column_dimensions['N'].width = 17
    ws.column_dimensions['O'].width = 21
    ws.column_dimensions['P'].width = 36
 
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "8DC858"
    ws.sheet_view.zoomScale = 80
 
    # === Font Styles ===
    title_bold_black_font = Font(name='Arial', bold=True, color="000000", size=14)
    blue_fill = PatternFill(start_color="050543", end_color="050543", fill_type="solid")
    white_bold_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    norma_black_font = Font(name="Arial", bold=False, color="000000", size=11)

    # === Alignment Styles ===
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # === Border Styles ===
    no_border = Border()
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    black_colored_border = Side(style="thin", color="000000")
    white_colored_border = Side(style="thin", color="FFFFFF")
    thick_black_colored_border = Side(style="thick", color="000000")

    right_white_rest_black_border = Border(left=black_colored_border, top=black_colored_border, right=white_colored_border, bottom=black_colored_border)
    left_right_white_rest_black_border = Border(left=white_colored_border, top=black_colored_border, right=white_colored_border, bottom=black_colored_border)
    left_white_rest_black_border = Border(left=white_colored_border, top=black_colored_border, right=black_colored_border, bottom=black_colored_border)
    top_bottom_left_white_rest_black_border = Border(left=white_colored_border, top=white_colored_border, right=black_colored_border, bottom=white_colored_border)

    # # # Table-1 Values
    for row in ws.iter_rows(min_row=start_row_1+1, max_row=end_row_1, min_col = 2, max_col=num_cols_1+1):
        for cell in row:
            cell.font = norma_black_font
            cell.alignment = center
            cell.border = border
            cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    for row in range(start_row_1+1, end_row_1+1):
        ws.row_dimensions[row].height = 16

    # formulas
    apply_arithmetic_formula(
        ws=ws, start_row=start_row_1, start_col=start_col_1, end_col=end_col_1,
        formula_str='=unit_price*qty',
        input_col_headers=['unit_price','qty'],
        mode='horizontal', output_col_name='extended_price'
    )

    apply_arithmetic_formula(
        ws=ws, start_row=start_row_1, start_col=start_col_1, end_col=end_col_1,
        formula_str='IF(discount_%="",0,extended_price*discount_%)',
        input_col_headers=['extended_price','discount_%'],
        mode='horizontal', output_col_name='discount_$'
    )
 

    apply_arithmetic_formula(
        ws=ws, start_row=start_row_1, start_col=start_col_1, end_col=end_col_1,
        formula_str='IF(discount_$="",extended_price,extended_price-discount_$)',
        input_col_headers=['extended_price','discount_$'],
        mode='horizontal', output_col_name='excluded_[o&a]_total_price'
    )

    col_index_of_total_price = 12
    col_letter = get_column_letter(col_index_of_total_price)
    output_cell = f"{col_letter}{end_row_1+1}"

    apply_arithmetic_formula(
        ws=ws,
        start_row=start_row_1,
        start_col=12,
        end_col=12,
        formula_str="sum(excluded_[o&a]_total_price)",
        input_col_headers=['excluded_[o&a]_total_price'],
        mode='vertical',
        output_cell=output_cell
    )

    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Item (SD)")     
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="PO/Repair Tag")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Qty")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Hour / CRD")     
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Unit Price")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Extended Price")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Discount %")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Discount $")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Excluded [O&A] Total Price")
    
    sub_total_1_row_idx = end_row_1 + 1
    sub_total_1_col_idx = 12
    col_letter_sub_total_1 = get_column_letter(sub_total_1_col_idx)
    sub_total_1_cell = ws.cell(row=sub_total_1_row_idx, column=sub_total_1_col_idx)

    # Create the SUM formula
    sum_formula = f"=SUM({col_letter_sub_total_1}{start_row_1 + 1}:{col_letter_sub_total_1}{end_row_1})"
    sub_total_1_cell.value = sum_formula
    sub_total_1_cell.font = white_bold_font 
    sub_total_1_cell.fill = blue_fill   
    sub_total_1_cell.alignment = center
    sub_total_1_cell.border = border
    sub_total_1_cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'   

    for col in ws.iter_cols(min_col=14, max_col=14, min_row=5, max_row=ws.max_row):
            for cell in col:
                cell.number_format = '0'   

    ws["L1"].value = "Excluded [O&A] Price"
    ws["L1"].fill = blue_fill
    ws["L1"].font = white_bold_font
    ws["L1"].alignment = center
    ws["L1"].border = border

    # Set header values
    ws["H2"].value = "GE Repair"
    ws["H2"].alignment = Alignment(horizontal="center", vertical="center")#, wrap_text=False)
    ws["H2"].font = title_bold_black_font

    ws["K2"].value = "Grand Total"
    ws["K2"].fill = blue_fill
    ws["K2"].font = white_bold_font
    ws["K2"].alignment = center
    ws["K2"].border = right_white_rest_black_border
     
    # Section headers - 1
    ws["B4"].value = "Item (SD)"
    ws["B4"].fill = blue_fill
    ws["B4"].font = white_bold_font
    ws["B4"].alignment = center
    ws["B4"].border = right_white_rest_black_border

    ws["C4"].value = "ATA Chapter Code"
    ws["C4"].fill = blue_fill
    ws["C4"].font = white_bold_font
    ws["C4"].alignment = center
    ws["C4"].border = left_right_white_rest_black_border

    ws["D4"].value = "Part Number"
    ws["D4"].fill = blue_fill
    ws["D4"].font = white_bold_font
    ws["D4"].alignment = center
    ws["D4"].border = left_right_white_rest_black_border

    ws["E4"].value = "Part Description"
    ws["E4"].fill = blue_fill
    ws["E4"].font = white_bold_font
    ws["E4"].alignment = center
    ws["E4"].border = left_right_white_rest_black_border

    ws["F4"].value = "Qty"
    ws["F4"].fill = blue_fill
    ws["F4"].font = white_bold_font
    ws["F4"].alignment = center
    ws["F4"].border = left_right_white_rest_black_border

    ws["G4"].value = "Hour / CRD"
    ws["G4"].fill = blue_fill
    ws["G4"].font = white_bold_font
    ws["G4"].alignment = center
    ws["G4"].border = left_right_white_rest_black_border

    ws["H4"].value = "Unit Price"
    ws["H4"].fill = blue_fill
    ws["H4"].font = white_bold_font
    ws["H4"].alignment = center
    ws["H4"].border = left_right_white_rest_black_border

    ws["I4"].value = "Extended Price"
    ws["I4"].fill = blue_fill
    ws["I4"].font = white_bold_font
    ws["I4"].alignment = center
    ws["I4"].border = left_right_white_rest_black_border

    ws["J4"].value = "Discount %"
    ws["J4"].fill = blue_fill
    ws["J4"].font = white_bold_font
    ws["J4"].alignment = center
    ws["J4"].border = left_right_white_rest_black_border

    ws["K4"].value = "Discount $"
    ws["K4"].fill = blue_fill
    ws["K4"].font = white_bold_font
    ws["K4"].alignment = center
    ws["K4"].border = left_right_white_rest_black_border

    ws["L4"].value = "Excluded [O&A] Total Price"
    ws["L4"].fill = blue_fill
    ws["L4"].font = white_bold_font
    ws["L4"].alignment = center
    ws["L4"].border = left_right_white_rest_black_border

    ws["M4"].value = "Comments"
    ws["M4"].fill = blue_fill
    ws["M4"].font = white_bold_font
    ws["M4"].alignment = center
    ws["M4"].border = left_right_white_rest_black_border

    ws["N4"].value = "PO/Repair Tag"
    ws["N4"].fill = blue_fill
    ws["N4"].font = white_bold_font
    ws["N4"].alignment = center
    ws["N4"].border = left_right_white_rest_black_border

    ws["O4"].value = "Repair Code / Service Level"
    ws["O4"].fill = blue_fill
    ws["O4"].font = white_bold_font
    ws["O4"].alignment = center
    ws["O4"].border = left_right_white_rest_black_border

    ws["P4"].value = "QPE_Highlight"
    ws["P4"].fill = blue_fill
    ws["P4"].font = white_bold_font
    ws["P4"].alignment = center
    ws["P4"].border = left_right_white_rest_black_border

    ws["L2"].value = f"=({sub_total_1_cell.coordinate})"
    ws["L2"].fill = blue_fill
    ws["L2"].font = white_bold_font
    ws["L2"].alignment = Alignment(horizontal="right", vertical="center")
    ws["L2"].border = top_bottom_left_white_rest_black_border
    ws["L2"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)' 

    wb.save(invoice_path)
    print(f"[INFO] ✅ Ge-Repair sheet filled and saved to {invoice_path}")

def fill_vendor_sheet(invoice_path, vendor_sheet_name, ven_info_refined):
    # print("ven_info_refined :", ven_info_refined)
    df_index_1 = ven_info_refined[0]["df_index"]

    start_row_1 = ven_info_refined[0]["start_row"]
    end_row_1 = ven_info_refined[0]["end_row"]
    
    num_rows_1 = ven_info_refined[0]["num_rows"]
    num_cols_1 = ven_info_refined[0]["num_cols"]
    
    start_col_1 = ven_info_refined[0]["start_col"]
    end_col_1 = ven_info_refined[0]["end_col"]
    
    wb = load_workbook(invoice_path)
    ws = wb[vendor_sheet_name]
 
     # === Insert GE Logo ===
    img = Image('ge_new_logo.png')
    img.height = int(0.56 * 96)
    img.width = int(2.32 * 96)
    ws.add_image(img, 'B1')

    # ---------- column widths, styles ----------
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 17
    ws.column_dimensions['I'].width = 17
    ws.column_dimensions['J'].width = 17
    ws.column_dimensions['K'].width = 11
    ws.column_dimensions['L'].width = 17
    ws.column_dimensions['M'].width = 10
    ws.column_dimensions['N'].width = 17
    ws.column_dimensions['O'].width = 21
 
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "8DC858"
    ws.sheet_view.zoomScale = 80
 
    # === Font Styles ===
    title_bold_black_font = Font(name='Arial', bold=True, color="000000", size=14)
    blue_fill = PatternFill(start_color="050543", end_color="050543", fill_type="solid")
    white_bold_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    norma_black_font = Font(name="Arial", bold=False, color="000000", size=11)

    # === Alignment Styles ===
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # === Border Styles ===
    no_border = Border()
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    black_colored_border = Side(style="thin", color="000000")
    white_colored_border = Side(style="thin", color="FFFFFF")
    thick_black_colored_border = Side(style="thick", color="000000")

    right_white_rest_black_border = Border(left=black_colored_border, top=black_colored_border, right=white_colored_border, bottom=black_colored_border)
    left_right_white_rest_black_border = Border(left=white_colored_border, top=black_colored_border, right=white_colored_border, bottom=black_colored_border)
    left_white_rest_black_border = Border(left=white_colored_border, top=black_colored_border, right=black_colored_border, bottom=black_colored_border)
    top_bottom_left_white_rest_black_border = Border(left=white_colored_border, top=white_colored_border, right=black_colored_border, bottom=white_colored_border)

    if (start_row_1 == end_row_1) | (num_rows_1 == 1) :  # Empty case
        fill_empty_df_headers(ws, start_row_1-1, start_col_1,
        pd.DataFrame(columns=["Item (SD)", "ATA Chapter Code", "Repair Source", "Part Number", "Part Description", "Qty", "Vendor Invoice Amount", 
        "Handling Fee %", "Handling Fee Price", "Excluded [O&A] Total Price", 'Comments', "PO #" , "QPE_Highlight"]),
                              white_bold_font, blue_fill, Alignment(horizontal="center"), border,
                              number_format_cols={"item_(SD)":"0", "qty": "0", "clp": "_($* #,##0_);_($* (#,##0);_(@_)",
                                                  "unit_price": "_($* #,##0_);_($* (#,##0);_(@_)",
                                                  "extended_price": "_($* #,##0_);_($* (#,##0);_(@_)",
                                                  "discount_%": "0%",
                                                  "discount_$": "_($* #,##0_);_($* (#,##0);_(@_)",
                                                  "handling_fee_%": "0%",
                                                  "handling_fee_price": "_($* #,##0_);_($* (#,##0);_(@_)",
                                                  "life_cicles_remaining": "0", "pro-rata_%": "0%",
                                                  "excluded_[o&a]_total_price": "_($* #,##0_);_($* (#,##0);_(@_)",
                                                  "mrb": "0"})
        sub_total_1_cell = ws.cell(row=start_row_1+1, column=start_col_1+9)
        set_style(sub_total_1_cell, font=white_bold_font, fill=blue_fill,
                  align=left, border=left_right_white_rest_black_border,
                  number_format='_($* #,##0_);_($* (#,##0);_(@_)')
        sub_total_1_cell.value = 0

        for row in range(start_row_1+1, end_row_1+1):
            ws.row_dimensions[row].height = 16


    else:
        # # # Table-1 Values
        for row in ws.iter_rows(min_row=start_row_1+1, max_row=end_row_1, min_col = 2, max_col=num_cols_1+1):
            for cell in row:
                cell.font = norma_black_font
                cell.alignment = center
                cell.border = border
                cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
            
        for row in range(start_row_1+1, end_row_1+1):
            ws.row_dimensions[row].height = 16

        # formulas
        apply_arithmetic_formula(
            ws=ws, start_row=start_row_1, start_col=start_col_1, end_col=end_col_1,
            formula_str='(qty)*(handling_fee_%)',
            input_col_headers=['qty','handling_fee_%'],
            mode='horizontal', output_col_name='handling_fee_price'
        )

        apply_arithmetic_formula(
            ws=ws, start_row=start_row_1, start_col=start_col_1, end_col=end_col_1,
            formula_str="=(vendor_invoice_amount)+(handling_fee_price)-(discount_$)",
            input_col_headers=['vendor_invoice_amount','handling_fee_price','discount_$'],
            mode='horizontal', output_col_name='excluded_[o&a]_total_price'
        )

        col_index_of_total_price = 13
        col_letter = get_column_letter(col_index_of_total_price)
        output_cell = f"{col_letter}{end_row_1+1}"

        apply_arithmetic_formula(
            ws=ws,
            start_row=start_row_1,
            start_col=13,
            end_col=13,
            formula_str="sum(excluded_[o&a]_total_price)",
            input_col_headers=['excluded_[o&a]_total_price'],
            mode='vertical',
            output_cell=output_cell
        )
        
        # -------------------------------------------------
        # SUBCON-NONGE-ACTUAL Business Rule Enforcement
        # -------------------------------------------------
        
        header_row = start_row_1
        
        def col_idx(col_name):
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=header_row, column=col).value == col_name:
                    return col
            return None
        
        repair_col   = col_idx("Repair Source")
        disc_pct_col = col_idx("Discount %")
        disc_amt_col = col_idx("Discount $")
        hf_pct_col   = col_idx("Handling Fee %")
        
        for row in range(start_row_1 + 1, end_row_1 + 1):
        
            repair_val = ws.cell(row=row, column=repair_col).value
            repair_val = str(repair_val).strip().upper() if repair_val else ""
        
            # ---------------------------------------------
            # CASE 1: SUBCON-NONGE-ACTUAL
            # → Discount must NOT apply
            # ---------------------------------------------
            if repair_val == "SUBCON-NONGE-ACTUAL":
                if disc_pct_col:
                    ws.cell(row=row, column=disc_pct_col).value = 0
                if disc_amt_col:
                    ws.cell(row=row, column=disc_amt_col).value = 0
        
            # ---------------------------------------------
            # CASE 2: GE / Others
            # → Handling Fee must NOT apply
            # ---------------------------------------------
            else:
                if hf_pct_col:
                    ws.cell(row=row, column=hf_pct_col).value = 0

        apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Item (SD)")     
        apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Qty")
        apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Vendor Invoice Amount")
        apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Handling Fee %")
        apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Handling Fee Price")
        apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Excluded [O&A] Total Price")
        apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Discount %")
 
        
        sub_total_1_row_idx = end_row_1 + 1
        sub_total_1_col_idx = 13
        col_letter_sub_total_1 = get_column_letter(sub_total_1_col_idx)
        sub_total_1_cell = ws.cell(row=sub_total_1_row_idx, column=sub_total_1_col_idx)

        # Create the SUM formula
        sum_formula = f"=SUM({col_letter_sub_total_1}{start_row_1 + 1}:{col_letter_sub_total_1}{end_row_1})"
        sub_total_1_cell.value = sum_formula
        sub_total_1_cell.font = white_bold_font 
        sub_total_1_cell.fill = blue_fill   
        sub_total_1_cell.alignment = left
        sub_total_1_cell.border = border
        sub_total_1_cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'   

    ws["K1"].value = "Excluded [O&A] Price"
    ws["K1"].fill = blue_fill
    ws["K1"].font = white_bold_font
    ws["K1"].alignment = center
    ws["K1"].border = border

    # Set header values
    ws["E2"].value = "Sub Contracted Repair"
    ws["E2"].alignment = Alignment(horizontal="center", vertical="center")#, wrap_text=False)
    ws["E2"].font = title_bold_black_font
    
    # Section headers - 1
    ws["B4"].value = "Item (SD)"
    ws["B4"].fill = blue_fill
    ws["B4"].font = white_bold_font
    ws["B4"].alignment = center
    ws["B4"].border = right_white_rest_black_border

    ws["C4"].value = "ATA Chapter Code"
    ws["C4"].fill = blue_fill
    ws["C4"].font = white_bold_font
    ws["C4"].alignment = center
    ws["C4"].border = left_right_white_rest_black_border

    ws["D4"].value = "Repair Source"
    ws["D4"].fill = blue_fill
    ws["D4"].font = white_bold_font
    ws["D4"].alignment = center
    ws["D4"].border = left_right_white_rest_black_border

    ws["E4"].value = "Part Number"
    ws["E4"].fill = blue_fill
    ws["E4"].font = white_bold_font
    ws["E4"].alignment = center
    ws["E4"].border = left_right_white_rest_black_border

    ws["F4"].value = "Part Description"
    ws["F4"].fill = blue_fill
    ws["F4"].font = white_bold_font
    ws["F4"].alignment = center
    ws["F4"].border = left_right_white_rest_black_border

    ws["G4"].value = "Qty"
    ws["G4"].fill = blue_fill
    ws["G4"].font = white_bold_font
    ws["G4"].alignment = center
    ws["G4"].border = left_right_white_rest_black_border

    ws["H4"].value = "Vendor Invoice Amount"
    ws["H4"].fill = blue_fill
    ws["H4"].font = white_bold_font
    ws["H4"].alignment = center
    ws["H4"].border = left_right_white_rest_black_border

    ws["I4"].value = "Discount %"
    ws["I4"].fill = blue_fill
    ws["I4"].font = white_bold_font
    ws["I4"].alignment = center
    ws["I4"].border = left_right_white_rest_black_border

    ws["J4"].value = "Discount $"
    ws["J4"].fill = blue_fill
    ws["J4"].font = white_bold_font
    ws["J4"].alignment = center
    ws["J4"].border = left_right_white_rest_black_border

    ws["K4"].value = "Handling Fee %"
    ws["K4"].fill = blue_fill
    ws["K4"].font = white_bold_font
    ws["K4"].alignment = center
    ws["K4"].border = left_right_white_rest_black_border

    ws["L4"].value = "Handling Fee Price"
    ws["L4"].fill = blue_fill
    ws["L4"].font = white_bold_font
    ws["L4"].alignment = center
    ws["L4"].border = left_right_white_rest_black_border

    ws["M4"].value = "Excluded [O&A] Total Price"
    ws["M4"].fill = blue_fill
    ws["M4"].font = white_bold_font
    ws["M4"].alignment = center
    ws["M4"].border = left_right_white_rest_black_border

    ws["N4"].value = "Comments"
    ws["N4"].fill = blue_fill
    ws["N4"].font = white_bold_font
    ws["N4"].alignment = center
    ws["N4"].border = left_right_white_rest_black_border

    ws["O4"].value = "PO #"
    ws["O4"].fill = blue_fill
    ws["O4"].font = white_bold_font
    ws["O4"].alignment = center
    ws["O4"].border = left_right_white_rest_black_border

    ws["P4"].value = "QPE_Highlight"
    ws["P4"].fill = blue_fill
    ws["P4"].font = white_bold_font
    ws["P4"].alignment = center
    ws["P4"].border = left_right_white_rest_black_border

    ws["J2"].value = "Grand Total"
    ws["J2"].fill = blue_fill
    ws["J2"].font = white_bold_font
    ws["J2"].alignment = center
    ws["J2"].border = right_white_rest_black_border
    
    ws["K2"].value = f"=({sub_total_1_cell.coordinate})"
    ws["K2"].fill = blue_fill
    ws["K2"].font = white_bold_font
    ws["K2"].alignment = Alignment(horizontal="right", vertical="center")
    ws["K2"].border = top_bottom_left_white_rest_black_border
    ws["K2"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)' 

    wb.save(invoice_path)
    print(f"[INFO] ✅ Sub-Contracted Repair sheet filled and saved to {invoice_path}")


def format_specific_row_header(ws, row_idx, start_col, end_col, font, fill, alignment, border):

    for col_idx in range(start_col, end_col+1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = font 
        cell.fill = fill   
        cell.alignment = alignment
        cell.border = border
    

def fill_sb_oa_summary(invoice_path, sb_sheet_name, metadata_sb):
    # print("mat_info_refined :", metadata_sb)
    df_index_1 = metadata_sb[0]["df_index"]

    start_row_1 = metadata_sb[0]["start_row"]
    end_row_1 = metadata_sb[0]["end_row"]
    
    num_rows_1 = metadata_sb[0]["num_rows"]
    num_cols_1 = metadata_sb[0]["num_cols"]
    
    start_col_1 = metadata_sb[0]["start_col"]
    end_col_1 = metadata_sb[0]["end_col"]
    
    wb = load_workbook(invoice_path)
    ws = wb[sb_sheet_name]

    # Set column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 23
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 40

    ws.row_dimensions[2].height = 20
    ws.row_dimensions[4].height = 15
    ws.row_dimensions[5].height = 15
    ws.row_dimensions[6].height = 15
    ws.row_dimensions[7].height = 15
    ws.row_dimensions[9].height = 28

    # === Insert GE Logo ===
    img = Image('ge_new_logo.png')
    img.height = int(0.56 * 96)
    img.width = int(2.32 * 96)
    ws.add_image(img, 'B1')

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "8DC858"
    ws.sheet_view.zoomScale = 80
 
    # === Font Styles ===
    title_bold_black_font = Font(name='Arial', bold=True, color="000000", size=14)
    blue_fill = PatternFill(start_color="050543", end_color="050543", fill_type="solid")
    white_bold_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    norma_black_font = Font(name="Arial", bold=False, color="000000", size=11)

    # === Alignment Styles ===
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # === Border Styles ===
    no_border = Border()
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    black_colored_border = Side(style="thin", color="000000")
    white_colored_border = Side(style="thin", color="FFFFFF")
    thick_black_colored_border = Side(style="thick", color="000000")

    right_white_rest_black_border = Border(left=black_colored_border, top=black_colored_border, right=white_colored_border, bottom=black_colored_border)
    left_right_white_rest_black_border = Border(left=white_colored_border, top=black_colored_border, right=white_colored_border, bottom=black_colored_border)
    left_white_rest_black_border = Border(left=white_colored_border, top=black_colored_border, right=black_colored_border, bottom=black_colored_border)
    top_bottom_left_white_rest_black_border = Border(left=white_colored_border, top=white_colored_border, right=black_colored_border, bottom=white_colored_border)
    left_bottom_white_rest_black_border = Border(left=white_colored_border, top=black_colored_border, right=black_colored_border, bottom=white_colored_border)

    # # # Table-1 Values
    for row in ws.iter_rows(min_row=start_row_1+1, max_row=end_row_1, min_col = 2, max_col=num_cols_1+1):
        for cell in row:
            cell.font = norma_black_font
            cell.alignment = center
            cell.border = border
            cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    for row in range(start_row_1+1, end_row_1+1):
        ws.row_dimensions[row].height = 16

    # formulas
    apply_arithmetic_formula(
        ws=ws, start_row=start_row_1, start_col=start_col_1, end_col=end_col_1,
        formula_str="=(labor)+(material)+(repair)",
        input_col_headers=['labor','material', 'repair'],
        mode='horizontal', output_col_name='sb_total'
    )

    col_index_of_total_price = 4
    col_letter = get_column_letter(col_index_of_total_price)
    output_cell = f"{col_letter}{end_row_1+1}"

    apply_arithmetic_formula(
        ws=ws,
        start_row=start_row_1,
        start_col=4,
        end_col=4,
        formula_str="sum(labor)",
        input_col_headers=['labor'],
        mode='vertical',
        output_cell=output_cell
    )

    apply_arithmetic_formula(
        ws=ws,
        start_row=start_row_1,
        start_col=5,
        end_col=5,
        formula_str="sum(material)",
        input_col_headers=['material'],
        mode='vertical',
        output_cell=output_cell
    )

    apply_arithmetic_formula(
        ws=ws,
        start_row=start_row_1,
        start_col=6,
        end_col=6,
        formula_str="sum(repair)",
        input_col_headers=['repair'],
        mode='vertical',
        output_cell=output_cell
    )

    apply_arithmetic_formula(
        ws=ws,
        start_row=start_row_1,
        start_col=7,
        end_col=7,
        formula_str="sum(sb total)",
        input_col_headers=['sb_total'],
        mode='vertical',
        output_cell=output_cell
    )

    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="SB Category")     
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Labor")     
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Material")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="Repair")
    apply_number_format_by_column_name(ws, header_row=start_row_1, start_row=start_row_1+1, end_row=end_row_1, target_col_name="SB Total")
    
    sub_total_1_cell = insert_subtotal(ws, end_row_1+1, 4, start_row_1+1, end_row_1,
                                       font=white_bold_font, fill=blue_fill, border=left_white_rest_black_border,
                                       number_format='_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)')

    sub_total_2_cell = insert_subtotal(ws, end_row_1+1, 5, start_row_1+1, end_row_1,
                                       font=white_bold_font, fill=blue_fill, border=left_white_rest_black_border,
                                       number_format='_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)')

    sub_total_3_cell = insert_subtotal(ws, end_row_1+1, 6, start_row_1+1, end_row_1,
                                       font=white_bold_font, fill=blue_fill, border=left_white_rest_black_border,
                                       number_format='_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)')

    sub_total_4_cell = insert_subtotal(ws, end_row_1+1, 7, start_row_1+1, end_row_1,
                                       font=white_bold_font, fill=blue_fill, border=left_white_rest_black_border,
                                       number_format='_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)')


    ws["E2"].value = "SB O&A Summary"
    ws["E2"].font = title_bold_black_font
    ws["E2"].alignment = center

    ws["F3"].value = "SB Total Labor"
    ws["F3"].fill = blue_fill
    ws["F3"].font = white_bold_font
    ws["F3"].alignment = left
    ws["F3"].border = right_white_rest_black_border

    ws["G3"].value = f"={sub_total_1_cell.coordinate}"
    ws["G3"].fill = blue_fill
    ws["G3"].font = white_bold_font
    ws["G3"].alignment = left
    ws["G3"].border = left_bottom_white_rest_black_border
    ws["G3"].number_format='_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    ws["F4"].value = "SB Total Material"
    ws["F4"].fill = blue_fill
    ws["F4"].font = white_bold_font
    ws["F4"].alignment = left
    ws["F4"].border = right_white_rest_black_border

    ws["G4"].value = f"={sub_total_2_cell.coordinate}"
    ws["G4"].fill = blue_fill
    ws["G4"].font = white_bold_font
    ws["G4"].alignment = left
    ws["G4"].border = left_bottom_white_rest_black_border
    ws["G4"].number_format ='_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    ws["F5"].value = "SB Total Repair"
    ws["F5"].fill = blue_fill
    ws["F5"].font = white_bold_font
    ws["F5"].alignment = left
    ws["F5"].border = right_white_rest_black_border

    ws["G5"].value = f"={sub_total_3_cell.coordinate}"
    ws["G5"].fill = blue_fill
    ws["G5"].font = white_bold_font
    ws["G5"].alignment = left
    ws["G5"].border = left_bottom_white_rest_black_border
    ws["G5"].number_format ='_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    ws["F6"].value = "SB Grand Total"
    ws["F6"].fill = blue_fill
    ws["F6"].font = white_bold_font
    ws["F6"].alignment = left
    ws["F6"].border = right_white_rest_black_border

    ws["G6"].value = "=SUM(G3:G5)"
    ws["G6"].fill = blue_fill
    ws["G6"].font = white_bold_font
    ws["G6"].alignment = left
    ws["G6"].border = left_bottom_white_rest_black_border
    ws["G6"].number_format='_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    # Set header values
    ws["B9"].value = "SB Number"
    ws["B9"].fill = blue_fill
    ws["B9"].font = white_bold_font
    ws["B9"].alignment = center
    ws["B9"].border = right_white_rest_black_border

    ws["C9"].value = "SB Category"
    ws["C9"].fill = blue_fill
    ws["C9"].font = white_bold_font
    ws["C9"].alignment = center
    ws["C9"].border = left_right_white_rest_black_border

    ws["D9"].value = "Labor"
    ws["D9"].fill = blue_fill
    ws["D9"].font = white_bold_font
    ws["D9"].alignment = center
    ws["D9"].border = left_right_white_rest_black_border

    ws["E9"].value = "Material"
    ws["E9"].fill = blue_fill
    ws["E9"].font = white_bold_font
    ws["E9"].alignment = center
    ws["E9"].border = left_right_white_rest_black_border

    ws["F9"].value = "Repair"
    ws["F9"].fill = blue_fill
    ws["F9"].font = white_bold_font
    ws["F9"].alignment = center
    ws["F9"].border = left_right_white_rest_black_border

    ws["G9"].value = "SB Total"
    ws["G9"].fill = blue_fill
    ws["G9"].font = white_bold_font
    ws["G9"].alignment = center
    ws["G9"].border = left_right_white_rest_black_border

    ws["H9"].value = "Comments"
    ws["H9"].fill = blue_fill
    ws["H9"].font = white_bold_font
    ws["H9"].alignment = center
    ws["H9"].border = left_right_white_rest_black_border
    
    # ===== Write DataFrame Data =====
    total_row = 9 + num_rows_1 + 1
    ws.cell(row=total_row, column=3).fill = blue_fill
    ws.cell(row=total_row, column=3, value="Subtotal").font = white_bold_font
    ws.cell(row=total_row, column=3).alignment = left
    ws.cell(row=total_row, column=3).border = left_right_white_rest_black_border
    ws.cell(row=total_row, column=3).number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    wb.save(invoice_path)
    print(f"[INFO] ✅ SB sheet filled and saved to {invoice_path}")


def to_number(x):
    """Robust: handles None, '', '-', '(1,234)', '1,234', '50%', arrays/Series."""
    if isinstance(x, (list, tuple, np.ndarray, pd.Series, pd.Index)):
        # reduce containers to a scalar (take first non-null)
        if isinstance(x, (np.ndarray, pd.Index)): x = x.tolist()
        if isinstance(x, pd.Series): x = x.tolist()
        for v in x:
            y = to_number(v)
            if not np.isnan(y):
                return y
        return 0.0
 
    if x is None: return 0.0
    if isinstance(x, (int, float, np.floating)): return float(x)
 
    s = str(x).strip()
    if s in {"", "-", "–", "—"}: return 0.0          # accounting dash
    if s.endswith("%"):
        try: return float(s[:-1].replace(",", ""))/100.0
        except: return 0.0
    # (1,234.56) -> -1234.56
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    # remove currency symbols, spaces, commas
    s = re.sub(r"[^\d\.\-]", "", s)
    try:    return float(s) if s not in {"", "-", ".", "-."} else 0.0
    except: return 0.0

def write_cover_sheet_cost_items(actual_draft_invoice_path, mat_sum_val, rep_sum_val, ven_sum_val, qec_mat_sum, qec_rep_sum, scrap_mat_sum, scrap_rep_sum, o_and_above_total_sum_val, other_total_sum, grand_total_sum, adder_creep_data):

    updated_wb = load_workbook(actual_draft_invoice_path, data_only=False)
    updated_cover_ws = updated_wb["Summary"]

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    updated_cover_ws.sheet_properties.tabColor = "FF0000"

    updated_cover_ws["H20"] = mat_sum_val #"=Material!O2"
    updated_cover_ws["H20"].alignment = left
    updated_cover_ws["H20"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    updated_cover_ws["H21"] = rep_sum_val #"='GE Repair'!L2" 
    updated_cover_ws["H21"].alignment = left
    updated_cover_ws["H21"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    updated_cover_ws["H22"].value = ven_sum_val #"='Sub Contracted Repair'!K2"   
    updated_cover_ws["H22"].alignment = left
    updated_cover_ws["H22"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    updated_cover_ws["H23"] = qec_rep_sum #"='QEC-LRU-ACC'!J2"  ##f"='QEC-LRU-ACC'!{sub_total_1_cell.coordinate}" 
    updated_cover_ws["H23"].alignment = left
    updated_cover_ws["H23"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    updated_cover_ws["H24"] = qec_mat_sum  #"='QEC-LRU-ACC'!J1" ##f"='QEC-LRU-ACC'!{sub_total_2_cell.coordinate}" 
    updated_cover_ws["H24"].alignment = left
    updated_cover_ws["H24"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    updated_cover_ws["H25"] = "TBD" 
    updated_cover_ws["H25"].alignment = center
    updated_cover_ws["H25"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    updated_cover_ws["H26"] = scrap_rep_sum 
    updated_cover_ws["H26"].alignment = center
    updated_cover_ws["H26"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    # updated_cover_ws["H27"] = o_and_above_total_sum_val #"=SUM(H19:H25)" 
    updated_cover_ws["H27"] = scrap_mat_sum #"=SUM(H19:H25)" 
    updated_cover_ws["H27"].alignment = center
    updated_cover_ws["H27"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    

    updated_cover_ws["H32"] = "TBD" 
    updated_cover_ws["H32"].alignment = center
    updated_cover_ws["H32"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    updated_cover_ws["H33"] = "TBD" 
    updated_cover_ws["H33"].alignment = center
    updated_cover_ws["H33"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    
    updated_cover_ws["H34"] = "TBD" 
    updated_cover_ws["H34"].alignment = center
    updated_cover_ws["H34"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    # Start point for the static rows after the creep items
    current_row = 35 + len(adder_creep_data)

    # H (current_row) -> Transportation Credit
    updated_cover_ws[f"H{current_row}"] = "TBD" 
    updated_cover_ws[f"H{current_row}"].alignment = center
    updated_cover_ws[f"H{current_row}"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    current_row += 1

    # H (current_row) -> Progress Invoice
    updated_cover_ws[f"H{current_row}"] = "TBD" 
    updated_cover_ws[f"H{current_row}"].alignment = center
    updated_cover_ws[f"H{current_row}"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    current_row += 1

    # H (current_row) -> Initial Invoice
    updated_cover_ws[f"H{current_row}"] = "TBD" 
    updated_cover_ws[f"H{current_row}"].alignment = center
    updated_cover_ws[f"H{current_row}"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    current_row += 1

    # H (current_row) -> Other Total
    updated_cover_ws[f"H{current_row}"] = other_total_sum 
    updated_cover_ws[f"H{current_row}"].alignment = left
    updated_cover_ws[f"H{current_row}"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    current_row += 2 # Skip one row (equivalent to the spacer row 38)

    # H (current_row) -> Grand Total
    updated_cover_ws[f"H{current_row}"] = grand_total_sum  
    updated_cover_ws[f"H{current_row}"].alignment = left
    updated_cover_ws[f"H{current_row}"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'


    updated_wb.save(actual_draft_invoice_path)
    return actual_draft_invoice_path

def write_formulas_back(actual_draft_invoice_path):

    updated_wb = load_workbook(actual_draft_invoice_path, data_only=False)
    updated_cover_ws = updated_wb["Summary"]

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    updated_cover_ws.sheet_properties.tabColor = "FF0000"

    updated_cover_ws["H20"] = "=Material!O2"
    updated_cover_ws["H20"].alignment = left
    updated_cover_ws["H20"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    updated_cover_ws["H21"] = "='GE Repair'!L2" 
    updated_cover_ws["H21"].alignment = left
    updated_cover_ws["H21"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    updated_cover_ws["H22"].value = "='Sub Contracted Repair'!K2"   
    updated_cover_ws["H22"].alignment = left
    updated_cover_ws["H22"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    updated_cover_ws["H23"] = "='QEC-LRU-ACC'!J2"  ##f"='QEC-LRU-ACC'!{sub_total_1_cell.coordinate}" 
    updated_cover_ws["H23"].alignment = left
    updated_cover_ws["H23"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    updated_cover_ws["H24"] = "='QEC-LRU-ACC'!J1" ##f"='QEC-LRU-ACC'!{sub_total_2_cell.coordinate}" 
    updated_cover_ws["H24"].alignment = left
    updated_cover_ws["H24"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    updated_cover_ws["H26"] = "='Scrap Table'!J2" 
    updated_cover_ws["H26"].alignment = left
    updated_cover_ws["H26"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    updated_cover_ws["H27"] = "='Scrap Table'!J1" 
    updated_cover_ws["H27"].alignment = left
    updated_cover_ws["H27"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    updated_cover_ws["H28"] = "=SUM(H19:H27)" 
    updated_cover_ws["H28"].alignment = center
    updated_cover_ws["H28"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    
    updated_cover_ws["H38"] = "=SUM(H32:H37)" 
    updated_cover_ws["H38"].alignment = left
    updated_cover_ws["H38"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    updated_cover_ws["H40"] = "=H16+H28+H38" 
    updated_cover_ws["H40"].alignment = left
    updated_cover_ws["H40"].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    updated_wb.save(actual_draft_invoice_path)
    return actual_draft_invoice_path


def extract_inscope_outscope_cost(path: str):
    updated_wb = load_workbook(path, data_only=True)
    updated_cover_ws = updated_wb["Summary"]
 
    # Extract raw values from Excel
    cells = {
        "MFP": updated_cover_ws["H16"].value,
        "material": updated_cover_ws["H20"].value,
        "ge_repair": updated_cover_ws["H21"].value,
        "sub_contracted_repair": updated_cover_ws["H22"].value,
        "qec_lru_repair": updated_cover_ws["H23"].value,
        "qec_lru_material": updated_cover_ws["H24"].value  
    }
 
    # Calculate total ignoring any non-numeric values
    total_sum = sum(
        value for value in cells.values()
        if isinstance(value, (int, float))
    )
 
    # Save if needed (optional, since we're not modifying anything)
    updated_wb.save(path)
 
    # Print individual values
    for name, value in cells.items():
        print(f"{name}: {value}")
 
    print("total_sum:", total_sum)
    return (*cells.values(), total_sum)


def fill_cover_sheet(cover_page, result, labor_amount, new_parts_non_llp, new_parts_llp, cfe):

    df = get_dataframe_by_file_and_sheet(
        all_dataframes=result["dataframes"],
        file_key="timeliness",
        sheet_name_substring="timeliness"
    )

    # Normalize columns
    df.columns = [col.strip().lower() for col in df.columns]

    # Try to identify engine family & esn column
    esn_col = next((col for col in df.columns if "esn" in col), None)

    esn_list = [int(result['esn'])]
    filt_df = df[df[esn_col].isin(esn_list)].copy()

    cust_name_col = next((col for col in filt_df.columns if "customer name" in col), None)
    inv_no_col = next((col for col in filt_df.columns if "invoice #" in col), None)
    entitle_date_col = next((col for col in filt_df.columns if "entitlement date" in col), None)
    ship_date_col = next((col for col in filt_df.columns if "ship date" in col), None)
    eng_model_col = next((col for col in filt_df.columns if "engine model" in col), None)
    contract_no_col = next((col for col in filt_df.columns if "contract number" in col), None)
    customer_order_num = next((col for col in filt_df.columns if "workorder" in col), None)
    shop_visit_induction_date = next((col for col in filt_df.columns if "induction date" in col), None)
    

    # Not in Timeliness
    work_performed_at = f"GE Celma Ltda\n356 Rua Alice Herve\nPetropolis - RJ\n25669-900BR\nTel No. - 55 24 2233-4000"
    foreign_aff_of = f"GE Engine Services, LLC\n1 Neumann Way\nCincinnati , OH 45215"
    send_pay_show = 'TBD'
    ship_add = f"AVIANCA, INC.\nHANGAR 1 - AEROPUERTO EL DORADO SN\nBOGOTÁ\n11001, BOGOTA\nColombia"
    bill_add = "TBD"
    
    cust_name = filt_df[cust_name_col].iloc[0]
    eng_ser_no = str(esn_list[0])
    inv_no = filt_df[inv_no_col].iloc[0]
    inv_date = (filt_df[entitle_date_col]
                .pipe(pd.to_datetime, errors="coerce")
                .dt.strftime("%d-%b-%Y")
                .fillna("NA")
                .iloc[0])
    inv_date_str = inv_date #inv_date.strftime("%d-%b-%Y")
    customer_ordnum = filt_df[customer_order_num].iloc[0]
    induction_date = (filt_df[shop_visit_induction_date]
                .pipe(pd.to_datetime, errors="coerce")
                .dt.strftime("%d-%b-%Y")
                .fillna("NA")
                .iloc[0])
    
    filt_df[ship_date_col] = pd.to_datetime(filt_df[ship_date_col])

    s = pd.to_datetime(filt_df[ship_date_col], errors = "coerce")
    ship_ts = s.iloc[0]

    ship_date = (filt_df[ship_date_col]
                .pipe(pd.to_datetime, errors="coerce")
                .dt.strftime("%d-%b-%Y")
                .fillna("NA")
                .iloc[0]) #filt_df[ship_date_col].iloc[0]
    ship_date_str = "NA" if pd.isna(ship_ts) else ship_ts.strftime("%Y-%m-%d")

    pay_terms = "30 days from Invoice Date"
    # pay_due_on = ship_date + timedelta(days=30)
    pay_due_on_str = "NA" if pd.isna(ship_ts) else (ship_ts + pd.Timedelta(days=30)).strftime("%Y-%m-%d") #pay_due_on.strftime("%d-%b-%Y")

    eng_model = filt_df[eng_model_col].iloc[0]
    contract_no = filt_df[contract_no_col].iloc[0]

    cover_page = "cover_invoice_sheet.xlsx"
    wb = load_workbook(cover_page)
    cover_sheet_name = "Summary"
    ws = wb[cover_sheet_name]
    # ws.sheet_properties.tabColor = "FF0000"

    # norma_black_font = Font(bold=False, color="000000", size=9)
    # norma_white_font = Font(bold=False, color="FFFFFF", size=9)
    # center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    ws["B4"].value = cust_name
    ws["D4"].value = contract_no
    ws["E4"].value = eng_ser_no
    
    ws["B6"].value = customer_ordnum
    ws["E6"].value = inv_no
    ws["F6"].value = inv_date_str

    ws["B8"].value = work_performed_at
    ws["E8"].value = foreign_aff_of
    ws["F8"].value = send_pay_show

    ws["B10"].value = ship_add
    ws["E10"].value = bill_add

    ws["B12"].value = induction_date
    ws["D12"].value = ship_date_str
    ws["E12"].value = pay_terms
    ws["G12"].value = pay_due_on_str

    wb.save("cover_invoice_attr.xlsx")

def safe_sheet_name(name, max_len=31):
    """Ensure name is valid for Excel (max 31 chars, no illegal chars)."""
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
    for ch in invalid_chars:
        name = name.replace(ch, '')
    return name.strip()[:max_len]
 
def create_sheets_from_range(input_path, output_path):
    try:
        wb = load_workbook(input_path)
        ws = wb.active
 
        sheet_names = []
        # sheet_names.append("MFP")
        for row in range(20, 27):
            cell_value = ws[f'B{row}'].value
            if cell_value and isinstance(cell_value, str):
                clean_name = safe_sheet_name(cell_value)
                if clean_name and clean_name not in sheet_names:
                    sheet_names.append(clean_name)
 
        for name in sheet_names:
            if name not in wb.sheetnames:
                try:
                    wb.create_sheet(title=name)
                    print(f"[INFO] ➕ Created sheet: {name}")
                except IllegalCharacterError:
                    print(f"[ERROR] ❌ Illegal characters in sheet name '{name}'. Skipping.")
            else:
                print(f"[WARNING] ⚠️ Sheet '{name}' already exists. Skipping.")
 
        # Safely set tab color using Color object with valid ARGB
        ws.sheet_properties.tabColor = Color(rgb="FF00B050")
 
        wb.save(output_path)
        print(f"[INFO] ✅ Updated workbook saved at: {output_path}")
    except Exception as e:
        print(f"[ERROR] ❌ Failed to process workbook: {e}")


def append_to_excel(print_statement, sheet_name="Data", column="A", header="Messages"):
    file_path = "InvCreationStepsLoader.xlsx"  # Hardcoded file path
    try:
        # Check if file exists
        if os.path.exists(file_path):
            # Load existing workbook
            workbook = load_workbook(file_path)
        else:
            # Create new workbook
            workbook = Workbook()
       
        # Get or create the specified sheet
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
       
        # If it's a new workbook, remove the default sheet if it's empty
        if "Sheet" in workbook.sheetnames and sheet_name != "Sheet":
            default_sheet = workbook["Sheet"]
            if default_sheet.max_row == 1 and default_sheet.max_column == 1:
                workbook.remove(default_sheet)
       
        # Check if the header exists in the specified column, if not, add it
        if sheet[f"{column}1"].value != header:
            sheet[f"{column}1"] = header
       
        # Find the next available row in the specified column
        current_row = 2  # Start from row 2 (after header)
        while sheet[f"{column}{current_row}"].value:
            current_row += 1
        target_cell = f"{column}{current_row}"
       
        # Append the print statement to the target cell
        sheet[target_cell] = print_statement
       
        # Save the workbook
        workbook.save(file_path)
        print(f"Successfully appended '{print_statement}' to {file_path} in sheet '{sheet_name}' at cell {target_cell} under '{header}' column")
        return True
   
    except Exception as e:
        print(f"Error: {str(e)}")
        return False

def generate_invoice_summary(mat_df, repair_df, vendor_df, output_path):
    def summarize_pipeline(df: pd.DataFrame, label: str) -> pd.DataFrame:
        
        if df is None or df.empty:
            return pd.DataFrame(columns=["Attribute", "Value"])
        df = df.copy()

        if "total_price" not in df.columns:
            df["total_price"] = 0
        else:
            df["total_price"] = pd.to_numeric(df["total_price"], errors = 'coerce').fillna(0)

        if label == "Material":
            summary = df.groupby("cost category")["total_price"].sum().reset_index()
            summary.columns = ["Attribute", "Value"]
            summary["Attribute"] = summary["Attribute"]
        else:
            summary = df.groupby("remarks")["total_price"].sum().reset_index()
            summary.columns = ["Attribute", "Value"]
            summary["Attribute"] = summary["Attribute"]
        return summary
 
    # Generate summaries
    mat_summary = summarize_pipeline(mat_df, "Material")
    repair_summary = summarize_pipeline(repair_df, "Repair")
    vendor_summary = summarize_pipeline(vendor_df, "Vendor")
 
    # Combine all
    dfs = [mat_summary] #, repair_summary, vendor_summary]
    non_empty_dfs = [df for df in dfs if not df.empty and not all(df.isna().all())]
    combined_summary = pd.concat(non_empty_dfs, ignore_index=True)
    
    # Calculate grand total
    grand_total = combined_summary["Value"].sum()
    combined_summary = pd.concat([combined_summary, pd.DataFrame([{"Attribute" : "Grand Total", "Value": grand_total}])],
                        ignore_index=True)
 
    # Save to Excel
    combined_summary.to_excel(output_path, index=False)
    # print(f"[INFO] ✅ Invoice summary saved at: {output_path}")


def write_single_df_to_excel(excel_path, target_sheet_name, df, start_row=1, gap_rows=1):
    # Load workbook and get sheet names
    wb = load_workbook(excel_path)
    file_sheets = wb.sheetnames

    # Normalize names (assuming you have a normalize_string function)
    normalized_file_sheets = [normalize_string(name) for name in file_sheets]
    normalized_target = normalize_string(target_sheet_name)

    # Find matching sheet
    matched_sheet = None
    for orig, norm in zip(file_sheets, normalized_file_sheets):
        if normalized_target in norm:
            matched_sheet = orig
            break

    if not matched_sheet:
        print(f"[❌] No matching sheet found for '{target_sheet_name}'.")
        return None

    # Add blank first column
    df_with_blank = pd.concat([pd.Series([""] * len(df), name=""), df], axis=1)
    #start_row = ws.max_row + 2 
    # Write DataFrame to Excel
    with pd.ExcelWriter(excel_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
        df_with_blank.to_excel(
            writer,
            sheet_name=matched_sheet,
            startrow=start_row,
            index=False,
            header=True
        )
    #start_row = ws.max_row + 2  # Leave space after merged header rows
    # Summary tracking logic
    df_rows = df.shape[0]
    df_cols = df.shape[1] + 1  # +1 for blank column
    col_start = 1  # Column A
    col_end = col_start + df_cols - 1
    row_end = start_row + df_rows + 1  # +1 for header row

    summary_info = {
        'row_start': start_row,
        'row_end': row_end,
        'num_rows': df_rows,
        'num_cols': df_cols,
        'col_start': col_start,
        'col_end': col_end
    }

    print(f"[✅] DataFrame written to sheet: '{matched_sheet}' in '{excel_path}'")
    print("[ℹ️] Summary Info:", summary_info)

    return summary_info